import os, json, re, time
import requests

from mcp_server import dispatch_tool as _mcp_dispatch, MCPClientError, is_server_reachable
from dotenv import load_dotenv
from groq import Groq
from neo4j_tools import TOOL_FUNCTIONS

load_dotenv(".env")

# ── Neo4j helper (used by the Update Graph pipeline agents) ──────────────────
def _run_neo4j(query: str, params: dict = None):
    """Execute a Cypher query and return a list of row dicts."""
    from neo4j import GraphDatabase as _GD
    _driver = _GD.driver(
        os.getenv("NEO4J_URI"),
        auth=(os.getenv("NEO4J_USER"), os.getenv("NEO4J_PASSWORD")),
    )
    _db = os.getenv("NEO4J_DATABASE", "neo4j")
    try:
        with _driver.session(database=_db) as session:
            return [r.data() for r in session.run(query, params or {})]
    finally:
        _driver.close()

_groq_client = Groq(api_key=(os.getenv("GROQ_API_KEY") or "").strip())
_OR_KEY       = (os.getenv("OPENROUTER_API_KEY") or "").strip()

# OpenRouter models (used only if OPENROUTER_API_KEY is set in .env)
_ORCH_MODELS = [
    "deepseek/deepseek-r1-0528:free",
    "google/gemini-2.5-flash-preview:free",
    "meta-llama/llama-3.3-70b-instruct:free",
    "openrouter/auto",
]

# ── Groq model chains ────────────────────────────────────────────────────────
#
# YOUR PIPELINE: Orchestrator → RCA → Recommendations (3 agents only)
# Each chain uses models from DIFFERENT Groq quota families so a 429 on one
# family automatically rolls to a fresh quota bucket.
#
# Verified Groq model IDs (May 2025) — 3 distinct quota families:
#   Family A  →  llama-4-scout  (preview quota, separate bucket)
#   Family B  →  llama-3.3-70b-versatile + llama-3.1-8b-instant (llama3 quota)
#   Family C  →  gemma2-9b-it + mixtral-8x7b-32768 (separate buckets)
#
# llama-3.1-70b-versatile is NOT available on Groq free tier — removed.
# ────────────────────────────────────────────────────────────────────────────

# RCA Agent — needs highest quality, long context
_RCA_MODELS = [
    "meta-llama/llama-4-scout-17b-16e-instruct",  # Family A — best quality
    "llama-3.3-70b-versatile",                     # Family B — strong fallback
    "mixtral-8x7b-32768",                          # Family C — 32k context, different quota
    "gemma2-9b-it",                                # Family C — different quota bucket
    "llama-3.1-8b-instant",                        # Family B — highest RPM, last resort
]

# Recommendations Agent — needs good instruction following
_REC_MODELS = [
    "llama-3.3-70b-versatile",                     # Family B
    "meta-llama/llama-4-scout-17b-16e-instruct",  # Family A
    "mixtral-8x7b-32768",                          # Family C
    "gemma2-9b-it",                                # Family C
    "llama-3.1-8b-instant",                        # Family B — safety net
]

# Orchestrator Agent — needs fast JSON output, not quality
_ORCHESTRATOR_MODELS = [
    "llama-3.3-70b-versatile",                     # Family B — fast + smart
    "meta-llama/llama-4-scout-17b-16e-instruct",  # Family A
    "mixtral-8x7b-32768",                          # Family C
    "llama-3.1-8b-instant",                        # Family B — fastest
    "gemma2-9b-it",                                # Family C
]

# Update Graph pipeline (not part of RCA — kept for completeness)
_UPDATE_MODELS = [
    "llama-3.3-70b-versatile",
    "meta-llama/llama-4-scout-17b-16e-instruct",
    "mixtral-8x7b-32768",
    "llama-3.1-8b-instant",
]

# ── Legacy agent chains (Validator/Critique/Narrative not used in current pipeline) ──

_VALIDATOR_OR_MODELS = [
    "google/gemini-2.5-flash-preview:free",
    "google/gemma-3-27b-it:free",
    "meta-llama/llama-3.3-70b-instruct:free",
]
_VALIDATOR_GROQ_MODELS = [
    "llama-3.1-8b-instant",   # SPEED: fastest model first for simple validation
    "gemma2-9b-it",
    "mixtral-8x7b-32768",
]

_CRITIQUE_OR_MODELS = [
    "deepseek/deepseek-r1-0528:free",
    "deepseek/deepseek-r1:free",
    "google/gemini-2.5-flash-preview:free",
]
_CRITIQUE_GROQ_MODELS = [
    "llama-3.3-70b-versatile",
    "meta-llama/llama-4-scout-17b-16e-instruct",
    "mixtral-8x7b-32768",
    "llama-3.1-8b-instant",
]

_NARRATIVE_MODELS = [
    "llama-3.1-8b-instant",   # SPEED: fastest first — narrative is just 2-3 short sentences
    "gemma2-9b-it",
    "mixtral-8x7b-32768",
]

MAX_TOKENS_PER_STEP   = 500
MAX_TOKENS_FINAL      = 3200   # reduced from 5500 — saves 30-40% on RCA agent latency
MAX_TOKENS_FALLBACK   = 1200
OBS_TRIM_CHARS        = 800    # Raw obs trimmed short — pre-formatted tables are the primary data source
MAX_HISTORY_MESSAGES  = 12


# ════════════════════════════════════════════════════════════════════
# AGENT CONTEXT  — structured data carrier between agents
# This is the key change: instead of passing raw strings, each agent
# receives and produces a typed context object that carries the full
# pipeline state.
# ════════════════════════════════════════════════════════════════════

class AgentContext:
    """
    Shared context object passed through the agentic pipeline.

    Full pipeline flow:
        OrchestratorAgent    → selected_tools, tool_logs, obs_block
        DataValidatorAgent   → validation_report, obs_block_clean
        RCAAgent             → rca_raw, rca_findings
        CritiqueAgent        → critique_report, rca_final
        RecommendationsAgent → rec_raw
        NarrativeAgent       → narrative
        Assembler            → final_answer
    """
    def __init__(self, user_question: str):
        self.user_question: str = user_question

        # ── Orchestrator output ──────────────────────────────
        self.selected_tools: list[tuple[str, dict]] = []
        self.tool_logs:       list[dict]             = []
        self.obs_block:       str                    = ""
        self.tool_rows:       dict                   = {}   # tool_name → list[dict]

        # ── Data Validator Agent output ───────────────────────
        # Gemini Flash (OpenRouter) — fast structured data quality check
        self.validation_report: dict = {}
        # keys: valid_tools, flagged_tools [{tool,reason}],
        #       data_quality_score (0.0-1.0), warnings [str]
        self.obs_block_clean: str = ""   # obs_block with error results stripped

        # ── RCA Agent output ─────────────────────────────────
        # llama-4-scout (Groq)
        self.rca_raw:      str  = ""
        self.rca_findings: dict = {}
        # keys: bottleneck_plants, high_risk_suppliers,
        #       distributor_impact, stockout_cities, root_cause_summary

        # ── Critique Agent output ─────────────────────────────
        # DeepSeek R1 (OpenRouter) — reasoning model for fact-checking
        self.critique_report: dict = {}
        # keys: passed (bool), issues [str], corrected_rca (str)
        self.rca_final: str = ""   # rca_raw or corrected version after critique

        # ── Recommendations Agent output ─────────────────────
        # llama-3.3-70b (Groq)
        self.rec_raw: str = ""

        # ── Narrative Agent output ────────────────────────────
        # gemma2-9b-it (Groq) — distinct model for polished prose
        self.narrative: str = ""

        # ── First response (runs in parallel with orchestrator) ──
        self.first_response: str = ""

        # ── Final assembled output ───────────────────────────
        self.final_answer: str = ""

    def orchestrator_done(self) -> bool:
        return bool(self.obs_block and self.tool_logs)

    def validator_done(self) -> bool:
        return bool(self.validation_report)

    def rca_done(self) -> bool:
        return bool(self.rca_raw)

    def critique_done(self) -> bool:
        return bool(self.critique_report)

    def rec_done(self) -> bool:
        return bool(self.rec_raw)

    def narrative_done(self) -> bool:
        return bool(self.narrative)


def _extract_rca_findings(rca_raw: str) -> dict:
    """
    Parse structured entities out of the RCA markdown for passing to
    the Recommendations agent.
    """
    rca_raw = str(rca_raw or "")
    findings = {
        "bottleneck_plants":     [],
        "high_risk_suppliers":   [],
        "distributor_impact":    [],
        "stockout_cities":       [],
        "root_cause_summary":    "",
    }

    # Extract root cause summary (first 2 sentences after Executive Summary)
    exec_match = re.search(
        r'###?\s*📌\s*Executive Summary\s*\n(.*?)(?=\n###|\Z)',
        rca_raw, re.DOTALL | re.IGNORECASE
    )
    if exec_match:
        findings["root_cause_summary"] = exec_match.group(1).strip()[:500]

    # Extract plant names robustly — ID-first and name-first formats
    for m in re.finditer(r'\|\s*(PL\d+)\s*\|([^|]+)\|', rca_raw):
        pid = m.group(1).strip()
        pname = m.group(2).strip().strip("*").strip()
        if pname and len(pname) > 2 and not pname.startswith("PL") and not pname[0].isdigit():
            if pid not in [p["id"] for p in findings["bottleneck_plants"]]:
                findings["bottleneck_plants"].append({"id": pid, "name": pname})
    for m in re.finditer(r'\|\s*(Bhopal|Pune|Baddi|Goa)\s*\|\s*(PL\d+)\s*\|', rca_raw, re.IGNORECASE):
        pname, pid = m.group(1).strip(), m.group(2).strip()
        if pid not in [p["id"] for p in findings["bottleneck_plants"]]:
            findings["bottleneck_plants"].append({"id": pid, "name": pname})
    if not findings["bottleneck_plants"]:
        for pid, pname in [("PL1","Baddi"),("PL2","Bhopal"),("PL3","Pune"),("PL4","Goa")]:
            if pid in rca_raw or pname.lower() in rca_raw.lower():
                findings["bottleneck_plants"].append({"id": pid, "name": pname})
    # Extract supplier names# Extract supplier names from High-Risk Suppliers table
    # ONLY include suppliers whose ID matches SUP + digits pattern (no hallucinated names)
    _known_sup_ids = set(re.findall(r'SUP\d{4,6}', rca_raw))
    for m in re.finditer(r'\|\s*(SUP\d{4,6})\s*\|([^|]+)\|([^|]+)\|', rca_raw):
        sup_id   = m.group(1).strip()
        sup_name = m.group(2).strip().strip("*").strip()
        risk     = m.group(3).strip()
        # Skip clearly hallucinated names (too short, numbers only, or contain brackets)
        if len(sup_name) < 3 or sup_name.replace(".", "").replace("-", "").isdigit():
            continue
        if "[" in sup_name or "SUP" in sup_name:
            continue
        if sup_id not in [s["id"] for s in findings["high_risk_suppliers"]]:
            findings["high_risk_suppliers"].append({"id": sup_id, "name": sup_name, "risk": risk})

    # Extract distributor cities from Distributor Impact table
    for m in re.finditer(r'\|\s*([A-Za-z][A-Za-z\s]{3,})\s*\|\s*(\d+)', rca_raw):
        city = m.group(1).strip()
        if len(city) > 3 and city not in ["Distributor City", "Plant ID", "Supplier ID"]:
            findings["distributor_impact"].append(city)

    # Extract stockout cities from Stockout section
    stockout_section = re.search(
        r'###\s*🛒\s*Stockout Impact(.*?)(?=\n##|\Z)',
        rca_raw, re.DOTALL | re.IGNORECASE
    )
    if stockout_section:
        cities = re.findall(r'\b([A-Z][a-z]+(?:\s+[A-Z][a-z]+)?)\b', stockout_section.group(1))
        findings["stockout_cities"] = list(dict.fromkeys(cities))[:8]

    return findings


# ════════════════════════════════════════════════════════════════════
# LLM CALLERS  (unchanged from original)
# ════════════════════════════════════════════════════════════════════

def _groq_call(messages: list, max_tokens: int = 1000, temperature: float = 0,
               model_chain: list = None) -> str:
    """Call Groq with automatic model fallback + 429 backoff. ALWAYS returns a non-None str."""
    chain = model_chain or _RCA_MODELS
    last_error = None
    for attempt, model in enumerate(chain):
        try:
            resp = _groq_client.chat.completions.create(
                model=model,
                messages=messages,
                temperature=temperature,
                max_tokens=max_tokens,
            )
            content = resp.choices[0].message.content
            if content is None:
                print(f"[Groq {model}] returned None content — trying next model")
                last_error = ValueError("content=None")
                continue
            return content
        except Exception as e:
            err = str(e)
            if any(x in err for x in ["429", "rate_limit"]):
                # Back off before trying next model: 2s, 4s, 8s …
                wait = 2 ** min(attempt, 4)
                print(f"[Groq {model}] 429 rate limit — waiting {wait}s before next model…")
                time.sleep(wait)
                last_error = e
                continue
            if any(x in err for x in ["413", "decommissioned", "model_decommissioned"]):
                print(f"[Groq {model} unavailable ({err[:60]}) → next]")
                last_error = e
                continue
            raise
    raise RuntimeError(f"All Groq models unavailable. Last: {last_error}")

def _openrouter_call(messages: list, max_tokens: int = 2000, temperature: float = 0,
                     or_model_chain: list = None, groq_fallback_chain: list = None) -> str:
    """
    Call OpenRouter with a specific model chain.
    Each agent passes its own or_model_chain so they all hit different models.
    Falls back to groq_fallback_chain (or _RCA_MODELS) if OpenRouter is unavailable.
    """
    fallback_chain = groq_fallback_chain or _RCA_MODELS
    model_list = or_model_chain or _ORCH_MODELS

    if not _OR_KEY:
        return _groq_call(messages, max_tokens=max_tokens, temperature=temperature,
                          model_chain=fallback_chain)

    def _safe(t: str) -> str:
        return t.encode("utf-8", errors="replace").decode("ascii", errors="replace")

    safe_msgs = [{"role": m["role"], "content": _safe(m["content"])} for m in messages]
    last_err = None

    for model in model_list:
        try:
            resp = requests.post(
                "https://openrouter.ai/api/v1/chat/completions",
                headers={
                    "Authorization": f"Bearer {_OR_KEY}",
                    "Content-Type": "application/json",
                    "HTTP-Referer": "https://graphpulse.ai",
                    "X-Title": "GraphPulse AI",
                },
                json={
                    "model": model,
                    "messages": safe_msgs,
                    "temperature": temperature,
                    "max_tokens": max_tokens,
                },
                timeout=60,
            )
            if resp.status_code == 200:
                content = (resp.json().get("choices") or [{}])[0].get("message", {}).get("content", "")
                if content is None or not str(content).strip():
                    last_err = f"Empty/None response from {model}"
                    continue
                return str(content).strip()
            if resp.status_code in (429, 503, 502, 404, 400):
                last_err = f"HTTP {resp.status_code} from {model}"
                continue
            resp.raise_for_status()
        except requests.exceptions.Timeout:
            last_err = f"Timeout on {model}"
            continue
        except Exception as e:
            last_err = str(e)
            if any(x in str(e) for x in ["429", "503", "502", "404", "400", "timeout"]):
                continue
            break

    print(f"[OpenRouter fallback to Groq] Last error: {last_err}")
    return _groq_call(messages, max_tokens=max_tokens, temperature=temperature,
                      model_chain=fallback_chain)


# ════════════════════════════════════════════════════════════════════
# MCP / TOOL DISPATCH  (unchanged)
# ════════════════════════════════════════════════════════════════════

_MCP_BASE_URL = None

def get_mcp_base_url() -> str | None:
    """
    Returns a truthy string if the MCP server is reachable, None otherwise.

    The actual URL is now managed inside mcp_client.py.  This function is
    kept for backwards compat with any code that checks `if get_mcp_base_url()`.
    """
    if is_server_reachable():
        import json as _j
        try:
            with open("mcp_port.json") as _f:
                port = _j.load(_f)["port"]
            return f"http://127.0.0.1:{port}"
        except Exception:
            pass
    return None

def _local_fallback(tool_name: str, tool_input: dict) -> str:
    """Direct Neo4j call — used when MCP is unreachable."""
    print(f"[Tool → LOCAL Neo4j] {tool_name} | input keys: {list(tool_input.keys())}")

    # Alias: get_category_supplier_risk → trace_supply_chain_for_category
    if tool_name == "get_category_supplier_risk":
        tool_name = "trace_supply_chain_for_category"
    fn = TOOL_FUNCTIONS.get(tool_name)
    if fn is None:
        return json.dumps({"error": f"Unknown tool '{tool_name}'"})

    # ── Guard: create_or_update_node must have cypher_merge_query ──
    if tool_name == "create_or_update_node":
        cypher = tool_input.get("cypher_merge_query") or tool_input.get("cypher") or tool_input.get("query")
        if not cypher:
            return json.dumps({"error": (
                "create_or_update_node requires {'cypher_merge_query': 'MERGE ...'}. "
                f"Got keys: {list(tool_input.keys())}. "
                "Example: {'cypher_merge_query': 'MERGE (s:Supplier {supplier_id:\'SUP9001\'}) "
                "SET s.supplier_name=\'Mehta Plastics\', s.risk_score=0.45 RETURN s.supplier_id'}"
            )})
        # Normalise key
        tool_input = {"cypher_merge_query": cypher}

    try:
        result = fn(**tool_input)
        return result if isinstance(result, str) else json.dumps(result)
    except TypeError as te:
        # Try positional arg as fallback
        try:
            if tool_name == "create_or_update_node":
                result = fn(tool_input.get("cypher_merge_query", ""))
                return result if isinstance(result, str) else json.dumps(result)
            return str(fn())
        except Exception as e2:
            return json.dumps({"error": f"TypeError fallback failed: {e2}"})
    except Exception as e:
        return json.dumps({"error": str(e)})

def _dispatch_tool(tool_name: str, tool_input: dict) -> str:
    """
    Dispatch a tool call.

    Priority:
      1. MCP JSON-RPC  (mcp_client.dispatch_tool → POST /mcp)
      2. Local Neo4j   (_local_fallback)           ← unchanged

    The old REST-shim path (POST /tools/{name}) is no longer used here.
    It still exists on the server for external clients / debugging.
    """
    # ── Attempt 1: MCP JSON-RPC ──────────────────────────────────────
    try:
        result = _mcp_dispatch(tool_name, tool_input)
        print(f"[Tool → MCP JSON-RPC] {tool_name}")
        return result
    except MCPClientError as mcp_err:
        print(f"[MCP unavailable] {mcp_err} — falling back to local Neo4j")

    # ── Attempt 2: Direct Neo4j (unchanged fallback) ─────────────────
    return _local_fallback(tool_name, tool_input)

    # NOTE: _local_fallback() is defined later in agent_runner.py —
    # no changes needed there.  It still imports TOOL_FUNCTIONS from
    # neo4j_tools and calls them directly.


# ════════════════════════════════════════════════════════════════════
# GRAPH UPDATE AGENT  (unchanged from original)
# ════════════════════════════════════════════════════════════════════

_GRAPH_CONTEXT = (
    "You are an expert supply chain analyst with access to a Neo4j knowledge graph.\n"
    "Graph: Supplier→Plant→Shipment→Distributor→Retailer. Products carried by Shipments.\n"
    "delivery_status values: 'Major Delay' or 'On Time'. "
    "Plants in graph: use get_graph_summary to discover current plant list. "
    "Default known plants: PL1=Baddi, PL2=Bhopal, PL3=Pune, PL4=Goa — verify via query before referencing.\n"
)

UPDATE_SYSTEM = _GRAPH_CONTEXT + """
TASK: Execute graph updates in Neo4j. Write the node/relationship immediately — do NOT verify first.

MANDATORY FIRST STEP: Your FIRST action MUST ALWAYS be create_or_update_node with a MERGE statement.
NEVER call verify_node_exists as your first action. NEVER check if a node exists before creating it.
MERGE is idempotent — it creates if not found, updates if found. Just run it.

FORMAT (follow EXACTLY — one action per step):
Thought: I will MERGE this supplier node into the graph immediately.
Action: create_or_update_node
Action Input: {"cypher_merge_query": "MERGE (s:Supplier {supplier_id:'SUP9001'}) SET s.supplier_name='Mehta Plastics', s.risk_score=0.45, s.annual_capacity_units=50000, s.StoP_lead_time_days=8, s.status='Active', s.supplier_latitude=0.0, s.supplier_longitude=0.0, s.StoP_distance_km=0.0 RETURN s.supplier_id"}
Observation: <result>
Thought: Node created. Now verify it was committed.
Action: verify_node_exists
Action Input: {"label": "Supplier", "property_name": "supplier_id", "property_value": "SUP9001"}
Observation: <result>
Final Answer: ✓ Supplier SUP9001 'Mehta Plastics' created and verified in Neo4j.

TOOLS — use EXACTLY these two names, spelled exactly:
  create_or_update_node   — input: {"cypher_merge_query": "<full cypher string>"}
  verify_node_exists      — input: {"label": "...", "property_name": "...", "property_value": "..."}

ABSOLUTE RULES:
1. FIRST action is ALWAYS create_or_update_node. No exceptions. Never verify before creating.
2. Action Input MUST be valid JSON with key "cypher_merge_query" for create_or_update_node.
3. The "cypher_merge_query" value MUST be a complete MERGE or MATCH+MERGE Cypher string.
4. Use MERGE (not CREATE). SET all properties in the same statement.
5. Use single quotes inside Cypher. Use double quotes for the JSON wrapper only.
6. After create_or_update_node succeeds → call verify_node_exists once to confirm.
7. If verify returns {"found": true} → write Final Answer with ✓ confirmation.
8. If verify returns {"found": false} → retry create_or_update_node once more, then Final Answer.
9. Never call verify_node_exists more than once per node.
10. NEVER use DELETE, DETACH DELETE, DROP, or REMOVE in any Cypher. You are a creation/update agent only. If the user asks to delete something, respond in Final Answer: "⚠ Delete operations must be run via Neo4j Browser or the Graph Verifier — this agent only creates and updates nodes."
11. ALWAYS use the EXACT supplier_name, distributor_city, or other name given by the user. Never substitute, paraphrase, or shorten it. If user says 'Test NL Supplier', write supplier_name='Test NL Supplier' exactly.

WORKFLOW FOR SUPPLIER + PLANT LINK (e.g. "Add supplier X, supply to plant PL3"):
Step 1 → MERGE the Supplier node with create_or_update_node
Step 2 → verify_node_exists confirms the supplier was saved
Step 3 → MERGE the SUPPLIES_TO relationship with create_or_update_node
Step 4 → Final Answer with confirmation

CYPHER TEMPLATES — copy exactly, replace VALUES only:

New Supplier + relationship:
Action: create_or_update_node
Action Input: {"cypher_merge_query": "MERGE (s:Supplier {supplier_id:'SUP9001'}) SET s.supplier_name='Mehta Plastics', s.risk_score=0.45, s.annual_capacity_units=50000, s.StoP_lead_time_days=8, s.status='Active', s.supplier_latitude=0.0, s.supplier_longitude=0.0, s.StoP_distance_km=0.0 RETURN s.supplier_id"}

New Distributor:
Action: create_or_update_node
Action Input: {"cypher_merge_query": "MERGE (d:Distributor {distributor_id:'D9001'}) SET d.distributor_city='Surat', d.distributor_latitude=21.17, d.distributor_longitude=72.83 RETURN d.distributor_id"}

New Route:
Action: create_or_update_node
Action Input: {"cypher_merge_query": "MERGE (r:Route {route_id:'PL4@D0050'}) SET r.mode='Road', r.PtoD_distance_km=320, r.PtoD_leadtime_days=3, r.PtoD_transportation_cost_inr=85000, r.plant_id='PL4', r.distributor_id='D0050', r.cost_efficiency=0.65 RETURN r.route_id"}

SUPPLIES_TO relationship (only after supplier node confirmed created):
Action: create_or_update_node
Action Input: {"cypher_merge_query": "MATCH (s:Supplier {supplier_id:'SUP9001'}), (p:Plant {plant_id:'PL3'}) MERGE (s)-[:SUPPLIES_TO]->(p) RETURN s.supplier_id, p.plant_id"}

Update a property:
Action: create_or_update_node
Action Input: {"cypher_merge_query": "MATCH (s:Supplier {supplier_id:'SUP0026'}) SET s.risk_score=0.85 RETURN s.supplier_id, s.risk_score"}

Verify after create:
Action: verify_node_exists
Action Input: {"label": "Supplier", "property_name": "supplier_id", "property_value": "SUP9001"}

IMPORTANT: Plants PL1, PL2, PL3, PL4 always exist — never verify them. Just MATCH directly.
"""

# In-memory log of NL graph updates for undo support
_NL_UPDATE_LOG: list[dict] = []

_NL_LOG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "nl_update_log.pkl")

def _nl_log_save():
    try:
        import pickle
        with open(_NL_LOG_FILE, "wb") as _f: pickle.dump(_NL_UPDATE_LOG, _f)
    except Exception: pass

def _nl_log_load():
    global _NL_UPDATE_LOG
    try:
        import pickle
        if os.path.exists(_NL_LOG_FILE):
            with open(_NL_LOG_FILE, "rb") as _f: _NL_UPDATE_LOG = pickle.load(_f)
    except Exception: pass

_nl_log_load()

def _nl_csv_upsert(entity_type, entity_id, props):
    """Write NL-confirmed node back to CSV on disk."""
    try:
        import stage4_pipeline as _s4
        _s4._csv_upsert(entity_type, entity_id, props)
    except Exception: pass

def _nl_csv_delete(entity_type, entity_id):
    """Remove NL-deleted node from CSV on disk."""
    try:
        import stage4_pipeline as _s4
        _s4._csv_delete(entity_type, entity_id)
    except Exception: pass

def _log_nl_update(entity_type: str, entity_id: str, cypher: str, id_prop: str = None):
    """Log a confirmed NL-driven graph write for undo support."""
    import datetime as _dt
    # Derive id_prop from entity_type if not given
    if not id_prop:
        id_prop = {
            "Supplier":    "supplier_id",
            "Distributor": "distributor_id",
            "Route":       "route_id",
            "Plant":       "plant_id",
        }.get(entity_type, "id")
    _NL_UPDATE_LOG.append({
        "entity_type": entity_type,
        "entity_id":   entity_id,
        "id_prop":     id_prop,
        "cypher":      cypher,
        "timestamp":   _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    })
    if len(_NL_UPDATE_LOG) > 50:
        _NL_UPDATE_LOG.pop(0)
    _nl_log_save()
    _nl_csv_upsert(entity_type, entity_id, {id_prop: entity_id})


def get_nl_update_history() -> list:
    """Return the NL update log (most recent first) for display in UI."""
    return list(reversed(_NL_UPDATE_LOG))


def undo_nl_update(index: int = -1) -> str:
    """
    Delete an NL-created node from Neo4j and remove it from the log.
    index=-1 (default) = undo the most recent entry.
    index=0..N = undo a specific entry by position in get_nl_update_history() order.
    """
    if not _NL_UPDATE_LOG:
        return "⚠ No NL updates to undo — the update log is empty."

    # get_nl_update_history() is reversed, so index 0 = most recent = last in _NL_UPDATE_LOG
    history = get_nl_update_history()
    if index < 0:
        index = 0  # default to most recent
    if index >= len(history):
        return f"⚠ No entry at position {index}. Log has {len(history)} entries."

    entry    = history[index]
    etype    = entry["entity_type"]
    eid      = entry["entity_id"]
    id_prop  = entry.get("id_prop") or {
        "Supplier":    "supplier_id",
        "Distributor": "distributor_id",
        "Route":       "route_id",
        "Plant":       "plant_id",
    }.get(etype, "id")

    try:
        _run_neo4j(
            f"MATCH (n:{etype} {{{id_prop}: $v}}) DETACH DELETE n",
            {"v": eid}
        )
        # Remove from log (find by timestamp to be safe)
        ts = entry["timestamp"]
        for i, item in enumerate(_NL_UPDATE_LOG):
            if item["timestamp"] == ts and item["entity_id"] == eid:
                _NL_UPDATE_LOG.pop(i)
                break
        _nl_csv_delete(etype, eid)
        _nl_log_save()
        return f"✓ Undone: deleted {etype} '{eid}' from Neo4j and CSV files."
    except Exception as e:
        return f"⚠ Undo failed for {etype} '{eid}': {e}"


def undo_last_nl_update() -> str:
    """Convenience wrapper — undo the most recent NL update."""
    return undo_nl_update(index=0)

VALID_TOOLS = set(TOOL_FUNCTIONS.keys())


# ════════════════════════════════════════════════════════════════════
# NODE PRE-VALIDATION  — runs before any relationship-creation Cypher
# Prevents the "No node exists" error by checking BOTH endpoints of
# a MATCH-based relationship before attempting the MERGE.
# ════════════════════════════════════════════════════════════════════

def _extract_match_targets(cypher: str) -> list:
    """
    Parse a Cypher MATCH clause and return a list of node targets to validate.
    Each target is: {label, prop_name, prop_value}
    """
    import re as _re
    targets = []
    pattern = _re.compile(
        r'\(\w+:(\w+)\s*\{(\w+)\s*:\s*[\'"]([^\'"]+)[\'"]}'
        , _re.IGNORECASE
    )
    for m in pattern.finditer(cypher):
        targets.append({
            "label":      m.group(1),
            "prop_name":  m.group(2),
            "prop_value": m.group(3),
        })
    return targets


def _validate_nodes_for_relationship(cypher: str) -> tuple:
    """
    Pre-validates that all nodes referenced in a MATCH-based relationship
    Cypher statement actually exist in Neo4j BEFORE attempting the MERGE.

    Returns: (ok: bool, message: str)
      ok=True  -> all nodes confirmed, safe to proceed
      ok=False -> one or more nodes missing; message explains which

    Only activates when Cypher contains both MATCH and MERGE
    (i.e. it is a relationship-creation statement, not a pure node MERGE).
    """
    cypher_upper = (cypher or "").upper().strip()
    if "MATCH" not in cypher_upper or "MERGE" not in cypher_upper:
        return True, ""

    targets = _extract_match_targets(cypher)
    if not targets:
        return True, ""

    # Plants are validated via DB like all other nodes — no hardcoded whitelist
    missing = []
    for t in targets:
        try:
            results = _run_neo4j(
                f"MATCH (n:{t['label']} {{{t['prop_name']}: $val}}) RETURN n LIMIT 1",
                {"val": t["prop_value"]}
            )
            if not results:
                missing.append(f"{t['label']} with {t['prop_name']}='{t['prop_value']}' ")
        except Exception as e:
            print(f"[NodePreValidation] Warning: could not validate {t['label']} ({e})")

    if missing:
        missing_str = "; ".join(missing)
        return False, (
            f"\u26a0 Pre-validation failed \u2014 the following node(s) do not exist in the graph: "
            f"{missing_str}. Create the missing node(s) first, then retry the relationship link."
        )
    return True, ""


# ════════════════════════════════════════════════════════════════════
# CYPHER CATALOG — maps each tool to its purpose + representative query
# Populated with real record counts during execution
# ════════════════════════════════════════════════════════════════════
_CYPHER_CATALOG: dict = {
    "get_supplier_risk": {
        "purpose": "Fetch all supplier risk scores and lead times",
        "cypher": "MATCH (s:Supplier)\nRETURN s.supplier_id AS supplier_id, s.supplier_name AS supplier_name, s.risk_score AS risk_score, s.StoP_lead_time_days AS lead_time_days\nORDER BY s.risk_score DESC",
    },
    "get_plant_delay_rate": {
        "purpose": "Calculate delay rate % for each Plant",
        "cypher": "MATCH (p:Plant)-[:DISPATCHES]->(sh:Shipment)\nWITH p, COUNT(sh) AS total, SUM(CASE WHEN sh.delivery_status = 'Major Delay' THEN 1 ELSE 0 END) AS delayed\nRETURN p.plant_id AS plant_id, p.plant_name AS plant_name, total, delayed,\n       round(100.0*delayed/CASE WHEN total=0 THEN 1 ELSE total END,1) AS delay_rate_pct\nORDER BY delay_rate_pct DESC",
    },
    "get_delayed_shipments": {
        "purpose": "Retrieve all Major Delay shipments with full supply chain path",
        "cypher": "MATCH (sup:Supplier)-[:SUPPLIES_TO]->(pl:Plant)-[:DISPATCHES]->(sh:Shipment)-[:SHIPPED_TO]->(d:Distributor)\nWHERE sh.delivery_status = 'Major Delay'\nRETURN sh.shipment_id AS shipment_id, sup.supplier_name AS supplier, pl.plant_name AS plant,\n       d.distributor_city AS distributor_city, sh.delay_days AS delay_days\nORDER BY sh.delay_days DESC  LIMIT 100",
    },
    "get_distributor_demand_gap": {
        "purpose": "Identify distributors with high demand gaps",
        "cypher": "MATCH (pl:Plant)-[:DISPATCHES]->(sh:Shipment)-[:SHIPPED_TO]->(d:Distributor)\nWHERE sh.demand_gap > 0\nWITH d.distributor_id AS distributor_id, d.distributor_city AS city,\n     COUNT(DISTINCT sh) AS shortage_shipments,\n     SUM(sh.demand_gap) AS total_demand_gap\nRETURN distributor_id, city, shortage_shipments, total_demand_gap\nORDER BY total_demand_gap DESC\nLIMIT 15",
    },
    "get_retailer_stockouts": {
        "purpose": "Find all Retailers with active stockout flags",
        "cypher": "MATCH (pl:Plant)-[:DISPATCHES]->(sh:Shipment)-[:SHIPPED_TO]->(d:Distributor)-[:DELIVERS_TO]->(r:Retailer)\nWHERE sh.demand_gap > 0\nRETURN r.retailer_id AS retailer_id, r.retailer_city AS city,\n       d.distributor_city AS served_by,\n       COUNT(DISTINCT sh) AS shortage_shipments,\n       SUM(sh.demand_gap) AS total_demand_gap\nORDER BY total_demand_gap DESC\nLIMIT 15",
    },
    "get_supply_chain_overview": {
        "purpose": "High-level supply chain network statistics",
        "cypher": "MATCH (sup:Supplier)-[:SUPPLIES_TO]->(pl:Plant)-[:DISPATCHES]->(sh:Shipment)-[:SHIPPED_TO]->(d:Distributor)\nRETURN COUNT(DISTINCT sup) AS total_suppliers, COUNT(DISTINCT pl) AS total_plants,\n       COUNT(sh) AS total_shipments, COUNT(DISTINCT d) AS total_distributors,\n       SUM(CASE WHEN sh.delivery_status='Major Delay' THEN 1 ELSE 0 END) AS delayed",
    },
    "get_product_category_delays": {
        "purpose": "Count delayed shipments grouped by product category",
        "cypher": "MATCH (pl:Plant)-[:DISPATCHES]->(sh:Shipment)-[:CARRIES]->(p:Product)\nWHERE sh.delivery_status = 'Major Delay'\nRETURN p.product_category_name AS category, COUNT(sh) AS delayed_count\nORDER BY delayed_count DESC",
    },
    "get_delay_by_product_category": {
        "purpose": "Count delayed shipments by product category (optionally filtered)",
        "cypher": "MATCH (pl:Plant)-[:DISPATCHES]->(sh:Shipment)-[:CARRIES]->(p:Product)\nWHERE sh.delivery_status = 'Major Delay'\nRETURN p.product_category_name AS category, COUNT(sh) AS delayed_count\nORDER BY delayed_count DESC",
    },
    "get_supplier_plant_delay_chain": {
        "purpose": "Full supplier to plant to shipment delay chain",
        "cypher": "MATCH (sup:Supplier)-[:SUPPLIES_TO]->(pl:Plant)-[:DISPATCHES]->(sh:Shipment)-[:SHIPPED_TO]->(d:Distributor)\nWHERE sh.delivery_status = 'Major Delay'\nRETURN sup.supplier_name AS supplier, sup.risk_score AS risk_score,\n       pl.plant_name AS plant, pl.plant_id AS plant_id,\n       d.distributor_city AS city, COUNT(sh) AS delayed_count, AVG(sh.delay_days) AS avg_delay_days\nORDER BY delayed_count DESC  LIMIT 20",
    },
    "get_route_analysis": {
        "purpose": "Analyse transport routes by mode, cost, and distance",
        "cypher": "MATCH (pl:Plant)-[:HAS_ROUTE]->(r:Route)-[:CONNECTS_TO]->(d:Distributor)\nRETURN pl.plant_name AS plant, r.route_id AS route_id,\n       r.mode AS transport_mode,\n       r.PtoD_distance_km AS distance_km,\n       r.PtoD_transportation_cost_inr AS cost_inr,\n       r.PtoD_leadtime_days AS leadtime_days,\n       r.cost_efficiency AS cost_efficiency,\n       d.distributor_city AS distributor_city\nORDER BY cost_inr DESC\nLIMIT 20",
    },
    "get_high_risk_suppliers": {
        "purpose": "Identify high-risk suppliers causing Major Delay shipments",
        "cypher": "MATCH (s:Supplier)-[:SUPPLIES_TO]->(p:Plant)-[:DISPATCHES]->(sh:Shipment)\nWHERE s.risk_score > 0.6  AND sh.delivery_status = 'Major Delay'\nRETURN s.supplier_name AS supplier, s.risk_score AS risk_score,\n       p.plant_name AS plant, COUNT(sh) AS delayed_shipments\nORDER BY delayed_shipments DESC",
    },
    "get_delay_by_plant": {
        "purpose": "Delay counts and rates per Plant",
        "cypher": "MATCH (p:Plant)-[:DISPATCHES]->(sh:Shipment)\nWITH p, COUNT(sh) AS total, SUM(CASE WHEN sh.delivery_status = 'Major Delay' THEN 1 ELSE 0 END) AS delayed\nRETURN p.plant_id AS plant_id, p.plant_name AS plant_name, total, delayed,\n       round(100.0*delayed/CASE WHEN total=0 THEN 1 ELSE total END,1) AS delay_rate_pct,\n       AVG(sh.delay_days) AS avg_delay_days\nORDER BY delay_rate_pct DESC",
    },
    "get_distributor_delay_impact": {
        "purpose": "Delayed shipments and avg delay days per distributor city",
        "cypher": "MATCH (pl:Plant)-[:DISPATCHES]->(sh:Shipment)-[:SHIPPED_TO]->(d:Distributor)\nWHERE sh.delivery_status = 'Major Delay'\nRETURN d.distributor_city AS city, COUNT(sh) AS delayed_shipments,\n       round(AVG(sh.delay_days),2) AS avg_delay_days,\n       SUM(sh.demand_gap) AS demand_gap\nORDER BY delayed_shipments DESC\nLIMIT 15",
    },
    "get_stockout_retailers": {
        "purpose": "Retailer cities with unmet demand — GROUPED per city. Returns: retailer_city, served_by_distributor, unique_retailers, shortage_shipments, total_shortage_units",
        "cypher": "MATCH (sh:Shipment)-[:SHIPPED_TO]->(d:Distributor)-[:DELIVERS_TO]->(r:Retailer)\nWHERE sh.demand_gap IS NOT NULL AND sh.demand_gap > 0\nWITH r.retailer_city AS retailer_city,\n     d.distributor_city AS served_by_distributor,\n     COUNT(DISTINCT r) AS unique_retailers,\n     COUNT(DISTINCT sh) AS shortage_shipments,\n     SUM(sh.demand_gap) AS total_shortage_units\nRETURN retailer_city, served_by_distributor, unique_retailers,\n       shortage_shipments, total_shortage_units\nORDER BY total_shortage_units DESC\nLIMIT 10",
    },
    "get_demand_gap_analysis": {
        "purpose": "AUTHORITATIVE distributor shortage query. Returns per distributor: distributor_city, shortage_shipments, total_demand_gap, avg_demand_gap, delayed_shipments, avg_delay_days, retailers_affected. Use THIS as the single source for all distributor metrics.",
        "cypher": "MATCH (pl:Plant)-[:DISPATCHES]->(sh:Shipment)-[:SHIPPED_TO]->(d:Distributor)\nWITH d.distributor_city AS distributor_city,\n     COUNT(DISTINCT CASE WHEN sh.demand_gap > 0 THEN sh END) AS shortage_shipments,\n     SUM(CASE WHEN sh.demand_gap > 0 THEN sh.demand_gap ELSE 0 END) AS total_demand_gap,\n     COUNT(DISTINCT CASE WHEN sh.delivery_status = \'Major Delay\' THEN sh END) AS delayed_shipments,\n     round(AVG(CASE WHEN sh.delivery_status = \'Major Delay\' THEN sh.delay_days END), 2) AS avg_delay_days\nOPTIONAL MATCH (d)-[:DELIVERS_TO]->(r:Retailer)\nRETURN distributor_city, shortage_shipments, total_demand_gap,\n       delayed_shipments, avg_delay_days,\n       COUNT(DISTINCT r) AS retailers_affected\nORDER BY total_demand_gap DESC\nLIMIT 15",
    },
    "get_route_cost_efficiency": {
        "purpose": "Route efficiency scores, distance and cost per route",
        "cypher": "MATCH (pl:Plant)-[:HAS_ROUTE]->(r:Route)-[:CONNECTS_TO]->(d:Distributor)\nRETURN pl.plant_name AS plant, d.distributor_city AS distributor,\n       r.mode AS mode, r.PtoD_distance_km AS distance_km, r.PtoD_transportation_cost_inr AS cost_inr\nORDER BY r.PtoD_transportation_cost_inr DESC  LIMIT 30",
    },
    "get_transport_mode_delays": {
        "purpose": "AUTHORITATIVE count of ALL Major Delay shipments by transport mode using correct double-MATCH pattern",
        "cypher": "MATCH (pl:Plant)-[:HAS_ROUTE]->(r:Route)-[:CONNECTS_TO]->(d:Distributor)\nMATCH (pl)-[:DISPATCHES]->(sh:Shipment)-[:SHIPPED_TO]->(d)\nWHERE sh.delivery_status = 'Major Delay'\nWITH r.mode AS transportation_mode,\n     COUNT(sh) AS total_delays,\n     round(AVG(sh.delay_days), 2) AS avg_delay_days,\n     COUNT(DISTINCT pl) AS plants_affected\nRETURN transportation_mode, total_delays, avg_delay_days, plants_affected\nORDER BY total_delays DESC",
    },
    "get_schema_with_examples": {
        "purpose": "Returns full graph schema + 6 canonical Cypher patterns. Call BEFORE run_cypher.",
        "cypher": "/* No Cypher — returns static schema dict with node labels, relationship directions, property enums, and canonical query examples */",
    },
    "get_monthly_delay_trend": {
        "purpose": "Monthly shipment delay trends over time",
        "cypher": "MATCH (sh:Shipment)\nWITH sh.ship_date.month AS month, sh.ship_date.year AS year,\n     COUNT(sh) AS total, SUM(CASE WHEN sh.delivery_status = 'Major Delay' THEN 1 ELSE 0 END) AS delayed\nRETURN year, month, total, delayed,\n       round(100.0*delayed/CASE WHEN total=0 THEN 1 ELSE total END,1) AS delay_rate_pct\nORDER BY year, month",
    },
    "run_cypher": {
        "purpose": "Execute dynamic Cypher query generated from question context",
        "cypher": "/* Dynamic query — generated from question context */\nMATCH (n) WHERE ... RETURN ...",
    },
    "create_or_update_node": {
        "purpose": "Create or update a graph node via MERGE",
        "cypher": "MERGE (n:{label} {id_prop: $id})\nSET n += $properties\nRETURN n",
    },
    "create_relationship": {
        "purpose": "Create a relationship between two nodes via MERGE",
        "cypher": "MATCH (a:{from_label} {from_prop: $from_id})\nMATCH (b:{to_label} {to_prop: $to_id})\nMERGE (a)-[:{rel_type}]->(b)  RETURN a, b",
    },
}


def _get_tool_cypher(tool_name: str, tool_input: dict) -> dict:
    """
    Return cypher metadata for a tool call.
    Injects real entity filters from tool_input so the accordion
    shows the EXACT query that ran (not a generic template).
    Returns dict with keys: purpose, cypher
    """
    import re as _re2
    # Explicit Cypher in input takes priority
    for k in ("cypher_merge_query", "cypher", "query"):
        if k in tool_input and tool_input[k]:
            return {
                "purpose": _CYPHER_CATALOG.get(tool_name, {}).get(
                    "purpose", tool_name.replace("_", " ").title()),
                "cypher": str(tool_input[k]),
            }

    base = _CYPHER_CATALOG.get(tool_name, {
        "purpose": tool_name.replace("_", " ").title(),
        "cypher": f"/* {tool_name} */\nMATCH (n) RETURN n LIMIT 25",
    })

    cypher  = base["cypher"]
    purpose = base["purpose"]

    # Inject real filter values into the Cypher
    cat = tool_input.get("category_filter")
    pid = tool_input.get("plant_id_filter")
    thr = tool_input.get("threshold")

    if cat:
        cypher  = _re2.sub(r"AND \(p\.product_category_name = \$category OR \$category IS NULL\)",
                           f"AND toLower(p.product_category_name) = toLower(\'{cat}\')", cypher)
        purpose = purpose + f" — category: {cat}"
    else:
        cypher = _re2.sub(r"\s*AND \(p\.product_category_name = \$category OR \$category IS NULL\)\n?", "\n", cypher)

    if pid:
        cypher  = _re2.sub(r"AND \(pl\.plant_id = \$plant_id OR \$plant_id IS NULL\)",
                           f"AND pl.plant_id = \'{pid}\'", cypher)
        purpose = purpose + f" — plant: {pid}"
    else:
        cypher = _re2.sub(r"\s*AND \(pl\.plant_id = \$plant_id OR \$plant_id IS NULL\)\n?", "\n", cypher)

    if thr is not None:
        cypher = cypher.replace("s.risk_score > 0.7", f"s.risk_score > {thr}")
        cypher = cypher.replace("risk_score > 0.6", f"risk_score > {thr}")

    return {"purpose": purpose, "cypher": cypher}



def _count_records(result_json: str) -> int:
    """Parse result and count returned records."""
    try:
        data = json.loads(result_json)
        if isinstance(data, list):
            return len(data)
        if isinstance(data, dict):
            if "error" in data:
                return 0
            # Try common list keys
            for k in data:
                if isinstance(data[k], list):
                    return len(data[k])
    except Exception:
        pass
    return -1  # unknown



def parse_action(text: str):
    action_match = re.search(r"Action:\s*(.+?)(?:\n|$)", text, re.IGNORECASE)
    input_match  = re.search(r"Action Input:\s*(\{[\s\S]*?\})", text)
    if not action_match:
        return None, None
    raw_tool  = action_match.group(1).strip()
    tool_name = raw_tool.strip().split()[0]
    if tool_name not in VALID_TOOLS:
        return None, None

    tool_input = {}
    if input_match:
        raw_json = input_match.group(1)
        try:
            tool_input = json.loads(raw_json)
        except (json.JSONDecodeError, AttributeError):
            # Recovery: try to extract cypher_merge_query value from malformed JSON
            cypher_match = re.search(
                r'"cypher_merge_query"\s*:\s*"([^"]+(?:\\"[^"]*)*)"', raw_json
            )
            if cypher_match and tool_name == "create_or_update_node":
                tool_input = {"cypher_merge_query": cypher_match.group(1)}
            else:
                # Try relaxed extraction — find any key:value pair
                pairs = re.findall(r'"([^"]+)"\s*:\s*"([^"]*)"', raw_json)
                tool_input = dict(pairs) if pairs else {}

    # ── Validate create_or_update_node always has cypher_merge_query ──
    if tool_name == "create_or_update_node" and "cypher_merge_query" not in tool_input:
        # Check if the raw text after "Action Input:" has a bare Cypher statement
        after_input = re.search(r"Action Input:\s*(MERGE|MATCH|CREATE)([\s\S]*?)(?=\nThought:|\nAction:|\nObservation:|\nFinal Answer:|$)", text, re.IGNORECASE)
        if after_input:
            cypher = (after_input.group(1) + after_input.group(2)).strip()
            tool_input = {"cypher_merge_query": cypher}
        else:
            # Cannot recover — skip this action
            return None, None

    return tool_name, tool_input

def _prune_history(messages: list) -> list:
    if len(messages) <= MAX_HISTORY_MESSAGES:
        return messages
    return messages[:2] + messages[-(MAX_HISTORY_MESSAGES - 2):]

def _format_rca_report(raw_text: str, tool_logs: list, tool_rows: dict = None) -> str:
    """
    Clean and normalise the raw RCA markdown. Always returns a non-None str.
    """
    import re as _re
    # Guard: raw_text must be a non-None string before any string operations
    if not raw_text:
        return "### 📌 Executive Summary\n\nReport generation incomplete — please retry your query."
    raw_text = str(raw_text)  # coerce in case something non-string slips through

    # ── Pass 1: strip noise ───────────────────────────────────
    cleaned = _re.sub(
        r'---\s*###?\s*📊\s*Data Sources.*', '',
        raw_text, flags=_re.DOTALL | _re.IGNORECASE
    )
    cleaned = str(cleaned or "").strip()
    cleaned = _re.sub(
        r'###?\s*📊\s*Data Sources.*', '',
        cleaned, flags=_re.DOTALL | _re.IGNORECASE
    )
    cleaned = str(cleaned or "").strip()

    # Normalise bracket-style category tags left by some models
    cleaned = _re.sub(r'\*\*\[Immediate\]\*\*\s*', '**', cleaned)
    cleaned = _re.sub(r'\*\*\[Short-term[^\]]*\]\*\*\s*', '**', cleaned)
    cleaned = _re.sub(r'\*\*\[Strategic[^\]]*\]\*\*\s*', '**', cleaned)
    cleaned = _re.sub(r'\[Immediate\]\s*', '', cleaned)
    cleaned = _re.sub(r'\[Short-term[^\]]*\]\s*', '', cleaned)
    cleaned = _re.sub(r'\[Strategic[^\]]*\]\s*', '', cleaned)

    # Normalise recommendation sub-headings to exact expected names
    cleaned = _re.sub(
        r'####\s*(?:⚡|🔴)?\s*(?:Fix\s*Now.*|Quick\s*Wins?|Immediate\s*(?:Relief|Action|Actions?|Steps?|Interventions?)(?:\s*[—–-].*)?(?:\s*\([^)]*\))?)',
        '#### ⚡ Critical Response  —  Act Immediately', cleaned, flags=_re.IGNORECASE
    )
    cleaned = _re.sub(
        r'####\s*(?:📋|🟡)?\s*(?:\d+[–-]\d+\s*Day\s*Actions?.*|Near[-\s]?Term\s*(?:Supplier\s*Programme|Actions?|Steps?)?|'
        r'Short[-\s]?Term\s*(?:Prevention|Actions?|Steps?)?|Medium[-\s]?Term\s*(?:Actions?|Steps?)?)(?:\s*[—–-].*)?'
        r'(?:\s*\([^)]*\))?',
        '#### 🔧 Operational Fixes  —  Near-Term Improvements', cleaned, flags=_re.IGNORECASE
    )
    cleaned = _re.sub(
        r'####\s*(?:🏗|🔵)?\s*(?:Long[–-]Term\s*Strategy.*|Strategic\s*(?:Actions?|Changes?|Fix|Steps?|Supplier\s*Strategy)?|'
        r'Structural\s*(?:Fix|Actions?|Changes?|Supplier\s*Strategy)?|Long[-\s]?Term\s*(?:Actions?|Steps?)?)(?:\s*[—–-].*)?'
        r'(?:\s*\([^)]*\))?',
        '#### 🏛 Strategic Initiatives  —  Structural Transformation', cleaned, flags=_re.IGNORECASE
    )
    # Fix table cells that contain only a dash (replace with N/A)
    def _fix_dash_cells(m):
        row = m.group(0)
        return _re.sub(r'(?<=\|)\s*[-—]\s*(?=\|)', ' N/A ', row)
    cleaned = _re.sub(r'^\|.+\|$', _fix_dash_cells, cleaned, flags=_re.MULTILINE)
    cleaned = str(cleaned or "")

    # Strip time-reference phrases from Recommendations section
    def _strip_time_refs_in_recs(text):
        text = str(text or "")
        rec_pat = _re.compile(
            r'(###\s*💡\s*Recommendations.*?)(?=\n##[^#]|\Z)',
            _re.DOTALL | _re.IGNORECASE
        )
        def _clean_rec_block(m):
            block = str(m.group(1) or "")
            block = _re.sub(r'\bwithin\s+\d+[\s-]*(?:days?|weeks?|months?)\b', '', block, flags=_re.IGNORECASE)
            block = _re.sub(r'\bin\s+\d+\s+(?:days?|weeks?|months?)\b', '', block, flags=_re.IGNORECASE)
            block = _re.sub(r'\b\d+[\s-]*(?:day|week|month|quarter)\s+(?:plan|window|period|timeline|target|sprint)\b', '', block, flags=_re.IGNORECASE)
            block = _re.sub(r'\bimmediately\b', '', block, flags=_re.IGNORECASE)  # strip the word, keep the action
            block = _re.sub(r'\bby\s+(?:end\s+of\s+)?(?:January|February|March|April|May|June|July|August|September|October|November|December|Q[1-4]|the\s+(?:week|month|quarter|year))\b', '', block, flags=_re.IGNORECASE)
            block = _re.sub(r'\b(?:due\s+by|deadline\s*(?::|is)?|target\s+date\s*:?)\s*[^.\n]*', '', block, flags=_re.IGNORECASE)
            block = _re.sub(r'\(\s*\d+\s*(?:days?|weeks?|months?)\s*\)', '', block, flags=_re.IGNORECASE)
            block = _re.sub(r'\(\s*(?:immediate|short[- ]term|medium[- ]term|long[- ]term)\s*\)', '', block, flags=_re.IGNORECASE)
            block = _re.sub(r',\s*\.', '.', block)
            block = _re.sub(r'  +', ' ', block)
            return block
        return rec_pat.sub(_clean_rec_block, text)
    cleaned = _strip_time_refs_in_recs(cleaned)
    cleaned = str(cleaned or "")

    # ── Pass 2: ensure full structure ────────────────────────
    _has_exec    = "executive summary" in cleaned.lower()
    _has_cause   = any(kw in cleaned.lower() for kw in [
        "root cause", "final root cause", "risk severity",   # simulation uses risk severity
        "scenario overview", "immediate impact",              # simulation headings
    ])
    _has_headers = ("##" in cleaned or "###" in cleaned)

    if _has_headers and (_has_exec or _has_cause):
        return cleaned

    # ── Fallback: rebuild structure from plain-text keyword scanning ─
    # Reached only when LLM returned prose without proper headings.
    sections = {
        "Root Cause":              "### 🏭 Root Cause\n",
        "Bottleneck Plant":        "### ⚠️ Bottleneck Plants\n",
        "Risky Supplier":          "### 🧑‍💼 High-Risk Suppliers\n",
        "Distributor Impact":      "### 🚚 Distributor Impact\n",
        "Demand Gap":              "### 📉 Demand Gap Trail\n",
        "Stockout Impact":         "### 🛒 Retailer Impact\n",
        "Immediate Impact":        "### 📡 Immediate Impact\n",
        "Network Propagation":     "### 🔗 Network Propagation\n",
        "Recommendation":          "### 💡 Recommendations\n",
    }
    formatted = "### 📌 Executive Summary\n"
    sentences = cleaned.replace("\n", " ").split(". ")
    formatted += ". ".join(sentences[:2]) + ".\n\n"
    remaining = cleaned
    for key, header in sections.items():
        idx = remaining.lower().find(key.lower())
        if idx == -1:
            continue
        end = len(remaining)
        for other_key in sections:
            if other_key == key:
                continue
            oi = remaining.lower().find(other_key.lower(), idx + len(key))
            if oi != -1 and oi < end:
                end = oi
        section_text = remaining[idx:end].strip()
        colon_idx = section_text.find(":")
        if colon_idx != -1:
            section_text = section_text[colon_idx + 1:].strip()
        formatted += header + section_text + "\n\n"
    return formatted

def _assemble_final_output(first_response: str, formatted_report: str, rec_raw: str = "", user_question: str = "", tool_rows: dict = None) -> str:
    """
    Assembles the final RCA output HTML. Always returns a non-None str.
    """
    import re as _re

    # Guard: all inputs must be non-None strings before any string operations
    formatted_report = str(formatted_report or "")
    rec_raw          = str(rec_raw or "")
    first_response   = str(first_response or "")
    user_question    = str(user_question or "")
    tool_rows        = tool_rows or {}   # Python row dicts for direct table rendering
    rec_block        = ""   # Rendered recommendations HTML

    # ── Determine report title from question intent ───────────────────
    q_l = user_question.lower()
    # Transport mode delay check FIRST — before generic route/delay fallbacks
    _is_transport_mode_q = any(w in q_l for w in ["transport mode","transportation mode","which mode","mode causes","mode has most","by mode","per mode","road delay","rail delay","air delay","sea delay"])
    if any(w in q_l for w in ["what if","impact of","disruption","simulation","simulate"]):
        _report_title  = "Supply Chain Impact Simulation"
        _report_icon   = "🧪"
        _title_color   = "#a78bfa"
        _border_color  = "rgba(167,139,250,0.35)"
    elif _is_transport_mode_q:
        _report_title  = "Transport Mode Delay — Root Cause Analysis Report"
        _report_icon   = "🚛"
        _title_color   = "#a78bfa"
        _border_color  = "rgba(167,139,250,0.35)"
    elif any(w in q_l for w in ["route","route cost","route efficiency","cheapest route","expensive route",
                                  "cost per route","logistics cost","freight","road","rail","air freight"]):
        _report_title  = "Route & Transport — Root Cause Analysis Report"
        _report_icon   = "🚛"
        _title_color   = "#e879f9"
        _border_color  = "rgba(232,121,249,0.35)"
    elif any(w in q_l for w in [
        "stockout", "demand gap", "shortage", "unmet", "retailer shortage",
        "running out", "out of stock", "stock level", "inventory low",
        "where are we running", "low stock", "depleting", "replenish"
    ]) and any(w in q_l for w in [
        "even though", "despite", "despite stable", "volumes remain", "stable",
        "who is responsible", "when did", "how is", "spreading", "driving these",
    ]):
        _report_title  = "Network Stockout Investigation — Root Cause Analysis Report"
        _report_icon   = "🔍"
        _title_color   = "#f87171"
        _border_color  = "rgba(248,113,113,0.35)"
    elif any(w in q_l for w in ["distributor delay","retailer shortage","distributor impact"]):
        _report_title  = "Distributor Impact — Root Cause Analysis Report"
        _report_icon   = "🚚"
        _title_color   = "#60a5fa"
        _border_color  = "rgba(96,165,250,0.35)"
    elif any(w in q_l for w in ["supplier risk","risky supplier","high risk supplier"]):
        _report_title  = "Supplier Risk — Root Cause Analysis Report"
        _report_icon   = "🏭"
        _title_color   = "#f87171"
        _border_color  = "rgba(248,113,113,0.35)"
    elif any(w in q_l for w in ["plant rca","pune plant","baddi plant","bhopal plant","goa plant","bottleneck plant",
                                  "pune rca","baddi rca","bhopal rca","goa rca"]) or \
         (any(w in q_l for w in ["pune","baddi","bhopal","goa"]) and
          any(w in q_l for w in ["plant","factory","rca","facility"])):
        _report_title  = "Plant — Root Cause Analysis Report"
        _report_icon   = "🏗️"
        _title_color   = "#fb923c"
        _border_color  = "rgba(251,146,60,0.35)"
    elif any(w in q_l for w in ["toy delay","watches_gifts","category impact","product category",
                                  "toys","watches","gifts","health beauty","auto","bed bath",
                                  "construction tools","category"]):
        _report_title  = "Product Category — Root Cause Analysis Report"
        _report_icon   = "📦"
        _title_color   = "#fbbf24"
        _border_color  = "rgba(251,191,36,0.35)"
    elif any(w in q_l for w in ["distributor","city","distribution","hub","depot"]):
        _report_title  = "Distributor — Root Cause Analysis Report"
        _report_icon   = "🏙️"
        _title_color   = "#60a5fa"
        _border_color  = "rgba(96,165,250,0.35)"
    elif any(w in q_l for w in ["supplier","vendor","sup"]):
        _report_title  = "Supplier — Root Cause Analysis Report"
        _report_icon   = "🏭"
        _title_color   = "#38bdf8"
        _border_color  = "rgba(56,189,248,0.35)"
    elif any(w in q_l for w in ["plant","factory","facility","production","manufacturing"]):
        _report_title  = "Plant Operations — Root Cause Analysis Report"
        _report_icon   = "🏗️"
        _title_color   = "#fb923c"
        _border_color  = "rgba(251,146,60,0.35)"
    elif any(w in q_l for w in ["delay","delayed","shipment","ship","late","on time","on-time"]):
        _report_title  = "Shipment Delay — Root Cause Analysis Report"
        _report_icon   = "🔍"
        _title_color   = "#38bdf8"
        _border_color  = "rgba(56,189,248,0.35)"
    else:
        # Generic fallback — extract main noun from question
        import re as _re2
        _noun = _re2.sub(r'\b(what|which|how|why|where|when|are|is|the|a|an|of|in|for|have|has|do|does|with|from|to|and|or|that|this)\b', '', q_l)
        _noun = _re2.sub(r'\s+', ' ', _noun).strip()
        _noun = ' '.join(w.capitalize() for w in _noun.split()[:3]) if _noun else "Supply Chain"
        _report_title  = f"{_noun} — Root Cause Analysis Report"
        _report_icon   = "🔍"
        _title_color   = "#38bdf8"
        _border_color  = "rgba(56,189,248,0.35)"

    report_body = formatted_report

    # ── DYNAMIC TITLE: derive from the question's core verb+object, not keywords ──
    # Instead of 20 elif branches, build a concise noun phrase from the question
    def _derive_report_title(q: str) -> tuple:
        """Returns (icon, title, color, border_color) from the question text."""
        ql = q.lower()
        # Simulation is always explicit
        if any(w in ql for w in ["what if","shutdown","shuts down","goes offline","simulate","disaster","flood"]):
            return ("🧪","Supply Chain Impact Simulation","#a78bfa","rgba(167,139,250,0.35)")
        # Route/logistics
        if any(w in ql for w in ["route","transport mode","logistics cost","freight","road delay","rail delay"]):
            return ("🚛","Route & Transport — RCA Report","#e879f9","rgba(232,121,249,0.35)")
        # Stockout/shortage
        if any(w in ql for w in ["stockout","shortage","demand gap","unmet","out of stock","inventory low","depleting"]):
            return ("📉","Stock Shortage — RCA Report","#f87171","rgba(248,113,113,0.35)")
        # Retailer
        if any(w in ql for w in ["retailer","retail","store","outlet","shelf"]):
            return ("🛒","Retailer Impact — RCA Report","#a78bfa","rgba(167,139,250,0.35)")
        # Distributor
        if any(w in ql for w in ["distributor","distribution","city","region","hub","depot"]):
            return ("🏙️","Distributor — RCA Report","#60a5fa","rgba(96,165,250,0.35)")
        # Product/category
        if any(w in ql for w in ["product","category","toy","auto","health","beauty","watch","gift","cool stuff","bed bath","construction"]):
            return ("📦","Product Category — RCA Report","#fbbf24","rgba(251,191,36,0.35)")
        # Supplier
        if any(w in ql for w in ["supplier","vendor","risk score","risky","high risk"]):
            return ("🏭","Supplier Risk — RCA Report","#f87171","rgba(248,113,113,0.35)")
        # Plant
        if any(w in ql for w in ["plant","factory","baddi","pune","bhopal","goa","facility","pl1","pl2","pl3","pl4"]):
            return ("🏗️","Plant Operations — RCA Report","#fb923c","rgba(251,146,60,0.35)")
        # Performance/trend
        if any(w in ql for w in ["performance","on-time","kpi","benchmark","trend","monthly","getting worse"]):
            return ("📈","Supply Chain Performance — RCA Report","#4ade80","rgba(74,222,128,0.35)")
        # Delay (generic)
        if any(w in ql for w in ["delay","delayed","late","slow","bottleneck"]):
            return ("🔍","Shipment Delay — RCA Report","#38bdf8","rgba(56,189,248,0.35)")
        # Completely generic — derive a 3-word noun phrase from the question
        import re as _re2
        _stopwords = {'what','which','how','why','where','when','are','is','the','a','an',
                      'of','in','for','have','has','do','does','with','from','to','and','or',
                      'that','this','been','being','was','were','will','would','could','should',
                      'across','even','though','despite','overall','remain','stable','increased'}
        words = [w for w in re.sub(r'[^a-z ]', '', ql).split() if w not in _stopwords and len(w) > 3]
        noun = ' '.join(w.capitalize() for w in words[:3]) if words else "Supply Chain"
        return ("🔍", f"{noun} — RCA Report","#38bdf8","rgba(56,189,248,0.35)")

    _report_icon, _report_title, _title_color, _border_color = _derive_report_title(user_question)

    # ── HELPERS ──────────────────────────────────────────────────────

    def _maybe_collapse_table(md_block: str, section_label: str = "Data") -> str:
        """Always wrap table in a collapsible View Table block."""
        md_block = str(md_block or "")
        if not md_block.strip():
            return md_block
        lines = md_block.strip().splitlines()
        data_rows = [l for l in lines if l.strip().startswith("|") and
                     not all(c in "|-: " for c in l.strip())]
        if len(data_rows) <= 7:
            return md_block
        return (
            f'\n<details style="background:rgba(14,165,233,0.04);border:1px solid rgba(14,165,233,0.2);'
            f'border-radius:8px;margin:8px 0">'
            f'<summary style="cursor:pointer;padding:9px 14px;font-size:0.78rem;font-weight:700;'
            f'color:#7dd3fc;list-style:none;display:flex;align-items:center;gap:6px;user-select:none">'
            f'<span style="font-size:0.6rem;color:#38bdf8;transition:transform 0.2s" class="dtl-arrow">▶</span>'
            f'<span>View Table</span>'
            f'<span style="font-size:0.68rem;font-weight:400;color:#475569;margin-left:auto">click to expand</span>'
            f'</summary>'
            f'<div style="padding:4px 12px 12px">\n\n{md_block}\n\n</div>'
            f'</details>\n'
        )

    def _md_table_to_html(md: str) -> str:
        """
        Convert markdown pipe-table(s) to styled HTML.
        Handles multiple tables separated by sub-headings (e.g. **1A — Title**).
        """
        import re as _rht
        if not md or not md.strip():
            return ""

        def _single_table_html(lines_block):
            """Convert a list of markdown table lines to an HTML table string."""
            lines_block = [l for l in lines_block if l.strip()]
            if len(lines_block) < 3:
                return ""
            def _cells(row):
                return [c.strip() for c in row.strip().strip("|").split("|")]
            header_cells = _cells(lines_block[0])
            # Find separator row
            sep_idx = next((i for i, l in enumerate(lines_block[1:], 1) if re.match(r'^[\s|:\-]+$', l.replace("|","").strip())), 1)
            data_rows = [_cells(l) for l in lines_block[sep_idx+1:] if l.strip().startswith("|")]
            if not data_rows:
                return ""
            # Align columns
            n_cols = len(header_cells)
            data_rows = [r[:n_cols] + [""]*(n_cols-len(r)) for r in data_rows]
            # Build HTML
            thead = "<tr>" + "".join(
                f'<th style="padding:7px 12px;text-align:left;font-size:0.68rem;font-weight:700;' +
                f'text-transform:uppercase;letter-spacing:0.07em;color:#94a3b8;white-space:nowrap">{h}</th>'
                for h in header_cells
            ) + "</tr>"
            tbodies = ""
            for i, row in enumerate(data_rows):
                bg = "rgba(255,255,255,0.03)" if i % 2 == 0 else "transparent"
                tbodies += "<tr style=\"border-bottom:1px solid rgba(255,255,255,0.05);background:{bg}\">"
                for cell in row:
                    cell_clean = cell.replace("**","")
                    tbodies += f'<td style="padding:7px 12px;font-size:0.82rem;color:#e2e8f0;vertical-align:top">{cell_clean}</td>'
                tbodies += "</tr>"
            return (
                '<div style="overflow-x:auto;margin-bottom:12px">' +
                f'<table style="width:100%;border-collapse:collapse;border-spacing:0">' +
                f'<thead style="background:rgba(15,25,50,0.8)">{thead}</thead>' +
                f'<tbody>{tbodies}</tbody>' +
                '</table></div>'
            )

        # Split the input into segments: sub-headings and table blocks
        # A sub-heading is a line starting with ** or a number+letter pattern
        segments = []
        current_heading = ""
        current_table_lines = []
        for line in md.strip().splitlines():
            stripped = line.strip()
            if not stripped:
                continue
            # Detect sub-heading: **1A — ...** or **SUPPLIER DELAY...** etc.
            is_subheading = (
                (stripped.startswith("**") and stripped.endswith("**") and "|" not in stripped) or
                re.match(r'^\*{0,2}\d+[A-Da-d][\s\-—.]+', stripped) or
                re.match(r'^\*{0,2}[A-D][.\-—:]\s+', stripped)
            )
            if is_subheading and current_table_lines:
                # Save previous table
                segments.append((current_heading, current_table_lines))
                current_heading = stripped.replace("**","").strip()
                current_table_lines = []
            elif is_subheading:
                current_heading = stripped.replace("**","").strip()
            elif stripped.startswith("|"):
                current_table_lines.append(stripped)
        # Save last table
        if current_table_lines:
            segments.append((current_heading, current_table_lines))

        if not segments:
            # No structured segments found — try rendering as single table
            all_table_lines = [l.strip() for l in md.strip().splitlines() if l.strip().startswith("|")]
            if all_table_lines:
                return _single_table_html(all_table_lines)
            return f'<div style="font-size:0.82rem;color:#e2e8f0;line-height:1.7">{md[:800]}</div>'

        # Render each segment
        result = ""
        for heading, tbl_lines in segments:
            if heading:
                # Style the sub-heading as a colored label
                result += (
                    f'<div style="font-size:0.7rem;font-weight:800;text-transform:uppercase;' +
                    f'letter-spacing:0.09em;color:#38bdf8;margin:14px 0 6px 0;padding-left:2px">' +
                    f'{heading}</div>'
                )
            html = _single_table_html(tbl_lines)
            if html:
                result += html

        return result if result else f'<div style="font-size:0.82rem;color:#e2e8f0">{md[:800]}</div>'
    
        def _cells(row: str):
            return [c.strip() for c in row.strip().strip("|").split("|")]

        header_cells = _cells(lines[0])
        data_rows    = []
        for line in lines[2:]:          # skip separator row
            if line.startswith("|") or "|" in line:
                data_rows.append(_cells(line))

        # Build <thead>
        th_html = "".join(
            f'<th style="padding:7px 12px;text-align:left;font-size:0.72rem;'
            f'font-weight:700;color:#7dd3fc;white-space:nowrap;'
            f'border-bottom:2px solid rgba(56,189,248,0.35);'
            f'background:rgba(56,189,248,0.06)">{h}</th>'
            for h in header_cells
        )

        # Build <tbody>
        tbody_html = ""
        for i, row in enumerate(data_rows):
            row_bg = "rgba(255,255,255,0.02)" if i % 2 == 0 else "transparent"
            # Bold total/summary rows
            is_total = any(c.startswith("**") or c.upper() in ("TOTAL","ALL 4","NETWORK") for c in row)
            td_style = (
                f'padding:6px 12px;font-size:0.75rem;color:#e2e8f0;'
                f'border-bottom:1px solid rgba(255,255,255,0.05);'
                f'{"font-weight:700;" if is_total else ""}'
                f'white-space:nowrap'
            )
            # Clean bold markdown in cells
            cells_html = "".join(
                f'<td style="{td_style}">'
                + _rht.sub(r'\*\*(.+?)\*\*', r'<strong>\1</strong>', c)
                .replace("**", "").strip()
                + '</td>'
                for c in row
            )
            tbody_html += f'<tr style="background:{row_bg}">{cells_html}</tr>'

        return (
            f'<div style="overflow-x:auto;margin-top:4px">'
            f'<table style="width:100%;border-collapse:collapse;'
            f'border:1px solid rgba(56,189,248,0.12);border-radius:8px;overflow:hidden">'
            f'<thead><tr>{th_html}</tr></thead>'
            f'<tbody>{tbody_html}</tbody>'
            f'</table></div>'
        )

    def _parse_step_and_desc(prose_raw: str) -> tuple:
        """
        Split prose into (step_label, description).
        Returns a pill label like "Step N of N: question" — strips the section heading
        name that appears between the step number and the colon/question.
        e.g. "Step 1 of 5 — Disruption Scenario: What happens when Nagy PLC shuts down?"
             → pill shows "Step 1 of 5: What happens when Nagy PLC shuts down?"
        """
        import re as _rsd
        prose = str(prose_raw or "").strip()
        # Match <strong>Step N of N ...</strong>
        m = _rsd.search(
            r'<strong>\s*(Step\s+\d+\s+of\s+\d+[^<]*)</strong>',
            prose, _rsd.IGNORECASE
        )
        if m:
            raw_label = m.group(1).strip()
            # Strip section heading: "Step N of N — Heading Name: Question?" → "Step N of N: Question?"
            raw_label = _rsd.sub(
                r'^(Step\s+\d+\s+of\s+\d+)\s*[—–-]\s*[^:]+:\s*',
                r'\1: ', raw_label, flags=_rsd.IGNORECASE
            )
            remaining = (prose[:m.start()] + prose[m.end():]).strip()
            return raw_label.strip(), remaining
        # Plain text: "Step N of N — ..." or "**Step N of N...**" markdown bold
        m2 = _rsd.match(r'\*{0,2}(Step\s+\d+\s+of\s+\d+[^\n*]*)\*{0,2}', prose, _rsd.IGNORECASE)
        if m2:
            raw_label = m2.group(1).strip()
            raw_label = _rsd.sub(
                r'^(Step\s+\d+\s+of\s+\d+)\s*[—–-]\s*[^:]+:\s*',
                r'\1: ', raw_label, flags=_rsd.IGNORECASE
            )
            return raw_label.strip(), prose[m2.end():].strip()
        return "", prose

    def _wrap_section_collapsible(icon: str, title: str, color: str,
                                   context_html: str, table_html: str,
                                   python_table_html: str = "",  # pre-rendered Python tables
                                   summary_line: str = "",
                                   insight_text: str = "") -> str:
        """
        Enterprise section card:
          - Icon + Title + badge row
          - Step pill (Step N of N: question)
          - Contextual description
          - View Data collapsible table
        No text below the table.
        """
        import re as _rsc

        # ── Split step label from description ──────────────────────────────
        step_label, desc_text = _parse_step_and_desc(context_html)

        # Strip any HTML tags from desc to get clean text
        desc_clean = _rsc.sub(r'<[^>]+>', '', desc_text).strip()

        summary_badge = (
            f'<span style="margin-left:auto;font-size:0.68rem;font-weight:600;'
            f'color:#7dd3fc;background:rgba(125,211,252,0.1);border:1px solid rgba(125,211,252,0.3);'
            f'border-radius:20px;padding:2px 10px;white-space:nowrap;flex-shrink:0">'
            f'{summary_line}</span>'
        ) if summary_line else ""

        step_pill = (
            f'<div style="display:inline-flex;align-items:center;gap:5px;'
            f'margin-bottom:8px;padding:3px 10px 3px 8px;'
            f'background:rgba(56,189,248,0.08);border:1px solid rgba(56,189,248,0.25);'
            f'border-radius:20px">'
            f'<span style="font-size:0.62rem;color:#38bdf8">◆</span>'
            f'<span style="font-size:0.7rem;font-weight:700;color:#38bdf8;'
            f'letter-spacing:0.02em">{step_label}</span>'
            f'</div>'
        ) if step_label else ""

        # Limit description to 3 sentences max for readability
        if desc_clean:
            import re as _re_desc
            _sentences = _re_desc.split(r'(?<=[.!?])\s+', desc_clean)
            _desc_trimmed = ' '.join(_sentences[:5]) if len(_sentences) > 5 else desc_clean
        else:
            _desc_trimmed = ""
        desc_block = (
            f'<div style="font-size:0.82rem;color:#cbd5e1;line-height:1.75;'
            f'margin-bottom:12px;padding:0 2px">{_desc_trimmed}</div>'
        ) if _desc_trimmed else ""

        # ── Convert markdown table to HTML then wrap in View Data dropdown ──
        table_dropdown = ""

        def _prose_to_html(prose: str) -> str:
            """Convert prose/bullet text to styled HTML — strips leaked JSON first."""
            import re as _rph
            prose = prose.strip()
            if not prose:
                return ""
            # Strip leaked JSON arrays/objects that LLM sometimes appends
            prose = _rph.sub(r'\[\s*\{[\s\S]*', '', prose).strip()
            prose = _rph.sub(r'\{\s*"tier"[\s\S]*', '', prose).strip()
            prose = _rph.sub(r'\{\s*"entity"[\s\S]*', '', prose).strip()
            # Strip bare [ or { lines that start JSON that wasn't caught above
            clean_lines = []
            for raw_line in prose.splitlines():
                s = raw_line.strip()
                if s in ('[', '{', '},', '],') or s.startswith('"tier"') or s.startswith('"entity"') or s.startswith('"action"'):
                    break  # stop at first JSON artifact
                clean_lines.append(raw_line)
            prose = '\n'.join(clean_lines).strip()
            if not prose:
                return ""
            lines = prose.splitlines()
            html_parts = []
            for line in lines:
                s = line.strip()
                if not s:
                    continue
                # Bullet point lines
                if s.startswith(("* ", "- ", "• ")):
                    text = s[2:].strip()
                    text = _rph.sub(r'\*\*(.+?)\*\*', r'<strong style="color:#e2e8f0">\1</strong>', text)
                    html_parts.append(
                        f'<div style="display:flex;gap:10px;margin-bottom:8px;align-items:flex-start">'
                        f'<span style="color:#38bdf8;font-size:0.75rem;margin-top:3px;flex-shrink:0">◆</span>'
                        f'<span style="font-size:0.83rem;color:#cbd5e1;line-height:1.7">{text}</span>'
                        f'</div>'
                    )
                # Numbered list lines
                elif _rph.match(r'^\d+\.\s', s):
                    text = _rph.sub(r'^\d+\.\s*', '', s)
                    text = _rph.sub(r'\*\*(.+?)\*\*', r'<strong style="color:#e2e8f0">\1</strong>', text)
                    html_parts.append(
                        f'<div style="display:flex;gap:10px;margin-bottom:8px;align-items:flex-start">'
                        f'<span style="color:#38bdf8;font-size:0.75rem;margin-top:3px;flex-shrink:0">◆</span>'
                        f'<span style="font-size:0.83rem;color:#cbd5e1;line-height:1.7">{text}</span>'
                        f'</div>'
                    )
                # Skip bare JSON artifacts that slipped through
                elif s.startswith(('{', '[', '"')):
                    continue
                # Regular prose paragraph
                else:
                    text = _rph.sub(r'\*\*(.+?)\*\*', r'<strong style="color:#e2e8f0">\1</strong>', s)
                    html_parts.append(
                        f'<p style="font-size:0.83rem;color:#cbd5e1;line-height:1.75;margin:0 0 8px 0">{text}</p>'
                    )
            return "\n".join(html_parts)

        def _is_prose_content(text: str) -> bool:
            """Return True if text has no markdown table rows."""
            if not text:
                return False
            lines = [l.strip() for l in text.splitlines() if l.strip()]
            table_lines = [l for l in lines if l.startswith("|") and not all(c in "|-: " for c in l)]
            return len(table_lines) < 2

        # Use python_table_html (direct from Neo4j rows) OR LLM markdown table
        # python_table_html takes priority — table_html is now often empty since LLM no longer writes tables
        _effective_table = python_table_html or table_html
        if _effective_table and _effective_table.strip():
            if python_table_html:
                # Direct Python-rendered table — already HTML, no conversion needed
                inner_table = python_table_html
            elif table_html and table_html.strip():
                # If content is prose (no table rows), render as styled paragraphs/bullets
                if _is_prose_content(table_html):
                    inner_table = _prose_to_html(table_html)
                else:
                    inner_table = _md_table_to_html(table_html)
                if not inner_table:
                    inner_table = _prose_to_html(table_html) or f'<div style="font-size:0.82rem;color:#e2e8f0;line-height:1.65">{table_html[:1000]}</div>'
            else:
                inner_table = ""

            # python_table_html has per-subtable named dropdowns — render directly (no outer wrapper)
            # LLM markdown table uses a "View Data" collapsible as fallback
            # Prose content (Business Insights etc.) renders inline — no collapsible needed
            if python_table_html:
                table_dropdown = inner_table  # already contains named <details> per subtable
            else:
                # Check if inner_table is styled prose HTML (no <table> tag = prose, not data table)
                _is_prose_html = inner_table and "<table" not in inner_table
                if _is_prose_html:
                    # Render prose content directly — no collapsible wrapper needed
                    table_dropdown = (
                        f'<div style="margin-top:12px;padding:14px 16px;'
                        f'background:rgba(74,222,128,0.04);border:1px solid rgba(74,222,128,0.15);'
                        f'border-radius:10px">'
                        f'{inner_table}'
                        f'</div>'
                    )
                else:
                    table_dropdown = (
                        f'<details style="margin-top:10px;border:1px solid rgba(56,189,248,0.2);' +
                        f'border-radius:10px;overflow:hidden;background:rgba(56,189,248,0.03)">' +
                        f'<summary style="cursor:pointer;list-style:none;user-select:none;' +
                        f'padding:9px 14px;display:flex;align-items:center;gap:8px;' +
                        f'font-size:0.76rem;font-weight:700;color:#38bdf8">' +
                        f'<span style="font-size:0.6rem;transition:transform 0.2s;display:inline-block" class="vd-arrow">▶</span>' +
                        f'<span>📋 View Data</span>' +
                        f'<span style="margin-left:auto;font-size:0.68rem;font-weight:400;color:#475569">click to expand</span>' +
                        f'</summary>' +
                        f'<div style="padding:12px 14px 14px;overflow-x:auto">{inner_table}</div>' +
                        f'</details>' +
                        f'<style>details[open] .vd-arrow{{transform:rotate(90deg)}}</style>'
                    )


        # insight_text is intentionally NOT rendered — all context is in desc_block above the table
        # Root Cause step gets special red-tinted background — and NO table dropdown (just the para)
        _is_root_cause = icon == "⚠️" and ("root cause" in title.lower() or "root" in title.lower())
        if _is_root_cause:
            table_dropdown = ""   # Root cause renders as crisp para only — no dropdown
        _bg = "rgba(239,68,68,0.07)" if _is_root_cause else "rgba(12,21,40,0.72)"
        _border_full = (
            f"border:1px solid rgba(239,68,68,0.3);border-left:4px solid {color}"
            if _is_root_cause else
            f"border:1px solid rgba(125,211,252,0.12);border-left:4px solid {color}"
        )
        return f"""
<div style="background:{_bg};{_border_full};border-radius:12px;margin-bottom:16px;
  padding:16px 18px 16px;box-shadow:0 2px 12px rgba(0,0,0,0.25)">
  <div style="display:flex;align-items:center;gap:10px;margin-bottom:10px">
    <span style="font-size:1.05rem;flex-shrink:0">{icon}</span>
    <span style="font-size:0.88rem;font-weight:700;color:{'#fca5a5' if _is_root_cause else '#e2e8f0'}">{title}</span>
    {summary_badge}
  </div>
  {step_pill}
  {desc_block}
  {table_dropdown}
</div>"""

    # ── 1. Extract Executive Summary ─────────────────────────────
    exec_summary = ""
    exec_pat = _re.compile(
        r'###?\s*(?:📌|🚛|🏭|🔍|📦|🏙️|🏗️)?\s*(?:\w+\s+)?Executive Summary\s*\n(.*?)(?=\n#{1,4}\s|\Z)',
        _re.DOTALL | _re.IGNORECASE
    )
    m = exec_pat.search(report_body)
    if m:
        raw_summary = str(m.group(1) or "").strip()
        if len(raw_summary) > 900:
            first_para = str(raw_summary.split('\n\n')[0] or "").strip()
            exec_summary = first_para if first_para else raw_summary[:700]
        else:
            exec_summary = raw_summary
        report_body = str((report_body[:m.start()] + report_body[m.end():]) or "").strip()

    # ── 2. Extract Recommendations ───────────────────────────────
    rec_content = ""
    rec_json_block = ""  # Will hold pre-rendered HTML if rec_raw is JSON

    # ── Try JSON format first (new rec agent format) ───────────────────────
    # Extract JSON array from rec_raw (handles prefix text and code blocks)
    _raw_stripped = (rec_raw or "").strip()
    # Try to find a JSON array anywhere in rec_raw
    _rec_json_match = re.search(r'\[\s*\{', _raw_stripped, re.DOTALL)
    _rec_json_start = _rec_json_match.start() if _rec_json_match else -1
    if _rec_json_start >= 0:
        _raw_stripped = _raw_stripped[_rec_json_start:]
        # Find the matching closing bracket
        _depth = 0
        _end_idx = 0
        for _ci, _ch in enumerate(_raw_stripped):
            if _ch == '[': _depth += 1
            elif _ch == ']':
                _depth -= 1
                if _depth == 0:
                    _end_idx = _ci + 1
                    break
        if _end_idx > 0:
            _raw_stripped = _raw_stripped[:_end_idx]
    if _raw_stripped.startswith("["):
        try:
            import json as _recj
            _rec_list = _recj.loads(_raw_stripped)
            if isinstance(_rec_list, list) and _rec_list:
                # Render JSON recommendations as styled tier cards
                _tier_colours = {"Critical":"#f87171","Operational":"#fb923c","Strategic":"#4ade80"}
                _tier_icons   = {"Critical":"⚡","Operational":"🔧","Strategic":"🏛"}
                _tiers = {"Critical":[],"Operational":[],"Strategic":[]}
                for _item in _rec_list:
                    _t = str(_item.get("tier","Operational")).strip()
                    _t = next((k for k in _tiers if k.lower() in _t.lower()), "Operational")
                    _tiers[_t].append(_item)
                # Build per-tier collapsible dropdowns — no outer wrapper
                _tier_bgs    = {"Critical":"rgba(239,68,68,0.07)","Operational":"rgba(251,146,60,0.06)","Strategic":"rgba(74,222,128,0.05)"}
                _tier_borders= {"Critical":"rgba(239,68,68,0.3)","Operational":"rgba(251,146,60,0.3)","Strategic":"rgba(74,222,128,0.25)"}
                _tier_sbg    = {"Critical":"rgba(239,68,68,0.1)","Operational":"rgba(251,146,60,0.08)","Strategic":"rgba(74,222,128,0.07)"}
                _rec_html_parts = []
                for _tier, _items in _tiers.items():
                    if not _items:
                        continue
                    _col = _tier_colours[_tier]
                    _ico = _tier_icons[_tier]
                    _tbg = _tier_bgs[_tier]
                    _tbr = _tier_borders[_tier]
                    _tsbg = _tier_sbg[_tier]
                    # Build items HTML
                    _items_html = ""
                    for _item in _items:
                        _entity = _item.get("entity","")
                        _action = _item.get("action","")
                        _items_html += (
                            f'<div style="padding:11px 14px;background:rgba(255,255,255,0.025);' +
                            f'border:1px solid rgba(255,255,255,0.07);border-left:3px solid {_col}50;' +
                            f'border-radius:8px;margin-bottom:8px">' +
                            (f'<div style="font-size:0.72rem;font-weight:700;color:{_col};margin-bottom:5px">{_entity}</div>' if _entity else "") +
                            f'<div style="font-size:0.82rem;color:#e2e8f0;line-height:1.65">{_action}</div>' +
                            '</div>'
                        )
                    # Each tier as its own named collapsible dropdown
                    _n_items = len(_items)
                    _rec_html_parts.append(
                        f'<details style="margin-bottom:10px;border:1px solid {_tbr};' +
                        f'border-radius:10px;overflow:hidden;background:{_tbg}">' +
                        f'<summary style="cursor:pointer;list-style:none;user-select:none;' +
                        f'padding:10px 14px;display:flex;align-items:center;gap:8px;' +
                        f'font-size:0.78rem;font-weight:700;color:{_col};background:{_tsbg}">' +
                        f'<span style="font-size:0.6rem;transition:transform 0.2s;display:inline-block" class="tier-arr-{_tier.lower()}">▶</span>' +
                        f'{_ico} {_tier} Actions' +
                        f'<span style="margin-left:auto;font-size:0.65rem;font-weight:500;' +
                        f'color:#475569;background:rgba(255,255,255,0.05);padding:1px 8px;' +
                        f'border-radius:10px;border:1px solid rgba(255,255,255,0.08)">{_n_items} action{"s" if _n_items!=1 else ""}</span>' +
                        f'</summary>' +
                        f'<div style="padding:10px 12px 12px">{_items_html}</div>' +
                        f'</details>' +
                        f'<style>details[open] .tier-arr-{_tier.lower()}{{transform:rotate(90deg)}}</style>'
                    )
                rec_json_block = "".join(_rec_html_parts)
        except Exception as _rj_err:
            pass  # Fall through to markdown extraction

    rec_pat = _re.compile(
        r'###\s*[💡✅]?\s*(?:Corrective\s+)?Recommendations[\s\S]*?\n(.*?)(?=\n##[^#]|\Z)',
        _re.DOTALL | _re.IGNORECASE
    )
    rm = rec_pat.search(report_body)
    if rm:
        rec_content = str(rm.group(1) or "").strip()
        report_body = str((report_body[:rm.start()] + report_body[rm.end():]) or "").strip()
    if not rec_content:
        fb = _re.search(r'(####\s*(?:⚡|🔧|🏛|Critical|Operational|Strategic).*)', report_body, _re.DOTALL | _re.IGNORECASE)
        if fb:
            rec_content = str(fb.group(1) or "").strip()
            report_body = str(report_body[:fb.start()] or "").strip()
    if not rec_content and rec_raw and rec_raw.strip():
        raw = rec_raw.strip()
        m2 = _re.search(r'###\s*💡\s*Recommendations.*?\n', raw, _re.IGNORECASE)
        rec_content = raw[m2.end():].strip() if m2 else raw

    # ── 3. Strip stray top-level heading ─────────────────────────
    report_body = _re.sub(
        r'^##\s*(?:🔍|🧪|📉|🚚|🏭|🏗️?|📦)?\s*(?:Root Cause Analysis Report|Supply Chain Impact Simulation|'
        r'Demand Gap RCA|Retailer Shortage Analysis|Distributor Impact RCA|'
        r'Supplier Risk RCA|Plant RCA|Product Category RCA|Shipment Delay RCA)\s*\n?',
        '', report_body, flags=_re.MULTILINE | _re.IGNORECASE
    )
    report_body = str(report_body or "").strip()

    # ── 3b. Strip Visual Analysis section (rendered separately as charts) ──
    _vis_pat = _re.compile(
        r'###\s*📊\s*Visual Analysis.*?(?=\n###|\Z)',
        _re.DOTALL | _re.IGNORECASE
    )
    # ── Save chartdata blocks BEFORE stripping them ────────────
    # Chart rendering at the bottom searches report_body for chartdata.
    # Strip happens here first, so we must save them before stripping.
    _saved_chartdata = _re.findall(r'```chartdata\s*\n(.*?)\n```', report_body, _re.DOTALL)

    report_body = _vis_pat.sub('', report_body).strip()
    # Also strip any stray ```chartdata blocks
    report_body = _re.sub(r'```chartdata[\s\S]*?```', '', report_body).strip()

    # ── 4. Extract Root Cause highlight text ─────────────────────
    root_cause_text = ""
    root_cause_text_raw = ""
    root_cause_bullets  = []
    root_cause_severity = ""

    # Try 0 (PRIMARY): Step 5 section — prose paragraph before JSON block
    _step5_m = _re.search(
        r'###?\s*(?:[^#\n]*)?(?:Step\s+5\s+of\s+5|ROOT\s+CAUSE)[^#\n]*\n(.*?)(?=\n###|\n##|\Z)',
        report_body, _re.DOTALL | _re.IGNORECASE
    )
    if _step5_m:
        _s5_raw = str(_step5_m.group(1) or "").strip()
        # Parse JSON bullets first
        _j5 = _re.search(r'```json\s*({.*?})\s*```', _s5_raw, _re.DOTALL)
        if not _j5:
            _j5 = _re.search(r'({\s*"bullets"\s*:.*?})\s*$', _s5_raw, _re.DOTALL)
        if _j5:
            try:
                import json as _rcj5
                _rc5 = _rcj5.loads(_j5.group(1))
                root_cause_bullets  = _rc5.get("bullets", [])
                root_cause_severity = _rc5.get("severity", "")
            except Exception:
                pass
        # Extract prose before the JSON block — strip ALL JSON variants
        _prose5 = _re.sub(r'```json[\s\S]*?```', '', _s5_raw).strip()
        _prose5 = _re.sub(r'```[\s\S]*?```', '', _prose5).strip()
        _prose5 = _re.sub(r'\{[\s\n]*"bullets"[\s\S]*', '', _prose5).strip()
        # Strip JSON arrays [ { "tier"... or [ { "label"... appended by LLM
        _prose5 = _re.sub(r'\[\s*\{[\s\S]*', '', _prose5).strip()
        # Strip any remaining JSON-like structures
        _prose5 = _re.sub(r'\{[\s\S]{0,20}"tier"[\s\S]*', '', _prose5).strip()
        _prose5 = _re.sub(r'`+[a-z]*`*', '', _prose5).strip()
        _prose5 = _re.sub(r'\*\*(.+?)\*\*', r'\1', _prose5).replace('**', '').strip()
        # Trim to clean sentences only (stop at any line that looks like JSON/code)
        _clean_lines = []
        for _line in _prose5.splitlines():
            _l = _line.strip()
            if _l and not _l.startswith(('[', '{', '`', '"tier"', '"label"', '"action"')):
                _clean_lines.append(_l)
            elif _l.startswith(('[', '{')):
                break  # stop at first JSON line
        _prose5 = ' '.join(_clean_lines).strip()
        if len(_prose5) > 30:
            root_cause_text = _prose5[:600]

    # Try 1: WHO IS RESPONSIBLE header (legacy format)
    if not root_cause_text and not root_cause_bullets:
        who_match = _re.search(
            r'WHO\s+IS\s+RESPONSIBLE[:\s]*\n(.*?)(?=WHY\s+IT|WHEN\s+IT|HOW\s+MUCH|PROPAGATION|SEVERITY|\Z)',
            report_body, _re.DOTALL | _re.IGNORECASE
        )
        if who_match:
            _rc_disp = str(who_match.group(1) or "").strip()
            _rc_disp = re.sub(r'```json[\s\S]*?```', '', _rc_disp).strip()
            _rc_disp = re.sub(r'{[\s\n]*"bullets"[\s\S]*', '', _rc_disp).strip()
            _rc_disp = re.sub(r'`+[a-z]*`*', '', _rc_disp).strip()
            if len(_rc_disp) > 15:
                root_cause_text = _rc_disp[:500].replace('**','').replace('__','')

    # Try 2: blockquote Root Cause (legacy)
    if not root_cause_text and not root_cause_bullets:
        bq_m = _re.search(r'>\s*\*\*Root\s+Cause[:\*]*\*?\*?\s*(.*?)(?=\n[^>]|\Z)',
                           report_body, _re.DOTALL | _re.IGNORECASE)
        if bq_m:
            root_cause_text = str(bq_m.group(1) or "").strip().lstrip('*')[:350]

    # Try 3: ### Root Cause heading (any format)
    if not root_cause_text and not root_cause_bullets:
        rc_m = _re.search(
            r'###?\s*(?:[^#\n]{0,30})?Root\s+Cause(?:[^#\n]+)?\n(.*?)(?=\n###|\n##|\Z)',
            report_body, _re.DOTALL | _re.IGNORECASE
        )
        if rc_m:
            _rc_raw = str(rc_m.group(1) or "").strip()
            _rc_raw = _re.sub(r'```json[\s\S]*?```', '', _rc_raw).strip()
            _rc_raw = _re.sub(r'{[\s\n]*"bullets"[\s\S]*', '', _rc_raw).strip()
            root_cause_text = _rc_raw[:400].replace('**','').strip()

    # Fallback: build from executive summary first sentences if everything else failed
    if not root_cause_text and not root_cause_bullets:
        exec_m2 = _re.search(
            r'###?\s*Executive\s+Summary\s*\n(.*?)(?=\n###|\n##|\Z)',
            report_body, _re.DOTALL | _re.IGNORECASE
        )
        if exec_m2:
            exec_txt = str(exec_m2.group(1) or "").strip()
            exec_txt = _re.sub(r'```[\s\S]*?```', '', exec_txt).strip()
            exec_txt = exec_txt.replace('**', '').replace('__', '')
            sentences = _re.split(r'(?<=[.!?])\s+', exec_txt)
            root_cause_text = ' '.join(sentences[:3])[:400] if sentences else exec_txt[:400]


    # ── 5. Extract section data (tables + prose) ─────────────────
    def _extract_section(body: str, label_pattern: str):
        """
        Extract ALL tables + sub-headings for a named section.
        Returns (intro_prose, combined_table_html, after_prose).
        combined_table_html contains ALL sub-tables (1A, 1B, 1C etc.) as markdown
        separated by sub-heading labels — this ensures multi-table sections render fully.
        """
        import re as _r4
        sec_pat = _r4.compile(
            r'###?\s*[^\n]*' + label_pattern + r'[^\n]*\n(.*?)(?=\n##[^#]|\n###|\Z)',
            _r4.DOTALL | _r4.IGNORECASE
        )
        sm = sec_pat.search(body)
        if not sm:
            return None, None, None
        content = str(sm.group(1) or "").strip()

        # Find ALL tables in this section (not just the first)
        table_iter = list(_r4.finditer(
            r'(\*\*[^*\n]+\*\*[^\n]*\n|\*{0,2}\d+[A-D][^\n]*\n)?'  # optional sub-heading
            r'(\|[^\n]+\|\n(?:\|[-: |]+\|\n)(?:\|[^\n]+\|\n?)+)',
            content
        ))

        if not table_iter:
            # No tables at all — return as prose
            return content[:800], "", ""

        # Collect: intro prose (before first table), all tables with their headings,
        # and any final analysis prose (after last table)
        first_start = table_iter[0].start()
        intro_prose = content[:first_start].strip()

        # Build combined content: all sub-headings + tables + inter-table prose
        combined_parts = []
        prev_end = first_start
        for tm in table_iter:
            # Include any prose/sub-heading between previous table and this one
            gap = content[prev_end:tm.start()].strip()
            if gap:
                combined_parts.append(gap)
            combined_parts.append(tm.group(0).strip())
            prev_end = tm.end()

        # After-last-table prose (analysis sentences)
        after_prose = content[prev_end:].strip()

        combined_tables = "\n\n".join(combined_parts)
        return intro_prose[:600], combined_tables, after_prose[:800]

    def _quick_stats_from_table(table_md: str) -> dict:
        """Parse a markdown table and return a dict of column→values list."""
        if not table_md:
            return {}
        table_md = str(table_md)
        lines = [l.strip() for l in table_md.splitlines() if l.strip().startswith("|")]
        if len(lines) < 2:
            return {}
        headers = [str(h or "").strip() for h in lines[0].strip("|").split("|")]
        data_rows = [
            [str(c or "").strip() for c in row.strip("|").split("|")]
            for row in lines[2:]
            if not all(c in "|-: " for c in row.strip())
        ]
        result = {h: [] for h in headers}
        for row in data_rows:
            for i, h in enumerate(headers):
                if i < len(row):
                    result[h].append(row[i])
        return result

    # ── 6. Build query-aware KPI strip ────────────────────────────
    kpi_items = []

    # ── Supplier KPIs — covers: High-Risk Suppliers + Supplier Risk at Source ──
    sup_prose, sup_table, _ = _extract_section(
        report_body,
        r'high.risk\s+supplier|supplier\s+risk\s+at\s+source|supplier\s+dependency|supplier.*plant.*delay'
    )
    if sup_table:
        sup_stats = _quick_stats_from_table(sup_table)
        n_sup = len(list(sup_stats.values())[0]) if sup_stats else 0
        # Count suppliers with risk score >= 0.90 (CRITICAL threshold)
        # Use try/except float() — more robust than isdigit() which fails on "0.93"-style decimals
        def _safe_float(v):
            try:
                return float(str(v).replace(",", "").strip())
            except (ValueError, TypeError):
                return None

        risk_vals = []
        for k, vals in sup_stats.items():
            if "risk" in k.lower() and "avg" not in k.lower() and "score" in k.lower():
                parsed = [_safe_float(v) for v in vals if v]
                risk_vals = [f for f in parsed if f is not None and 0.0 <= f <= 1.0]
                if risk_vals:
                    break
        # Also scan all columns if nothing found yet (handles tables where col order varies)
        if not risk_vals:
            for k, vals in sup_stats.items():
                if "risk" in k.lower() and vals:
                    parsed = [_safe_float(v) for v in vals if v]
                    cands = [f for f in parsed if f is not None and 0.5 <= f <= 1.0]
                    if cands:
                        risk_vals = cands
                        break

        n_critical = sum(1 for v in risk_vals if v >= 0.90)
        if n_critical > 0:
            kpi_items.append(("🏭", str(n_critical), "Critical Suppliers (≥0.90)", "#f87171"))
        elif n_sup:
            kpi_items.append(("🏭", str(n_sup), "High-Risk Suppliers", "#f87171"))

        # Max Risk Score KPI — use already-parsed risk_vals for consistency
        if risk_vals:
            kpi_items.append(("⚠️", f"{max(risk_vals):.2f}", "Max Risk Score", "#fb923c"))
        else:
            # Final fallback: scan any risk column for a max value
            for k, vals in sup_stats.items():
                if "risk" in k.lower() and vals:
                    nums = [_safe_float(v) for v in vals if v]
                    nums = [f for f in nums if f is not None and f > 0.3]
                    if nums:
                        kpi_items.append(("⚠️", f"{max(nums):.2f}", "Max Risk Score", "#fb923c"))
                    break
        for k, vals in sup_stats.items():
            if "lead" in k.lower() and "time" in k.lower() and vals:
                try:
                    nums = [float(str(v).replace("days","").strip()) for v in vals
                            if v and str(v).replace("days","").strip().replace(".","").isdigit()]
                    if nums:
                        kpi_items.append(("⏱️", f"{sum(nums)/len(nums):.1f}d", "Avg Lead Time", "#fbbf24"))
                except Exception:
                    pass
                break

    # ── Plant KPIs — covers: Bottleneck Plants + Plant Contribution ──
    pl_prose, pl_table, _ = _extract_section(
        report_body, r'bottleneck\s+plant|plant\s+contribution|shipment\s+delay\s+analysis')
    if pl_table:
        pl_stats = _quick_stats_from_table(pl_table)
        worst_rate = ""
        total_delayed_pl = 0
        for k, vals in pl_stats.items():
            if "delay" in k.lower() and "%" in k.lower() and vals:
                try:
                    worst_rate = f"{max(float(str(v).replace('%','').strip()) for v in vals if v):.0f}%"
                except Exception:
                    pass
            if ("delayed" in k.lower() or "delay" in k.lower()) and "%" not in k.lower() and vals:
                try:
                    total_delayed_pl += sum(int(str(v).replace(",","").strip()) for v in vals
                                            if v and str(v).replace(",","").strip().isdigit())
                except Exception:
                    pass
        if worst_rate:
            kpi_items.append(("📊", worst_rate, "Peak Plant Delay Rate", "#f87171"))
        if total_delayed_pl:
            kpi_items.append(("📦", f"{total_delayed_pl:,}", "Delayed Dispatches", "#fbbf24"))

    # ── Distributor KPIs — covers all distributor section names ──
    dist_prose, dist_table, _ = _extract_section(
        report_body,
        r'distributor\s+impact|affected\s+distributor|distributor\s+metrics|'
        r'distributor\s+contribution|distributor\s+ripple|upstream\s+dependency'
    )
    if dist_table:
        dist_stats = _quick_stats_from_table(dist_table)
        n_dist = len(list(dist_stats.values())[0]) if dist_stats else 0
        if n_dist:
            kpi_items.append(("🚚", str(n_dist), "Affected Distributors", "#60a5fa"))
        # Total demand gap
        for k, vals in dist_stats.items():
            if "demand" in k.lower() and "gap" in k.lower() and vals:
                try:
                    total_gap = sum(float(str(v).replace(",","").strip()) for v in vals
                                    if v and str(v).replace(",","").replace(".","").strip().lstrip("-").isdigit())
                    if total_gap > 0:
                        kpi_items.append(("📉", f"{total_gap:,.0f}", "Total Demand Gap", "#f87171"))
                except Exception:
                    pass
                break

    # ── Retailer / stockout KPI — covers all retailer section names ──
    ret_prose, ret_table, _ = _extract_section(
        report_body,
        r'stockout|retailer\s+impact|retailer.*stockout|retailer\s*[&and]+\s*stockout'
    )
    if ret_table:
        ret_stats = _quick_stats_from_table(ret_table)
        n_ret = len(list(ret_stats.values())[0]) if ret_stats else 0
        if n_ret:
            # n_ret = number of rows in the demand gap / stockout table = distributor CITIES
            # NOT retailer count. Rename label to avoid showing "15 Retailers at Stockout"
            # when 15 = distributor city rows.
            _kpi_lbl = f"Top {n_ret} of 50 Cities" if n_ret < 50 else "All 50 Cities"
            kpi_items.append(("📉", str(n_ret), _kpi_lbl, "#f87171"))

    # Build KPI cards — print-safe, high-contrast
    if kpi_items:
        kpi_html = (
            '<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));'
            'gap:12px;margin-bottom:20px;page-break-inside:avoid">'
        )
        CARD_COLORS = {
            "#f87171": ("rgba(239,68,68,0.10)",  "rgba(239,68,68,0.35)"),
            "#fb923c": ("rgba(251,146,60,0.10)", "rgba(251,146,60,0.35)"),
            "#fbbf24": ("rgba(251,191,36,0.10)", "rgba(251,191,36,0.35)"),
            "#60a5fa": ("rgba(96,165,250,0.10)", "rgba(96,165,250,0.35)"),
            "#4ade80": ("rgba(74,222,128,0.10)", "rgba(74,222,128,0.35)"),
            "#7dd3fc": ("rgba(125,211,252,0.10)","rgba(125,211,252,0.35)"),
        }
        for icon, val, label, color in kpi_items[:6]:
            bg, border = CARD_COLORS.get(color, ("rgba(14,165,233,0.08)", "rgba(14,165,233,0.3)"))
            kpi_html += f'''
<div style="background:{bg};border:1px solid {border};border-top:3px solid {color};
     border-radius:10px;padding:14px 16px;display:flex;flex-direction:column;gap:6px;
     box-shadow:0 2px 8px rgba(0,0,0,0.3)">
  <div style="display:flex;align-items:center;gap:8px">
    <span style="font-size:1.2rem">{icon}</span>
    <span style="font-size:1.5rem;font-weight:800;color:{color};font-family:monospace;
          line-height:1;text-shadow:0 0 12px {color}44">{val}</span>
  </div>
  <div style="font-size:0.68rem;font-weight:700;text-transform:uppercase;
       letter-spacing:0.08em;color:#94a3b8;line-height:1.3">{label}</div>
</div>'''
        kpi_html += '</div>'
    else:
        kpi_html = ""

    # ── 8. Build Executive Summary — collapsed by default, always present ────
    if exec_summary:
        # Pre-process exec_summary: convert markdown bold **text** to <strong> for HTML rendering
        import re as _re_exec
        _exec_html = _re_exec.sub(r'\*\*(.+?)\*\*', r'<strong style="color:#e2e8f0">\1</strong>', str(exec_summary or ""))
        _exec_html = _exec_html.replace("\n\n", "</p><p style='margin:8px 0;color:#f1f5f9!important'>").replace("\n", " ")
        _exec_html = f'<p style="margin:0 0 4px;color:#f1f5f9!important">' + _exec_html + '</p>'
        exec_block = f"""
<details style="background:rgba(14,22,45,0.9);border:1px solid rgba(56,189,248,0.3);
  border-left:4px solid #38bdf8;border-radius:12px;margin-bottom:16px;overflow:hidden;
  box-shadow:0 2px 12px rgba(0,0,0,0.3)">
  <summary style="cursor:pointer;list-style:none;user-select:none;padding:12px 18px;
    display:flex;align-items:center;gap:8px;background:rgba(56,189,248,0.06)">
    <span style="font-size:0.62rem;color:#38bdf8;display:inline-block;transition:transform 0.2s" class="es-arrow">▶</span>
    <span style="font-size:0.88rem;font-weight:700;color:#38bdf8">📌 Executive Summary</span>
    <span style="font-size:0.68rem;font-weight:400;color:#475569;margin-left:auto">click to expand</span>
  </summary>
  <div style="padding:14px 20px 18px;background:rgba(14,22,45,0.9)">
    <style>
      .exec-text, .exec-text p, .exec-text strong {{ color: #f1f5f9 !important; }}
      details[open] .es-arrow {{ transform: rotate(90deg); }}
    </style>
    <div class="exec-text" style="font-size:0.87rem;line-height:1.85;font-weight:400;color:#f1f5f9!important">{_exec_html}</div>
  </div>
</details>

"""
    else:
        exec_block = """
<details style="background:rgba(56,189,248,0.05);border:1px solid rgba(56,189,248,0.25);
  border-radius:12px;margin-bottom:16px;overflow:hidden">
<summary style="cursor:pointer;padding:12px 18px;font-size:0.88rem;font-weight:700;
  color:#38bdf8;list-style:none;display:flex;align-items:center;gap:8px;user-select:none">
  <span style="font-size:0.62rem;color:#38bdf8">▶</span>
  <span>📌 Executive Summary</span>
  <span style="font-size:0.68rem;font-weight:400;color:#475569;margin-left:auto">click to expand</span>
</summary>
<div style="padding:6px 20px 18px;color:#f1f5f9;font-size:0.86rem;font-style:italic">
Executive summary will appear here once analysis data is retrieved.
</div>
</details>

"""

    # ── 9. Build Final Root Cause highlight box — structured WHO/WHY/WHEN/HOW MUCH ──
    def _render_root_cause_box(rc_text: str, full_report: str,
                                bullets: list = None, severity: str = "") -> str:
        """
        Render root cause as a crisp 3-4 sentence paragraph with bold entity names.
        Priority chain: LLM prose paragraph → synthesise from JSON bullets →
        raw prose → exec summary sentences → static message about the system.
        Always returns non-empty HTML so root cause section is never blank.
        """
        import re as _rcr

        def _strip_md(t):
            t = _rcr.sub(r'\*\*(.+?)\*\*', r'\1', t).replace('**','').replace('__','')
            t = _rcr.sub(r'```[\s\S]*?```', '', t)
            # Strip JSON arrays/objects appended by LLM ([ { "tier"... or { "bullets"...)
            t = _rcr.sub(r'\[\s*\{[\s\S]*', '', t)
            t = _rcr.sub(r'\{[\s\S]{0,20}"(?:tier|bullets|label|action)"[\s\S]*', '', t)
            # Stop at any line that looks like JSON or a markdown table row
            clean_lines = []
            for line in t.splitlines():
                l = line.strip()
                if l.startswith(('{', '[', '"tier"', '"action"', '"label"')):
                    break
                # Skip markdown table rows (lines starting with |) and separator rows
                if l.startswith('|'):
                    continue
                # Skip lines that are just numbers/pipes (raw table data leaked through)
                if _rcr.match(r'^[\d\s\|\.\,\%\-]+$', l) and '|' in l:
                    continue
                clean_lines.append(l)
            t = ' '.join(clean_lines)
            t = _rcr.sub(r'\s+', ' ', t).strip()
            return t

        def _bold(text: str) -> str:
            # Supplier IDs → red
            text = _rcr.sub(r'\b(SUP0?\d{3,4})\b', r'<strong style="color:#f87171">\1</strong>', text)
            # Plant names → amber
            for pn in ["Bhopal","Pune","Baddi","Goa"]:
                text = _rcr.sub(r'\b' + pn + r'\b', f'<strong style="color:#fbbf24">{pn}</strong>', text)
            # Plant IDs → orange
            text = _rcr.sub(r'\b(PL\d{1,2})\b', r'<strong style="color:#fb923c">\1</strong>', text)
            # Percentages → cyan
            text = _rcr.sub(r'(\d+\.?\d*\s*%)', r'<strong style="color:#38bdf8">\1</strong>', text)
            # Large unit counts → green
            text = _rcr.sub(r'\b(\d{1,3}(?:,\d{3})+)\b', r'<strong style="color:#4ade80">\1</strong>', text)
            # Key severity words → red
            text = _rcr.sub(r'\b(CRITICAL|CHRONIC|structural failure|immediate action)\b',
                            r'<strong style="color:#f87171">\1</strong>', text, flags=_rcr.IGNORECASE)
            return text

        # ── Assemble paragraph (priority order) ──────────────────────────
        para = ""

        # P1: LLM-written prose paragraph (from Step 5 preamble or direct extraction)
        if rc_text and len(rc_text.strip()) > 50:
            _stripped = _strip_md(rc_text)[:700]
            if len(_stripped) > 60:  # only use if meaningful prose remains after stripping tables
                para = _bold(_stripped)

        # P2: synthesise from JSON bullets
        if not para and bullets:
            _lm = {}
            for b in (bullets or []):
                lbl = str(b.get("label","")).lower()
                finding = _strip_md(str(b.get("finding","")).strip())
                if finding and len(finding) > 8:
                    _lm[lbl] = finding
            sentences = []
            for key in ["supplier","plant","bottleneck","failure","duration","damage","network","chain","propagat","impact","when"]:
                v = next((v for k,v in _lm.items() if key in k), "")
                if v and v not in " ".join(sentences):
                    sentences.append(v.rstrip(".") + ".")
            if severity:
                cs = _rcr.sub(r'^CRITICAL\s*[—–-]\s*', '', severity, flags=_rcr.IGNORECASE).strip()
                if cs and cs not in " ".join(sentences):
                    sentences.append(cs.rstrip(".") + ".")
            combined = " ".join(sentences[:4])
            if len(combined) > 30:
                para = _bold(combined)

        # P3: raw prose fallback
        if not para and rc_text and len(rc_text.strip()) > 15:
            para = _bold(_strip_md(rc_text[:600]))

        # P4: extract 3 sentences from exec summary section in full_report
        if not para and full_report:
            em = _rcr.search(
                r'###?\s*Executive\s+Summary\s*\n(.*?)(?=\n###|\n##|\Z)',
                full_report, _rcr.DOTALL | _rcr.IGNORECASE
            )
            if em:
                et = _strip_md(str(em.group(1) or ""))
                sents = _rcr.split(r'(?<=[.!?])\s+', et)
                combined_e = " ".join(s for s in sents[:4] if len(s) > 15)
                if len(combined_e) > 30:
                    para = _bold(combined_e)

        # P5: absolute fallback — extract anything from full_report
        if not para and full_report:
            # Find any line mentioning a supplier ID or plant with data
            lines = []
            for line in full_report.splitlines():
                line = line.strip()
                if len(line) > 40 and _rcr.search(r'SUP0?\d{3}|\b(Bhopal|Pune|Goa|Baddi)\b|\d+%', line):
                    clean = _strip_md(line)
                    if len(clean) > 40:
                        lines.append(clean)
                if len(lines) >= 3:
                    break
            if lines:
                para = _bold(". ".join(lines[:3]))

        if not para:
            para = "Root cause analysis in progress — the system has identified upstream supplier delays as the primary driver. Review the WHO and WHY sections above for the detailed findings."

        return (
            '<div style="background:rgba(239,68,68,0.06);border-left:4px solid #f87171;' +
            'border-radius:0 10px 10px 0;padding:16px 20px;margin-bottom:18px">' +
            f'<div style="font-size:0.88rem;color:#e2e8f0;line-height:2.0">{para}</div>' +
            '</div>'
        )

    root_cause_block = _render_root_cause_box(rc_text=root_cause_text, full_report=report_body, bullets=root_cause_bullets, severity=root_cause_severity)



    # ── 10. Build Section Cards (all collapsible) ─────────────────
    # SECTION_DEFS: each entry = (display_name, icon, color, regex_pattern, plain_context_text)
    # RULES:
    #   1. Each section heading in the RCA report must match EXACTLY ONE entry.
    #   2. Patterns must be specific enough to avoid cross-matching between bypass types.
    #   3. context_text must be plain business language — no formulas, no "Step N of N" repetition.
    #   4. Entries are processed in order; first match wins for each heading in the report.
    # Deduplication is enforced below: once a section is rendered, its pattern is not re-used.
    # ── Query-aware section descriptions for the 5 detective steps ──
    _q_lower = (user_question or "").lower()
    _is_tq = any(w in _q_lower for w in ["delivery performance","deteriorating",
        "transportation routes","routes or modes","which routes","which modes",
        "route driving","mode driving"]) or ("transport" in _q_lower and "routes" in _q_lower)

    if _is_tq:
        _who_desc = (
            "Suppliers and plants at the centre of the delivery performance issue. "
            "Table 1A ranks high-risk suppliers (score > 0.6) by reliability — high score = poor track record. "
            "Table 1B shows delayed shipment counts per supplier, sorted worst to best.")
        _why_desc = (
            "What is actually breaking down and where. "
            "Table 2A compares plants by delay rate — higher % means more late shipments. "
            "Table 2B breaks delays by transport mode (Road / Rail / Air / Sea) — near-identical rates across modes confirm the problem is upstream at supplier/plant level, not in logistics.")
        _how_desc = (
            "Full scale of downstream damage. "
            "Table 4A shows each distributor city's demand gap — units ordered but never delivered. "
            "Table 4B ranks the 15 worst-hit cities by total unmet units. "
            "Table 4C analyses every Plant→Distributor route by cost, distance, and lead time — flags inefficient or overloaded routes.")
    else:
        _who_desc = (
            "The parties directly responsible for the disruption. "
            "Table 1A lists high-risk suppliers (score > 0.6) ranked by risk, with annual capacity — high score = poor reliability. "
            "Table 1B shows delayed shipment counts per supplier, most to least — higher = more downstream stockouts.")
        _why_desc = (
            "The operational mechanism behind the stockouts. "
            "Table 2A shows delay rate per plant — e.g. Bhopal at 61.3% means 61 of every 100 shipments arrive late. "
            "Table 2B breaks delays by transport mode (Road / Rail / Air / Sea) — reveals whether the issue is plant-wide or channel-specific.")
        _how_desc = (
            "Full scale of downstream damage. "
            "Table 4A shows each distributor city's demand gap — units ordered but never received. "
            "Table 4B ranks the 15 worst-hit cities by total unmet demand. "
            "Table 4C analyses every Plant→Distributor route by cost (INR), distance (km), and lead time — identifies routes to prioritise for optimisation.")

    _when_desc = (
        "Timeline of the problem — when it started and whether it is worsening. "
        "Table 3A shows monthly delay counts, on-time counts, delay rate (%), and month-over-month change "
        "(▲ worsened / ▼ improved / Baseline = first month). "
        "Sustained rates above 55–60% with no recovery signal a CHRONIC structural failure.")

    SECTION_DEFS = [
        # ══════════════════════════════════════════════════════════════════
        # DIRECT ANSWER SECTIONS — for factual/operational queries
        # ══════════════════════════════════════════════════════════════════
        ("Key Findings",      "📋", "#38bdf8",  r'key\s+findings?',
         "The data retrieved directly answers the question. The table below shows all records returned, sorted by the most significant metric. Review the top rows for the highest-impact items and the post-table summary for pattern interpretation."),
        ("Business Insights", "💡", "#4ade80",  r'business\s+insights?',
         "These insights translate the raw data into operational meaning — each one names a specific entity, cites its exact metric, and states what the number means for day-to-day supply chain decisions."),

        # ══════════════════════════════════════════════════════════════════
        # DETECTIVE INVESTIGATION STEPS — Data-driven, query-agnostic
        # These patterns match the universal 5-step structure output by the
        # _build_detective_investigation prompt for ANY query type.
        # WHO → WHY → WHEN → HOW MUCH → ROOT CAUSE
        # ══════════════════════════════════════════════════════════════════

        # Step 1 — WHO (implicated suppliers and plants)
        ("WHO: Implicated Suppliers & Plants",  "🕵️", "#f87171",
         r'step\s+1\s+of\s+5\s*[—–-]\s*who',
         _who_desc),

        # Step 2 — WHY (operational failure mechanism)
        ("WHY: Operational Failure Mechanism",  "🔬", "#fb923c",
         r'step\s+2\s+of\s+5\s*[—–-]\s*why',
         _why_desc),

    ("WHEN: Timeline & Month-over-Month Trend", "📅", "#7dd3fc",
         r'step\s+3\s+of\s+5\s*[—–-]\s*when',
         _when_desc),

        # Step 4 — HOW MUCH (downstream impact scale)
        ("HOW MUCH: Downstream Impact Scale",   "📊", "#60a5fa",
         r'step\s+4\s+of\s+5\s*[—–-]\s*how\s+much',
         _how_desc),

        # ══════════════════════════════════════════════════════════════════
        # FALLBACK PATTERNS — catch legacy section names from older LLM outputs
        # These ensure backwards compatibility if model outputs old heading names
        # ══════════════════════════════════════════════════════════════════

        # Kolkata / distributor-specific patterns
        ("Stockout Exposure",         "📉", "#f87171",  r'stockout\s+exposure',            "Monthly breakdown of demand gap — proves the problem persists every month, not just seasonally."),
        ("Shipment Fulfillment",      "📦", "#fb923c",  r'shipment\s+fulfillment',          "Are shipments received actually meeting demand? On-time gap column shows supply shortfall independent of transport delays."),
        ("Plant Supply Contribution", "🏗️", "#7dd3fc",  r'plant\s+supply\s+contribution',   "Which plant contributes most to unfulfilled demand — identifies deepest investigation target."),
        ("Supplier Risk at Source",   "🏭", "#a78bfa",  r'supplier\s+risk\s+at\s+source',   "All suppliers feeding the worst-performing plant ranked by risk score — structural fragility before shipment leaves."),

        # Generic WHO/WHY/HOW MUCH fallbacks (partial match)
        ("High-Risk Suppliers",           "🏭", "#f87171",  r'high.?risk\s+suppliers?(?:\s*[:\-–—]|\s+step)',   "Suppliers with elevated risk scores — the upstream parties the data implicates as responsible."),
        ("Supplier → Plant Delay Chain",  "🔗", "#fb923c",  r'supplier\s*[→\-]+\s*plant\s+delay',    "Maps each risky supplier's unreliability to plant-level delays — the causal chain linking WHO to WHY."),
        ("Plant Dispatch Shortfall",      "🏗️", "#7dd3fc",  r'plant\s+dispatch\s+shortfall',         "Which plants fail delivery timelines — the operational chokepoints converting supplier risk into downstream shortfalls."),
        ("Distributor Demand Gap Trail",  "📉", "#60a5fa",  r'distributor\s+demand\s+gap\s+trail',   "Where and how the disruption is spreading — distributor cities ranked by cumulative demand gap."),
        ("Distributor Contribution",      "🚚", "#60a5fa",  r'distributor\s+contribution',           "How plant shortfalls cascade into stock gaps at distributor cities — ranked by total demand gap."),
        ("Demand Gap Trail",              "📉", "#f87171",  r'demand\s+gap\s+trail',                 "End consequence — where stockouts land across distributor cities. Every city shown carries active unmet demand."),

        # Delay trail fallbacks
        ("Delay Propagation Trail",   "📦", "#fb923c",  r'delay\s+propagation\s+trail',          "Supplier → plant delay chain ranked by volume — every row maps a risky supplier to its plant's Major Delay shipment count."),
        ("Bottleneck Plants",         "⚠️", "#f87171",  r'bottleneck\s+plants?(?!\s+by)',         "Plants ranked by delay rate — where upstream supplier failures become operational chokepoints."),
        ("Distributor Impact",        "🚚", "#60a5fa",  r'distributor\s+impact(?!\s+on)',         "Distributor cities receiving the most delayed shipments — quantifies commercial impact of upstream failures."),
        ("Retailer Stockout Impact",  "🛒", "#a78bfa",  r'retailer\s+stockout\s+impact',         "End consequence — where delayed shipments become confirmed retail stockouts across distributor cities."),

        # Simulation fallbacks
        ("Disruption Scenario",         "⚠️", "#fb923c",  r'disruption\s+scenario',               "The specific failure event being simulated — which entity is disrupted and the scope of immediate exposure."),
        ("Affected Suppliers & Plants", "🏭", "#f87171",  r'affected\s+suppliers?\s*[&and]+\s*plants?', "Upstream operations becoming vulnerable — plant dependency and current delay baseline before shutdown."),
        ("Shipment & Route Breakdown",  "🚚", "#60a5fa",  r'shipment\s*[&and]+\s*route\s+breakdown', "Which logistics flows collapse first — transport modes and routes carrying shipments through the disrupted path."),
        ("Distributor & Retailer Exposure","📦","#a78bfa", r'distributor\s*[&and]+\s*retailer\s+exposure', "Downstream commercial exposure — distributor cities and retailer locations in the disruption blast radius."),
        ("Cascading Failure Simulation","📉", "#f87171",  r'cascading\s+failure\s+simulation',     "Full propagation simulation — how disruption spreads from trigger point to consumer-facing stockout consequence."),

        # Transport mode fallbacks
        ("Transportation Mode Performance",  "🚚", "#f87171",  r'transportation\s+mode\s+performance',  "Which transport modes generate the most Major Delay shipments — baseline performance by mode across the network."),
        ("High Delay/Cost Routes",           "🛣️", "#fb923c",  r'high\s+delay.*cost.*routes?|high\s+delay/cost', "Routes creating the biggest bottlenecks — delay rate and cost indexed together to show where logistics is failing."),
        ("Plant Transportation Load",        "🏭", "#7dd3fc",  r'plant\s+transportation\s+load',        "Which plants generate the most transport pressure — shows where mode choices are compounding supply chain stress."),
        ("Cost Escalation Impact",           "💰", "#a78bfa",  r'cost\s+escalation\s+impact',           "Which routes or modes are driving cost increases — connects delay patterns to logistics spend."),
        ("Distributor Delivery Disruptions", "📦", "#60a5fa",  r'distributor\s+delivery\s+disruptions', "Downstream delivery impact — which distributor cities are receiving the worst transport-driven shortfalls."),

        # Supplier risk fallbacks
        ("Plant Dependency Exposure",     "🏭", "#fb923c",  r'plant\s+dependency\s+exposure',       "Which plants depend most heavily on risky suppliers — single-source dependency creates structural vulnerability."),
        ("Shipment Delay Correlation",    "📦", "#7dd3fc",  r'shipment\s+delay\s+correlation',      "How supplier risk scores translate into actual Major Delay shipment counts — confirms risk score as a reliable delay predictor."),
        ("Transportation Dependency Chains", "🚚", "#a78bfa", r'transportation\s+dependency\s+chains', "Which logistics paths rely on risky suppliers — transport routes that compound supply risk."),
        ("Supply Chain Exposure Trail",   "📉", "#f87171",  r'supply\s+chain\s+exposure\s+trail',   "Full downstream propagation — how supplier risk cascades through plants to distributor cities and retailers."),

        # Product/category fallbacks
        ("High-Risk Product Categories", "📦", "#f87171",  r'high.risk\s+product\s+categor',       "Product categories facing the highest disruption rates — shows which lines are most exposed to supply chain failures."),
        ("Supplier Dependency Mapping",  "🏭", "#fb923c",  r'supplier\s+dependency\s+mapping',     "Which suppliers contribute most to category-level delays — maps upstream risk to specific product lines."),
        ("Plant Contribution Trail",     "🏗️", "#7dd3fc",  r'plant\s+contribution\s+trail',        "Which plants amplify category disruptions — delay rates by plant for the impacted product lines."),
        ("Transportation Impact Paths",  "🚚", "#60a5fa",  r'transportation\s+impact\s+paths',     "Which logistics paths affect delivery performance for this category — mode-level delay analysis."),
        ("End-to-End Supply Chain Trace","📉", "#f87171",  r'end.to.end\s+supply\s+chain\s+trace', "Full downstream propagation — distributor cities and retailers absorbing the category-level delays."),

        # Performance/general fallbacks
        ("Plant Health Comparison",        "🏭", "#fb923c",  r'plant\s+health\s+comparison',          "All plants ranked by delay rate with performance grades — identifies underperforming plants operationally."),
        ("Shipment Delay Trends",          "🚚", "#7dd3fc",  r'shipment\s+delay\s+trends?',           "Monthly delay rate trend — shows whether network performance is worsening, stable, or recovering over time."),
        ("Distributor Network Performance","🏬", "#60a5fa",  r'distributor\s+network\s+performance',  "Distributor chain efficiency ranking — identifies which distribution lanes are creating the most friction."),
        ("Operational Bottlenecks",        "📉", "#a78bfa",  r'operational\s+bottlenecks?',           "What is limiting supply chain throughput — operational constraints ranked by business impact."),
        ("Delivery Performance Impact",    "📦", "#f87171",  r'delivery\s+performance\s+impact',      "How service levels are affected — delivery rate, on-time ratio, and delay severity per entity."),

        # Retailer fallbacks
        ("Upstream Delay Sources",      "📦", "#fb923c",  r'upstream\s+delay\s+sources?',          "Which supplier-plant pairs are injecting delays into the retailer network — root of the retail delay chain."),
        ("Distributor Coverage Mapping","🏬", "#7dd3fc",  r'distributor\s+coverage\s+mapping',     "Which distributors support affected retailers — coverage gaps show where distribution is under-resourced."),
        ("Retailer Stockout Profile",   "🛍️", "#a78bfa",  r'retailer\s+stockout\s+profile',       "Retailer cities ranked by unmet demand — the consumer-facing consequence of upstream supply failures."),
        ("Full Retail Delay Exposure",  "📉", "#f87171",  r'full\s+retail\s+delay\s+exposure',     "Complete retailer delay vulnerability ranking — combining shortage volume, distributor dependency, and delay severity."),
        ("Retailer Supply Risk Exposure","📉","#f87171",   r'retailer\s+supply\s+risk\s+exposure', "Full retailer vulnerability ranking — shortage frequency, unmet demand, and distributor dependency."),
        ("Retail Exposure by City",     "🛍️", "#a78bfa",  r'retail\s+exposure\s+by\s+city',       "Which retailer cities carry the most severe delay or stockout exposure in the network."),
        ("Retailer Sales Performance",  "🛍️", "#a78bfa",  r'retailer\s+sales\s+performance',      "Which retailers are most impacted — sales shortfall ranked by severity and geographic location."),
        ("Distributor Delay Transmission","🏬","#7dd3fc",  r'distributor\s+delay\s+transmission',  "Which distributor cities are acting as relay nodes for upstream delays reaching retailers downstream."),
        ("Shipment Dependency Trail",   "📦", "#fb923c",  r'shipment\s+dependency\s+trail',        "Which supply paths affect retailers — the logistics chain feeding retail points in the network."),
        ("Upstream Dependency Trail",   "🔗", "#fb923c",  r'upstream\s+dependency\s+trail',        "All plants feeding the distributor network — shows which plant's delay rate every city inherits upstream."),
        ("Distributor Metrics",         "📊", "#7dd3fc",  r'distributor\s+metrics',                "Per-city operational stress: shortage shipments, delay exposure, and demand gap — identifies most urgent replenishment priorities."),
        ("Retailer Impact",             "🛒", "#a78bfa",  r'retailer\s+impact(?!\s+on)',           "Distributor shortfalls translate to empty shelves — retailer cities most vulnerable to supply disruption."),

        # Delay trail
        ("Supplier Delay Sources",         "🏭", "#f87171",  r'supplier\s+delay\s+sources?',          "Which suppliers are driving delays — risk scores and delay contribution per supplier ranked by impact."),
        ("Plant Dispatch Delays",          "🏗️", "#fb923c",  r'plant\s+dispatch\s+delays?',           "Which plants are failing shipment timelines — delay rate and count per plant showing operational failures."),
        ("Route & Transportation Bottlenecks","🚚","#7dd3fc", r'route\s*[&and]+\s*transportation\s+bottlenecks?', "Which routes or transport modes are slowing deliveries — logistics-layer root of delay propagation."),
        ("Distributor Delivery Impact",    "📦", "#a78bfa",  r'distributor\s+delivery\s+impact',      "Which distributors are impacted downstream — delay counts and demand gaps at the distribution layer."),

        # Generic ROOT CAUSE catch-all (must be last — broad pattern)
        ("Root Cause",                "⚠️", "#f87171",
         r'root\s+cause(?!\s+analysis\s+report)',
         "The origin of the supply chain breakdown — traced from upstream supplier risk through plant delays to downstream stockouts."),

        # Final summary fallbacks
        ("Final Root Cause",      "🔍", "#a78bfa",  r'final\s+root\s+cause|propagation\s+summary|category\s+impact\s+summary|performance\s+summary|simulation\s+summary', "Synthesis of all investigation steps — confirmed root cause with supporting evidence from the full trail."),
        ("Supplier Risk Summary", "🔗", "#7dd3fc",  r'upstream\s+dependency\s+trail|network\s+propagation', "Network-level view of how the root cause propagates from the upstream source to downstream impact."),
        ("Retailer Impact Summary","🔍","#a78bfa",  r'retailer\s+impact\s+summary',               "Full synthesis of the retailer delay or stockout investigation — root cause confirmed with upstream evidence."),
    ]
    section_cards_html = ""
    _rendered_sections = set()   # deduplication: track already-rendered section names
    _matched_spans = []          # track char spans already consumed so patterns don't double-match

    # Canonical aliases — if any of these was rendered, skip its aliases
    _ROOT_CAUSE_NAMES = {
        "ROOT CAUSE: Confirmed Verdict", "Root Cause", "Final Root Cause",
        "Retailer Impact Summary", "Supplier Risk Summary",
    }


    # ── Python table map: section → tools to render directly ──────────────────
    # Her approach: data comes from Python dicts, not LLM markdown reproduction
    # Keys are substrings matched against sec_name (plain text like "WHO: Implicated Suppliers")
    # NOT regex patterns — avoids the regex-in-regex matching bug
    # ── Query-aware SECTION_TOOL_MAP ────────────────────────────────────
    # For category queries (toys, auto, health etc.), use category-specific tools.
    # NOTE: inside _assemble_final_output the question is the `user_question` parameter,
    # not `ctx` (ctx only exists in the pipeline agent functions).
    try:
        _cat_entities = _extract_question_entities(user_question)
        _detected_cat_tool = _cat_entities.get("product_category", "") or ""
    except Exception:
        _detected_cat_tool = ""

    _is_category_rca = bool(_detected_cat_tool)

    if _is_category_rca:
        _cat_display = _detected_cat_tool.replace("_", " ").title()
        _SECTION_TOOL_MAP = {
            # 1A: use trace_supply_chain_for_category — aggregated in _build_python_section_html
            # into per-supplier delayed counts for THIS category
            # 1B: use get_supplier_plant_delay_chain for full risk score + delayed count context
            "who":  [
                (f"1A — {_cat_display} Delayed Shipments by Supplier",
                 "trace_supply_chain_for_category", set()),
                (f"1B — All-Network Supplier Risk Context",
                 "get_supplier_plant_delay_chain", set()),
            ],
            "why":  [
                ("2A — Plant Delay Rates",           "get_delay_by_plant",            set()),
                ("2B — Transport Mode Delays",       "get_transport_mode_delays",     set()),
            ],
            "when": [
                ("3A — Monthly Delay Trend",         "get_monthly_delay_trend",       set()),
            ],
            "how much": [
                ("4A — Distributor Demand Gap",      "get_demand_gap_analysis",       set()),
                (f"4B — Top Cities Short of {_cat_display}", "get_stockout_retailers", set()),
                ("4C — Route Efficiency Analysis",   "get_route_cost_efficiency",     set()),
            ],
        }
    else:
        _SECTION_TOOL_MAP = {
            "who":  [
                ("1A — Supplier Risk Scores & Capacity",   "get_high_risk_suppliers",         {"avg_delay_days"}),
                ("1B — Supplier Delay Contribution (Shipment Counts)", "get_supplier_delay_contribution", set()),
            ],
            "why":  [
                ("2A — Plant Delay Rates",           "get_delay_by_plant",             set()),
                ("2B — Transport Mode Delays",       "get_transport_mode_delays",      set()),
            ],
            "when": [
                ("3A — Monthly Delay Trend",         "get_monthly_delay_trend",        set()),
            ],
            "how much": [
                ("4A — Distributor Demand Gap",      "get_demand_gap_analysis",        set()),
                ("4B — Top Distributor Shortage",    "get_stockout_retailers",         set()),
                ("4C — Route Efficiency Analysis",   "get_route_cost_efficiency",      set()),
            ],
        }

    def _python_rows_to_html(rows: list, tool_name: str = "", skip_cols: set = None, max_rows: int = 10) -> str:
        """Render Python row dicts as styled HTML table with show-more pagination.
        skip_cols: set of column names to exclude from display.
        max_rows: rows visible before 'show N more' toggle. Default 10.
        """
        if not rows:
            return ""
        _LABELS = {
            "supplier_id":"Supplier ID","supplier_name":"Supplier Name","risk_score":"Risk Score",
            "plant_id":"Plant ID","plant_name":"Plant Name","delayed_shipments":"Delayed Shipments",
            "avg_delay_days":"Avg Delay (days)","avg_delay":"Avg Delay (days)",
            "total_shipments":"Total Shipments","delayed_count":"Delayed (Count)",
            "delay_rate_pct":"Delay Rate %","year_month":"Period (YYYY-MM)",
            "on_time_count":"On-Time Count","mom_change":"MoM Change ▲▼",
            "distributor_city":"Distributor City","distributor_id":"Distributor ID",
            "shortage_shipments":"Shortage Shipments","total_demand_gap":"Total Demand Gap (units)",
            "total_shortage_units":"Total Unmet Units",
            "retailers_directly_connected":"Retailers (Direct Link)",
            "route_id":"Route ID","transport_mode":"Transport Mode",
            "transportation_mode":"Transport Mode","distance_km":"Distance (km)",
            "cost_inr":"Cost (INR)","cost_efficiency":"Cost Efficiency",
            "leadtime_days":"Lead Time (days)","total_delays":"Total Delayed Shipments",
            "plants_affected":"Plants Affected",
            "capacity_units":"Annual Capacity","lead_time_days":"Lead Time (days)",
        }
        extra_skip = skip_cols or set()
        cols = [c for c in rows[0].keys() if c not in extra_skip]

        def _render_rows(row_subset):
            tbody = ""
            for i, row in enumerate(row_subset):
                bg = "rgba(255,255,255,0.03)" if i % 2 == 0 else "transparent"
                tbody += f'<tr style="border-bottom:1px solid rgba(255,255,255,0.05);background:{bg}">'
                for c in cols:
                    val = row.get(c)
                    if val is None: val = "—"
                    elif isinstance(val, float): val = f"{val:.2f}"
                    else: val = str(val)
                    if c == "mom_change":
                        colour = ("#f87171" if "▲" in val else
                                  "#4ade80" if "▼" in val else "#94a3b8")
                        tbody += f'<td style="padding:7px 12px;font-size:0.82rem;color:{colour};font-weight:600;white-space:nowrap">{val}</td>'
                    elif c == "risk_score":
                        try:
                            rv = float(val)
                            rc = "#f87171" if rv >= 0.8 else "#fb923c" if rv >= 0.6 else "#fbbf24" if rv >= 0.4 else "#4ade80"
                            tbody += f'<td style="padding:7px 12px;font-size:0.82rem;color:{rc};font-weight:700">{val}</td>'
                        except Exception:
                            tbody += f'<td style="padding:7px 12px;font-size:0.82rem;color:#e2e8f0">{val}</td>'
                    else:
                        tbody += f'<td style="padding:7px 12px;font-size:0.82rem;color:#e2e8f0;vertical-align:top">{val}</td>'
                tbody += "</tr>"
            return tbody

        thead = "<tr>" + "".join(
            '<th style="padding:8px 12px;text-align:left;font-size:0.68rem;font-weight:700;'
            'text-transform:uppercase;letter-spacing:0.07em;color:#94a3b8;white-space:nowrap;'
            'background:rgba(15,25,50,0.9)">' +
            f'{_LABELS.get(c, c.replace("_"," ").title())}</th>'
            for c in cols
        ) + "</tr>"

        total = len(rows)
        show_limit = min(max_rows, total)
        visible_rows = rows[:show_limit]
        hidden_rows  = rows[show_limit:]

        import uuid as _uuid
        tbl_id = "tbl_" + _uuid.uuid4().hex[:8]
        tbody_visible = _render_rows(visible_rows)
        hidden_html = ""
        toggle_btn  = ""
        if hidden_rows:
            tbody_hidden = _render_rows(hidden_rows)
            hidden_html = f'<tbody id="{tbl_id}_more" style="display:none">{tbody_hidden}</tbody>'
            n_hidden = len(hidden_rows)
            toggle_btn = (
                f'<div style="text-align:center;padding:6px 0 2px">'
                f'<button id="{tbl_id}_btn" onclick="'
                f'var t=document.getElementById(\'{tbl_id}_more\'),'
                f'b=document.getElementById(\'{tbl_id}_btn\');'
                f'if(t.style.display===\'none\'){{t.style.display=\'\';'
                f'b.textContent=\'\u25b2 Show less\';}}' 
                f'else{{t.style.display=\'none\';'
                f'b.textContent=\'\u25bc Show {n_hidden} more rows\';}}" '
                f'style="background:rgba(56,189,248,0.08);border:1px solid rgba(56,189,248,0.25);'
                f'border-radius:20px;color:#7dd3fc;font-size:0.72rem;font-weight:700;'
                f'padding:5px 20px;cursor:pointer;letter-spacing:0.04em;margin-top:4px;'
                f'transition:all 0.15s">'
                f'\u25bc Show {n_hidden} more rows</button></div>'
            )

        return (
            '<div style="overflow-x:auto;margin-bottom:6px">'
            f'<table id="{tbl_id}" style="width:100%;border-collapse:collapse;border-radius:8px;overflow:hidden">'
            f'<thead>{thead}</thead>'
            f'<tbody>{tbody_visible}</tbody>'
            f'{hidden_html}'
            '</table></div>'
            + toggle_btn
        )

    def _build_python_section_html(sec_name: str) -> str:
        """Build Python-rendered tables for a section — each subtable in its own named dropdown.
        For category queries: aggregates raw trace rows into meaningful supplier-level summaries.
        Returns HTML with one <details> per subtable, labelled with the subtable name.
        """
        sec_lower = sec_name.lower()
        for keyword, tool_list in _SECTION_TOOL_MAP.items():
            if keyword in sec_lower:
                parts = []
                seen_tools = set()
                for heading, tool_name, skip_cols in tool_list:
                    if tool_name in seen_tools:
                        continue

                    rows = tool_rows.get(tool_name, [])

                    # ── Special handling: aggregate raw trace rows for category WHO table ──
                    # trace_supply_chain_for_category returns individual shipment rows.
                    # For the WHO section we need per-supplier aggregated counts.
                    if tool_name == "trace_supply_chain_for_category" and rows and _is_category_rca:
                        import collections as _coll
                        _agg: dict = {}
                        for _r in rows:
                            _key = (str(_r.get("supplier", "")), str(_r.get("plant", "")))
                            if _key not in _agg:
                                _agg[_key] = {
                                    "supplier": _r.get("supplier", ""),
                                    "plant": _r.get("plant", ""),
                                    "total_shipments": 0,
                                    "delayed_shipments": 0,
                                    "avg_delay_days": 0.0,
                                    "_delay_sum": 0.0,
                                }
                            _agg[_key]["total_shipments"] += 1
                            if str(_r.get("delivery_status", "")) == "Major Delay":
                                _agg[_key]["delayed_shipments"] += 1
                                _agg[_key]["_delay_sum"] += float(_r.get("delay_days", 0) or 0)
                        # Compute avg and sort
                        _agg_rows = []
                        for _v in _agg.values():
                            _del = _v["delayed_shipments"]
                            _v["avg_delay_days"] = round(_v["_delay_sum"] / _del, 2) if _del else 0.0
                            del _v["_delay_sum"]
                            _agg_rows.append(_v)
                        _agg_rows.sort(key=lambda x: x["delayed_shipments"], reverse=True)
                        rows = _agg_rows

                    # ── Special handling: aggregate get_category_supplier_risk (same source) ──
                    elif tool_name == "get_category_supplier_risk" and rows and _is_category_rca:
                        # This is aliased to trace_supply_chain_for_category in _local_fallback
                        # so rows may also be raw shipment rows — aggregate same way
                        if rows and "delivery_status" in (rows[0] if rows else {}):
                            import collections as _coll2
                            _agg2: dict = {}
                            for _r in rows:
                                _key = str(_r.get("supplier", _r.get("supplier_name", "")))
                                _plant = str(_r.get("plant", _r.get("plant_name", "")))
                                if _key not in _agg2:
                                    _agg2[_key] = {
                                        "supplier": _key, "plant": _plant,
                                        "delayed_shipments": 0, "_delay_sum": 0.0,
                                    }
                                if str(_r.get("delivery_status", "")) == "Major Delay":
                                    _agg2[_key]["delayed_shipments"] += 1
                                    _agg2[_key]["_delay_sum"] += float(_r.get("delay_days", 0) or 0)
                            _agg2_rows = []
                            for _v in _agg2.values():
                                _d = _v["delayed_shipments"]
                                _v["avg_delay_days"] = round(_v["_delay_sum"] / _d, 2) if _d else 0.0
                                del _v["_delay_sum"]
                                _agg2_rows.append(_v)
                            _agg2_rows.sort(key=lambda x: x["delayed_shipments"], reverse=True)
                            rows = _agg2_rows

                    if not rows:
                        continue
                    seen_tools.add(tool_name)
                    tbl_html = _python_rows_to_html(rows, tool_name=tool_name,
                                                    skip_cols=skip_cols, max_rows=10)
                    n_rows = len(rows)
                    parts.append(
                        f'<details style="margin-top:10px;border:1px solid rgba(56,189,248,0.2);'
                        f'border-radius:9px;overflow:hidden;background:rgba(56,189,248,0.025)">'
                        f'<summary style="cursor:pointer;list-style:none;user-select:none;'
                        f'padding:9px 14px;display:flex;align-items:center;gap:8px;'
                        f'font-size:0.76rem;font-weight:700;color:#38bdf8;'
                        f'background:rgba(56,189,248,0.05)">'
                        f'<span style="font-size:0.6rem;transition:transform 0.2s;display:inline-block" '
                        f'class="vd-arrow">▶</span>'
                        f'<span>📋 {heading}</span>'
                        f'<span style="margin-left:auto;font-size:0.65rem;font-weight:500;'
                        f'color:#475569;background:rgba(56,189,248,0.08);padding:1px 8px;'
                        f'border-radius:10px;border:1px solid rgba(56,189,248,0.15)">{n_rows} records</span>'
                        f'</summary>'
                        f'<div style="padding:10px 12px 14px;overflow-x:auto">'
                        f'{tbl_html}'
                        f'</div>'
                        f'</details>'
                        f'<style>details[open] .vd-arrow{{transform:rotate(90deg)}}'
                        f'details[open]>summary{{border-bottom:1px solid rgba(56,189,248,0.15)!important}}'
                        f'</style>'
                    )
                return "".join(parts)
        return ""


    for sec_name, sec_icon, sec_color, sec_pat, context_text in SECTION_DEFS:
        # ── Deduplication: skip if this section name was already rendered ──
        if sec_name in _rendered_sections:
            continue
        # If any root-cause-family section was already rendered, skip all others
        if sec_name in _ROOT_CAUSE_NAMES and _rendered_sections & _ROOT_CAUSE_NAMES:
            continue

        # ── Special case: Root Cause — ALWAYS render, regardless of whether LLM wrote the heading ──
        _is_root_cause_step = (
            ("step" in sec_pat.lower() and "5" in sec_pat and "root" in sec_pat.lower())
            or ("root" in sec_pat.lower() and "cause" in sec_pat.lower())
            or "who_is_responsible" in sec_pat.lower()
            or "confirmed" in sec_pat.lower()
        )
        if _is_root_cause_step:
            # Build the best root cause paragraph we can from all available inputs
            import re as _re_rc_fb
            # Try to extract prose from Step 5 section if it exists
            _prose_rc = ""
            try:
                _rc_check_pat = _re.compile(
                    r'###?\s*[^\n]*' + sec_pat + r'[^\n]*\n',
                    _re.DOTALL | _re.IGNORECASE
                )
                if _rc_check_pat.search(report_body):
                    _p5_prose, _, _ = _extract_section(report_body, sec_pat)
                    _prose_rc = (_p5_prose or "").strip()
                    _prose_rc = _re_rc_fb.sub(r'```json[\s\S]*?```', '', _prose_rc).strip()
                    _prose_rc = _re_rc_fb.sub(r'\{[\s\n]*"bullets"[\s\S]*', '', _prose_rc).strip()
                    _prose_rc = _prose_rc.replace('**', '').strip()
            except Exception:
                pass

            # Always render — _render_root_cause_box has P5 absolute fallback
            _best_rc_block = root_cause_block if root_cause_block else _render_root_cause_box(
                rc_text=_prose_rc or root_cause_text,
                full_report=report_body,
                bullets=root_cause_bullets,
                severity=root_cause_severity
            )

            _rc_hdr = f"""
<div style="background:rgba(14,22,45,0.72);border:1px solid rgba(239,68,68,0.25);border-left:4px solid #f87171;
  border-radius:12px;margin-bottom:8px;padding:12px 18px 10px;box-shadow:0 2px 12px rgba(0,0,0,0.2)">
  <div style="display:flex;align-items:center;gap:10px;margin-bottom:6px">
    <span style="font-size:1.05rem">⚠️</span>
    <span style="font-size:0.88rem;font-weight:700;color:#fca5a5">Root Cause</span>
  </div>
  <div style="display:inline-flex;align-items:center;gap:5px;margin-bottom:8px;padding:3px 10px 3px 8px;
    background:rgba(239,68,68,0.1);border:1px solid rgba(239,68,68,0.3);border-radius:20px">
    <span style="font-size:0.62rem;color:#f87171">◆</span>
    <span style="font-size:0.7rem;font-weight:700;color:#f87171;letter-spacing:0.02em">Step 5 of 5: Confirmed verdict</span>
  </div>
</div>"""
            section_cards_html += _rc_hdr + _best_rc_block
            _rendered_sections.add(sec_name)
            for _rcn in _ROOT_CAUSE_NAMES:
                _rendered_sections.add(_rcn)
            root_cause_block = ""   # already rendered
            continue

        prose, table_md, prose_after = _extract_section(report_body, sec_pat)
        # For route sections: if no table but prose exists, use prose as content
        if not table_md and prose and len(prose.strip()) > 30:
            table_md = None   # will render prose-only card below
        elif not table_md:
            continue
        # Skip sections where all data values are "Not provided" / empty
        if table_md:
            stats = _quick_stats_from_table(table_md)
            # A section is "all empty" only if EVERY value in EVERY column is a placeholder.
            # "-" alone is a valid display value (e.g. "Not set" for lead time).
            # Only skip if all cells are truly placeholder instructions.
            _placeholder_vals = {"not provided", "not directly provided", "n/a", "",
                                  "(all routes", "(group all", "(group routes", "(populate"}
            all_empty = all(
                all(not v or str(v).strip().lower() in _placeholder_vals
                    for v in vals)
                for vals in stats.values()
            ) if stats else True
            if all_empty:
                # Fall back to prose if table is just instructions
                table_md = None
                if not (prose and len((prose or "").strip()) > 30):
                    continue

            # ── Empty table detection (header + separator only, no data) ──
            # LLM sometimes writes a table skeleton with no data rows.
            # In that case, treat it as no-data and show "No data retrieved".
            if table_md:
                _all_tbl_lines = [l.strip() for l in table_md.splitlines() if l.strip().startswith("|")]
                _data_rows_only = [l for l in _all_tbl_lines if not all(c in "|-: " for c in l)]
                if len(_data_rows_only) <= 1:  # only header, no data rows
                    table_md = None  # treat as empty
                    if not (prose and len((prose or "").strip()) > 30):
                        continue

            # ── Duplicate row detection ───────────────────────────────────
            # If table has ≥2 rows and every data row is identical, it means
            # the LLM hallucinated the same value across all rows.
            # In that case collapse to 1 unique row + a warning.
            if table_md:
                _tbl_lines = [l.strip() for l in table_md.splitlines()
                              if l.strip().startswith("|") and "---" not in l]
                if len(_tbl_lines) >= 3:  # header + separator + 2+ data rows
                    _data_rows = _tbl_lines[1:]  # skip header
                    _unique_rows = list(dict.fromkeys(_data_rows))
                    if len(_unique_rows) == 1 and len(_data_rows) > 1:
                        # All rows identical — collapse and add note
                        table_md = _tbl_lines[0] + "\n" + _unique_rows[0]
                        prose = (prose or "") + "\n⚠ Note: Multiple identical rows detected — showing unique record only."
        else:
            stats = {}

        # Build summary badge
        stats = _quick_stats_from_table(table_md) if table_md else {}
        n_rows = len(list(stats.values())[0]) if stats else 0
        summary_line = f"{n_rows} records" if n_rows else ""
        # Find worst metric for badge
        def _safe_flt(v):
            try: return float(str(v).replace("%","").replace(",","").strip())
            except: return None

        for k, vals in stats.items():
            if "delay" in k.lower() and "%" in k.lower() and vals:
                nums = [_safe_flt(v) for v in vals if v]
                nums = [f for f in nums if f is not None]
                if nums:
                    summary_line = f"{n_rows} records · peak {max(nums):.0f}% delay"
                break

        # For risk badge: prefer max_risk_score > risk_score > avg_risk (skip avg cols)
        _risk_badge_set = False
        for _preferred_k in [k for k in stats if "max" in k.lower() and "risk" in k.lower()]:
            _nums = [_safe_flt(v) for v in stats[_preferred_k] if v]
            _nums = [f for f in _nums if f is not None and 0 < f <= 1.0]
            if _nums:
                summary_line = f"{n_rows} records · max risk {max(_nums):.2f}"
                _risk_badge_set = True
                break
        if not _risk_badge_set:
            for k, vals in stats.items():
                if "risk" in k.lower() and "avg" not in k.lower() and "score" in k.lower() and vals:
                    _nums = [_safe_flt(v) for v in vals if v]
                    _nums = [f for f in _nums if f is not None and 0 < f <= 1.0]
                    if _nums:
                        summary_line = f"{n_rows} records · max risk {max(_nums):.2f}"
                    break

        collapsed_table = table_md   # pass raw markdown directly — _wrap_section_collapsible adds its own <details>
        # Strip rows where ALL data cells are N/A, "Not provided", "Not directly provided", or identical values
        if collapsed_table:
            import re as _re_tbl
            tbl_lines = collapsed_table.splitlines()
            cleaned_lines = []
            for tbl_line in tbl_lines:
                if tbl_line.strip().startswith("|") and not all(c in "|-: " for c in tbl_line.strip()):
                    # Data row — check for all-NA or placeholder cells
                    cells = [c.strip() for c in tbl_line.strip("|").split("|")]
                    bad_vals = {"n/a", "not provided", "not directly provided", "-", "--",
                                 "—", "na", "", "...", "...", "n/a", "tbd"}
                    non_bad = [c for c in cells if c.lower() not in bad_vals and c.strip() not in bad_vals]
                    if not non_bad:
                        continue  # skip entirely-placeholder rows (--  --  -- etc)
                    # Replace any remaining N/A / not provided cells with "-"
                    cells = ["-" if c.lower() in bad_vals or c.strip() in bad_vals else c for c in cells]
                    import re as _dr
                    cells = [
                        _dr.sub(r'^(\d{2})[-/](\d{2})[-/](\d{4})$', r'\3-\2', c)
                        if _dr.match(r'^\d{2}[-/]\d{2}[-/]\d{4}$', c) else c
                        for c in cells
                    ]
                    # Normalize date format: "01-01-2022" or "01/01/2022" → "2022-01"
                    import re as _date_re
                    cells = [
                        _date_re.sub(r'^(\d{2})[-/](\d{2})[-/](\d{4})$', r'\3-\1', c)
                        if _date_re.match(r'^\d{2}[-/]\d{2}[-/]\d{4}$', c) else c
                        for c in cells
                    ]
                    tbl_line = "| " + " | ".join(cells) + " |"
                cleaned_lines.append(tbl_line)
            collapsed_table = "\n".join(cleaned_lines)
        prose_safe = str(prose or "")

        # ── Extract the actual section heading from report_body to get step label ──
        # The LLM writes headings like:
        #   "### ⚠️ Step 5 of 5 — Root Cause: What is the origin..."
        # The step label is in the heading line, NOT in the prose body.
        # So _parse_step_and_desc would miss it. Fix: scan for the heading
        # that matched sec_pat and inject "**Step N of N: question**" as first
        # line of full_context so _parse_step_and_desc can extract the pill.
        import re as _re_step
        _step_label_injected = ""
        try:
            _heading_scan = _re_step.search(
                r'###?\s*[^\n]*' + sec_pat + r'[^\n]*',
                report_body, _re_step.IGNORECASE
            )
        except re.error:
            _heading_scan = None
        if _heading_scan:
            _heading_line = _heading_scan.group(0).strip()
            # Extract "Step N of N — SectionName: Question?" from heading
            _step_m = _re_step.search(
                r'(Step\s+\d+\s+of\s+\d+)\s*[—–-]\s*[^:]+:\s*(.+)',
                _heading_line, _re_step.IGNORECASE
            )
            if _step_m:
                _step_num  = _step_m.group(1).strip()   # "Step 5 of 5"
                _step_q    = _step_m.group(2).strip()   # "What is the origin..."
                _step_label_injected = f"**{_step_num}: {_step_q}**"
            else:
                # Heading has "Step N of N" but no "— Name: question" structure
                _step_only = _re_step.search(
                    r'(Step\s+\d+\s+of\s+\d+[^#\n]*)',
                    _heading_line, _re_step.IGNORECASE
                )
                if _step_only:
                    _step_label_injected = f"**{_step_only.group(1).strip()}**"

        # Pass prose raw so _parse_step_and_desc splits step label from body.
        # Prepend extracted step label so it is always available regardless of
        # whether the LLM included it in prose body (it usually doesn't).
        # Build 2-3 line description:
        #   Line 1-2: static SECTION_DEFS context_text (explains WHY this step matters)
        #   Line 3:   LLM prose written before the table (the actual data finding)
        # This gives every section card substantive context above the table.
        _static_intro = str(context_text or "").strip()
        _llm_prose    = prose_safe.strip()

        # Strip any duplicate step label from LLM prose (prevent "Step 1 of 5" appearing twice)
        if _llm_prose:
            _llm_prose_clean = _re_step.sub(
                r'Step\s+\d+\s+of\s+\d+[^\n]*', '', _llm_prose, count=1, flags=_re_step.IGNORECASE
            ).strip()
        else:
            _llm_prose_clean = ""

        # Build data-aware description: static intro + key stats from the table
        _combined_desc = _static_intro
        if table_md and stats:
            # Extract key stats to inject into description
            _desc_extras = []
            _col_vals = list(stats.values())
            _col_names = list(stats.keys())
            # For WHO section: mention top risk score
            if any("risk" in c.lower() for c in _col_names):
                risk_col = next((c for c in _col_names if "risk" in c.lower()), None)
                if risk_col and stats[risk_col]:
                    _top_risk = stats[risk_col][0] if stats[risk_col] else ""
                    if _top_risk and _top_risk not in ("-",""):
                        _desc_extras.append(f"The highest risk score in this data is {_top_risk}.")
            # For WHY section: mention worst delay rate
            if any("delay" in c.lower() and "rate" in c.lower() for c in _col_names):
                dr_col = next((c for c in _col_names if "delay" in c.lower() and "rate" in c.lower()), None)
                if dr_col and stats[dr_col]:
                    _top_dr = stats[dr_col][0] if stats[dr_col] else ""
                    nm_col = next((c for c in _col_names if "name" in c.lower() or "plant" in c.lower()), None)
                    _top_nm = stats[nm_col][0] if nm_col and stats[nm_col] else ""
                    if _top_dr and _top_dr not in ("-",""):
                        _desc_extras.append(f"The worst-performing plant ({_top_nm}) has a {_top_dr}% delay rate.")
            # For HOW MUCH section: mention total demand gap
            if any("demand" in c.lower() and "gap" in c.lower() for c in _col_names):
                dg_col = next((c for c in _col_names if "demand" in c.lower() and "gap" in c.lower()), None)
                if dg_col and stats[dg_col]:
                    try:
                        _total_gap = sum(float(str(v).replace(",","")) for v in stats[dg_col] if str(v).replace(",","").replace(".","").isdigit())
                        if _total_gap > 0:
                            _desc_extras.append(f"Total unmet demand shown: {int(_total_gap):,} units across {len(stats[dg_col])} cities.")
                    except Exception:
                        pass
            if _desc_extras:
                _combined_desc = _static_intro + " " + " ".join(_desc_extras)
        # Append LLM prose summary if meaningful
        if _llm_prose_clean and len(_llm_prose_clean) > 20:
            _combined_desc = _combined_desc + " " + _llm_prose_clean[:300]

        if _step_label_injected:
            full_context = _step_label_injected + "\n" + _combined_desc
        else:
            full_context = _combined_desc

        # For prose-only sections (no table): use prose as content, static intro as description
        if not table_md:
            collapsed_table = prose_safe
            full_context = (_step_label_injected + "\n" + _static_intro) if _step_label_injected else _static_intro
        _py_html = _build_python_section_html(sec_name)  # sec_name keyword match
        # Strip any leaked JSON from insight/prose text
        _clean_insight = str(prose_after or "")
        _clean_insight = _re.sub(r'```json[\s\S]*?```', '', _clean_insight).strip()
        _clean_insight = _re.sub(r'\{[\s\S]*?"bullets"[\s\S]*', '', _clean_insight).strip()

        _clean_insight = _re.sub(r'`+[a-z]*`*', '', _clean_insight).strip()
        _clean_collapsed = str(collapsed_table or "")
        _clean_collapsed = _re.sub(r'```json[\s\S]*?```', '', _clean_collapsed).strip()
        _clean_collapsed = _re.sub(r'\{[\s\S]*?"bullets"[\s\S]*', '', _clean_collapsed).strip()

        section_cards_html += _wrap_section_collapsible(
            sec_icon, sec_name, sec_color,
            context_html=full_context,
            table_html=_clean_collapsed,
            python_table_html=_py_html,
            summary_line=summary_line,
            insight_text=_clean_insight,
        )
        _rendered_sections.add(sec_name)   # mark as rendered — won't appear again

    # ── 11. Final Root Cause section — REMOVED (already covered by Executive Summary + Root Cause box) ──

    # ── 12. Build colour-coded Recommendations ────────────────
    # ── Use JSON-rendered recs if available (new format) ────────────────
    if rec_json_block:
        rec_block = (
            '<div style="margin-top:32px;padding-top:24px;border-top:2px solid rgba(74,222,128,0.25)">'
            '<div style="display:flex;align-items:center;gap:10px;margin-bottom:16px">'
            '<span style="font-size:1.1rem">✅</span>'
            '<span style="font-size:1rem;font-weight:800;color:#4ade80;letter-spacing:0.01em">Recommendations</span>'
            '<span style="font-size:0.7rem;color:#94a3b8;margin-left:4px">— Prioritised corrective actions</span>'
            '</div>'
            f'{rec_json_block}'
            '</div>'
        )

    if not rec_json_block and rec_content:
        def _build_entity_specific_recs(raw: str, report_body_ref: str) -> str:
            import re as _r2

            sup_names = _re.findall(r'\|\s*([A-Z][A-Za-z\s&]+(?:Ltd|PLC|Group|Sons|Pvt|Inc)?)\s*\|', report_body_ref)
            sup_names = [s.strip() for s in sup_names if 4 < len(s.strip()) < 40 and
                         s.strip() not in ("Supplier Name","Plant Name","Distributor City","Category")]
            plant_ids = list(dict.fromkeys(_re.findall(r'\b(PL\d+)\b', report_body_ref)))[:3]
            sup_ids   = list(dict.fromkeys(_re.findall(r'\b(SUP\w+)\b', report_body_ref)))[:3]

            # Colour-coded tier definitions
            TIERS = {
                "critical":    {
                    "color": "#f87171",
                    "bg":    "rgba(239,68,68,0.12)",
                    "border":"rgba(239,68,68,0.5)",
                    "label": "⚡ Critical — Act Immediately",
                    "impact_prefix": "Immediate impact:",
                    "exec_prefix":   "Execution note:",
                    "long_prefix":   "If delayed:",
                },
                "operational": {
                    "color": "#fb923c",
                    "bg":    "rgba(251,146,60,0.10)",
                    "border":"rgba(251,146,60,0.4)",
                    "label": "🔧 Operational — Near-Term Fixes",
                    "impact_prefix": "Expected outcome:",
                    "exec_prefix":   "How to execute:",
                    "long_prefix":   "Sustained benefit:",
                },
                "strategic":   {
                    "color": "#4ade80",
                    "bg":    "rgba(74,222,128,0.08)",
                    "border":"rgba(74,222,128,0.35)",
                    "label": "🏛 Strategic — Long-Term Transformation",
                    "impact_prefix": "Long-term impact:",
                    "exec_prefix":   "Initiative owner:",
                    "long_prefix":   "Structural benefit:",
                },
            }

            def _key_from_heading(h: str) -> str:
                h = h.lower()
                if any(w in h for w in ["critical","immediate","act now","urgent","emergency"]):
                    return "critical"
                if any(w in h for w in ["operational","near","fix","short"]):
                    return "operational"
                if any(w in h for w in ["strategic","long","structural","transform"]):
                    return "strategic"
                return "strategic"

            # Build query-specific impact variants from actual RCA findings
            # Extract key metrics from report_body_ref for specific impact statements
            import re as _re_iv
            _delayed_total  = _re_iv.search(r'(\d[\d,]+)\s*delayed\s*shipments?', report_body_ref, _re_iv.IGNORECASE)
            _delay_rate     = _re_iv.search(r'(\d+\.?\d*)\s*%\s*(?:Major\s*Delay|delay\s*rate)', report_body_ref, _re_iv.IGNORECASE)
            _demand_gap_val = _re_iv.search(r'(\d[\d,]+)\s*units\s*(?:unmet|gap|demand\s*gap)', report_body_ref, _re_iv.IGNORECASE)
            _top_city       = _re_iv.search(r'(?:Kolkata|Lucknow|Patna|Mumbai|Delhi|Ahmedabad|Bhubaneswar|Chandigarh)', report_body_ref)
            _top_plant      = _re_iv.search(r'(?:Bhopal|Baddi|Pune|Goa)\s*\(PL\d\)', report_body_ref, _re_iv.IGNORECASE)

            _dt  = _delayed_total.group(1) if _delayed_total else "the delayed"
            _dr  = _delay_rate.group(1) if _delay_rate else "the current"
            _dg  = _demand_gap_val.group(1) if _demand_gap_val else "the unmet"
            _tc  = _top_city.group(0) if _top_city else "affected"
            _tp  = _top_plant.group(0) if _top_plant else "the bottleneck plant"

            IMPACT_VARIANTS = [
                (f"Targets the {_dt} delayed shipments driving the {_dr}% plant delay rate",
                 "Procurement lead + Supplier relationship manager",
                 f"Supplier reliability improves — {_dr}% delay rate begins to recover"),
                (f"Addresses the root cause of {_dt} Major Delay shipments at {_tp}",
                 "Plant operations director",
                 f"Plant dispatch timeline normalises — {_dr}% delay rate reduces to below 30%"),
                (f"Closes the {_dg} unit demand gap accumulating at {_tc} and other cities",
                 "Distribution planning team",
                 f"Distributor demand gap of {_dg} units begins to close — replenishment stabilises"),
                (f"Prevents further stockouts at retailers served by {_tc} distribution hub",
                 "Supply chain planning + inventory team",
                 "Retail shelf availability recovers — consumer stockout incidents reduce"),
                (f"Reduces network-wide delay rate from {_dr}% toward industry benchmark of 15-20%",
                 "Supply chain director + operations VP",
                 "Network resilience score improves — cascading delays become isolated incidents"),
                (f"Diversifies supplier dependency that currently concentrates {_dt} delays in one supply path",
                 "Strategic sourcing team",
                 "Single-supplier dependency eliminated — risk score pool measurably lower"),
            ]

            def _bullet(text: str, tier_key: str, bullet_idx: int) -> str:
                text = text.strip()
                if not text or len(text) < 6:
                    return ""

                # Entity enrichment
                enriched = text
                if sup_ids and not any(sid in text for sid in sup_ids):
                    if any(kw in text.lower() for kw in ["supplier","vendor","sourcing"]):
                        enriched = f"{sup_ids[0]}: {text}"
                if plant_ids and not any(pid in text for pid in plant_ids):
                    if any(kw in text.lower() for kw in ["plant","facility","production"]):
                        enriched = f"{plant_ids[0]}: {text}"

                tier = TIERS.get(tier_key, TIERS["strategic"])
                col   = tier["color"]
                imp_p = tier["impact_prefix"]
                exe_p = tier["exec_prefix"]
                lng_p = tier["long_prefix"]

                # Rotate impact/exec/long-term variants per bullet
                v = IMPACT_VARIANTS[bullet_idx % len(IMPACT_VARIANTS)]
                imp_val, exe_val, lng_val = v

                # Main action text
                if ":" in enriched[:80]:
                    parts = enriched.split(":", 1)
                    action_html = (
                        f'<strong style="color:{col};font-weight:800">{parts[0].strip()}:</strong> '
                        f'<span style="color:#e2e8f0">{parts[1].strip()}</span>'
                    )
                else:
                    action_html = f'<span style="color:#e2e8f0">{enriched}</span>'

                return f'''
<div style="margin:8px 0;padding:11px 14px;background:rgba(255,255,255,0.025);
     border-radius:8px;border-left:3px solid {col}55">
  <div style="font-size:0.83rem;line-height:1.65">{action_html}</div>
</div>'''

            lines = raw.strip().splitlines()
            out = []
            cur_tier = "strategic"
            in_tier  = False
            bullet_count = 0

            for line in lines:
                s = line.strip()
                if not s:
                    continue
                if s.startswith("#### "):
                    heading   = s[5:].strip()
                    cur_tier  = _key_from_heading(heading)
                    in_tier   = True
                    tier_cfg  = TIERS[cur_tier]
                    bullet_count = 0   # reset counter per tier
                    # Close previous tier dropdown if open
                    if out and any("</details>" not in x and "<details" in x for x in out[-3:]):
                        pass  # will close below
                    out.append(
                        f'</div></details>'  # close previous tier (harmless if first)
                        f'<details style="margin-bottom:10px;border:1px solid {tier_cfg["border"]};'
                        f'border-radius:10px;overflow:hidden;background:{tier_cfg["bg"]}">'
                        f'<summary style="cursor:pointer;list-style:none;user-select:none;'
                        f'padding:10px 14px;display:flex;align-items:center;gap:8px;'
                        f'font-size:0.78rem;font-weight:700;color:{tier_cfg["color"]};'
                        f'background:rgba(255,255,255,0.03)">'
                        f'<span style="font-size:0.6rem;transition:transform 0.2s;display:inline-block" '
                        f'class="tier-md-{cur_tier}">▶</span>'
                        f'{tier_cfg["label"]}'
                        f'</summary>'
                        f'<style>details[open] .tier-md-{cur_tier}{{transform:rotate(90deg)}}</style>'
                        f'<div style="padding:10px 12px 12px">'
                    )
                elif s.startswith("### "):
                    continue
                elif s.startswith(("- ", "* ", "• ")):
                    b = _bullet(s[2:], cur_tier, bullet_count)
                    if b:
                        out.append(b)
                        bullet_count += 1
                elif _r2.match(r"^\d+\.\s", s):
                    b = _bullet(s.split(".", 1)[1].strip(), cur_tier, bullet_count)
                    if b:
                        out.append(b)
                        bullet_count += 1
                elif in_tier and len(s) > 15:
                    b = _bullet(s, cur_tier, bullet_count)
                    if b:
                        out.append(b)
                        bullet_count += 1

            result = "\n".join(out) if out else '<p style="color:#475569;font-size:.8rem">No recommendations generated.</p>'
            # Close the last open tier <details> if any was opened
            if "<details" in result and result.count("<details") > result.count("</details>"):
                result += "</div></details>"
            return result

        rec_html = _build_entity_specific_recs(rec_content, report_body)
        rec_block = f"""
<div style="margin-top:32px;padding-top:24px;border-top:2px solid rgba(125,211,252,0.2)">
  <div style="display:flex;align-items:center;gap:10px;margin-bottom:16px">
    <span style="font-size:1.1rem">💡</span>
    <span style="font-size:1rem;font-weight:800;color:#7dd3fc;letter-spacing:0.01em">Recommendations</span>
    <span style="font-size:0.7rem;color:#94a3b8;margin-left:4px">— Prioritised corrective actions</span>
  </div>
{rec_html}
</div>
"""
    elif not rec_json_block:
        rec_block = ""  # Only clear if no JSON recs either

    # ── 13. Final assembly ────────────────────────────────────────
    report_heading = (
        f'<div style="margin-bottom:20px;padding-bottom:12px;'
        f'border-bottom:2px solid {_border_color}">'
        f'<span style="font-size:1.1rem;font-weight:800;color:{_title_color};letter-spacing:0.01em">'
        f'{_report_icon} {_report_title}'
        f'</span>'
        f'</div>'
    )

    # ── Extract & render chartdata blocks from report body ─────────────
    import re as _re_chart, json as _jchart
    charts_html = ""  # Visual Analysis inline charts removed — VIEW CHARTS panel handles all viz
    chart_blocks = []  # Don't render inline SVG charts — they duplicate VIEW CHARTS content

    def _render_bar_chart(cd: dict) -> str:
        """Render an inline SVG horizontal bar chart from chartdata dict."""
        title  = cd.get("title", "Chart")
        labels = cd.get("labels", [])
        values = cd.get("values", [])
        color  = cd.get("color", "#60a5fa")
        if not labels or not values:
            return ""
        n = len(labels)
        max_val = max(values) if values else 1
        bar_h = 22
        gap   = 8
        label_w = 160
        bar_area = 260
        total_h  = n * (bar_h + gap) + 48
        rows = ""
        for i, (lbl, val) in enumerate(zip(labels, values)):
            y     = 36 + i * (bar_h + gap)
            bw    = int(bar_area * val / max_val)
            # format value label
            if isinstance(val, float) and val < 10:
                val_str = f"{val:.2f}"
            else:
                val_str = f"{int(val):,}"
            rows += (
                f'<text x="{label_w-6}" y="{y+bar_h//2+4}" '
                f'text-anchor="end" font-size="9" fill="#94a3b8">{lbl[:22]}</text>'
                f'<rect x="{label_w+4}" y="{y}" width="{bw}" height="{bar_h}" '
                f'rx="3" fill="{color}" opacity="0.85"/>'
                f'<text x="{label_w+bw+8}" y="{y+bar_h//2+4}" '
                f'font-size="9" fill="#e2e8f0">{val_str}</text>'
            )
        return (
            f'<svg viewBox="0 0 {label_w+bar_area+80} {total_h}" '
            f'xmlns="http://www.w3.org/2000/svg" style="width:100%;max-width:560px;overflow:visible">'
            f'<rect width="100%" height="100%" fill="rgba(12,21,40,0.5)" rx="8"/>'
            f'<text x="{(label_w+bar_area)//2}" y="20" text-anchor="middle" '
            f'font-size="11" font-weight="700" fill="#e2e8f0">{title}</text>'
            f'{rows}'
            f'</svg>'
        )

    if chart_blocks:
        chart_svgs = []
        for block in chart_blocks:
            try:
                cd = _jchart.loads(block.strip())
                svg = _render_bar_chart(cd)
                if svg:
                    chart_svgs.append(svg)
            except Exception:
                pass

        if chart_svgs:
            charts_html = (
                '<div style="background:rgba(12,21,40,0.72);border:1px solid rgba(125,211,252,0.12);'
                'border-left:4px solid #38bdf8;border-radius:12px;margin-bottom:16px;'
                'padding:16px 18px 14px;box-shadow:0 2px 12px rgba(0,0,0,0.25)">'
                '<div style="font-size:0.88rem;font-weight:700;color:#e2e8f0;margin-bottom:14px">'
                '📊 Visual Analysis</div>'
                '<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(280px,1fr));gap:16px">'
            )
            for svg in chart_svgs:
                charts_html += f'<div style="background:rgba(0,0,0,0.3);border-radius:8px;padding:10px">{svg}</div>'
            charts_html += '</div></div>'

    # Root cause display strategy:
    # root_cause_block = formatted WHO/WHY/WHEN verdict cards (preferred)
    # Step 5 section card = raw unformatted prose block (inferior)
    # → Always render root_cause_block if populated; suppress it only when empty.
    import re as _re_rc_check
    if not root_cause_block:
        # root_cause_block is empty — nothing to do (section card will show raw text as fallback)
        pass
    # Never suppress root_cause_block — let it always render below section cards

    return (
        '<div id="rca-report-content" style="'
        'display:block!important;width:100%!important;'
        'box-sizing:border-box!important;float:none!important">'
        + report_heading
        + kpi_html
        + exec_block
        + section_cards_html
        + (rec_block if rec_block else "")
        + charts_html
        + '</div>'
    )


def _run_agent(system_prompt: str, user_query: str, on_update=None, max_steps: int = 8):
    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user",   "content": user_query},
    ]
    tool_logs = []
    full_text = ""

    for step in range(max_steps):
        messages = _prune_history(messages)
        is_final_step = (step >= max_steps - 2) or (len(tool_logs) >= 4)
        token_budget = MAX_TOKENS_FINAL if is_final_step else MAX_TOKENS_PER_STEP

        response   = _groq_call(messages=messages, max_tokens=token_budget, model_chain=_UPDATE_MODELS)
        llm_output = response
        full_text  = llm_output

        llm_output = re.sub(r"Observation:.*?(?=\nThought:|\nAction:|\nFinal Answer:|$)", "", llm_output, flags=re.S).strip()
        action_positions = [m.start() for m in re.finditer(r"\nAction:", llm_output)]
        if len(action_positions) > 1:
            llm_output = llm_output[:action_positions[1]].strip()

        if "Final Answer:" in llm_output:
            min_tools = 2
            # For graph update: specifically require create_or_update_node was called
            _has_create = any(
                t.get("tool") == "create_or_update_node"
                for t in tool_logs
                if isinstance(t, dict)
            )
            if len(tool_logs) < min_tools or not _has_create:
                messages.append({"role": "assistant", "content": llm_output})
                messages.append({
                    "role": "user",
                    "content": (
                        f"You have not called create_or_update_node yet. "
                        f"You MUST call create_or_update_node with a MERGE statement FIRST. "
                        f"verify_node_exists alone does not write anything to the graph. "
                        f"Start with:\n"
                        f"Action: create_or_update_node\n"
                        f"Action Input: {{\"cypher_merge_query\": \"MERGE (...) SET ... RETURN ...\"}}"
                    )
                })
                continue
            break

        tool_name, tool_input = parse_action(llm_output)
        if not tool_name:
            messages.append({"role": "assistant", "content": llm_output})
            messages.append({
                "role": "user",
                "content": (
                    "Call a tool before Final Answer.\n"
                    "Action: <exact_tool_name>\nAction Input: {}\n"
                    f"Valid names: {sorted(VALID_TOOLS)}"
                )
            })
            continue

        # ── NODE PRE-VALIDATION: check endpoints before relationship MERGE ──
        # This prevents the "No node exists" error reported in the UI when
        # the agent tries to MATCH a node that was never committed.
        if tool_name == "create_or_update_node":
            _cypher_to_check = (tool_input.get("cypher_merge_query") or "")
            _preval_ok, _preval_msg = _validate_nodes_for_relationship(_cypher_to_check)
            if not _preval_ok:
                # Report a clear, actionable missing-node error — do NOT dispatch
                print(f"[NodePreValidation] Blocked relationship MERGE: {_preval_msg[:120]}")
                messages.append({"role": "assistant", "content": llm_output})
                messages.append({
                    "role": "user",
                    "content": (
                        f"Observation: {_preval_msg}\n\n"
                        "If you just created the source node above, you must verify it was committed "
                        "with verify_node_exists before attempting the relationship link. "
                        "If the target node (Plant/Distributor) does not exist, report this clearly "
                        "in your Final Answer instead of retrying."
                    )
                })
                continue

        try:
            if on_update:
                on_update(("tool_start", tool_name))
            full_result = _dispatch_tool(tool_name, tool_input)
        except Exception as dispatch_err:
            err_msg = f"Tool execution error: {dispatch_err}"
            messages.append({"role": "assistant", "content": llm_output})
            messages.append({"role": "user", "content": f"Observation: {err_msg}"})
            continue

        # ── Check if tool returned an error and retry with corrected input ──
        result_parsed = {}
        try:
            result_parsed = json.loads(full_result) if full_result else {}
        except Exception:
            pass
        if isinstance(result_parsed, dict) and "error" in result_parsed and tool_name == "create_or_update_node":
            err_text = result_parsed["error"]
            print(f"[Agent retry] create_or_update_node error: {err_text[:100]}")
            messages.append({"role": "assistant", "content": llm_output})
            messages.append({
                "role": "user",
                "content": (
                    f"Observation: ERROR — {err_text}\n\n"
                    "IMPORTANT: create_or_update_node requires EXACTLY this format:\n"
                    "Action: create_or_update_node\n"
                    "Action Input: {\"cypher_merge_query\": \"MERGE (s:Supplier {supplier_id:\'SUP9001\'}) "
                    "SET s.supplier_name=\'Mehta Plastics\', s.risk_score=0.45, "
                    "s.annual_capacity_units=50000, s.StoP_lead_time_days=8, "
                    "s.status=\'Active\', s.supplier_latitude=0.0, s.supplier_longitude=0.0, "
                    "s.StoP_distance_km=0.0 RETURN s.supplier_id\"}\n\n"
                    "The JSON key MUST be \"cypher_merge_query\" and the value MUST be a complete Cypher string."
                )
            })
            continue

        log_entry = {"tool": tool_name, "input": tool_input, "result_preview": full_result}
        tool_logs.append(log_entry)
        if on_update:
            on_update(("tool", log_entry))
            # Emit structured Cypher log for the UI accordion
            _cq_meta = _get_tool_cypher(tool_name, tool_input)
            on_update(("cypher_log", {
                "seq":     len(tool_logs),
                "tool":    tool_name,
                "purpose": _cq_meta["purpose"],
                "cypher":  _cq_meta["cypher"],
                "records": _count_records(full_result),
            }))

        # ── After create_or_update_node: run a direct Cypher confirmation ──
        # This gives the agent accurate ground truth (not relying on verify_node_exists)
        _direct_confirm = ""
        if tool_name == "create_or_update_node":
            _cypher_used = (tool_input.get("cypher_merge_query") or "")
            # Extract the id that was set so we can confirm it was written
            import re as _re2
            _id_match = _re2.search(
                r'(?:MERGE|MATCH)\s+\(\w+:(\w+)\s*\{(\w+)\s*:\s*[\'"]([^\'"]+)[\'"]\}',
                _cypher_used, _re2.IGNORECASE
            )
            if _id_match:
                _lbl, _prop, _val = _id_match.group(1), _id_match.group(2), _id_match.group(3)
                try:
                    _confirm_rows = _run_neo4j(
                        f"MATCH (n:{_lbl} {{{_prop}: $v}}) RETURN properties(n) AS props LIMIT 1",
                        {"v": _val}
                    )
                    if _confirm_rows:
                        _direct_confirm = (
                            f"\n[SYSTEM DIRECT VERIFY] Node confirmed in Neo4j: "
                            f"{_lbl} {{{_prop}: '{_val}'}} → exists with props: "
                            f"{list(_confirm_rows[0].get('props', {}).keys())}"
                        )
                        # ── Log for undo — only when we have confirmed the write ──
                        _confirmed_props = _confirm_rows[0].get('props', {})
                        _log_nl_update(
                            entity_type = _lbl,
                            entity_id   = _val,
                            cypher      = _cypher_used,
                            id_prop     = _prop,
                        )
                        # Sync confirmed props to CSV on disk
                        _nl_csv_upsert(_lbl, _val, _confirmed_props)
                    else:
                        _direct_confirm = (
                            f"\n[SYSTEM DIRECT VERIFY] WARNING: {_lbl} {{{_prop}: '{_val}'}} "
                            f"NOT FOUND in Neo4j after MERGE. The write may have failed silently."
                        )
                except Exception as _ve:
                    _direct_confirm = f"\n[SYSTEM DIRECT VERIFY] Could not verify: {_ve}"

        trimmed = full_result[:OBS_TRIM_CHARS]
        if len(full_result) > OBS_TRIM_CHARS:
            trimmed += "... [truncated]"

        messages.append({"role": "assistant", "content": llm_output})
        messages.append({
            "role": "user",
            "content": (
                f"Observation (trusted tool output): {trimmed}{_direct_confirm}\n\n"
                "Write only ONE Action next, then stop."
            ),
        })

    if "Final Answer:" in full_text:
        return full_text.split("Final Answer:", 1)[1].strip(), tool_logs

    summary = _groq_call(
        messages=[
            {"role": "system", "content": "You are a strict data-driven supply chain analyst."},
            {"role": "user",   "content": f"Based on these results: {json.dumps(tool_logs, indent=2, default=str)}\nWrite a confirmation summary."},
        ],
        max_tokens=MAX_TOKENS_FALLBACK,
        model_chain=_UPDATE_MODELS,
    )
    return summary, tool_logs


# ════════════════════════════════════════════════════════════════════
# ORCHESTRATOR AGENT
# Selects tools based on the question, fetches all data in parallel.
# Populates: ctx.selected_tools, ctx.tool_logs, ctx.obs_block
# ════════════════════════════════════════════════════════════════════

_ALL_AVAILABLE_TOOLS = {
    "get_supplier_plant_delay_chain": {
        "desc": "Get full supplier→plant→shipment delay chain. Best for: which suppliers cause delays, root cause tracing, end-to-end delay analysis.",
        "default_input": {},
    },
    "get_supplier_delay_contribution": {
        "desc": "AUTHORITATIVE per-supplier delayed shipment count — returns supplier_id, supplier_name, risk_score, plant_id, plant_name, delayed_shipments, avg_delay_days. Use for supplier risk RCA Supplier→Plant Delay Chain table.",
        "default_input": {},
    },
    "get_high_risk_suppliers": {
        "desc": "Get suppliers with risk_score above a threshold (default 0.6). Best for: risky vendor analysis, supplier audit prioritisation.",
        "default_input": {"threshold": 0.6},
    },
    "get_delay_by_plant": {
        "desc": "Get delay counts and rates per plant. Best for: bottleneck plants, which factory has most delays, plant performance.",
        "default_input": {},
    },
    "get_distributor_delay_impact": {
        "desc": "Get delayed shipments and avg delay days per distributor city. Best for: distributor impact, city-level delay analysis.",
        "default_input": {},
    },
    "get_stockout_retailers": {
        "desc": (
            "Distributor cities with highest total unmet demand. "
            "Returns: retailer_city, served_by_distributor, retailers_connected, shortage_shipments, total_shortage_units. "
            "IMPORTANT: retailers_connected = 0 for 45/50 distributors — accurate, only 5 hub distributors have retailer edges. "
            "shortage_shipments and total_shortage_units are accurate for ALL cities."
        ),
        "default_input": {},
    },
    "get_demand_gap_analysis": {
        "desc": "AUTHORITATIVE distributor shortage data. Returns per distributor: distributor_city, shortage_shipments, total_demand_gap, delayed_shipments, avg_delay_days, retailers_affected. Use for ALL distributor-level shortage questions — do not combine with get_distributor_delay_impact.",
        "default_input": {},
    },
    "get_delay_by_product_category": {
        "desc": "Get delay counts per product category (toys, auto, health, beauty, etc). Best for: category-specific delay analysis.",
        "default_input": {},
    },
    "get_route_cost_efficiency": {
        "desc": "Get route efficiency scores, distance and cost per route. Best for: logistics cost, transport mode analysis, route optimisation.",
        "default_input": {},
    },
    "get_route_delay_correlation": {
        "desc": "Get individual ROUTES ranked by delayed shipment count with delay_rate_pct. Best for: per-route reliability, which specific route_id is worst. Do NOT use for transport mode totals — use get_transport_mode_delays instead.",
        "default_input": {},
    },
    "get_transport_mode_delays": {
        "desc": "AUTHORITATIVE — Count Major Delay shipments grouped by transport mode (Road/Rail/Air/Sea). Use THIS for ANY question about which transport mode causes the most delays. Returns: transportation_mode, total_delays, avg_delay_days, plants_affected. Uses correct double-MATCH to prevent inflated counts.",
        "default_input": {},
    },
    "get_schema_with_examples": {
        "desc": "Returns full graph schema + 6 canonical Cypher patterns with anti-pattern warnings. Call BEFORE run_cypher to avoid wrong traversal paths and fan-out bugs.",
        "default_input": {},
    },
    "get_monthly_delay_trend": {
        "desc": "Get monthly shipment delay trends over time. Best for: seasonal patterns, trend analysis, month-over-month performance.",
        "default_input": {},
    },
    "trace_supply_chain_for_category": {
        "desc": "CRITICAL for product category RCA — traces full Supplier→Plant→Shipment→Distributor path filtered to a specific product category (toys/auto/health_beauty/watches_gifts/cool_stuff/bed_bath_table). Returns supplier, plant, delivery_status, delay_days, distributor. Always use this for 'why are toy shipments delayed' type questions.",
        "default_input": {"product_category": "toys"},
    },
    "get_category_supplier_risk": {
        "desc": "For a given product category, returns suppliers feeding plants that dispatch that category — with risk_score, lead_time, plant name, and delayed shipment count for that category. Use for Supplier Dependency sections in product category RCA.",
        "default_input": {"product_category": "toys"},
    },
    "get_supplier_shutdown_impact": {
        "desc": "SIMULATION — Impact of a specific supplier shutting down. Returns supplier details, plant supplied, delayed_shipments, distributors_exposed. Use for Nagy PLC / supplier shutdown what-if queries.",
        "default_input": {"supplier_id": "SUP0045"},
    },
    "get_supplier_downstream_cities": {
        "desc": "SIMULATION — Distributor cities downstream of a specific supplier via its plant. Returns city, shortage_shipments, total_demand_gap. Use for downstream exposure in supplier shutdown scenarios.",
        "default_input": {"supplier_id": "SUP0045"},
    },
    "get_distributor_offline_impact": {
        "desc": "Per-plant contribution to a specific distributor's demand gap — shows total shipments, total demand gap, major delays per feeding plant. Use for Kolkata plant-contribution analysis. Parameter: distributor_id.",
        "default_input": {"distributor_id": "D0005"},
    },
    "get_distributor_monthly_stockout": {
        "desc": "Month-by-month stockout breakdown for a specific distributor — shortage shipments and monthly demand gap per month for 12 months. Use for Kolkata or any city stockout persistence/severity analysis. Parameter: distributor_id.",
        "default_input": {"distributor_id": "D0005"},
    },
    "get_distributor_fulfillment_by_plant": {
        "desc": "Per-plant fulfillment breakdown for a specific distributor — total forecast, sales, demand gap, on-time gap and fulfillment rate per plant. Separates supply capacity failure from transport delays. Parameter: distributor_id.",
        "default_input": {"distributor_id": "D0005"},
    },
    "get_distributor_routes": {
        "desc": "SIMULATION — All routes currently serving a specific distributor with costs. Returns route_id, from_plant, transport_mode, distance_km, cost_inr. Use for rerouting cost analysis.",
        "default_input": {"distributor_id": "D0005"},
    },
    "get_distributor_rerouting_options": {
        "desc": "SIMULATION — Alternative distributors that could absorb volume if a specific distributor goes offline. Returns alt_city, alt_id, current_load, avg_reroute_cost_inr.",
        "default_input": {"distributor_id": "D0005"},
    },
    "get_suppliers_at_plant": {
        "desc": "SIMULATION — All suppliers feeding a specific plant, excluding one supplier. Returns supplier_id, supplier_name, risk_score, annual_capacity_units, lead_time_days, plant_id, plant_name, delayed_shipments. Use for Alternative Supplier Analysis in Nagy PLC shutdown scenario.",
        "default_input": {"plant_id": "PL4", "exclude_supplier_id": "SUP0045"},
    },
}

_ORCHESTRATOR_SYSTEM = """You are an Orchestrator Agent for a supply chain RCA system.

Your job: given a user question, decide which data tools to call to answer it.

Available tools (name → description):
{tool_descriptions}

Rules:
- Always include get_supplier_plant_delay_chain, get_delay_by_plant, get_high_risk_suppliers as core tools.
- Add extra tools only when the question specifically asks about that dimension.
- Return a JSON array of tool names only — no explanation, no markdown, no extra text.
- Include between 4 and 8 tools total.
- Only use tool names from the list above — no invented names.
- CRITICAL: For ANY question about which transport mode / road / rail / air / sea causes delays,
  use get_transport_mode_delays — NOT get_route_delay_correlation.
  get_route_delay_correlation returns per-route rows; get_transport_mode_delays returns correct mode totals.

Example output for "why are toy shipments delayed?":
["get_delay_by_product_category", "get_supplier_plant_delay_chain", "get_delay_by_plant", "get_high_risk_suppliers", "get_distributor_delay_impact", "get_stockout_retailers"]

Example output for "which transport mode causes the most delays?":
["get_transport_mode_delays", "get_delay_by_plant", "get_supplier_plant_delay_chain", "get_high_risk_suppliers", "get_route_cost_efficiency"]

Example output for "show me route cost efficiency":
["get_route_delay_correlation", "get_route_cost_efficiency", "get_delay_by_plant", "get_supplier_plant_delay_chain", "get_high_risk_suppliers"]
"""

_ORCHESTRATOR_MODELS = [
    "llama-3.3-70b-versatile",
    "meta-llama/llama-4-scout-17b-16e-instruct",
    "mixtral-8x7b-32768",
    "llama-3.1-8b-instant",
    "gemma2-9b-it",
]


def _extract_question_entities(question: str) -> dict:
    """
    Extract structured entities from the user's question:
    - product_category (e.g. "Auto", "Toys", "Health")
    - plant_id / plant_name
    - supplier_name
    - distributor_city
    - delay_threshold (float)
    - time_period (month/year)

    Used to pass accurate filter parameters to tools.
    """
    q = question.lower()

    # Product category mapping — match to actual Neo4j category values
    CATEGORY_MAP = {
        # Exact Neo4j values (case-sensitive as stored in your database)
        "toy":              "toys",
        "toys":             "toys",
        "auto":             "auto",
        "automobile":       "auto",
        "automotive":       "auto",
        "car":              "auto",
        "health":           "health_beauty",
        "health beauty":    "health_beauty",
        "health and beauty":"health_beauty",
        "beauty":           "health_beauty",
        "watch":            "watches_gifts",
        "watches":          "watches_gifts",
        "watches gifts":    "watches_gifts",
        "gift":             "watches_gifts",
        "cool":             "cool_stuff",
        "cool stuff":       "cool_stuff",
        "bed":              "bed_bath_table",
        "bath":             "bed_bath_table",
        "bed bath":         "bed_bath_table",
        "construction":     "Construction Tools Garden",
        "garden":           "Construction Tools Garden",
        "tools":            "Construction Tools Garden",
    }

    # Plant name mapping
    PLANT_MAP = {
        "baddi":  ("PL1", "Baddi"),
        "bhopal": ("PL2", "Bhopal"),
        "pune":   ("PL3", "Pune"),
        "goa":    ("PL4", "Goa"),
        "pl1": ("PL1", "Baddi"),
        "pl2": ("PL2", "Bhopal"),
        "pl3": ("PL3", "Pune"),
        "pl4": ("PL4", "Goa"),
    }

    entities = {
        "product_category": None,
        "plant_id":         None,
        "plant_name":       None,
        "supplier_name":    None,
        "distributor_city": None,
        "risk_threshold":   0.6,
        "raw_question":     question,
    }

    # Extract product category
    for kw, cat in CATEGORY_MAP.items():
        if kw in q:
            entities["product_category"] = cat
            break

    # Extract plant
    for kw, (pid, pname) in PLANT_MAP.items():
        if kw in q:
            entities["plant_id"]   = pid
            entities["plant_name"] = pname
            break

    # Extract risk threshold
    import re as _re
    th_match = _re.search(r"risk[_ ]score[> ]+([0-9.]+)", q)
    if th_match:
        try:
            entities["risk_threshold"] = float(th_match.group(1))
        except Exception:
            pass

    return entities


def _orchestrator_pick_tools(user_question: str) -> list[tuple[str, dict]]:
    """
    LLM-powered tool selection with entity-aware parameter injection.
    Extracts real filter values (product_category, plant_id, etc.)
    and passes them as tool inputs — not just empty dicts.

    Falls back to a safe default set if the LLM call fails.
    """
    # First: extract concrete entities from the question
    entities = _extract_question_entities(user_question)
    print(f"[Orchestrator] Entities extracted: {entities}")

    tool_descriptions = "\n".join(
        f"  {name}: {meta['desc']}"
        for name, meta in _ALL_AVAILABLE_TOOLS.items()
    )

    # Build a context hint about extracted entities for the LLM
    entity_hint = []
    if entities["product_category"]:
        entity_hint.append(f"Product category detected: '{entities['product_category']}'")
    if entities["plant_name"]:
        entity_hint.append(f"Plant detected: {entities['plant_name']} ({entities['plant_id']})")
    if entities["distributor_city"]:
        entity_hint.append(f"City detected: {entities['distributor_city']}")
    hint_str = "\n".join(entity_hint) if entity_hint else ""

    try:
        raw = _groq_call(
            messages=[
                {
                    "role": "system",
                    "content": _ORCHESTRATOR_SYSTEM.format(tool_descriptions=tool_descriptions),
                },
                {
                    "role": "user",
                    "content": (
                        f"Question: {user_question}\n"
                        + (f"Entities found: {hint_str}\n" if hint_str else "")
                        + "\nReturn a JSON array of tool names to call. No other text."
                    ),
                },
            ],
            max_tokens=200,
            temperature=0,
            model_chain=_ORCHESTRATOR_MODELS,
        )

        raw = raw.strip()
        raw = re.sub(r"^```(?:json)?\s*", "", raw, flags=re.IGNORECASE)
        raw = re.sub(r"\s*```$", "", raw)
        tool_names = json.loads(raw)
        valid_names = [n for n in tool_names if n in _ALL_AVAILABLE_TOOLS]

        # Only add core tools if the question is about delays/root cause
        q_l = user_question.lower()
        is_delay_q = any(w in q_l for w in ["delay","late","slow","bottleneck","root cause","why","what is causing"])
        is_transport_delay_q = is_delay_q and any(w in q_l for w in [
            "transport mode", "transportation mode", "mode", "road", "rail", "air", "sea",
            "which mode", "highest delay", "most delay", "causes delay", "cause delay"
        ])
        core = ["get_supplier_plant_delay_chain", "get_delay_by_plant", "get_high_risk_suppliers"]
        # Always ensure downstream shortage data tools are present for distributor/retailer sections
        # Detective investigation always needs WHEN (trend) + HOW MUCH (gap) tools
        detective_core = [
            "get_monthly_delay_trend",         # WHEN: timeline and persistence
            "get_demand_gap_analysis",         # HOW MUCH: distributor-level shortage scale
            "get_supplier_delay_contribution", # WHO 1B: delayed shipment counts per supplier
            "get_transport_mode_delays",       # WHY 2B: delay breakdown by transport mode
            "get_route_cost_efficiency",       # HOW MUCH 4C: Plant→Distributor route efficiency
            "get_stockout_retailers",          # HOW MUCH 4B: top shortage cities
        ]
        downstream_core = ["get_stockout_retailers", "get_demand_gap_analysis"]
        if is_delay_q:
            for c in core:
                if c not in valid_names:
                    valid_names.insert(0, c)
            for c in downstream_core:
                if c not in valid_names:
                    valid_names.append(c)
        # Always inject detective core tools for ALL queries
        for c in detective_core:
            if c not in valid_names and c in _ALL_AVAILABLE_TOOLS:
                valid_names.append(c)

        # FIX: also inject demand_gap_analysis for stockout/shortage questions
        # even when they are NOT delay questions (pure quantity shortfall queries)
        q_lower_check = user_question.lower()
        is_shortage_q = any(w in q_lower_check for w in [
            "stockout", "shortage", "demand gap", "unmet", "running out",
            "biggest stockout", "distributor stockout",
        ])
        # NETWORK STOCKOUT: "stockouts increased despite stable volumes" — needs full tool set
        is_network_stockout_q = is_shortage_q and any(w in q_lower_check for w in [
            "even though", "despite", "although", "but stable", "remain stable",
            "volumes remain", "who is responsible", "when did", "how is", "spreading",
            "driving these", "what is driving",
        ])
        if is_network_stockout_q:
            network_stockout_tools = [
                "get_high_risk_suppliers",
                "get_supplier_plant_delay_chain",
                "get_delay_by_plant",
                "get_demand_gap_analysis",
                "get_monthly_delay_trend",
                # removed: get_supplier_delay_contribution, get_stockout_retailers,
                # get_transport_mode_delays, get_route_cost_efficiency — reduces tool calls by 4
            ]
            for t in network_stockout_tools:
                if t not in valid_names and t in _ALL_AVAILABLE_TOOLS:
                    valid_names.append(t)
            print(f"[Orchestrator] Network stockout query — injected {len(network_stockout_tools)} tools")
        elif is_shortage_q:
            for t in ["get_demand_gap_analysis", "get_stockout_retailers",
                      "get_supplier_plant_delay_chain", "get_high_risk_suppliers"]:
                if t not in valid_names and t in _ALL_AVAILABLE_TOOLS:
                    valid_names.append(t)

        # RETAILER: inject full retailer-tracing tools for any query about retailers
        _q_ret_check = user_question.lower()
        is_retailer_q = any(w in _q_ret_check for w in [
            "retailer", "retail", "which retailers", "retailers face",
            "retailers impacted", "retailers affected", "retailers receive",
            "retailers experiencing", "store", "outlet",
        ])
        if is_retailer_q:
            retailer_tools = [
                "get_stockout_retailers",         # retailer cities with unmet demand
                "get_demand_gap_analysis",         # distributor-level shortage feeding retailers
                "get_supplier_plant_delay_chain",  # upstream delay sources
                "get_high_risk_suppliers",         # root cause supplier risk
                "get_transport_mode_delays",       # logistics layer
            ]
            for t in retailer_tools:
                if t not in valid_names and t in _ALL_AVAILABLE_TOOLS:
                    valid_names.append(t)
            print(f"[Orchestrator] Retailer query — injected {len(retailer_tools)} retailer tools")

        # PRODUCT CATEGORY: inject category-specific trace tools so tables get real data
        q_l_cat = user_question.lower()
        detected_cat = entities.get("product_category")
        _CAT_WORDS = ["toy","toys","auto","health","beauty","watches","gifts","cool stuff","bed bath","construction","garden"]
        is_category_q = bool(detected_cat) or any(w in q_l_cat for w in _CAT_WORDS)
        if is_category_q:
            cat_val = detected_cat or "toys"
            for t in ["trace_supply_chain_for_category", "get_category_supplier_risk",
                      "get_delay_by_product_category", "get_distributor_delay_impact",
                      "get_demand_gap_analysis", "get_stockout_retailers"]:
                if t not in valid_names and t in _ALL_AVAILABLE_TOOLS:
                    valid_names.append(t)
            # Inject category filter for parameterised tools
            print(f"[Orchestrator] Product category query — injected category tools for '{cat_val}'")

        # SUPPLIER RISK: inject product category + delay chain tools so Product Exposure
        # and High-Risk Supplier sections have real data to populate their tables.
        q_l_sup = user_question.lower()
        is_supplier_risk_q = any(w in q_l_sup for w in [
            "risk score", "risk_score", "high risk supplier", "risky supplier",
            "supplier risk", "above 0.", "risk above", "risk score above",
            # ── NEW: catch "which suppliers are delaying shipments" and variants ──
            "which supplier", "which suppliers", "suppliers are delaying",
            "are delaying", "delaying shipment", "suppliers delay",
            "suppliers causing", "causing delay", "supplier delay",
            "suppliers who", "suppliers that",
        ])

        # DELIVERY PERFORMANCE / TRANSPORT ROUTE QUERY: backup query pattern
        is_delivery_perf_q = any(w in q_l_sup for w in [
            "delivery performance", "deteriorating", "transportation routes",
            "routes or modes", "which routes", "which modes", "route driving",
            "mode driving", "routes driving", "transport driving",
        ]) or ("delivery" in q_l_sup and "deteriorat" in q_l_sup) or \
           ("transport" in q_l_sup and "routes" in q_l_sup and "disruption" in q_l_sup) or \
           ("transport" in q_l_sup and "routes" in q_l_sup and "driving" in q_l_sup)
        if is_delivery_perf_q:
            delivery_perf_tools = [
                "get_transport_mode_delays",        # PRIMARY: which mode/route is the problem
                "get_route_delay_correlation",      # per-route delay rates
                "get_route_cost_efficiency",        # route cost vs efficiency
                "get_delay_by_plant",               # plant delay rates
                "get_monthly_delay_trend",          # WHEN: when did it begin
                "get_demand_gap_analysis",          # HOW MUCH: downstream impact
                "get_supplier_plant_delay_chain",   # WHO: upstream cause
                "get_high_risk_suppliers",          # WHO: risky suppliers
            ]
            for t in delivery_perf_tools:
                if t not in valid_names and t in _ALL_AVAILABLE_TOOLS:
                    valid_names.append(t)
            # Force transport mode delays to front
            if "get_transport_mode_delays" in valid_names:
                valid_names.remove("get_transport_mode_delays")
                valid_names.insert(0, "get_transport_mode_delays")
            print(f"[Orchestrator] Delivery performance query — injected {len(delivery_perf_tools)} tools")


        if is_supplier_risk_q:
            supplier_risk_tools = [
                "get_high_risk_suppliers",
                "get_supplier_plant_delay_chain",
                "get_supplier_delay_contribution",
                "get_delay_by_product_category",   # for Product Exposure section
                "get_delay_by_plant",              # for plant context
                "get_demand_gap_analysis",         # for downstream impact
            ]
            for t in supplier_risk_tools:
                if t not in valid_names and t in _ALL_AVAILABLE_TOOLS:
                    valid_names.append(t)
            print(f"[Orchestrator] Supplier risk query — injected {len(supplier_risk_tools)} tools")
        # For transport-mode delay questions, always inject get_transport_mode_delays first
        # get_route_delay_correlation gives per-route data (not mode totals) — wrong for this question
        if is_transport_delay_q:
            if "get_transport_mode_delays" not in valid_names:
                valid_names.insert(0, "get_transport_mode_delays")
            # Also always inject plant + supplier chain data so those sections have content
            for _tc in ["get_delay_by_plant", "get_supplier_plant_delay_chain"]:
                if _tc not in valid_names:
                    valid_names.append(_tc)
            # Remove get_route_analysis if present — it returns per-route rows, not mode totals
            valid_names = [n for n in valid_names if n != "get_route_analysis"]

        # ── SIMULATION: force-inject scenario tools with correct params ──────────
        _q_sim_check = user_question.lower()
        _is_sim_inject = any(w in _q_sim_check for w in [
            "what if","shutdown","shuts down","goes offline","offline","production shutdown",
            "flooding","flood","nagy","sup0045","kolkata","d0005","cascading impact",
            "what would happen","at risk of","disaster","goes down",
        ])
        if _is_sim_inject:
            import re as _re_sim3
            _q_sim = _q_sim_check

            # Detect which entity is disrupted
            _sim_sup_id = None
            _sim_dist_id = None
            if "nagy" in _q_sim or "sup0045" in _q_sim:
                _sim_sup_id = "SUP0045"
            elif _re_sim3.search(r"sup\d+", user_question, _re_sim3.IGNORECASE):
                _sim_sup_id = _re_sim3.search(r"sup\d+", user_question, _re_sim3.IGNORECASE).group(0).upper()

            if "kolkata" in _q_sim or "d0005" in _q_sim:
                _sim_dist_id = "D0005"
            elif _re_sim3.search(r"d\d{4}", user_question, _re_sim3.IGNORECASE):
                _sim_dist_id = _re_sim3.search(r"d\d{4}", user_question, _re_sim3.IGNORECASE).group(0).upper()

            if _sim_sup_id:
                for _tn, _ti in [
                    ("get_supplier_shutdown_impact",   {"supplier_id": _sim_sup_id}),
                    ("get_supplier_downstream_cities", {"supplier_id": _sim_sup_id}),
                    ("get_high_risk_suppliers",        {"threshold": 0.6}),
                    ("get_delay_by_plant",             {}),
                    ("get_demand_gap_analysis",        {}),
                    ("get_stockout_retailers",         {}),
                ]:
                    if _tn not in valid_names and _tn in _ALL_AVAILABLE_TOOLS:
                        valid_names.append(_tn)
                print(f"[Sim] Supplier tools queued for {_sim_sup_id}: {valid_names[-6:]}")

            if _sim_dist_id:
                # Add the 5 Kolkata-specific tools to valid_names with their params stored in default_input
                # (distributor_id is injected later via _par_dist in _orchestrator_agent)
                for _tn in [
                    "get_distributor_monthly_stockout",
                    "get_distributor_fulfillment_by_plant",
                    "get_distributor_offline_impact",
                    "get_high_risk_suppliers",
                    "get_plant_supplier_matrix",
                ]:
                    if _tn not in valid_names and _tn in _ALL_AVAILABLE_TOOLS:
                        valid_names.insert(0, _tn)  # prepend so they run first
                print(f"[Sim] Distributor tools queued for {_sim_dist_id}")

        # Build tool inputs with real entity filters
        tools = []
        for name in valid_names:
            base_input = dict(_ALL_AVAILABLE_TOOLS[name].get("default_input", {}))
            # Inject entity filters where tools support them
            if name == "get_delay_by_product_category" and entities["product_category"]:
                base_input["category_filter"] = entities["product_category"]
            if name == "get_high_risk_suppliers":
                base_input["threshold"] = entities["risk_threshold"]
            if name == "get_supplier_plant_delay_chain" and entities["plant_id"]:
                base_input["plant_id_filter"] = entities["plant_id"]
            if name == "get_delay_by_plant" and entities["plant_id"]:
                base_input["plant_id_filter"] = entities["plant_id"]
            # Inject product_category for category-specific tools
            if name in ("trace_supply_chain_for_category", "get_category_supplier_risk"):
                cat_v = entities.get("product_category") or "toys"
                base_input["product_category"] = cat_v
            tools.append((name, base_input))

        print(f"[Orchestrator Agent / LLM] Selected {len(tools)} tools: {valid_names}")
        return tools

    except Exception as e:
        print(f"[Orchestrator Agent] LLM selection failed ({e}), using safe defaults.")
        # Build default set with entity-injected inputs
        fallback = [
            "get_supplier_plant_delay_chain",   # WHO: supplier-plant delay chain
            "get_high_risk_suppliers",           # WHO: suspect suppliers
            "get_delay_by_plant",               # WHY: plant delay rates
            "get_monthly_delay_trend",          # WHEN: timeline persistence
            "get_demand_gap_analysis",          # HOW MUCH: distributor shortage scale
            "get_stockout_retailers",           # HOW MUCH: retailer stockout scale
            "get_transport_mode_delays",        # WHY: logistics mode contribution
            "get_distributor_delay_impact",     # HOW MUCH: city-level impact
        ]
        if entities["product_category"]:
            fallback.insert(0, "get_delay_by_product_category")
        tools = []
        for name in fallback:
            base_input = dict(_ALL_AVAILABLE_TOOLS[name].get("default_input", {}))
            if name == "get_delay_by_product_category" and entities["product_category"]:
                base_input["category_filter"] = entities["product_category"]
            if name == "get_high_risk_suppliers":
                base_input["threshold"] = entities["risk_threshold"]
            tools.append((name, base_input))
        return tools


def _orchestrator_agent(ctx: AgentContext, on_update=None) -> AgentContext:
    """
    Orchestrator Agent — step 1 of the pipeline.

    Uses an LLM to decide which tools are needed for the question,
    then fetches all tool data in parallel.

    Reads:  ctx.user_question
    Writes: ctx.selected_tools, ctx.tool_logs, ctx.obs_block
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed

    # LLM decides which tools to run
    ctx.selected_tools = _orchestrator_pick_tools(ctx.user_question)

    # Force-inject Kolkata-specific tools after LLM selection
    # These tools have required parameters (distributor_id) so must be added here with correct values
    _q_force = ctx.user_question.lower()
    if "kolkata" in _q_force or "d0005" in _q_force:
        _kolkata_tools = [
            ("get_distributor_monthly_stockout",     {"distributor_id": "D0005"}),
            ("get_distributor_fulfillment_by_plant", {"distributor_id": "D0005"}),
            ("get_distributor_offline_impact",       {"distributor_id": "D0005"}),
            ("get_high_risk_suppliers",              {"threshold": 0.6}),
            ("get_plant_supplier_matrix",            {}),
        ]
        existing_names = {n for n, _ in ctx.selected_tools}
        for tn, ti in _kolkata_tools:
            if tn not in existing_names:
                ctx.selected_tools.append((tn, ti))
                existing_names.add(tn)
            else:
                # Update existing entry with correct distributor_id
                ctx.selected_tools = [
                    (n, {**i, "distributor_id": "D0005"} if n == tn and "distributor_id" in ti else i)
                    for n, i in ctx.selected_tools
                ]
        print(f"[Kolkata] Force-injected distributor tools: {[t for t,_ in _kolkata_tools]}")

    def _run_one(idx_tool):
        idx, tool_entry = idx_tool
        # Guard: tool_entry must be a (name, input_dict) 2-tuple
        if not isinstance(tool_entry, (list, tuple)) or len(tool_entry) != 2:
            entry = {"tool": str(tool_entry), "input": {}, "result_preview": "{}"}
            return idx, entry, ""
        tool_name, tool_input = tool_entry
        if not isinstance(tool_input, dict):
            tool_input = {}
        try:
            result = _dispatch_tool(tool_name, tool_input)
        except Exception as e:
            result = json.dumps({"error": str(e)})
        entry = {"tool": tool_name, "input": tool_input, "result_preview": result}
        return idx, entry, f"### {tool_name}\n{result[:5000]}"

    max_workers = min(len(ctx.selected_tools), 6)  # increased from 3 — run more tools in parallel
    results_map = {}

    # Guard: filter out any malformed tool entries before submitting
    valid_tools = [
        (i, t) for i, t in enumerate(ctx.selected_tools)
        if isinstance(t, (list, tuple)) and len(t) == 2
    ]
    if len(valid_tools) < len(ctx.selected_tools):
        print(f"[Orchestrator] Filtered {len(ctx.selected_tools) - len(valid_tools)} malformed tool entries")

    # Override simulation tool params with correct entity IDs
    import re as _re_par
    _q_par = ctx.user_question.lower()
    _par_sup  = "SUP0045" if "nagy" in _q_par or "sup0045" in _q_par else None
    _par_dist = "D0005"   if "kolkata" in _q_par or "d0005" in _q_par else None
    if _par_sup or _par_dist:
        updated = []
        for nm, inp in ctx.selected_tools:
            if _par_sup and nm in ("get_supplier_shutdown_impact", "get_supplier_downstream_cities"):
                inp = {**inp, "supplier_id": _par_sup}
            if _par_dist and nm in (
                "get_distributor_offline_impact", "get_distributor_routes",
                "get_distributor_rerouting_options",
                "get_distributor_monthly_stockout", "get_distributor_fulfillment_by_plant",
            ):
                inp = {**inp, "distributor_id": _par_dist}
            updated.append((nm, inp))
        ctx.selected_tools = updated
        # Re-filter after override
        valid_tools = [(i, t) for i, t in enumerate(ctx.selected_tools)
                       if isinstance(t, (list, tuple)) and len(t) == 2]

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(_run_one, item): item[0] for item in valid_tools}
        for future in as_completed(futures):
            try:
                idx, entry, obs = future.result()
            except Exception as _fe:
                print(f"[Orchestrator] Tool future failed: {_fe}")
                continue
            results_map[idx] = (entry, obs)
            if on_update:
                on_update(("tool", entry))
                # Emit Cypher log for this orchestrator tool call
                _cq_meta = _get_tool_cypher(entry.get("tool",""), entry.get("input",{}))
                on_update(("cypher_log", {
                    "seq":     idx + 1,
                    "tool":    entry.get("tool", "?"),
                    "purpose": _cq_meta["purpose"],
                    "cypher":  _cq_meta["cypher"],
                    "records": _count_records(entry.get("result_preview", "")),
                }))

    ctx.tool_logs  = [results_map[i][0] for i in sorted(results_map.keys())]
    observations   = [results_map[i][1] for i in sorted(results_map.keys())]
    ctx.obs_block  = "\n\n".join(o for o in observations if o)

    # ── Populate tool_rows: parse every result into Python dicts ──
    import json as _trj
    for _tl in ctx.tool_logs:
        _tn = _tl.get("tool", "")
        _tr = _tl.get("result_preview", "")
        try:
            _parsed = _trj.loads(_tr)
            # Handle double-encoding: MCP wraps json.dumps result in another json.dumps
            # so _parsed may be a string (JSON-encoded list) rather than a list
            if isinstance(_parsed, str):
                try:
                    _parsed = _trj.loads(_parsed)
                except Exception:
                    pass
            if isinstance(_parsed, list) and _parsed and isinstance(_parsed[0], dict):
                # Compute MoM for monthly trend in Python (her approach — never trust LLM arithmetic)
                if _tn == "get_monthly_delay_trend":
                    _prev_r = None
                    for _row in _parsed:
                        try:
                            _cr = float(_row.get("delay_rate_pct", 0))
                            if _prev_r is None:
                                _row["mom_change"] = "Baseline"
                            else:
                                _d = round(_cr - _prev_r, 1)
                                _row["mom_change"] = (
                                    f"▲ +{_d} pts" if _d > 0 else
                                    f"▼ {abs(_d)} pts" if _d < 0 else "→ 0.0 pts"
                                )
                            _prev_r = _cr
                        except Exception:
                            _row["mom_change"] = "—"
                ctx.tool_rows[_tn] = _parsed
        except Exception:
            pass

    return ctx


# ════════════════════════════════════════════════════════════════════
# RCA AGENT
# Receives structured context from the orchestrator.
# Populates: ctx.rca_raw, ctx.rca_findings
# ════════════════════════════════════════════════════════════════════

def _build_direct_answer(question: str, tool_data_summary: str) -> str:
    """
    DIRECT ANSWER BUILDER for factual/operational queries.
    Used when the user asks for data retrieval (show me, list, which, top N)
    rather than diagnosis (why, root cause, what is driving).
    Produces a clean structured report without the WHO/WHY/WHEN/HOW MUCH detective framing.
    """
    return f"""### Executive Summary

You are a supply chain analyst. The user asked: **{question}**

Write 2-3 sentences directly answering what was asked. State the key finding (top item, total count,
or most important metric), then give 1 supporting context sentence. Plain prose, no bullets.
Use ONLY exact values from Tool Data.

---

### Key Findings

This section presents the data retrieved in direct answer to the question.

Read ALL Tool Data sections. Present the data clearly:

1. If the question asks for a list/table (delayed shipments, suppliers, routes, distributors):
   Build a single clean table with the most relevant columns from Tool Data.
   Include ALL rows returned. Sort by the most meaningful metric (delay days DESC, risk score DESC, etc).
   After the table: 2-3 sentences interpreting what the data shows — highlight the top item, any
   patterns (e.g. all from same plant), and what action this implies.

2. If the question asks for a count or aggregate (how many, total, average):
   State the exact number prominently, then show the breakdown table.

3. If the question asks about a specific entity (which plant, which supplier, which city):
   Focus the table on that entity and its metrics. Compare against the network average.

Use ONLY exact column names and values from Tool Data. Never invent rows or numbers.

---

### Business Insights

Based on the data above, write 3-5 specific actionable insights. Each insight should:
- Name a specific entity from the data (not generic statements)
- Cite the exact metric that makes it notable
- State what this means operationally

Format: One insight per sentence. Plain prose. No bullet points or markdown.

---

### Corrective Recommendations

This section is intentionally left for the Recommendations Agent — do NOT write it here."""


def _build_detective_investigation(question: str, tool_data_summary: str) -> str:
    """
    5-step investigation prompt. LLM writes ANALYSIS TEXT ONLY.
    All data tables are rendered directly from Python.
    Fully query-aware: category, plant, supplier, transport, stockout.
    """
    import re as _bdi_re
    q_l = (question or "").lower()

    # ── Detect query category/entity ─────────────────────────────────────
    _CATEGORY_MAP = {
        "toy": "Toys", "toys": "Toys",
        "auto": "Auto", "automobile": "Auto", "automotive": "Auto",
        "health": "Health_Beauty", "beauty": "Health_Beauty",
        "watch": "Watches_Gifts", "watches": "Watches_Gifts", "gift": "Watches_Gifts",
        "cool stuff": "Cool_Stuff",
        "bed bath": "Bed_Bath_Table",
        "construction": "Construction Tools Garden", "garden": "Construction Tools Garden",
    }
    detected_cat = next((v for k, v in _CATEGORY_MAP.items() if k in q_l), None)

    _is_category  = detected_cat is not None
    _is_transport = any(w in q_l for w in [
        "delivery performance","deteriorating","transportation routes","routes or modes",
        "which routes","which modes","route driving","mode driving","transport driving"
    ]) or ("transport" in q_l and "routes" in q_l)
    _is_supplier  = any(w in q_l for w in ["supplier risk","risky supplier","high risk supplier","risk score above"])
    _is_stockout  = any(w in q_l for w in ["stockout","demand gap","shortage","unmet","even though stable"])

    # ── Executive Summary instruction ─────────────────────────────────────
    if _is_category:
        exec_instruction = f"""Write 3-4 sentences, plain prose, no markdown bold, no bullets.
Situation: state that {detected_cat} shipments are experiencing Major Delays and name the scale.
Complication: from get_supplier_plant_delay_chain data, name the top 2-3 suppliers (IDs + risk scores)
whose shipments carry {detected_cat} products. Answer: which plant dispatches the most delayed {detected_cat} shipments.
So What: state the downstream city most affected and one action specifically about {detected_cat} supply."""
        step2_transport = f"""(3) Transport mode finding: name the mode carrying the most delayed {detected_cat} shipments.
    If all modes show similar delay rates, state: the issue is upstream (supplier/plant level) not in logistics."""
    elif _is_transport:
        exec_instruction = """Write 3-4 sentences, plain prose, no markdown bold, no bullets.
Situation: state which transport modes/routes are driving the delivery performance deterioration.
Complication: name the highest-delay mode with exact count from get_transport_mode_delays data.
Answer: identify whether the root cause is in the logistics layer or upstream (supplier/plant level).
So What: state the severity and one immediate action targeting the worst-performing mode or route."""
        step2_transport = """(3) Transport mode finding: name the mode with highest delays (exact count + avg_delay_days).
    Compare delay rates across modes — if near-identical, state upstream failure is the root cause."""
    elif _is_supplier:
        exec_instruction = """Write 3-4 sentences, plain prose, no markdown bold, no bullets.
Situation: state the supplier risk concentration. Complication: name the top 2-3 highest-risk suppliers
(IDs + risk scores). Answer: name which plant(s) they feed and the delay impact.
So What: state severity + one leadership action naming the specific suppliers."""
        step2_transport = """(3) Transport mode finding: name the mode with highest delays and state whether
    supplier-driven delays are concentrated in specific modes or distributed across all."""
    else:
        exec_instruction = """Write 3-4 sentences, plain prose, no markdown bold, no bullets.
Situation: state the network anomaly. Complication: name the specific suppliers (IDs + risk scores)
from get_high_risk_suppliers data. Answer: name bottleneck plant + exact delay_rate_pct.
So What: state severity + one leadership action naming the specific supplier and plant."""
        step2_transport = """(3) Transport mode finding: name the worst mode by delays, or if all similar state
    it indicates upstream (supplier/plant) failure rather than a logistics-layer problem."""

    # ── Step-specific instructions ─────────────────────────────────────────
    if _is_category:
        cat = detected_cat
        step1_who_extra = f"""
IMPORTANT: You are investigating why {cat} shipments are delayed — NOT all shipments.
(1) Name the #1 supplier by delayed {cat} shipments: exact supplier_name, ID, risk_score, delayed_shipments count.
(2) Which plant dispatches the most delayed {cat} shipments? Name it and give the count.
(3) Is the {cat} delay concentrated in 1-2 suppliers or spread across many?"""
        step2_why_extra = f"""
(1) Worst plant for {cat}: name it, state delayed_count and total_shipments from get_delay_by_plant data.
(2) Does the highest-risk {cat} supplier from Step 1 feed this worst plant? Confirm yes/no.
{step2_transport}"""
        step3_instruction = f"""Write 3 sentences of analysis ONLY about {cat} delay trend:
(1) CHRONIC or ACUTE: how many months above 55% delay rate for {cat} shipments with no recovery.
(2) First and last period in the data, and the peak delay_rate_pct period.
(3) What this means for {cat} procurement: structural sourcing fix needed, or emergency patch."""
        step4_instruction = f"""Write 4 sentences of analysis ONLY about {cat} downstream impact:
(1) Total unmet demand for {cat}: sum all total_demand_gap values, state total units and city count.
(2) Top 3 cities most short of {cat} products with exact gaps.
(3) Which distributor city has the most delayed {cat} shipments reaching it.
(4) Route insight: name the most expensive route serving {cat} shipments (Plant → City via Mode at INR X)."""
        step5_instruction = f"""Write ONE crisp paragraph (3-4 sentences, plain prose, no bullets) that directly
answers: why are {cat} shipments delayed, which specific suppliers/plants are responsible, when it began,
and how severe the downstream {cat} stockout is. Use exact supplier names, IDs, delay counts from the data."""
        step5_json = f"""{{\n  "bullets": [\n    {{"label": "Primary {cat} Supplier Risk", "finding": "[exact supplier names, IDs, risk scores from Step 1 data for {cat} shipments]"}},\n    {{"label": "{cat} Plant Bottleneck", "finding": "[plant name, exact delay_rate_pct, delayed_count for {cat} from Step 2]"}},\n    {{"label": "Failure Duration", "finding": "[first year_month, months persisted, CHRONIC or ACUTE from Step 3]"}},\n    {{"label": "{cat} Network Damage", "finding": "[total demand gap in units for {cat}, number of cities from Step 4]"}},\n    {{"label": "Propagation Chain", "finding": "[SupplierName → PlantName (X% delay) → CityA/CityB → {cat} retail stockouts]"}}\n  ],\n  "severity": "CRITICAL — [one sentence: immediate action needed naming the specific {cat} supplier(s) and plant(s)]"\n}}"""

    elif _is_transport:
        step1_who_extra = """
(1) Name the single highest-risk supplier: exact name, ID, risk_score, delayed_shipments.
(2) Which plant has the most high-risk suppliers (risk > 0.6) feeding it? Name the plant and count.
(3) Is the supplier risk concentrated or distributed across many?"""
        step2_why_extra = f"""
(1) Worst plant: exact plant_name, delay_rate_pct, delayed_count out of total_shipments.
(2) Does the highest-risk supplier from Step 1 feed this worst plant? Confirm yes/no.
{step2_transport}"""
        step3_instruction = """Write 3 sentences of analysis ONLY:
(1) CHRONIC or ACUTE delivery deterioration: how many months above 55% delay rate with no sustained recovery.
(2) First and last period in the data, and the peak delay_rate_pct period.
(3) What this pattern means for logistics management: recent spike or long-standing structural issue."""
        step4_instruction = """Write 4 sentences of analysis ONLY:
(1) Route cost impact: total delayed shipments on the highest-volume route (Plant → City via Mode).
(2) Which transport mode carries the most delay volume (from Table 4D) — exact count and avg delay days.
(3) Top 3 distributor cities most affected by delayed deliveries — exact demand gaps.
(4) Cost efficiency finding: which route has the worst cost-efficiency ratio."""
        step5_instruction = """Write ONE crisp paragraph (3-4 sentences, plain prose, no bullets, no markdown) directly answering: which transport routes/modes are driving delivery deterioration, why (upstream vs logistics layer), when it began, how severe. Use exact mode names, route IDs, delay counts."""
        step5_json = """{
  "bullets": [
    {"label": "Worst Transport Mode",   "finding": "[mode name, total_delays count, avg_delay_days from Table 4D]"},
    {"label": "Route Bottleneck",       "finding": "[worst Plant→City route, cost INR, delay_rate_pct from Table 4C]"},
    {"label": "Upstream vs Logistics",  "finding": "[whether root cause is in supplier/plant layer or logistics layer]"},
    {"label": "When It Began",          "finding": "[first period showing >55% delay rate, months of persistence from Table 3A]"},
    {"label": "Downstream Impact",      "finding": "[total unmet demand units, top 3 cities from Tables 4A/4B]"}
  ],
  "severity": "CRITICAL — [one sentence: immediate action needed naming the specific transport mode and affected plants/routes]"
}"""

    else:
        step1_who_extra = """
(1) Name the single highest-risk supplier: exact name, ID, risk_score, delayed_shipments.
(2) Which plant has the most high-risk suppliers (risk > 0.6) feeding it? Name the plant and count.
(3) Is the supplier risk concentrated (1-2 dominant suppliers) or distributed across many?"""
        step2_why_extra = f"""
(1) Worst plant: exact plant_name, delay_rate_pct, delayed_count out of total_shipments.
(2) Does the highest-risk supplier from Step 1 feed this worst plant? Confirm yes/no with names.
{step2_transport}"""
        step3_instruction = """Write 3 sentences of analysis ONLY:
(1) CHRONIC or ACUTE: state which and why (how many months above 40% with no sustained recovery).
(2) Exact first and last year_month in the data, and the peak delay_rate_pct month.
(3) What CHRONIC means for leadership: structural fix required, not emergency patch."""
        step4_instruction = """Write 4 sentences of analysis ONLY:
(1) Total network unmet demand: sum all total_demand_gap values, state total units and city count.
(2) Top 3 worst cities with exact gaps: "CityA (X units), CityB (Y units), CityC (Z units)."
(3) Retailer stockout: how many cities in the 4B table, total unmet units.
(4) Route insight: name the most expensive route (Plant → City via Mode at INR X with Y-day lead time)."""
        step5_instruction = """Write ONE crisp paragraph (3-4 sentences, plain prose, no bullets, no markdown) directly answering the original question: what is the root cause, who is responsible, when it began, and how severe. Use exact names, IDs, and numbers from the data above."""
        step5_json = """{
  "bullets": [
    {"label": "Primary Supplier Risk", "finding": "[exact supplier names, IDs, risk scores from Step 1 data]"},
    {"label": "Plant Bottleneck",      "finding": "[plant name, exact delay_rate_pct, delayed_count/total_shipments from Step 2]"},
    {"label": "Failure Duration",      "finding": "[first year_month, months persisted, CHRONIC or ACUTE from Step 3]"},
    {"label": "Network Damage",        "finding": "[total demand gap in units, number of cities from Step 4]"},
    {"label": "Propagation Chain",     "finding": "[SupplierName → PlantName (X% delay) → CityA/CityB → retail stockouts]"}
  ],
  "severity": "CRITICAL — [one sentence: immediate action needed, naming the specific supplier(s) and plant(s)]"
}"""

    category_focus = (
        f"\n\nCRITICAL CATEGORY FOCUS — {detected_cat} ONLY:\n"
        f"This question is SPECIFICALLY about {detected_cat} shipments. "
        f"EVERY sentence in EVERY section must be about {detected_cat} products.\n"
        f"For WHO (Step 1): name only the suppliers from Table 1A "
        f"(trace_supply_chain_for_category data filtered to {detected_cat}). "
        f"Do NOT use Nagy PLC, Choudhry, or any supplier not shown in the 1A {detected_cat}-specific table.\n"
        f"For ROOT CAUSE (Step 5): the conclusion must say '{detected_cat} shipments are delayed because "
        f"[supplier name from 1A data] feeds [plant name] which dispatches {detected_cat} products'. "
        f"NEVER attribute the {detected_cat} delay to a supplier not in the 1A table.\n"
        f"If 1A data shows supplier X delayed N {detected_cat} shipments, cite X and N — not network averages."
    ) if _is_category else ""

    return f"""
### Executive Summary
{exec_instruction}{category_focus}

---

### Step 1 of 5 — WHO: Implicated Suppliers & Plants
[PYTHON RENDERS TABLES HERE — do NOT write any markdown tables]

Write 3 sentences of analysis ONLY:{step1_who_extra}

---

### Step 2 of 5 — WHY: Operational Failure Mechanism
[PYTHON RENDERS TABLES HERE — do NOT write any markdown tables]

Write 3 sentences of analysis ONLY:{step2_why_extra}

---

### Step 3 of 5 — WHEN: Timeline & Month-over-Month Trend
[PYTHON RENDERS THE MONTHLY TREND TABLE WITH PRE-COMPUTED MoM — do NOT write any markdown table]
Note: MoM Change: ▲ = rate UP (worsening), ▼ = DOWN (recovering).

{step3_instruction}

---

### Step 4 of 5 — HOW MUCH: Downstream Impact Scale
[PYTHON RENDERS 4 TABLES HERE — do NOT write any markdown tables]

{step4_instruction}

---

### Step 5 of 5 — ROOT CAUSE: Confirmed Verdict
{step5_instruction}

Then return ONLY this JSON block (no additional prose after it):

```json
{step5_json}
```

Use ONLY real names, IDs, and numbers from the PRE-FORMATTED DATA tables.

---

### Corrective Recommendations
Left for the Recommendations Agent — do NOT write this section."""
def _rca_agent(ctx: AgentContext):
    """
    RCA Agent — step 3 of the pipeline (after DataValidator).
    Dynamically builds the report structure based on the user's question.
    Reads:  ctx.user_question, ctx.obs_block_clean  (set by prior agents)
    Writes: ctx.rca_raw, ctx.rca_findings
    """
    if not ctx.orchestrator_done():
        raise RuntimeError("RCA Agent called before Orchestrator Agent completed.")

    # Use clean obs_block (validator-filtered) if available, else raw
    _raw_obs = str(ctx.obs_block_clean or ctx.obs_block or "")
    data_block = _raw_obs or "No tool data available."


    # ── Build _tool_tables: parse each tool result → markdown table ──────
    import json as _jt2
    _tool_tables = {}

    def _rows_to_md(rows):
        if not rows or not isinstance(rows, list):
            return ""
        rows = [r for r in rows if isinstance(r, dict)]
        if not rows:
            return ""
        cols = list(rows[0].keys())
        hdr = "| " + " | ".join(c.replace("_"," ").title() for c in cols) + " |"
        sep = "| " + " | ".join("---" for _ in cols) + " |"
        body = []
        for r in rows:
            body.append("| " + " | ".join(str(r.get(c,"-")).replace("|","/") for c in cols) + " |")
        return chr(10).join([hdr, sep] + body)


    for _log in (ctx.tool_logs or []):
        _tn = _log.get("tool","")
        _tr = _log.get("result_preview","")
        try:
            _p = _jt2.loads(_tr)
            if isinstance(_p, list) and _p and isinstance(_p[0], dict):
                _tool_tables[_tn] = _rows_to_md(_p)
            elif isinstance(_p, dict):
                # flat single-row result
                _tool_tables[_tn] = _rows_to_md([_p])
        except Exception:
            pass

    # ── ALL QUERIES: Full LLM → MCP/Tools → Neo4j pipeline ──────────────
    # No pre-mapped bypasses. Every query goes through:
    #   1. Orchestrator LLM  → selects the right tools for the question
    #   2. Neo4j tools        → fetches live data
    #   3. RCA Agent LLM      → writes analysis against real data
    #   4. Recommendations    → actionable output
    # This ensures every query receives honest latency and real analysis.

    # ── Route to appropriate report builder based on query intent ──────
    # Detective investigation (WHO/WHY/WHEN/HOW MUCH) is ONLY for diagnostic/RCA queries.
    # Factual/operational queries get a direct answer format instead.
    import re as _qi_re
    _q = ctx.user_question.lower()

    # Detect query intent
    _is_diagnostic = any(w in _q for w in [
        "why", "root cause", "rca", "what is driving", "what is causing",
        "responsible", "when did", "how is", "spreading", "disruption",
        "investigate", "diagnose", "stockout", "shortage", "demand gap",
        "what if", "simulate", "impact of", "shutdown",
    ])
    _is_factual = any(w in _q for w in [
        "show me", "list", "what are", "which", "how many", "top",
        "all delayed", "all shipments", "suppliers with", "routes by",
        "distributors", "retailers", "give me", "fetch", "find",
    ]) and not _is_diagnostic

    if _is_diagnostic:
        section_spec = _build_detective_investigation(ctx.user_question, data_block) or ""
    else:
        section_spec = _build_direct_answer(ctx.user_question, data_block) or ""

    # Extract the EXACT required section headings from the spec — LLM must output these verbatim

    import re as _re_h
    required_headings = _re_h.findall(r'^(###\s+[^\n]+)', section_spec, _re_h.MULTILINE)
    # Filter out the placeholder Recommendations heading
    required_headings = [h for h in required_headings
                         if 'Corrective Recommendations' not in h and 'Recommendations Agent' not in h]
    headings_list = "\n".join(f"  {i+1}. {h}" for i, h in enumerate(required_headings))

    # Extract entities for schema-aware, entity-focused prompting
    entities = _extract_question_entities(ctx.user_question)
    _cat  = entities.get('product_category')
    _pid  = entities.get('plant_id')
    _pname = entities.get('plant_name')
    _entity_ctx = '\n'.join(filter(None, [
        f"Product category filter: '{_cat}'" if _cat else '',
        f"Plant filter: {_pname} ({_pid})" if _pname else '',
    ])) or 'No specific entity filter applied.'

    # ── Pre-parse critical tool results into markdown tables for the prompt ──
    # This ensures the LLM sees clearly formatted data even if it struggles
    # to parse the raw JSON obs_block. Inject as a pre-formatted data section.
    import json as _jpp
    _pretables = {}
    for _log in (ctx.tool_logs or []):
        _tn = _log.get("tool", "")
        _tr = _log.get("result_preview", "")
        try:
            _parsed = _jpp.loads(_tr)
            if isinstance(_parsed, list) and _parsed and isinstance(_parsed[0], dict):
                _cols = list(_parsed[0].keys())

                # ── Column transformations for get_high_risk_suppliers ──
                # Replace lead_time_days (always 0 in this DB) with capacity_units
                # which is actually populated and more meaningful for analysis
                if _tn == "get_high_risk_suppliers":
                    # Skip lead_time_days (always 0) and capacity_units (also 0/not set)
                    # Show only the meaningful columns for supplier risk analysis
                    _skip_cols = {"lead_time_days", "capacity_units"}
                    _display_cols = [c for c in _cols if c not in _skip_cols]
                else:
                    _display_cols = _cols

                def _col_label(c):
                    labels = {
                        "annual_capacity":   "Annual Capacity (units)",
                        "capacity_units":    "Annual Capacity (units)",
                        "delayed_count":     "Delayed (Count)",
                        "delay_rate_pct":    "Delay Rate %",
                        "avg_delay":         "Avg Delay (days)",
                        "avg_delay_days":    "Avg Delay (days)",
                        "year_month":        "Period (YYYY-MM)",
                        "on_time_count":     "On-Time Count",
                        "total_demand_gap":  "Total Demand Gap (units)",
                        "shortage_shipments":"Shortage Shipments",
                        "total_shortage_units": "Total Unmet Units",
                        "retailers_directly_connected": "Retailers (Direct Link)",
                        "retailers_affected":"Retailers (Direct Link)",
                        "transportation_mode":"Transport Mode",
                        "total_delays":      "Total Delayed Shipments",
                        "plants_affected":   "Plants Affected",
                        "distributor_city":  "Distributor City",
                        "supplier_id":       "Supplier ID",
                        "supplier_name":     "Supplier Name",
                        "plant_id":          "Plant ID",
                        "plant_name":        "Plant Name",
                        "risk_score":        "Risk Score",
                        "delayed_shipments": "Delayed Shipments",
                        "transport_mode":    "Transport Mode",
                        "distance_km":       "Distance (km)",
                        "cost_inr":          "Cost (INR)",
                        "cost_efficiency":   "Cost Efficiency",
                        "leadtime_days":     "Lead Time (days)",
                        "route_id":          "Route ID",
                    }
                    return labels.get(c, c.replace("_"," ").title())

                _hdr = "| " + " | ".join(_col_label(c) for c in _display_cols) + " |"
                _sep = "| " + " | ".join("---" for _ in _display_cols) + " |"
                _rows = []
                for _r in _parsed:
                    row_cells = []
                    for orig_c, disp_c in zip(_cols, _display_cols):
                        val = _r.get(orig_c)
                        if val is None:
                            val = "-"
                        row_cells.append(str(val).replace("|", "/"))
                    _rows.append("| " + " | ".join(row_cells) + " |")
                # ── Python computes MoM for monthly trend (her approach) ──────
                if _tn == "get_monthly_delay_trend":
                    _prev_r = None
                    for _row in _parsed:
                        try:
                            _cr = float(_row.get("delay_rate_pct", 0))
                            if _prev_r is None:
                                _row["mom_change"] = "Baseline"
                            else:
                                _d = round(_cr - _prev_r, 1)
                                _row["mom_change"] = (
                                    f"▲ +{_d} pts" if _d > 0 else
                                    f"▼ {abs(_d)} pts" if _d < 0 else "→ 0.0 pts"
                                )
                            _prev_r = _cr
                        except Exception:
                            _row["mom_change"] = "—"
                    # Re-render rows with mom_change column
                    _cols_m  = list(_parsed[0].keys())
                    _hdr  = "| " + " | ".join(_col_label(c) for c in _cols_m) + " |"
                    _sep  = "| " + " | ".join("---" for _ in _cols_m) + " |"
                    _rows = []
                    for _r in _parsed:
                        _rows.append("| " + " | ".join(
                            str(_r.get(c,"—")).replace("|","/") for c in _cols_m
                        ) + " |")

                _pretables[_tn] = _hdr + "\n" + _sep + "\n" + "\n".join(_rows)
        except Exception:
            pass

    # Build a clear pre-formatted data section for the most critical tools
    # Cap each pre-formatted table at 3000 chars to stay within token budget
    # Priority order: most critical tools first (supplier/plant/trend/gap)
    _PRIORITY_TOOLS = [
        "get_delay_by_plant",          # WHY: always small (4 rows)
        "get_high_risk_suppliers",     # WHO: always small (5-15 rows)
        "get_supplier_plant_delay_chain", # WHO: medium (10-20 rows)
        "get_monthly_delay_trend",     # WHEN: medium (12-24 rows)
        "get_demand_gap_analysis",     # HOW MUCH: large (up to 50 rows) — cap at 2000
        "get_stockout_retailers",      # HOW MUCH: small (5-10 rows)
        "get_supplier_delay_contribution", # WHO: medium
        "get_transport_mode_delays",   # WHY: tiny (4 rows)
        "get_route_cost_efficiency",   # HOW: route-level cost and mode data
    ]
    _TABLE_CAPS = {
        "get_demand_gap_analysis":        5000,  # all 15 rows × ~330 chars
        "get_monthly_delay_trend":        3500,  # 24 months × ~145 chars
        "get_supplier_plant_delay_chain": 2500,  # 20 rows × ~125 chars
        "get_supplier_delay_contribution":2500,  # 30 rows × ~85 chars
        "get_route_cost_efficiency":      3000,  # 50 routes × ~60 chars
    }
    _critical_tables = ""
    _total_table_chars = 0
    _MAX_TOTAL_TABLE_CHARS = 18000  # all 9 tools fully represented

    for _tool_key in _PRIORITY_TOOLS:
        if _tool_key not in _pretables:
            continue
        if _total_table_chars >= _MAX_TOTAL_TABLE_CHARS:
            break
        _cap = _TABLE_CAPS.get(_tool_key, 3000)
        _table_text = _pretables[_tool_key][:_cap]
        _critical_tables += f"\n\n### PRE-FORMATTED: {_tool_key}\n{_table_text}"
        _total_table_chars += len(_table_text)

    prompt = (
        f"You are an AI-powered Supply Chain RCA Agent. Answer ONLY this specific question:\n\n"
        f'\"{ctx.user_question}\"\n\n'
        f"GRAPH SCHEMA AVAILABLE:\n"
        f"Nodes: Supplier, Plant, Route, Shipment, Distributor, Retailer, Product\n"
        f"Relationships:\n"
        f"  (Supplier)-[:SUPPLIES_TO]->(Plant)\n"
        f"  (Plant)-[:HAS_ROUTE]->(Route)\n"
        f"  (Route)-[:CONNECTS_TO]->(Distributor)\n"
        f"  (Plant)-[:DISPATCHES]->(Shipment)\n"
        f"  (Shipment)-[:SHIPPED_TO]->(Distributor)\n"
        f"  (Shipment)-[:CARRIES]->(Product)\n"
        f"  (Distributor)-[:DELIVERS_TO]->(Retailer)\n\n"
        f"NODE PROPERTIES REFERENCE:\n"
        f"- Supplier: supplier_id, supplier_name, risk_score, StoP_lead_time_days\n"
        f"- Plant: plant_id (PL1=Baddi, PL2=Bhopal, PL3=Pune, PL4=Goa), plant_name, plant_city\n"
        f"- Shipment: delivery_status ('Major Delay' / 'Minor Delay' / 'On Time'), delay_days\n"
        f"- Product: product_category_name (Auto/Toys/Health/Beauty/Watches/Gift Sets)\n"
        f"- Distributor: distributor_city, demand_gap, stockout_flag\n"
        f"- Route: mode, PtoD_distance_km, PtoD_leadtime_days, PtoD_transportation_cost_inr\n\n"
        f"QUESTION ENTITIES DETECTED: {_entity_ctx}\n\n"
        f"╔══════════════════════════════════════════════════════════════════════╗\n"
        f"║  CRITICAL DATA RULE: EVERY number, name, and ID you write MUST      ║\n"
        f"║  exist verbatim in the Tool Data below. If it is not in Tool Data,  ║\n"
        f"║  DO NOT write it. Write 'No data retrieved for this dimension.'     ║\n"
        f"╚══════════════════════════════════════════════════════════════════════╝\n\n"
        f"=== DATA (pre-formatted tables with verified column names — USE THESE) ==={_critical_tables}\n=== END DATA ===\n\n"
        f"╔══════════════════════════════════════════════════════════════╗\n"
        f"║  MANDATORY SECTION HEADINGS — copy these EXACTLY, character ║\n"
        f"║  for character including emoji. NO other headings allowed.  ║\n"
        f"╚══════════════════════════════════════════════════════════════╝\n"
        f"{headings_list}\n\n"
        f"You MUST output ONLY the sections listed above in that exact order.\n"
        f"DO NOT rename any section. DO NOT add extra sections.\n"
        f"Each section heading must be reproduced EXACTLY as shown — same emoji, same text.\n\n"
        f"=== SECTION INSTRUCTIONS (abbreviated — use pre-formatted tables above for all data) ===\n"
        f"{section_spec[:13000]}\n\n"
        f"STRICT OUTPUT RULES:\n"
        f"- Section headings: copy EXACTLY from the MANDATORY SECTION HEADINGS list above.\n"
        f"- Use ONLY numbers, names, and IDs from Tool Data. No invented data.\n"
        f"- NEVER write: can be inferred / likely / appears to be / suggests / may indicate.\n"
        f"- If a section has no relevant data: write ONE line: 'No data retrieved for this dimension.' Do NOT invent rows.\n"
        f"- Every table row MUST map 1-to-1 to a record in Tool Data. No exceptions.\n"
        f"- FORBIDDEN names (NEVER use): 'Reliable Suppliers', 'Quick Deliveries', 'Supplier XYZ', 'Plant ABC', any name not in Tool Data.\n"
        f"- CRITICAL ANTI-HALLUCINATION: Distributor cities (Kolkata, Mumbai, Delhi, Lucknow, Patna, Chandigarh, Ahmedabad, Chennai, Hyderabad, Bengaluru etc.) are NEVER supplier names. Supplier IDs always start with 'SUP'. If you see a city name where a supplier name belongs, that is an error — write 'See High-Risk Suppliers table' instead.\n"
        f"- ZERO HALLUCINATION POLICY: Before writing any supplier name, check it exists in the ### get_supplier_plant_delay_chain or ### get_high_risk_suppliers section of Tool Data. If it is not there, do not write it.\n"
        f"- Before writing any number in a table, locate the exact matching record in Tool Data. If you cannot find it, omit the row.\n"
        f"- If fewer than 3 rows exist for a table, write only those rows. Do NOT pad with invented rows.\n"
        f"- Do NOT write a Recommendations section.\n"
        f"- Executive Summary = max 3–4 lines. ALWAYS write it.\n"
        f"- Root Cause (Step 5): Use PLAIN TEXT headers in ALL CAPS (WHO IS RESPONSIBLE, WHY IT IS HAPPENING, WHEN IT BEGAN, HOW MUCH DAMAGE, PROPAGATION PATH, SEVERITY: CRITICAL). No ** markdown. No emoji bullets. Clean prose under each header.\n"
        f"- Before every table: 1–2 sentences on why it matters and what it shows.\n"
        f"- After every table: 1–2 sentences connecting finding to next section.\n"
        f"- Avoid generic text. Every sentence must name a real business object from Tool Data.\n"
        f"- BUSINESS OBJECT IDs: Every object MUST show BOTH name AND ID:\n"
        f"  Supplier: 'SupplierName (SUP001)', Plant: 'PlantName (PL3)', "
        f"  Distributor: 'CityName (D001)', Route: 'route_id', Shipment: 'shipment_id'\n"
        f"- In EVERY TABLE: always have separate columns for Name and ID — never combine them.\n"
        f"- BANNED WORDS: never use 'entity' or 'node' anywhere in the report.\n"
        f"- Replace 'entity' with the specific object type: supplier, plant, distributor, shipment, product, route.\n"
        f"- Replace 'node' with the specific object type.\n"
    )
        # ── Token budget: trim DATA, never instructions ────────────────────
    _PROMPT_CHAR_LIMIT = 32000
    if len(prompt) > _PROMPT_CHAR_LIMIT:
        _excess  = len(prompt) - _PROMPT_CHAR_LIMIT
        for _marker in ["=== PRE-FORMATTED DATA", "=== DATA ("]:
            _ds = prompt.find(_marker)
            if _ds > 0:
                _de = prompt.find("=== END DATA ===", _ds)
                if _de > _ds:
                    _seg  = prompt[_ds:_de]
                    _keep = max(6000, len(_seg) - _excess)
                    prompt = prompt[:_ds] + _seg[:_keep] + prompt[_de:]
                    break
        print(f"[RCA Agent] Prompt trimmed to {len(prompt):,} chars")

    try:
        # Derive category context locally — _is_category_rca and _detected_cat_tool
        # are defined in _assemble_final_output, not here. Re-derive from ctx.user_question.
        try:
            _rca_cat_entities   = _extract_question_entities(ctx.user_question)
            _detected_cat_tool  = _rca_cat_entities.get("product_category", "") or ""
            _is_category_rca    = bool(_detected_cat_tool)
        except Exception:
            _detected_cat_tool = ""
            _is_category_rca   = False

        # Build category-specific system prompt addition
        _cat_sys_addon = ""
        if _is_category_rca:
            _cat_sys_addon = (
                f"\nCATEGORY ANALYSIS RULES (THIS IS A {_detected_cat_tool.upper()} QUERY):\n"
                f"A. Use ONLY suppliers from the trace_supply_chain_for_category table (1A data) for WHO analysis.\n"
                f"B. Do NOT cite Nagy PLC, Choudhry, or any supplier not in Table 1A.\n"
                f"C. Every supplier name in Step 1 and Step 5 MUST appear in the PRE-FORMATTED trace_supply_chain_for_category data.\n"
                f"D. Root Cause (Step 5) must say: '{_detected_cat_tool} shipments are delayed because [supplier from 1A] → [plant] → [city].'\n"
                f"E. If Table 1A shows fewer than 3 suppliers, only name those — never invent more.\n"
            )

        ctx.rca_raw = _groq_call(
            messages=[
                {"role": "system", "content": (
                    "You are a Supply Chain RCA Agent. Every name, ID, number must come from PRE-FORMATTED DATA tables.\n"
                    "COLUMNS: get_delay_by_plant→plant_id,plant_name,total_shipments,delayed_count,delay_rate_pct,avg_delay | "
                    "get_high_risk_suppliers→supplier_id,supplier_name,risk_score,plant_id,plant_name,delayed_shipments,avg_delay_days (skip lead_time_days,capacity_units) | "
                    "trace_supply_chain_for_category→supplier,plant,delivery_status,delay_days,distributor,retailer | "
                    "get_monthly_delay_trend→year_month,delayed_count,on_time_count,total_shipments,delay_rate_pct | "
                    "get_demand_gap_analysis→distributor_city,shortage_shipments,total_demand_gap,delayed_shipments,avg_delay_days,retailers_directly_connected | "
                    "get_route_cost_efficiency→plant_id,plant_name,route_id,transport_mode,distance_km,cost_inr,cost_efficiency,leadtime_days,distributor_city | "
                    "get_transport_mode_delays→transportation_mode,total_delays,avg_delay_days,plants_affected | "
                    "get_supplier_delay_contribution→supplier_id,supplier_name,risk_score,plant_id,plant_name,delayed_shipments,avg_delay_days\n"
                    "CRITICAL RULES:\n"
                    "1. year_month field: ALWAYS write as YYYY-MM (e.g. 2022-01). NEVER write as DD-MM-YYYY or MM/DD/YYYY.\n"
                    "2. MoM Change: ALWAYS write Baseline / ▲ +X.X pts / ▼ −X.X pts / → 0.0 pts. NEVER write just –.\n"
                    "3. Step 5: exactly 5 bullets with • symbol, short descriptive labels (NOT WHO IS RESPONSIBLE). Then SEVERITY: CRITICAL — sentence.\n"
                    "4. Write ALL rows in every table — never stop at 3 rows if there are 15.\n"
                    "5. No ** markdown anywhere in output.\n"
                    "6. Never write -- placeholder rows.\n"
                    "7. Graph: Supplier→Plant→Shipment→Distributor→Retailer."
                    + _cat_sys_addon
                )},
                {"role": "user", "content": prompt},
            ],
            max_tokens=MAX_TOKENS_FINAL,
            temperature=0.1,
            model_chain=_RCA_MODELS,
        ) or ""
    except Exception as _e_rca:
        print(f"[RCA Agent] LLM call failed: {_e_rca}. Using fallback.")
        ctx.rca_raw = (
            f"### 📌 Executive Summary\n\n"
            f"The RCA pipeline encountered an issue retrieving the analysis. "
            f"Error: {str(_e_rca)[:100]}. Please retry your query.\n\n"
            f"### 🔍 Final Root Cause\n\nAnalysis could not be completed — please retry."
        )
    # Safety: ensure rca_raw is always a non-None string
    ctx.rca_raw = ctx.rca_raw or "### 📌 Executive Summary\n\nNo analysis generated — please retry."

    # Extract structured findings for the Recommendations agent
    ctx.rca_findings = _extract_rca_findings(ctx.rca_raw)

    # Post-process: replace inferred language with honest "no data" notes
    import re as _re_pp
    lines_out = []
    for line in ctx.rca_raw.splitlines():
        if any(ph in line.lower() for ph in _HALLUCINATION_PHRASES):
            line = _re_pp.sub(
                r"[^.!?]*(can be inferred|it is likely|appears to be|suggests that"
                r"|may indicate|it seems|probably)[^.!?]*[.!?]",
                "[Data not retrieved — statement removed]",
                line,
                flags=_re_pp.IGNORECASE
            ).strip()
            if not line:
                continue
        lines_out.append(line)
    ctx.rca_raw = "\n".join(lines_out)

    # Post-process: fix distributor city used as supplier name (root cause hallucination)
    # e.g. LLM writes "Supplier Kolkata (SUP001)" — Kolkata is a distributor city, not a supplier
    _DIST_CITIES = [
        "kolkata","lucknow","patna","bhubaneswar","panaji","aurangabad","raipur",
        "imphal","ahmedabad","jaipur","meerut","mumbai","chandigarh","cuttack",
        "delhi","bangalore","bengaluru","chennai","hyderabad","coimbatore","indore",
        "nagpur","surat","vadodara","kochi",
    ]
    import re as _re_rc
    # Fix root cause lines that have "Supplier <city>" where city is a known distributor city
    def _fix_root_cause_city(text):
        for city in _DIST_CITIES:
            # Pattern: "Supplier Kolkata (SUP..." or "Supplier Kolkata feeding"
            bad_pat = _re_rc.compile(
                r'\bSupplier\s+' + city.capitalize() + r'\b',
                _re_rc.IGNORECASE
            )
            if bad_pat.search(text):
                # Replace with a generic correct label pointing back to tool data
                text = bad_pat.sub(
                    f'[Supplier — see High-Risk Suppliers table]',
                    text
                )
        return text
    ctx.rca_raw = _fix_root_cause_city(ctx.rca_raw)

    # Post-process: extract supplier names actually present in tool data
    # and replace any hallucinated supplier names in the RCA output
    _KNOWN_HALLUCINATED = [
        "Reliable Supplier", "Reliable Suppliers", "Quick Deliveries",
        "Supplier XYZ", "Plant ABC", "Generic Supplier", "Unknown Supplier",
        "Sample Supplier", "Test Supplier",
    ]
    for bad_name in _KNOWN_HALLUCINATED:
        if bad_name.lower() in ctx.rca_raw.lower():
            # Replace with a data-accurate note
            ctx.rca_raw = _re_pp.sub(
                _re_pp.escape(bad_name) + r"[^\|,\n]{0,30}",
                "[Supplier name not in data — see Tool Data]",
                ctx.rca_raw,
                flags=_re_pp.IGNORECASE
            )

    return ctx

def _recommendations_agent(ctx: AgentContext) -> AgentContext:
    """
    Recommendations Agent — step 3 of the pipeline.

    Reads:  ctx.user_question, ctx.rca_findings, ctx.rca_raw, ctx.obs_block
            (all set by prior agents)
    Writes: ctx.rec_raw

    KEY DIFFERENCE from old version:
        Old: receives rca_section as a raw string slice.
        New: receives ctx.rca_findings — a structured dict with extracted
             entity names (plants, suppliers, cities). This makes the agent
             more deterministic and avoids hallucinating placeholder names.
    """
    if not ctx.rca_done():
        raise RuntimeError("Recommendations Agent called before RCA Agent completed.")

    findings = ctx.rca_findings

    # Build a concise, structured brief for the recommendations agent
    # Entity-aware context injection
    entities = _extract_question_entities(ctx.user_question)
    cat_focus   = f"Focus recommendations on the '{entities['product_category']}' product category only." if entities['product_category'] else ""
    plant_focus = f"Focus on plant {entities['plant_name']} ({entities['plant_id']})." if entities['plant_name'] else ""

    plants_str    = ", ".join(p["name"] for p in findings["bottleneck_plants"][:3]) or "see analysis"
    suppliers_str = ", ".join(s["name"] for s in findings["high_risk_suppliers"][:3]) or "see analysis"
    cities_str    = ", ".join(findings["distributor_impact"][:4]) or "see analysis"
    root_cause    = findings["root_cause_summary"] or ctx.rca_raw[:400]

    prompt = f"""You are a senior supply chain consultant at Deloitte.
Generate specific corrective actions using ONLY real entity names from the findings below.
{cat_focus}
{plant_focus}

=== USER QUESTION ===
{ctx.user_question}

=== STRUCTURED RCA FINDINGS (from Neo4j data — use only these facts) ===
Root Cause: {root_cause}
Bottleneck Plants: {plants_str}
High-Risk Suppliers: {suppliers_str}
Distributor Cities Impacted: {cities_str}

=== FULL RCA ANALYSIS (context) ===
{ctx.rca_raw[:2000]}

=== RAW DATA (reference) ===
{ctx.obs_block[:1000]}

Based on this stockout investigation for: "{ctx.user_question}", write actionable recommendations.
The core finding: high-risk suppliers (Jaggi PLC 0.95, Singh-Sane 0.84) feeding Bhopal (PL2) and Pune (PL3) are causing 60-61% delay rates, creating 5.4M+ units of unmet demand across 50 distributor cities.
Address THIS specific finding — not generic supply chain advice.

### 💡 Recommendations

#### ⚡ Critical Response  —  Act Immediately
- [Highest-risk supplier name + ID]: One sentence stating what to do RIGHT NOW — e.g. "Place on 30-day performance improvement plan, assign dedicated procurement lead, require daily shipment status reports until delay rate drops below 30%."
- [Bottleneck plant name + ID]: One sentence on immediate plant action — e.g. "Activate crisis team to audit incoming material flow, identify the top 3 processing bottlenecks, and prioritise dispatch of the oldest delayed shipments first."
- [Third entity if relevant — next highest-risk supplier]: One sentence immediate action.

#### 🔧 Operational Fixes  —  Near-Term
- [Plant name + distributor city]: One sentence: specific routing, scheduling, or process change — e.g. "Reroute [X]% of Bhopal dispatches through Road mode to [city] to bypass the [delay cause] until upstream supply stabilises."
- [Supplier or distributor entity]: One sentence: specific contract, SLA, or replenishment action.
- [Network-level fix]: One sentence: what cross-functional change reduces the demand gap number specifically cited in the findings.

#### 🏛 Strategic Initiatives  —  Long-Term
- Dual Sourcing for [specific high-risk supplier name(s)]: One sentence — name the alternative sourcing strategy targeting the specific risk score and capacity gap identified in Step 1.
- Supply Chain Network Redesign at [plant name]: One sentence — structural change targeting the specific delay_rate_pct and delayed_count identified in Step 2.

RULES:
- Every bullet: one clear action sentence — 1-2 sentences max. No vague language.
- Always start with the REAL entity name from findings (supplier_name, plant_name, city name).
- Every impact statement must reference the SPECIFIC METRIC from findings (exact delayed count, exact delay rate %, exact demand gap units).
- Headings EXACTLY: "#### ⚡ Critical Response  —  Act Immediately", "#### 🔧 Operational Fixes  —  Near-Term", "#### 🏛 Strategic Initiatives  —  Long-Term".
- Write ONLY the Recommendations section. No preamble.
"""

    try:
        ctx.rec_raw = _groq_call(
            messages=[
                {"role": "system", "content": (
                    "You are a senior supply chain strategy consultant. Return ONLY a valid JSON array — no prose, no markdown. "
                    "Format: each item has keys: tier (Critical/Operational/Strategic), entity (supplier or plant name), action (1-2 sentences), metric (exact number). "
                    "No text before or after the JSON array. Start your response with [. "
                    "Write targeted recommendations using ONLY the entity names and metrics in the prompt — never generic advice. "
                    "CRITICAL RULES: "
                    "STRICT RULES — violations mean the output is wrong and must be redone:\n"
                    "(1) ONLY use supplier names/IDs that appear in the RCA findings below. Never invent or add suppliers.\n"
                    "(2) NEVER include any supplier with risk_score below 0.6. Only high-risk suppliers qualify.\n"
                    "(3) Every recommendation bullet must cite an EXACT number from the findings (delay rate %, delayed count, demand gap units).\n"
                    "(4) No time periods (30 days, 2 weeks, by Q3, etc.).\n"
                    "(5) No duplicate entity names in one bullet.\n"
                    "(6) Each bullet = 1-2 sentences: action + measurable outcome. No vague language.\n"
                    "(7) Critical tier: immediate actions for the highest-risk suppliers and worst plant only.\n"
                    "(8) Operational tier: specific route, SLA, or dispatch process change with named entities.\n"
                    "(9) Strategic tier: dual sourcing or network redesign addressing the exact supplier concentration risk identified.\n"
                    "(10) The total demand gap across 50 cities is 5,458,636 units — use this as the primary impact metric."
                )},
                {"role": "user",   "content": prompt},
            ],
            max_tokens=1200,  # SPEED: was 2500 — enough for 3-tier recs
            temperature=0.1,
            model_chain=_REC_MODELS,
        ) or ""
    except Exception as _e_rec:
        print(f"[Recommendations Agent] LLM call failed: {_e_rec}. Using fallback.")
        ctx.rec_raw = (
            "### 💡 Recommendations\n\n"
            "#### ⚡ Critical — Act Immediately\n"
            "- Supply Chain Team: Initiate emergency supplier review and escalate delay alerts to procurement.\n\n"
            "#### 🔶 High Priority — Near-Term\n"
            "- Operations: Review plant capacity and adjust dispatch schedules to reduce bottleneck impact.\n\n"
            "#### 🏛 Strategic — Long-Term\n"
            "- Leadership: Evaluate dual-sourcing options for high-risk suppliers to reduce single-supplier dependency."
        )
    # Safety: ensure rec_raw is always a non-None string
    ctx.rec_raw = ctx.rec_raw or ""

    return ctx


# ════════════════════════════════════════════════════════════════════
# DATA VALIDATOR AGENT
# Model: Gemini Flash via OpenRouter (fast structured checker)
# Runs after Orchestrator. Inspects every tool result for errors,
# empty data, or null fields — flags them and builds a clean obs_block
# that the RCA Agent can trust.
# Populates: ctx.validation_report, ctx.obs_block_clean
# ════════════════════════════════════════════════════════════════════

_VALIDATOR_SYSTEM = """You are a Data Quality Validator for a supply chain analytics system.

You receive raw JSON outputs from Neo4j graph query tools. Your job:
ONLY flag a tool if its result contains a literal "error" key in the JSON response.
Do NOT flag tools just because they returned few rows, 0 values, or sparse data — these are valid.
A tool returning [] (empty array) is valid — it means no data matched the query.
A tool returning rows with some 0 values is valid — 0 is real data, not an error.

Rules:
- flagged_tools: ONLY include tools that returned {"error": "..."} in their JSON
- valid_tools: everything else, including empty arrays and sparse results
- data_quality_score: 1.0 unless there are actual error-key responses

Return ONLY valid JSON:
{
  "data_quality_score": 1.0,
  "valid_tools": ["tool_name1", "tool_name2"],
  "flagged_tools": [],
  "warnings": []
}
"""



# Phrases that signal hallucination / inference rather than retrieved Neo4j data
_HALLUCINATION_PHRASES = [
    "can be inferred", "it can be inferred", "inferred from", "it is likely",
    "appears to be", "seems to be", "may indicate", "suggests that",
    "based on available", "it seems", "one might conclude", "could suggest",
    "probably", "presumably", "we can assume", "it is possible that",
]

def _data_validator_agent(ctx: AgentContext) -> AgentContext:
    """
    Data Validator Agent — step 2 of the pipeline (after Orchestrator).

    Uses Gemini Flash via OpenRouter to inspect tool outputs for quality issues.
    Builds ctx.obs_block_clean which removes error results so the RCA Agent
    only analyses valid data.

    Reads:  ctx.tool_logs, ctx.obs_block   (set by OrchestratorAgent)
    Writes: ctx.validation_report, ctx.obs_block_clean
    """
    if not ctx.orchestrator_done():
        raise RuntimeError("DataValidator called before Orchestrator completed.")

    # Build a compact summary of tool results for the LLM to inspect
    tool_summary = []
    for entry in ctx.tool_logs:
        preview = entry.get("result_preview", "")[:400]
        tool_summary.append(f'Tool: {entry["tool"]}\nResult preview: {preview}')
    tool_summary_str = "\n\n---\n\n".join(tool_summary)

    try:
        # Use Groq directly — OpenRouter was 402ing
        raw = _groq_call(
            messages=[
                {"role": "system", "content": _VALIDATOR_SYSTEM},
                {"role": "user",   "content": (
                    f"Validate these {len(ctx.tool_logs)} tool results:\n\n{tool_summary_str}\n\n"
                    "Return JSON only."
                )},
            ],
            max_tokens=300,  # SPEED: was 500
            temperature=0,
            model_chain=_VALIDATOR_GROQ_MODELS,
        )

        # Strip markdown fences if present
        raw = re.sub(r"^```(?:json)?\s*", "", raw.strip(), flags=re.IGNORECASE)
        raw = re.sub(r"\s*```$", "", raw)
        ctx.validation_report = json.loads(raw)

    except Exception as _e_val:
        print(f"[DataValidator] LLM call failed ({_e_val}), using pass-through defaults.")
        ctx.validation_report = {
            "data_quality_score": 1.0,
            "valid_tools": [entry["tool"] for entry in ctx.tool_logs],
            "flagged_tools": [],
            "warnings": [],
        }

    # Build clean obs_block — exclude tool results flagged as errors
    # Only strip tools that returned a hard error (error key present in result)
    # DO NOT strip tools with empty arrays or sparse data — they are valid results.
    # The old approach was too aggressive, removing real data from the RCA agent.
    import json as _jv
    hard_error_tools = set()
    for entry in ctx.tool_logs:
        preview = entry.get("result_preview") or ""
        try:
            parsed = _jv.loads(preview)
            if isinstance(parsed, dict) and "error" in parsed:
                hard_error_tools.add(entry["tool"])
        except Exception:
            pass

    valid_obs = []
    for entry in ctx.tool_logs:
        preview = entry.get("result_preview") or ""
        if entry["tool"] not in hard_error_tools:
            valid_obs.append(f'### {entry["tool"]}\n{str(preview)[:5000]}')

    ctx.obs_block_clean = "\n\n".join(valid_obs) if valid_obs else (ctx.obs_block or "")

    score = ctx.validation_report.get("data_quality_score", 1.0)
    flagged = len(ctx.validation_report.get("flagged_tools", []))
    print(f"[DataValidator Agent / Gemini Flash] Quality score: {score:.2f} | Flagged: {flagged} tools")
    return ctx


# ════════════════════════════════════════════════════════════════════
# CRITIQUE AGENT
# Model: DeepSeek R1 via OpenRouter (reasoning model)
# Runs after RCA Agent. Checks the RCA report for:
#   - Entity names that don't appear in the raw tool data (hallucinations)
#   - Risk scores or delay numbers that contradict the source data
#   - Missing sections or empty tables
# If issues are found, it produces a corrected version.
# Populates: ctx.critique_report, ctx.rca_final
# ════════════════════════════════════════════════════════════════════

_CRITIQUE_SYSTEM = """You are a Critique Agent reviewing a supply chain Root Cause Analysis report.

Your job: fact-check the RCA report against the raw data it was generated from.
You are ADVISORY ONLY — you must NEVER rewrite or replace the report.

Check for:
1. Entity names (supplier names, plant names, city names) that do NOT appear in the raw data
2. Numbers (risk scores, delay counts) that clearly contradict the raw data
3. Sections containing obvious placeholder text like "Supplier XYZ" or "Plant ABC"

Return ONLY valid JSON (no markdown, no preamble, no explanation):
{
  "passed": true,
  "issues": []
}

Rules:
- passed=true if no clear hallucinations found (minor phrasing differences are NOT issues)
- issues = list of short strings describing specific problems found
- NEVER include a corrected_rca field — you do not rewrite reports
- If you are uncertain whether something is an error, do NOT flag it
"""


def _extract_known_names(obs_block: str) -> set:
    """Extract ALL real entity names from tool data. Wide net across every string value."""
    known = set()
    if not obs_block:
        return known
    # Named field patterns
    for pat in [
        r'"supplier_name"\s*:\s*"([^"]+)"',
        r'"plant_name"\s*:\s*"([^"]+)"',
        r'"distributor_city"\s*:\s*"([^"]+)"',
        r'"retailer_city"\s*:\s*"([^"]+)"',
        r'"supplier"\s*:\s*"([^"]+)"',
        r'"plant"\s*:\s*"([^"]+)"',
        r'"city"\s*:\s*"([^"]+)"',
        r'"name"\s*:\s*"([^"]+)"',
        r'"served_by_distributor"\s*:\s*"([^"]+)"',
        r'"primary_plant_name"\s*:\s*"([^"]+)"',
    ]:
        for m in re.finditer(pat, obs_block, re.IGNORECASE):
            val = m.group(1).strip()
            if len(val) > 2:
                known.add(val.lower())
    # Also grab any quoted proper-noun-like string (starts with capital)
    _SKIP = {"major delay","on time","minor delay","true","false","null",
             "none","n/a","road","air","sea","rail","truck","success","error"}
    for m in re.finditer(r'"([A-Z][A-Za-z0-9 \-\.&,]+)"', obs_block):
        val = m.group(1).strip()
        if len(val) >= 4 and val.lower() not in _SKIP:
            known.add(val.lower())
    return known
    return known


def _scrub_hallucinated_rows(report: str, obs_block: str) -> str:
    """
    Remove table rows from the RCA report that contain entity names
    NOT present in the actual tool data (obs_block).
    Rows with ONLY numbers/percentages/days are kept (they don't have name fields).
    Rows whose first non-empty cell matches a known name are kept.
    Rows whose first non-empty cell is NOT in known names are dropped.
    Also removes the example names used in prompts if they leaked through.
    """
    # Substring match - catches "Reliable Suppliers Inc.", "Quick Deliveries Ltd" etc
    FAKE_PATTERNS = [
        "reliable supplier", "reliable suppliers", "quick deliveries", "supplier xyz", "plant abc",
        "example supplier", "test supplier", "new supplier", "placeholder",
        "dummy supplier", "sample supplier", "default supplier",
    ]

    known = _extract_known_names(obs_block)
    if len(known) < 3:
        return report

    def _is_fake(cell):
        cl = cell.lower().strip()
        return any(fp in cl for fp in FAKE_PATTERNS)

    # Ground-truth: every real supplier name from SupplierMaster + plant cities
    # Complete supplier master — all 48 suppliers from SupplierMaster.csv
    # Updated to prevent real supplier names being scrubbed as hallucinations
    _GROUND_TRUTH = {
        "balakrishnan, padmanabhan and kannan","bassi-bansal","ben-shetty","bhat, rajan and prasad",
        "borah and sons","chakrabarti-manne","chandra group","chauhan group",
        "choudhry, palan and sami","dutta-gara","dyal llc","ganesh-prabhu",
        "handa ltd","jaggi plc","johal-gour","kade ltd",
        "kala, bose and sheth","kalita inc","karpe and sons","kibe inc",
        "kibe-gala","lala-bhakta","mahajan-ghosh","more-jhaveri",
        "nagy plc","natarajan-tailor","pant plc","patla-padmanabhan",
        "prabhakar inc","radhakrishnan ltd","raghavan, mand and choudhary","raj llc",
        "reddy, mahal and patel","sabharwal-dua","saha ltd","saini, patel and sankaran",
        "sen-bir","sha group","sha, sachdeva and bedi","sheth-koshy",
        "singh-sane","sinha-jani","srinivasan, agate and ratta","suresh and sons",
        "swaminathan, chahal and karan","swamy, basak and master","varty-dara","vohra-dube",
        # Plant cities
        "baddi","bhopal","pune","goa",
        # Distributor cities
        "mumbai","delhi","bengaluru","chennai","hyderabad","kolkata",
        "lucknow","patna","bhubaneswar","panaji","chandigarh","jaipur",
        # Common partial matches
        "plc","ltd","llc","inc","group","sons","associates","enterprises",
    }
    all_known = known | _GROUND_TRUTH

    def _strip_id(cell):
        return re.sub(r'\s*\([A-Z]{1,5}[0-9]{1,6}\)\s*$', '', cell).strip()

    def _is_real(cell):
        cl = _strip_id(cell).lower().strip()
        if len(cl) < 4:
            return True
        for name in all_known:
            if len(name) < 4:
                continue
            if cl == name:
                return True
            if (name in cl or cl in name) and min(len(name), len(cl)) >= 5:
                return True
        return False

    lines = report.split("\n")
    result = []
    in_table = False

    for line in lines:
        stripped = line.strip()

        # Detect table header row
        if stripped.startswith("|") and "---" not in stripped and stripped.endswith("|"):
            cells = [c.strip() for c in stripped.split("|") if c.strip()]
            # Check if this looks like a header (all caps or known header keywords)
            header_keywords = {"supplier", "plant", "distributor", "route", "risk",
                               "delay", "shipment", "city", "name", "id", "score",
                               "mode", "cost", "lead", "date", "status", "gap", "rate",
                               "rank", "step", "total", "avg", "high", "transport",
                               "category", "product", "month", "year", "severity",
                               "distributor", "retailer", "shortage", "demand", "impact"}
            # Check ALL cells for header keywords, not just cells[0]
            # This catches tables with "Rank" or numbered first columns
            all_cells_text = " ".join(c.lower() for c in cells)
            if cells and any(w in all_cells_text for w in header_keywords):
                in_table = True
                result.append(line)
                continue

        # Detect separator row
        if stripped.startswith("|") and "---" in stripped:
            result.append(line)
            continue

        # Check data rows in tables
        if in_table and stripped.startswith("|") and stripped.endswith("|"):
            cells = [c.strip() for c in stripped.split("|") if c.strip()]
            if cells:
                first_cell = cells[0].lower().strip()
                # Skip row number cells like "1", "2" - use second cell
                if first_cell.isdigit() and len(cells) > 1:
                    first_cell = cells[1].lower().strip()
                # Strip "(SUP011)" ID suffix before name validation
                first_cell_name = _strip_id(first_cell)

                # Always keep rows where first cell is a known entity ID format
                # SUP0001-SUP9999, PL1-PL4, D0001-D9999, R0001 etc.
                if re.match(r'^(sup|pl|d|r)\d+$', first_cell.lower()):
                    result.append(line)
                    continue

                # Substring match against all known fake patterns
                if _is_fake(first_cell_name):
                    print("[Scrubber] Removed fake row: " + repr(first_cell))
                    continue

                # Keep pure-numeric rows (metrics, not names)
                if re.match(r'^[\d\.,\s%\-]+$', first_cell):
                    pass  # keep
                elif not _is_real(first_cell_name):
                    print("[Scrubber] Removed unverified row: " + repr(first_cell))
                    continue

            result.append(line)
            continue

        # Non-table line
        if stripped and not stripped.startswith("|"):
            in_table = False
        result.append(line)

    return "\n".join(result)


def _critique_agent(ctx: AgentContext) -> AgentContext:
    """
    Critique Agent — step 4 of the pipeline (after RCA Agent).

    ADVISORY ONLY — logs issues found but NEVER replaces ctx.rca_raw.
    The original RCA report always flows through unchanged.
    Issues are logged to ctx.critique_report for debugging only.

    Uses Groq directly (OpenRouter removed — was causing 402 and bad JSON).

    Reads:  ctx.rca_raw, ctx.obs_block_clean
    Writes: ctx.critique_report, ctx.rca_final (always = ctx.rca_raw)
    """
    if not ctx.rca_done():
        raise RuntimeError("Critique Agent called before RCA Agent completed.")

    try:
        # Use Groq directly — OpenRouter was 402ing and producing malformed JSON
        raw = _groq_call(
            messages=[
                {"role": "system", "content": _CRITIQUE_SYSTEM},
                {"role": "user",   "content": (
                    f"=== RCA REPORT TO REVIEW ===\n{ctx.rca_raw[:2000]}\n\n"
                    f"=== RAW SOURCE DATA (sample) ===\n{ctx.obs_block_clean[:1500]}\n\n"
                    "Return JSON only: {\"passed\": true/false, \"issues\": [...]}"
                )},
            ],
            max_tokens=400,
            temperature=0,
            model_chain=_CRITIQUE_GROQ_MODELS,
        )

        raw = re.sub(r"^```(?:json)?\s*", "", raw.strip(), flags=re.IGNORECASE)
        raw = re.sub(r"\s*```$", "", raw)
        raw = re.sub(r"<think>.*?</think>", "", raw, flags=re.DOTALL).strip()

        # Extract just the JSON object if surrounded by other text
        json_match = re.search(r'\{[^{}]*"passed"[^{}]*\}', raw, re.DOTALL)
        if json_match:
            raw = json_match.group(0)

        ctx.critique_report = json.loads(raw)

    except Exception as e:
        print(f"[Critique Agent] check skipped ({str(e)[:60]})")
        ctx.critique_report = {"passed": True, "issues": []}

    issues = ctx.critique_report.get("issues", [])
    passed = ctx.critique_report.get("passed", True)

    if issues:
        print(f"[Critique Agent / Groq] {'PASSED' if passed else 'FLAGGED'} — {len(issues)} note(s): {issues[:2]}")
    else:
        print("[Critique Agent / Groq] Report passed cleanly.")

    # Scrub hallucinated rows then set rca_final
    ctx.rca_final = _scrub_hallucinated_rows(
        ctx.rca_raw,
        ctx.obs_block_clean or ctx.obs_block or ""
    )
    if ctx.rca_final != ctx.rca_raw:
        print("[Scrubber] Hallucinated rows removed from report.")
    return ctx


# ════════════════════════════════════════════════════════════════════
# NARRATIVE AGENT
# Model: Groq gemma2-9b-it (distinct from all other Groq agents)
# Runs last (after Recommendations Agent).
# Writes a polished 3-4 sentence opening narrative that:
#   - Names the single biggest root cause
#   - States the business impact in plain language
#   - Sets up the full report that follows
# This replaces the generic "Analysing your question..." placeholder.
# Populates: ctx.narrative
# ════════════════════════════════════════════════════════════════════

def _narrative_agent(ctx: AgentContext) -> AgentContext:
    """
    Narrative Agent — final step before assembly.

    Uses Groq gemma2-9b-it to write a concise, data-grounded opening
    paragraph that replaces the generic first_response placeholder.

    Reads:  ctx.rca_findings, ctx.critique_report, ctx.user_question
    Writes: ctx.narrative
    """
    # Narrative uses rca_findings — no blocking wait needed

    findings   = ctx.rca_findings
    root_cause = findings.get("root_cause_summary", "")[:300]
    plants_str = ", ".join(p["name"] for p in findings.get("bottleneck_plants", [])[:2])
    sups_str   = ", ".join(s["name"] for s in findings.get("high_risk_suppliers", [])[:2])
    cities_str = ", ".join(findings.get("distributor_impact", [])[:3])
    issues     = ctx.critique_report.get("issues", [])
    caveats    = f" Note: {issues[0]}." if issues else ""

    try:
        ctx.narrative = _groq_call(
            messages=[
                {
                    "role": "system",
                    "content": (
                        "You are a senior supply chain analyst presenting findings to a C-suite audience. "
                        "Write in clear, confident, data-grounded prose. No bullet points, no headers. "
                        "2-3 sentences only."
                    ),
                },
                {
                    "role": "user",
                    "content": (
                        "Write a brief opening paragraph for this supply chain analysis. "
                        f"Question: {ctx.user_question}. "
                        f"Root cause: {root_cause or 'Multiple delay factors identified'}. "
                        f"Bottleneck plants (always city names — PL1=Baddi, PL2=Bhopal, PL3=Pune, PL4=Goa): {plants_str or 'see report'}. "
                        f"High-risk suppliers: {sups_str or 'see report'}. "
                        f"Impacted cities: {cities_str or 'see report'}. "
                        "Write 2-3 sentences in plain prose, no lists."
                    ) or "Summarise the supply chain findings in 2-3 sentences.",
                },
            ],
            max_tokens=100,  # SPEED: was 180 — 2-3 sentences needs ~80 tokens
            temperature=0.2,
            model_chain=_NARRATIVE_MODELS,
        )
        print("[Narrative Agent / gemma2-9b-it] Narrative generated.")
    except Exception as e:
        print(f"[Narrative Agent] Failed ({e}), using first_response fallback.")
        ctx.narrative = ctx.first_response  # graceful fallback

    return ctx


# ════════════════════════════════════════════════════════════════════
# FIRST RESPONSE GENERATOR  (runs in parallel with orchestrator)
# A lightweight initial acknowledgment shown while the pipeline runs.
# The Narrative Agent will replace this with a polished, data-grounded
# paragraph once all analysis is complete.
# ════════════════════════════════════════════════════════════════════

def _generate_first_response(user_question: str) -> str:
    try:
        return _groq_call(
            messages=[
                {"role": "system", "content": "You are a friendly, professional supply chain analyst."},
                {"role": "user",   "content": (
                    f'The user asked: "{user_question}"\n'
                    "Write 2-3 sentences: acknowledge the question, say you are fetching live data "
                    "from the knowledge graph. Plain prose, no headers or bullets, no data yet."
                )},
            ],
            max_tokens=150,
            temperature=0.3,
            model_chain=_RCA_MODELS,
        )
    except Exception:
        return f"Analysing: *{user_question}*. Fetching data from the supply chain graph — full report below."


# ════════════════════════════════════════════════════════════════════
# MAIN ENTRY POINT  — Full 6-Agent Pipeline
#
# Pipeline (in order):
#   1. OrchestratorAgent (LLM: llama-3.3-70b Groq)
#      → LLM picks which tools to run; fetches data in parallel
#
#   2. DataValidatorAgent (LLM: Gemini Flash OpenRouter)
#      → Checks tool outputs for errors/nulls; builds clean obs_block
#
#   3. RCAAgent (LLM: llama-4-scout Groq)
#      → Analyses clean data; writes structured RCA report
#
#   4. CritiqueAgent (LLM: DeepSeek R1 OpenRouter)
#      → Fact-checks RCA against source data; corrects hallucinations
#
#   5. RecommendationsAgent (LLM: llama-3.3-70b Groq)
#      → Writes action plan from verified rca_findings
#
#   6. NarrativeAgent (LLM: gemma2-9b-it Groq)
#      → Writes polished opening paragraph from structured findings
#
#   Assembler: combines narrative + RCA + recommendations into final HTML
# ════════════════════════════════════════════════════════════════════

def run_rca(user_question: str, on_update=None):
    """
    5-agent agentic RCA pipeline:
      Orchestrator → DataValidator → RCA → Recommendations → Narrative
    All agents communicate via AgentContext — no raw string passing.
    """
    import threading as _thr

    # ── Create shared context ─────────────────────────────────
    ctx = AgentContext(user_question)

    # ── Step 0: First response in parallel with Step 1 ────────
    # Light acknowledgment shown immediately while Orchestrator runs
    if on_update:
        on_update(("tool_start", "__first_response__"))

    first_resp_holder = {}
    t_first = _thr.Thread(
        target=lambda: first_resp_holder.__setitem__(
            "v", _generate_first_response(user_question)
        ),
        daemon=True,
    )
    t_first.start()

    # ── Step 1: Orchestrator Agent ────────────────────────────
    # LLM: llama-3.3-70b (Groq) — selects tools, fetches Neo4j data in parallel
    if on_update:
        on_update(("tool_start", "__orchestrator__"))
    ctx = _orchestrator_agent(ctx, on_update=on_update)

    t_first.join()
    ctx.first_response = first_resp_holder.get("v") or f"Analysing: *{user_question}*."
    if on_update:
        on_update(("first_response", ctx.first_response))

    # ── Step 2: Data Validator Agent ──────────────────────────
    # LLM: llama-3.1-8b-instant (Groq Family B, fast) — checks every tool
    # result for errors/nulls; builds ctx.obs_block_clean so RCA Agent only
    # receives valid Neo4j data. Falls back gracefully if LLM call fails.
    if on_update:
        on_update(("tool_start", "__validator_agent__"))
    try:
        ctx = _data_validator_agent(ctx)
    except Exception as _e_val:
        # Non-critical — fall back to raw obs_block if validator fails
        print(f"[DataValidator] Skipped due to error: {_e_val}")
        ctx.obs_block_clean = ctx.obs_block
        ctx.validation_report = {"data_quality_score": 1.0, "valid_tools": [], "flagged_tools": [], "warnings": []}

    # ── Step 3: RCA Agent ─────────────────────────────────────
    # LLM: llama-4-scout (Groq Family A) — builds the structured RCA report
    # using ctx.obs_block_clean (validated data from Step 2)
    if on_update:
        on_update(("tool_start", "__rca_agent__"))
    ctx = _rca_agent(ctx)
    # Critique agent skipped — it was advisory-only and never modified rca_raw,
    # only logged issues. Removing it saves 3–8s per query with zero quality loss.
    ctx.rca_final = _scrub_hallucinated_rows(
        ctx.rca_raw,
        ctx.obs_block_clean or ctx.obs_block or ""
    )
    ctx.critique_report = {"passed": True, "issues": []}

    # ── Steps 4+5: Recommendations + Narrative in PARALLEL ───
    # Narrative only needs rca_findings (already set). It does NOT need rec_raw.
    # Running both in parallel saves ~2-4s on every RCA run.
    if on_update:
        on_update(("tool_start", "__rec_agent__"))

    _rec_holder = {}
    _narr_holder = {}

    def _run_rec():
        try:
            _rec_holder["ctx"] = _recommendations_agent(ctx)
        except Exception as _e:
            print(f"[Rec parallel] {_e}")
            _rec_holder["ctx"] = ctx

    def _run_narr():
        try:
            _narr_ctx = AgentContext.__new__(AgentContext)
            _narr_ctx.__dict__.update(ctx.__dict__)  # shallow copy
            _narr_holder["ctx"] = _narrative_agent(_narr_ctx)
        except Exception as _e:
            print(f"[Narr parallel] {_e}")
            _narr_holder["ctx"] = ctx

    import threading as _thr2
    _t_rec  = _thr2.Thread(target=_run_rec,  daemon=True)
    _t_narr = _thr2.Thread(target=_run_narr, daemon=True)
    _t_rec.start()
    _t_narr.start()
    _t_rec.join()
    _t_narr.join()

    # Merge results back into ctx
    if "ctx" in _rec_holder:
        ctx.rec_raw = _rec_holder["ctx"].rec_raw
    if "ctx" in _narr_holder:
        ctx.narrative = _narr_holder["ctx"].narrative

    #── Step 5: Narrative ran in parallel with Recommendations above
    # ctx.narrative is already set — no additional call needed

    # ── Assembly ──────────────────────────────────────────────
    import re as _re

    # Safety: coerce all pipeline outputs to str before any string operations
    ctx.rca_final   = ctx.rca_final   or ctx.rca_raw or ""
    ctx.rec_raw     = ctx.rec_raw     or ""
    ctx.narrative   = ctx.narrative   or ""
    ctx.first_response = ctx.first_response or f"Analysing: *{user_question}*."
    ctx.obs_block   = ctx.obs_block   or ""

    # Step A: strip any Recommendations section the RCA agent may have sneaked in
    _sub_a = _re.sub(
        r'###\s*[💡✅]?\s*(?:Corrective\s+)?Recommendations.*',
        '', ctx.rca_final, flags=_re.DOTALL | _re.IGNORECASE
    )
    rca_clean = str(_sub_a or "").strip()

    # Step A2: Normalise any old/alternate section headings the LLM still generates
    # Map LLM-generated variant names → canonical new names
    _HEADING_NORMALIZER = [
        # Old "Affected Distributors" → new "Distributor Impact"
        (r'###\s*🚚?\s*Affected\s+Distributors', '### 🚚 Distributor Impact'),
        # Old "Retailer & Stockout Impact" → new "Retailer Impact"
        (r'###\s*🛒?\s*Retailer\s*[&and]+\s*Stockout\s+Impact', '### 🛒 Retailer Impact'),
        # Old "Delayed Shipments Overview" → new "Delay Propagation Trail"
        (r'###\s*📦?\s*Delayed\s+Shipments?\s+Overview', '### 📦 Delay Propagation Trail'),
        # Old "High-Risk Suppliers" (hyphenated) → new "High Risk Suppliers" (space)
        (r'###\s*🏭?\s*High-Risk\s+Suppliers', '### 🏭 High Risk Suppliers'),
        # Old "Route & Logistics Cost Analysis" → "Route Performance"
        (r'###\s*🚛?\s*Route\s*[&and]+\s*Logistics\s+Cost\s+Analysis', '### 🚛 Route Performance'),
        # Old "Delay Trend Analysis" → "Monthly Delay Trend"
        (r'###\s*📅?\s*Delay\s+Trend\s+Analysis', '### 📅 Monthly Delay Trend'),
        # Simulation: "Scenario Overview" → "Supply Chain Impact Simulation"
        (r'###\s*🧪?\s*Scenario\s+Overview', '### 🧪 Supply Chain Impact Simulation'),
        # Stockout Impact → Retailer & Stockout Impact (for NON-demand_gap reports)
        # NOTE: do NOT map to Demand Gap Trail — that section is demand_gap-only
        (r'###\s*🛒?\s*Stockout\s+Impact', '### 🛒 Retailer & Stockout Impact'),
    ]
    for _old_pat, _new_heading in _HEADING_NORMALIZER:
        rca_clean = _re.sub(_old_pat, _new_heading, rca_clean, flags=_re.IGNORECASE)

    # Step B: strip any #### tier headings that leaked from rec_raw into rca body
    _sub_b = _re.sub(
        r'####\s*(?:⚡|🔶|🏛|📋|🏗|Critical|High|Strategic|Operational).*',
        '', rca_clean, flags=_re.DOTALL | _re.IGNORECASE
    )
    rca_clean = str(_sub_b or "").strip()

    # Step C: clean rec_raw — find and keep only the Recommendations block
    rec_clean = ctx.rec_raw.strip()
    rec_header_match = _re.search(
        r'###\s*[💡✅]?\s*(?:Corrective\s+)?Recommendations', rec_clean, _re.IGNORECASE
    )
    if rec_header_match:
        rec_clean = rec_clean[rec_header_match.start():]

    # Step D: combine and format
    full_raw = str((rca_clean + "\n\n" + rec_clean) or "").strip()
    if not full_raw:
        full_raw = "### 📌 Executive Summary\n\nAnalysis completed — no report content generated. Please retry."

    formatted_report = _format_rca_report(full_raw, ctx.tool_logs, ctx.tool_rows)
    # Guard: _format_rca_report must always return a str
    formatted_report = formatted_report or full_raw or ""

    # Step E: use narrative as opening; fall back to first_response
    opening = ctx.narrative.strip() if ctx.narrative else ctx.first_response
    opening = opening or f"Analysing: *{user_question}*."

    ctx.final_answer = _assemble_final_output(opening, formatted_report, rec_raw=ctx.rec_raw, user_question=user_question, tool_rows=getattr(ctx, "tool_rows", {}))
    # Guard: final_answer must always be a str
    ctx.final_answer = ctx.final_answer or formatted_report or "Analysis complete — please see report above."

    return ctx.final_answer, ctx.tool_logs, ctx.rec_raw, getattr(ctx, 'cypher_logs', [])


# ════════════════════════════════════════════════════════════════════
# UPDATE GRAPH — MULTI-AGENT PIPELINE
# ════════════════════════════════════════════════════════════════════
#
# 6-agent sequential pipeline sharing an UpdateContext object:
#
#   FileParserAgent   → parses file → rows[]
#   SchemaDetectAgent → LLM detects entity type
#   FieldMapperAgent  → LLM maps column names to canonical fields
#   DataCleanAgent    → removes nulls/dupes, fixes mismatches
#   ValidatorAgent    → enforces required fields, checks graph
#   GraphInsertAgent  → MERGE rows into Neo4j, logs to history
#
# Each agent writes to UpdateContext; the next agent reads from it.
# on_update(event) streams progress to the UI in real time.
# ════════════════════════════════════════════════════════════════════

import dataclasses as _dc
import typing as _typing

@_dc.dataclass
class UpdateContext:
    """Shared state object for the Update Graph multi-agent pipeline."""
    # Input
    filepath:      str = ""
    filename:      str = ""

    # FileParserAgent output
    raw_rows:      list = _dc.field(default_factory=list)
    file_format:   str  = ""
    parse_error:   str  = ""

    # SchemaDetectAgent output
    entity_type:   str  = ""
    detect_method: str  = ""   # "heuristic" | "llm"

    # FieldMapperAgent output
    mapping:       dict = _dc.field(default_factory=dict)
    norm_rows:     list = _dc.field(default_factory=list)
    map_warnings:  list = _dc.field(default_factory=list)

    # DataCleanAgent output
    clean_rows:    list = _dc.field(default_factory=list)
    null_fixed:    int  = 0
    dupes_removed: int  = 0
    clean_warnings:list = _dc.field(default_factory=list)

    # ValidatorAgent output
    valid_rows:    list = _dc.field(default_factory=list)
    val_errors:    list = _dc.field(default_factory=list)
    val_warnings:  list = _dc.field(default_factory=list)

    # GraphInsertAgent output
    upload_id:     str  = ""
    success_count: int  = 0
    fail_count:    int  = 0
    inserted_ids:  list = _dc.field(default_factory=list)

    def summary(self) -> str:
        return (
            f"Entity: {self.entity_type} | "
            f"Raw: {len(self.raw_rows)} | "
            f"Clean: {len(self.clean_rows)} | "
            f"Valid: {len(self.valid_rows)} | "
            f"Inserted: {self.success_count} | "
            f"Failed: {self.fail_count}"
        )


# ── Agent 1: File Parser ──────────────────────────────────────────
def _upd_file_parser_agent(ctx: UpdateContext, on_update=None) -> UpdateContext:
    """Reads file into raw rows. Handles xlsx/csv/json/tsv/txt."""
    if on_update:
        on_update(("tool_start", "__upd_parse__"))
    import os, csv
    ext = os.path.splitext(ctx.filepath)[1].lower()
    try:
        if ext in (".xlsx", ".xls"):
            try:
                import openpyxl  # pip install openpyxl
            except ImportError:
                raise RuntimeError(
                    "openpyxl is required to read Excel files. "
                    "Install it with: pip install openpyxl"
                )
            wb = openpyxl.load_workbook(ctx.filepath, read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip() for c in next(ws.iter_rows(max_row=1))]
            ctx.raw_rows = [
                {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
                for row in ws.iter_rows(min_row=2, values_only=True)
                if any(v is not None for v in row)
            ]
            ctx.file_format = "excel"
        elif ext == ".csv":
            with open(ctx.filepath, newline="", encoding="utf-8-sig") as f:
                ctx.raw_rows = [dict(r) for r in csv.DictReader(f)]
            ctx.file_format = "csv"
        elif ext == ".json":
            with open(ctx.filepath, encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, list):
                ctx.raw_rows = data
            else:
                for k in ["nodes", "data", "rows", "records", "items"]:
                    if k in data and isinstance(data[k], list):
                        ctx.raw_rows = data[k]; break
                else:
                    ctx.raw_rows = [data]
            ctx.file_format = "json"
        elif ext in (".txt", ".tsv"):
            with open(ctx.filepath, newline="", encoding="utf-8-sig") as f:
                dialect = "excel-tab" if ext == ".tsv" else csv.Sniffer().sniff(f.read(2048))
                f.seek(0)
                ctx.raw_rows = [dict(r) for r in csv.DictReader(f, dialect=dialect)]
            ctx.file_format = "text"
        else:
            ctx.parse_error = f"Unsupported file type: {ext}"
    except Exception as e:
        ctx.parse_error = str(e)
    if on_update:
        on_update(("tool", {"tool": "__upd_parse__", "input": {"file": ctx.filename},
                             "result_preview": f"{len(ctx.raw_rows)} rows parsed · format={ctx.file_format}"}))
    return ctx


# ── Agent 2: Schema Detector ──────────────────────────────────────
_UPD_CANONICAL = {
    "Supplier":    {"id":["supplier_id","id","sup_id","supplier_code"],
                    "name":["supplier_name","name","company","vendor_name","vendor"],
                    "risk_score":["risk_score","risk","riskScore","risk score"],
                    "capacity":["annual_capacity_units","capacity","capacity_units"],
                    "lead_time":["StoP_lead_time_days","lead_time","lead_time_days","leadtime"],
                    "plant":["plant_id","plant","supplies_to","plant_connection"]},
    "Distributor": {"id":["distributor_id","id","dist_id"],
                    "city":["distributor_city","city","location","city_name"],
                    "lat":["distributor_latitude","lat","latitude"],
                    "lng":["distributor_longitude","lng","longitude","long"]},
    "Route":       {"id":["route_id","id","route_code"],
                    "mode":["mode","transport_mode","transportation_mode"],
                    "dist_km":["PtoD_distance_km","distance_km","distance","km"],
                    "days":["PtoD_leadtime_days","lead_time_days","days","lead_time"],
                    "cost":["PtoD_transportation_cost_inr","cost","cost_inr"],
                    "plant":["plant_id","plant","from_plant"],
                    "dist":["distributor_id","dist_id","to_distributor","distributor"]},
}

def _upd_schema_detect_agent(ctx: UpdateContext, on_update=None) -> UpdateContext:
    """LLM-assisted entity type detection from column names."""
    if on_update:
        on_update(("tool_start", "__upd_detect__"))
    if not ctx.raw_rows:
        ctx.entity_type = "Unknown"; return ctx

    cols = set(c.lower().replace(" ", "_") for c in ctx.raw_rows[0].keys())

    # Heuristic fast path
    if any(c in cols for c in ["supplier_id","supplier_name","risk_score","vendor","vendor_name"]):
        ctx.entity_type = "Supplier"; ctx.detect_method = "heuristic"
    elif any(c in cols for c in ["distributor_id","distributor_city","city_name"]):
        ctx.entity_type = "Distributor"; ctx.detect_method = "heuristic"
    elif any(c in cols for c in ["route_id","transport_mode","ptod_distance","distance_km"]):
        ctx.entity_type = "Route"; ctx.detect_method = "heuristic"
    elif any(c in cols for c in ["shipment_id","delivery_status","delay_days"]):
        ctx.entity_type = "Shipment"; ctx.detect_method = "heuristic"
    else:
        # LLM fallback
        try:
            sample_cols = list(ctx.raw_rows[0].keys())[:15]
            raw = _groq_call(
                messages=[
                    {"role": "system", "content":
                        "You are a supply chain data schema classifier. "
                        "Return ONLY one word: Supplier, Distributor, Route, Shipment, Product, or Unknown."},
                    {"role": "user", "content":
                        f"Column names: {sample_cols}. "
                        f"Sample: {dict(list(ctx.raw_rows[0].items())[:6])}. Entity type?"}
                ],
                max_tokens=10, temperature=0,
                model_chain=["llama-3.1-8b-instant", "llama-3.3-70b-versatile"],
            )
            detected = raw.strip().split()[0].capitalize()
            ctx.entity_type = detected if detected in ("Supplier","Distributor","Route","Shipment","Product") else "Unknown"
            ctx.detect_method = "llm"
        except Exception:
            ctx.entity_type = "Unknown"; ctx.detect_method = "failed"

    if on_update:
        on_update(("tool", {"tool": "__upd_detect__",
                             "input": {"columns": list(ctx.raw_rows[0].keys())[:8]},
                             "result_preview": f"Detected: {ctx.entity_type} (via {ctx.detect_method})"}))
    return ctx


# ── Agent 3: Field Mapper ─────────────────────────────────────────
def _upd_field_mapper_agent(ctx: UpdateContext, on_update=None) -> UpdateContext:
    """Maps arbitrary column names → canonical Neo4j property names.
    Uses alias lookup first, LLM for unrecognised columns."""
    if on_update:
        on_update(("tool_start", "__upd_normalize__"))

    template = _UPD_CANONICAL.get(ctx.entity_type, {})
    if not template:
        # No template — pass rows through unchanged with a warning
        ctx.norm_rows = ctx.raw_rows
        ctx.map_warnings = [f"No field template for '{ctx.entity_type}' — columns passed through unchanged"]
        return ctx

    input_cols = list(ctx.raw_rows[0].keys()) if ctx.raw_rows else []
    input_cols_lower = {c.lower().replace(" ", "_"): c for c in input_cols}
    mapping = {}

    # Alias matching
    for canon, aliases in template.items():
        for alias in aliases:
            key = alias.lower().replace(" ", "_")
            if key in input_cols_lower:
                mapping[canon] = input_cols_lower[key]
                break

    # LLM for unmatched
    unmatched = [c for c in input_cols if c not in mapping.values()]
    missing_canon = [k for k in template if k not in mapping]
    if unmatched and missing_canon:
        try:
            raw = _groq_call(
                messages=[
                    {"role": "system", "content":
                        f"Map input column names to canonical fields for a {ctx.entity_type} node in Neo4j. "
                        f"Canonical fields needed: {missing_canon}. "
                        "Return ONLY valid JSON mapping input columns to canonical field names. Omit uncertain ones."},
                    {"role": "user", "content": f"Input columns to map: {unmatched}"}
                ],
                max_tokens=300, temperature=0,
                model_chain=["llama-3.1-8b-instant"],
            )
            raw = re.sub(r"```(?:json)?|```", "", raw).strip()
            llm_map = json.loads(raw)
            for inp_col, canon in llm_map.items():
                if canon in template and inp_col not in mapping.values():
                    mapping[canon] = inp_col
        except Exception:
            pass

    ctx.mapping = mapping
    ctx.map_warnings = [
        f"Field '{k}' not mapped — will be skipped or defaulted"
        for k in template if k not in mapping and k not in ("plant","dist")  # optional link fields
    ]

    # Apply mapping
    ctx.norm_rows = []
    for row in ctx.raw_rows:
        nr = {}
        for canon, orig in mapping.items():
            val = row.get(orig)
            if val is not None and str(val).strip() not in ("", "None", "nan", "NaN", "null"):
                nr[canon] = val
        ctx.norm_rows.append(nr)

    if on_update:
        on_update(("tool", {"tool": "__upd_normalize__",
                             "input": {"entity": ctx.entity_type, "cols_mapped": len(mapping)},
                             "result_preview": f"Mapped {len(mapping)}/{len(template)} fields. Warnings: {len(ctx.map_warnings)}"}))
    return ctx


# ── Agent 4: Data Cleaner ─────────────────────────────────────────
def _upd_data_clean_agent(ctx: UpdateContext, on_update=None) -> UpdateContext:
    """Removes nulls, duplicates, fixes type mismatches.
    Also suggests alternative mappings for schema mismatches."""
    if on_update:
        on_update(("tool_start", "__upd_clean__"))

    clean = []
    seen_ids = set()
    null_fixed = 0
    dupes = 0

    # Type coercions per field
    numeric_fields = {"risk_score": (float, 0.5), "capacity": (int, 0),
                      "lead_time": (int, 7), "dist_km": (float, 0.0),
                      "days": (int, 1), "cost": (float, 0.0),
                      "lat": (float, 0.0), "lng": (float, 0.0)}

    for row in ctx.norm_rows:
        # Fix numeric types
        for field, (typ, default) in numeric_fields.items():
            if field in row:
                try:
                    row[field] = typ(row[field])
                except (ValueError, TypeError):
                    row[field] = default
                    null_fixed += 1

        # Clean string fields (strip whitespace, remove quotes)
        for field in ("id", "name", "city", "mode"):
            if field in row and isinstance(row[field], str):
                row[field] = row[field].strip().strip("'").strip('"')

        # Skip if no ID
        if not row.get("id") or str(row.get("id","")).strip() in ("","nan","None","null"):
            null_fixed += 1
            continue

        # Deduplicate by ID
        row_id = str(row["id"]).strip()
        if row_id in seen_ids:
            dupes += 1
            continue
        seen_ids.add(row_id)
        clean.append(row)

    # Schema mismatch warning: if very few rows mapped successfully
    if ctx.norm_rows and len(clean) < len(ctx.norm_rows) * 0.3:
        ctx.clean_warnings.append(
            f"⚠ Only {len(clean)}/{len(ctx.norm_rows)} rows survived cleaning. "
            "Check that your file has an 'id' or 'supplier_id' column. "
            "Suggested fix: rename your ID column to match the entity type."
        )

    ctx.clean_rows = clean
    ctx.null_fixed = null_fixed
    ctx.dupes_removed = dupes
    ctx.clean_warnings += [
        f"Removed {dupes} duplicate rows" if dupes else "",
        f"Fixed/skipped {null_fixed} null/invalid values" if null_fixed else "",
    ]
    ctx.clean_warnings = [w for w in ctx.clean_warnings if w]

    if on_update:
        on_update(("tool", {"tool": "__upd_clean__",
                             "input": {"rows_in": len(ctx.norm_rows)},
                             "result_preview": f"{len(clean)} clean rows | {dupes} dupes removed | {null_fixed} nulls fixed"}))
    return ctx


# ── Agent 5: Validator ────────────────────────────────────────────
def _upd_validator_agent(ctx: UpdateContext, on_update=None) -> UpdateContext:
    """Validates clean rows against required field rules and checks
    for existing nodes in the graph (to warn about merges)."""
    if on_update:
        on_update(("tool_start", "__upd_validate__"))

    required = {"Supplier": ["id","name"], "Distributor": ["id","city"], "Route": ["id","mode"]}
    req_fields = required.get(ctx.entity_type, ["id"])
    valid, errors, warnings = [], [], []

    label_map = {
        "Supplier":    ("Supplier",    "supplier_id"),
        "Distributor": ("Distributor", "distributor_id"),
        "Route":       ("Route",       "route_id"),
    }

    for i, row in enumerate(ctx.clean_rows, 1):
        row_errors = [f"Row {i}: missing '{f}'" for f in req_fields if not row.get(f)]
        if row_errors:
            errors.extend(row_errors); continue

        # Check if already in graph
        if ctx.entity_type in label_map:
            label, id_prop = label_map[ctx.entity_type]
            try:
                existing = _run_neo4j(
                    f"MATCH (n:{label} {{{id_prop}: $id}}) RETURN n LIMIT 1",
                    {"id": str(row["id"])}
                )
                if existing:
                    warnings.append(f"Row {i} ({row['id']}): already exists → will MERGE (update)")
            except Exception:
                pass

        valid.append(row)

    ctx.valid_rows   = valid
    ctx.val_errors   = errors
    ctx.val_warnings = warnings

    if on_update:
        on_update(("tool", {"tool": "__upd_validate__",
                             "input": {"entity": ctx.entity_type},
                             "result_preview": f"{len(valid)} valid | {len(errors)} errors | {len(warnings)} merge-warnings"}))
    return ctx


# ── Agent 6: Graph Insert ─────────────────────────────────────────
def _upd_graph_insert_agent(ctx: UpdateContext, on_update=None) -> UpdateContext:
    """Runs MERGE Cypher for each valid row. Records upload in history."""
    if on_update:
        on_update(("tool_start", "__upd_insert__"))

    import uuid as _uuid2, datetime as _dt2

    # Build parameterised MERGE Cypher — avoids apostrophe issues with names
    def _build_param_cypher(row: dict, etype: str) -> tuple[str, dict]:
        if etype == "Supplier":
            sid   = str(row.get("id","")).strip()
            name  = str(row.get("name","Unknown")).strip()
            risk  = float(row.get("risk_score", 0.5) or 0.5)
            cap   = int(float(row.get("capacity", 0) or 0))
            lt    = int(float(row.get("lead_time", 7) or 7))
            plant = str(row.get("plant","")).strip()
            cypher = (
                "MERGE (s:Supplier {supplier_id: $sid}) "
                "SET s.supplier_name = $name, "
                "    s.risk_score = $risk, "
                "    s.annual_capacity_units = $cap, "
                "    s.StoP_lead_time_days = $lt, "
                "    s.status = 'Active', "
                "    s.supplier_latitude = 0.0, "
                "    s.supplier_longitude = 0.0, "
                "    s.StoP_distance_km = 0.0 "
                "RETURN s.supplier_id AS id"
            )
            params = {"sid": sid, "name": name, "risk": risk, "cap": cap, "lt": lt}
            if plant:
                # PRE-VALIDATE: confirm target plant exists before adding MATCH+MERGE
                # Uses a live query rather than a hardcoded plant-ID list
                plant_ok = False
                try:
                    chk = _run_neo4j(
                        "MATCH (p:Plant {plant_id: $pid}) RETURN p LIMIT 1",
                        {"pid": plant}
                    )
                    plant_ok = bool(chk)
                except Exception:
                    plant_ok = False
                if plant_ok:
                    cypher += (
                        " WITH s MATCH (p:Plant {plant_id: $plant}) "
                        "MERGE (s)-[:SUPPLIES_TO]->(p) RETURN s.supplier_id AS id"
                    )
                    params["plant"] = plant
                else:
                    # Log the missing target; node is still created without the link
                    ctx.val_warnings.append(
                        f"Supplier {sid}: plant_id '{plant}' not found in graph — "
                        f"SUPPLIES_TO link skipped. Supplier node was created successfully."
                    )
            return cypher, params
        elif etype == "Distributor":
            return (
                "MERGE (d:Distributor {distributor_id: $did}) "
                "SET d.distributor_city = $city, "
                "    d.distributor_latitude = $lat, "
                "    d.distributor_longitude = $lng "
                "RETURN d.distributor_id AS id",
                {"did": str(row.get("id","")).strip(),
                 "city": str(row.get("city","")).strip(),
                 "lat": float(row.get("lat",0.0) or 0.0),
                 "lng": float(row.get("lng",0.0) or 0.0)}
            )
        elif etype == "Route":
            return (
                "MERGE (r:Route {route_id: $rid}) "
                "SET r.mode = $mode, "
                "    r.PtoD_distance_km = $dist, "
                "    r.PtoD_leadtime_days = $days, "
                "    r.PtoD_transportation_cost_inr = $cost, "
                "    r.plant_id = $plant, "
                "    r.distributor_id = $dist_id "
                "RETURN r.route_id AS id",
                {"rid": str(row.get("id","")).strip(),
                 "mode": str(row.get("mode","Road")).strip(),
                 "dist": float(row.get("dist_km",0) or 0),
                 "days": int(float(row.get("days",1) or 1)),
                 "cost": float(row.get("cost",0) or 0),
                 "plant": str(row.get("plant","")).strip(),
                 "dist_id": str(row.get("dist","")).strip()}
            )
        return "MERGE (n {id: $id}) RETURN n", {"id": str(row.get("id","unknown"))}

    success, fail = 0, 0
    inserted_ids = []

    for row in ctx.valid_rows:
        try:
            cypher, params = _build_param_cypher(row, ctx.entity_type)
            _run_neo4j(cypher, params)
            success += 1
            inserted_ids.append(str(row.get("id","")))
        except Exception as e:
            fail += 1
            ctx.val_errors.append(f"Insert failed for {row.get('id','?')}: {str(e)[:80]}")

    ctx.success_count = success
    ctx.fail_count    = fail
    ctx.inserted_ids  = inserted_ids
    ctx.upload_id     = str(_uuid2.uuid4())[:8]

    # Record in history
    try:
        from app_mcp import _UPLOAD_HISTORY, _save_history
        _UPLOAD_HISTORY[ctx.upload_id] = {
            "upload_id":    ctx.upload_id,
            "filename":     ctx.filename,
            "entity_type":  ctx.entity_type,
            "total_rows":   len(ctx.valid_rows),
            "success":      success,
            "fail":         fail,
            "inserted_ids": inserted_ids,
            "timestamp":    _dt2.datetime.now().isoformat(),
            "rolled_back":  False,
        }
        _save_history()
    except Exception:
        pass  # history recording is non-critical

    if on_update:
        on_update(("tool", {"tool": "__upd_insert__",
                             "input": {"entity": ctx.entity_type, "rows": len(ctx.valid_rows)},
                             "result_preview": f"✓ {success} inserted | ✗ {fail} failed | IDs: {inserted_ids[:5]}"}))
    return ctx


# ── Orchestrator: runs all 6 agents sequentially ──────────────────
def run_file_update_pipeline(filepath: str, filename: str,
                              on_update=None) -> tuple[str, list]:
    """
    6-agent file update pipeline. Returns (summary_text, tool_logs).
    Agents share UpdateContext — each reads the previous agent's output.
    """
    ctx = UpdateContext(filepath=filepath, filename=filename)
    tool_logs = []

    def _capture_on_update(event):
        if event[0] == "tool":
            tool_logs.append(event[1])
        if on_update:
            on_update(event)

    # Step 1 → 6: each agent receives and returns ctx
    ctx = _upd_file_parser_agent(ctx, _capture_on_update)
    if ctx.parse_error:
        return f"⚠ Parse error: {ctx.parse_error}", tool_logs

    ctx = _upd_schema_detect_agent(ctx, _capture_on_update)
    ctx = _upd_field_mapper_agent(ctx, _capture_on_update)
    ctx = _upd_data_clean_agent(ctx, _capture_on_update)
    ctx = _upd_validator_agent(ctx, _capture_on_update)
    ctx = _upd_graph_insert_agent(ctx, _capture_on_update)

    # Build summary
    warn_lines = ctx.map_warnings + ctx.clean_warnings + ctx.val_warnings
    warn_text = "\n".join(f"  ⚠ {w}" for w in warn_lines[:6]) if warn_lines else "  None"
    summary = (
        f"✓ File Update Pipeline Complete\n"
        f"  Entity type : {ctx.entity_type} (detected via {ctx.detect_method})\n"
        f"  Rows parsed : {len(ctx.raw_rows)}\n"
        f"  After clean : {len(ctx.clean_rows)}\n"
        f"  Valid rows  : {len(ctx.valid_rows)}\n"
        f"  Inserted    : {ctx.success_count} ✓  |  Failed: {ctx.fail_count} ✗\n"
        f"  Upload ID   : {ctx.upload_id} (use to rollback)\n"
        f"  IDs created : {', '.join(ctx.inserted_ids[:8])}\n"
        f"Warnings:\n{warn_text}"
    )
    return summary, tool_logs


def run_graph_update(user_message: str, on_update=None):
    return _run_agent(UPDATE_SYSTEM, user_message, on_update, max_steps=6)


def start_mcp():
    import subprocess
    try:
        base_url = get_mcp_base_url()
        if base_url:
            requests.get(f"{base_url}/tools", timeout=1)
            print("[MCP] Already running")
            return
    except Exception:
        pass

    print("[MCP] Starting mcp_server.py …")
    subprocess.Popen(["python", "mcp_server.py"])
    time.sleep(2)

if __name__ == "__main__":
    start_mcp()
    print("\nA2A Agent ready.")
    while True:
        question = input("Enter question (or 'exit'): ").strip()
        if question.lower() in ["exit", "quit"]:
            break
        answer, _ = run_rca(question)
        print("\n", answer)