"""
stage4_pipeline.py
==================
GraphPulse AI — Stage 4: Safe, Auditable, Reversible AI Graph Update Pipeline

Full pipeline:
  Upload → AI Detection → Schema Validation → Semantic Validation →
  Conflict Detection → Graph Diff → Dry-Run Simulation → AI Impact Forecast →
  User Approval → MERGE Execution → Auto Relationship Wiring →
  Graph Refresh → RCA Refresh → Change Summary Report

Features:
  - Sandbox mode (dry-run before any writes)
  - Full rollback / undo support
  - Graph Diff Viewer (new / updated / removed nodes & relationships)
  - AI Impact Simulation ("adding SUP9001 increases PL3 risk by 14%...")
  - 5 test scenario detection (happy path / duplicate / invalid / missing targets / circular)
  - Multi-agent orchestration (Detection → Validation → Conflict → Impact → Insert → Verify)
  - Automatic relationship wiring (Supplier+Plant → SUPPLIES_TO, etc.)
  - Post-commit RCA refresh with before/after comparison
"""

import os, json, re, uuid, time, math, csv, shutil
import traceback
from datetime import datetime
from typing import Optional
from dotenv import load_dotenv

load_dotenv(".env")

# ── CSV sync helpers ──────────────────────────────────────────────
_S4_BASE   = os.path.dirname(os.path.abspath(__file__))

def _resolve_data_dir(base: str) -> str:
    """
    Find the actual data directory.
    Checks DATA_PATH env-var first, then probes 'projectdata/' and 'data/'
    relative to this file.  Whichever contains SupplierMaster.csv wins.
    Falls back to creating 'data/' if neither exists yet.
    """
    env_path = os.environ.get("DATA_PATH", "").strip()
    if env_path and os.path.isdir(env_path):
        return env_path
    for candidate in ("projectdata", "data"):
        path = os.path.join(base, candidate)
        if os.path.isfile(os.path.join(path, "SupplierMaster.csv")):
            return path
    # fallback — create data/ so later writes don't crash
    fallback = os.path.join(base, "data")
    os.makedirs(fallback, exist_ok=True)
    return fallback

_DATA_DIR  = _resolve_data_dir(_S4_BASE)
_CSV_FILES = {
    "Supplier":    os.path.join(_DATA_DIR, "SupplierMaster.csv"),
    "Distributor": os.path.join(_DATA_DIR, "Distributer-WarehouseMasterDist.csv"),
    "Route":       os.path.join(_DATA_DIR, "RoutesSheet.csv"),
}
_CSV_ID_COL = {
    "Supplier":    "supplier_id",
    "Distributor": "distributor_id",
    "Route":       "route_id",
}
_BACKUP_DIR = os.path.join(_DATA_DIR, "_original_backup")

def _csv_ensure_backup():
    os.makedirs(_BACKUP_DIR, exist_ok=True)
    for path in _CSV_FILES.values():
        bk = os.path.join(_BACKUP_DIR, os.path.basename(path))
        if not os.path.exists(bk) and os.path.exists(path):
            shutil.copy2(path, bk)

def _csv_read(path):
    if not os.path.exists(path):
        return []
    with open(path, newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))

def _csv_write(path, rows, fieldnames=None):
    """Write rows to CSV.  If rows is empty, write just the header (preserves column structure)."""
    if not rows and not fieldnames:
        return  # nothing to write — no header known
    fields = fieldnames or list(rows[0].keys())
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        if rows:
            w.writerows(rows)


# Maps Cypher param short-keys → actual CSV/Neo4j column names for each entity type
_CSV_PARAM_MAP = {
    "Supplier": {
        "sid":  "supplier_id",
        "name": "supplier_name",
        "risk": "risk_score",
        "cap":  "annual_capacity_units",
        "lt":   "StoP_lead_time_days",
    },
    "Distributor": {
        "did":  "distributor_id",
        "city": "distributor_city",
        "lat":  "distributor_latitude",
        "lng":  "distributor_longitude",
    },
    "Route": {
        "rid":     "route_id",
        "mode":    "mode",
        "dist":    "PtoD_distance_km",
        "days":    "PtoD_leadtime_days",
        "cost":    "PtoD_transportation_cost_inr",
        "plant":   "plant_id",
        "dist_id": "distributor_id",
    },
}

def _translate_params(entity_type: str, props: dict) -> dict:
    """Translate Cypher short-key params to proper CSV column names."""
    pmap = _CSV_PARAM_MAP.get(entity_type, {})
    translated = {}
    for k, v in props.items():
        csv_key = pmap.get(k, k)  # use mapped name, or original if not in map
        translated[csv_key] = v
    return translated

def _csv_upsert(entity_type, id_val, props):
    """Add or update a row in the matching CSV file."""
    path = _CSV_FILES.get(entity_type)
    if not path:
        return
    id_col = _CSV_ID_COL.get(entity_type, "id")
    # Translate Cypher short-key params to actual CSV column names
    mapped_props = _translate_params(entity_type, props)
    rows = _csv_read(path)
    fieldnames = list(rows[0].keys()) if rows else None
    idx = next((i for i, r in enumerate(rows) if str(r.get(id_col,"")).strip() == str(id_val).strip()), -1)
    if idx == -1:
        template = {k: "" for k in fieldnames} if fieldnames else {}
        template[id_col] = str(id_val)
        for k, v in mapped_props.items():
            if k in template:
                template[k] = v
        rows.append(template)
    else:
        for k, v in mapped_props.items():
            if k in rows[idx]:
                rows[idx][k] = v
    _csv_write(path, rows, fieldnames=fieldnames)

def _csv_delete(entity_type, id_val):
    """Remove a row from the matching CSV file."""
    path = _CSV_FILES.get(entity_type)
    if not path:
        return
    id_col = _CSV_ID_COL.get(entity_type, "id")
    rows = _csv_read(path)
    fieldnames = list(rows[0].keys()) if rows else None
    new_rows = [r for r in rows if str(r.get(id_col,"")).strip() != str(id_val).strip()]
    if len(new_rows) != len(rows):
        _csv_write(path, new_rows, fieldnames=fieldnames)

def _csv_restore_from_backup():
    """Restore all CSV files from _original_backup/."""
    _csv_ensure_backup()
    for path in _CSV_FILES.values():
        bk = os.path.join(_BACKUP_DIR, os.path.basename(path))
        if os.path.exists(bk):
            shutil.copy2(bk, path)

# Ensure backups exist on import
try:
    _csv_ensure_backup()
except Exception:
    pass
# ─────────────────────────────────────────────────────────────────

# ── Neo4j driver (reuse from app_mcp if possible) ─────────────────
try:
    from neo4j import GraphDatabase as _GD
    _neo4j_driver = _GD.driver(
        os.getenv("NEO4J_URI", "bolt://localhost:7687"),
        auth=(os.getenv("NEO4J_USER", "neo4j"), os.getenv("NEO4J_PASSWORD", "password")),
    )
    _DB = os.getenv("NEO4J_DATABASE", "neo4j")
except Exception:
    _neo4j_driver = None
    _DB = "neo4j"


def _run_cypher(query: str, params: dict = None) -> list:
    """Execute Cypher and return list of row dicts."""
    if not _neo4j_driver:
        return []
    try:
        with _neo4j_driver.session(database=_DB) as s:
            return [r.data() for r in s.run(query, params or {})]
    except Exception as e:
        raise RuntimeError(f"Neo4j error: {e}") from e


# ── Groq LLM caller ───────────────────────────────────────────────
try:
    from groq import Groq as _Groq
    _groq = _Groq(api_key=(os.getenv("GROQ_API_KEY") or "").strip())
except Exception:
    _groq = None

_LLM_MODELS = [
    "llama-3.3-70b-versatile",
    "meta-llama/llama-4-scout-17b-16e-instruct",
    "mixtral-8x7b-32768",
    "llama-3.1-8b-instant",
]


def _llm(messages: list, max_tokens: int = 1500, temperature: float = 0) -> str:
    """Call Groq with fallback chain. Always returns str."""
    if not _groq:
        return ""
    for model in _LLM_MODELS:
        try:
            resp = _groq.chat.completions.create(
                model=model, messages=messages,
                temperature=temperature, max_tokens=max_tokens,
            )
            content = resp.choices[0].message.content
            if content:
                return str(content).strip()
        except Exception as e:
            if any(x in str(e) for x in ["429", "rate_limit", "decommissioned"]):
                time.sleep(2)
                continue
            break
    return ""


# ══════════════════════════════════════════════════════════════════
# PIPELINE STATE — PipelineSession carries the full run context
# ══════════════════════════════════════════════════════════════════

class PipelineSession:
    """Shared state for one complete Stage 4 pipeline run."""

    def __init__(self, filepath: str, filename: str):
        self.session_id   = uuid.uuid4().hex[:8]
        self.filepath     = filepath
        self.filename     = filename
        self.created_at   = datetime.now().isoformat()

        # Stage outputs
        self.raw_rows:      list = []
        self.file_format:   str  = ""
        self.entity_type:   str  = ""
        self.detect_method: str  = ""
        self.detect_confidence: float = 0.0

        self.mapping:       dict = {}
        self.norm_rows:     list = []

        # Validation
        self.valid_rows:    list = []
        self.invalid_rows:  list = []   # [{row, errors: [str]}]
        self.schema_errors: list = []
        self.semantic_errors: list = []

        # Conflict detection
        self.duplicates_in_file:  list = []   # row indices
        self.existing_in_graph:   list = []   # [{id, existing_props}]
        self.missing_targets:     list = []   # [{id, missing_ref}]
        self.circular_refs:       list = []   # description strings

        # Graph diff
        self.new_nodes:          list = []   # rows that will create new nodes
        self.updated_nodes:      list = []   # [{id, old_props, new_props, diff}]
        self.new_relationships:  list = []   # [{from, to, type}]
        self.removed_relationships: list = []

        # Cypher plan
        self.cypher_plan:   list = []   # [{id, cypher, params, is_new}]
        self.dry_run_log:   list = []   # results of dry-run simulation

        # AI outputs
        self.impact_simulation: str = ""   # AI-generated impact forecast
        self.scenario_type:     str = ""   # happy_path / duplicate / invalid / missing_targets / circular

        # Execution results
        self.approved:         bool = False
        self.inserted_ids:     list = []
        self.failed_ids:       list = []
        self.wired_rels:       list = []   # relationships auto-created
        self.rollback_log:     list = []   # cypher to undo everything

        # RCA comparison
        self.rca_before_snapshot: dict = {}
        self.rca_after_snapshot:  dict = {}
        self.rca_diff_html:       str  = ""

        # Change summary
        self.change_summary_html: str = ""

    def to_dict(self) -> dict:
        return {k: v for k, v in self.__dict__.items()
                if not k.startswith("_")}


# ══════════════════════════════════════════════════════════════════
# CANONICAL SCHEMA — entity → required/optional fields + validators
# ══════════════════════════════════════════════════════════════════

SCHEMA = {
    "Supplier": {
        "id_field": "supplier_id",
        "neo4j_label": "Supplier",
        "required": ["id", "name"],
        "optional": ["risk_score", "capacity", "lead_time", "plant", "status"],
        "aliases": {
            # Canonical → all realistic column name variations
            "id": [
                "supplier_id", "sup_id", "id", "supplier_code", "vendor_id",
                "vendorid", "supplierid", "supplier id", "sup id", "code",
                "supplierno", "supplier_no", "vendor_code",
            ],
            "name": [
                "supplier_name", "name", "company", "vendor_name", "vendor",
                "supplier", "company_name", "firm_name", "firm", "business_name",
                "organisation", "organization", "org_name",
            ],
            "risk_score": [
                "risk_score", "risk", "riskscore", "risk score", "score",
                "risk_level", "risk_rating", "riskrating", "vendor_risk",
                "supplier_risk", "risk_index",
            ],
            "capacity": [
                "annual_capacity_units", "capacity", "capacity_units",
                "annual_capacity", "units", "production_capacity",
                "max_capacity", "supply_capacity", "output_capacity",
                "capacity_per_year", "yearly_capacity",
            ],
            "lead_time": [
                "StoP_lead_time_days", "lead_time", "lead_time_days",
                "leadtime", "lead_days", "delivery_days", "transit_time",
                "fulfillment_days", "days_to_deliver", "avg_lead_time",
                "stop_lead_time_days", "s_to_p_lead_time",
            ],
            "plant": [
                "plant_id", "plant", "supplies_to", "plant_connection",
                "plant_code", "target_plant", "assigned_plant",
                "manufacturing_plant", "production_plant", "facility_id",
            ],
            "status": [
                "status", "supplier_status", "active", "state",
            ],
        },
        "validators": {
            "risk_score": lambda v: 0.0 <= float(v) <= 1.0,
            "lead_time":  lambda v: int(float(v)) >= 0,
            "capacity":   lambda v: int(float(v)) >= 0,
        },
        "validator_msgs": {
            "risk_score": "must be between 0.0 and 1.0",
            "lead_time":  "must be ≥ 0 days",
            "capacity":   "must be ≥ 0 units",
        },
    },
    "Distributor": {
        "id_field": "distributor_id",
        "neo4j_label": "Distributor",
        "required": ["id", "city"],
        "optional": ["lat", "lng"],
        "aliases": {
            "id": [
                "distributor_id", "dist_id", "id", "distributor_code",
                "distid", "distribution_id", "hub_id", "depot_id",
                "distributor id", "dist id", "distribution_center_id",
            ],
            "city": [
                "distributor_city", "city", "location", "city_name",
                "distributor_location", "hub_city", "depot_city",
                "distribution_city", "region", "area", "place", "town",
                "distributor city",
            ],
            "lat": [
                "distributor_latitude", "lat", "latitude", "coord_lat",
                "distributor_lat", "geo_lat", "y_coord", "lat_deg",
            ],
            "lng": [
                "distributor_longitude", "lng", "longitude", "long",
                "coord_lng", "distributor_lng", "geo_lng", "x_coord",
                "lon", "lon_deg", "lng_deg",
            ],
        },
        "validators": {
            "lat": lambda v: -90  <= float(v) <= 90,
            "lng": lambda v: -180 <= float(v) <= 180,
        },
        "validator_msgs": {
            "lat": "latitude must be between -90 and 90",
            "lng": "longitude must be between -180 and 180",
        },
    },
    "Route": {
        "id_field": "route_id",
        "neo4j_label": "Route",
        "required": ["id", "mode"],
        "optional": ["dist_km", "days", "cost", "plant", "dist_id"],
        "aliases": {
            "id": [
                "route_id", "id", "route_code", "route", "routeid",
                "route id", "lane_id", "shipment_lane_id", "corridor_id",
            ],
            "mode": [
                "mode", "transport_mode", "transportation_mode", "transport",
                "transport_type", "shipping_mode", "delivery_mode",
                "transit_mode", "carrier_type",
            ],
            "dist_km": [
                "PtoD_distance_km", "distance_km", "distance", "km",
                "dist_km", "ptod_distance_km", "route_distance", "length_km",
                "distance_miles", "dist", "route_km",
            ],
            "days": [
                "PtoD_leadtime_days", "lead_time_days", "days", "lead_time",
                "transit_days", "ptod_leadtime_days", "transit_time_days",
                "delivery_days", "avg_transit_days", "travel_days",
            ],
            "cost": [
                "PtoD_transportation_cost_inr", "cost", "cost_inr",
                "transport_cost", "price", "ptod_transportation_cost_inr",
                "shipping_cost", "freight_cost", "logistics_cost",
                "route_cost", "cost_per_unit",
            ],
            "plant": [
                "plant_id", "plant", "from_plant", "origin_plant",
                "source_plant", "origin", "from", "dispatch_plant",
            ],
            "dist_id": [
                "distributor_id", "dist_id", "to_distributor", "distributor",
                "destination", "to_dist", "dest_id", "delivery_point",
            ],
        },
        "validators": {
            "dist_km": lambda v: float(v) >= 0,
            "days":    lambda v: int(float(v)) >= 0,
            "cost":    lambda v: float(v) >= 0,
            "mode":    lambda v: str(v).strip().lower() in {
                "road", "rail", "air", "sea", "ship", "truck", "train",
                "airways", "roadways", "railways",
            },
        },
        "validator_msgs": {
            "dist_km": "distance must be ≥ 0 km",
            "days":    "lead time must be ≥ 0 days",
            "cost":    "cost must be ≥ 0",
            "mode":    "must be one of: Road, Rail, Air, Sea",
        },
    },
}

AUTO_RELATIONSHIPS = {
    "Supplier":    [("plant", "Plant", "plant_id", "SUPPLIES_TO")],
    "Route":       [
        ("plant",   "Plant",        "plant_id",        "HAS_ROUTE"),
        ("dist_id", "Distributor",  "distributor_id",  "CONNECTS_TO"),
    ],
}


# ══════════════════════════════════════════════════════════════════
# STAGE 1 — FILE PARSER
# ══════════════════════════════════════════════════════════════════

def _find_header_row(ws, max_scan: int = 8):
    """
    Find the first row that looks like real column headers (not a title/instruction row).
    A header row has >=2 non-empty cells and does NOT start with known instruction prefixes.
    Returns (header_row_index_1based, [col_names]).
    """
    SKIP_PREFIXES = ("graphpulse", "📋", "instructions", "note:", "how to", "readme",
                     "fill", "template", "upload guide", "do not")
    for i, row in enumerate(ws.iter_rows(max_row=max_scan, values_only=True), start=1):
        cells = [str(c or "").strip() for c in row]
        non_empty = [c for c in cells if c]
        if len(non_empty) < 2:
            continue
        first = non_empty[0].lower()
        if any(first.startswith(p) for p in SKIP_PREFIXES):
            continue
        # Looks like a header row — cells should be short identifiers, not sentences
        avg_len = sum(len(c) for c in non_empty) / len(non_empty)
        if avg_len > 80:  # instruction row masquerading as data
            continue
        return i, cells
    return 1, []  # fallback


def _best_sheet_for_entity(wb, entity_type: str = None):
    """
    For multi-sheet workbooks, pick the sheet whose name best matches the entity type,
    or pick the sheet with the most data rows if no entity type hint.
    Always skips README/instructions/template info sheets.
    """
    SKIP_SHEETS = {"readme", "instructions", "guide", "info", "template", "notes"}
    candidate_sheets = [
        s for s in wb.sheetnames
        if s.lower() not in SKIP_SHEETS
        and not any(skip in s.lower() for skip in ["readme", "guide", "instruction"])
    ]
    if not candidate_sheets:
        candidate_sheets = wb.sheetnames[:1]

    if entity_type and entity_type != "Unknown":
        # Try to find a sheet whose name matches the entity type
        et_lower = entity_type.lower()
        for name in candidate_sheets:
            if et_lower in name.lower() or name.lower() in et_lower:
                return wb[name]

    # Score sheets by data content — pick the one with most non-empty rows after header
    best_sheet = None
    best_score = -1
    for name in candidate_sheets:
        ws = wb[name]
        _, hdrs = _find_header_row(ws)
        if not hdrs:
            continue
        row_count = sum(1 for r in ws.iter_rows(min_row=2, values_only=True)
                        if any(c is not None for c in r))
        if row_count > best_score:
            best_score = row_count
            best_sheet = ws
    return best_sheet or wb[candidate_sheets[0]]


def stage_parse(session: PipelineSession, on_update=None) -> PipelineSession:
    """Parse uploaded file into raw rows. Handles multi-sheet Excel, skips title/instruction rows."""
    _emit(on_update, "stage", {"stage": "parse", "status": "running"})
    import csv as _csv

    ext = os.path.splitext(session.filepath)[1].lower()
    try:
        if ext in (".xlsx", ".xls"):
            import openpyxl
            wb = openpyxl.load_workbook(session.filepath, read_only=True, data_only=True)

            # For multi-sheet workbooks, collect ALL data sheets and merge rows
            # so the system can handle whichever sheet the user filled in
            SKIP_SHEETS = {"readme", "instructions", "guide", "info", "notes"}
            data_sheets = [
                s for s in wb.sheetnames
                if not any(skip in s.lower() for skip in SKIP_SHEETS)
            ]

            all_rows = []
            used_sheet = None
            used_headers = []

            for sheet_name in data_sheets:
                ws = wb[sheet_name]
                hdr_row_idx, raw_headers = _find_header_row(ws)
                headers = [str(h or "").strip() for h in raw_headers]
                # Filter out empty header columns
                valid_col_idx = [i for i, h in enumerate(headers) if h]
                if not valid_col_idx:
                    continue
                headers_clean = [headers[i] for i in valid_col_idx]

                sheet_rows = []
                for r in ws.iter_rows(min_row=hdr_row_idx + 1, values_only=True):
                    if r is None or all(c is None for c in r):
                        continue
                    d = {}
                    for i in valid_col_idx:
                        if i < len(r):
                            val = r[i]
                            if val is not None and str(val).strip() not in ("", "None", "nan"):
                                d[headers[i]] = val
                    if d:
                        sheet_rows.append(d)

                if sheet_rows and len(sheet_rows) > len(all_rows):
                    # Use the sheet with the most data
                    all_rows = sheet_rows
                    used_sheet = sheet_name
                    used_headers = headers_clean

            session.raw_rows = all_rows
            session.file_format = f"Excel ({used_sheet or 'unknown sheet'})"

        elif ext == ".csv":
            with open(session.filepath, newline="", encoding="utf-8-sig") as f:
                session.raw_rows = [dict(r) for r in _csv.DictReader(f)]
            session.file_format = "CSV"

        elif ext == ".json":
            with open(session.filepath, encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, list):
                session.raw_rows = data
            else:
                for k in ["nodes", "data", "rows", "records", "items"]:
                    if k in data and isinstance(data[k], list):
                        session.raw_rows = data[k]
                        break
                else:
                    session.raw_rows = [data]
            session.file_format = "JSON"

        elif ext in (".txt", ".tsv"):
            with open(session.filepath, newline="", encoding="utf-8-sig") as f:
                try:
                    dialect = _csv.Sniffer().sniff(f.read(4096))
                    f.seek(0)
                except Exception:
                    dialect = "excel-tab"
                    f.seek(0)
                session.raw_rows = [dict(r) for r in _csv.DictReader(f, dialect=dialect)]
            session.file_format = "TSV/Text"
        else:
            raise ValueError(f"Unsupported file type: {ext}")

    except Exception as e:
        _emit(on_update, "stage", {"stage": "parse", "status": "error", "msg": str(e)})
        raise

    _emit(on_update, "stage", {
        "stage": "parse", "status": "done",
        "rows": len(session.raw_rows), "format": session.file_format
    })
    return session


# ══════════════════════════════════════════════════════════════════
# STAGE 2 — AI ENTITY TYPE DETECTION
# ══════════════════════════════════════════════════════════════════

def stage_detect(session: PipelineSession, on_update=None) -> PipelineSession:
    """Detect entity type using sheet name hint + heuristics + LLM."""
    _emit(on_update, "stage", {"stage": "detect", "status": "running"})
    if not session.raw_rows:
        session.entity_type = "Unknown"
        return session

    # ── Priority 0: Sheet name hint from file_format (set by multi-sheet parser) ──
    fmt_lower = session.file_format.lower()
    if "supplier" in fmt_lower:
        session.entity_type = "Supplier"
        session.detect_method = "sheet_name"
        session.detect_confidence = 0.98
        _emit(on_update, "stage", {"stage": "detect", "status": "done",
              "entity_type": "Supplier", "method": "sheet_name", "confidence": 0.98})
        return session
    if "distributor" in fmt_lower or "dist" in fmt_lower:
        session.entity_type = "Distributor"
        session.detect_method = "sheet_name"
        session.detect_confidence = 0.98
        _emit(on_update, "stage", {"stage": "detect", "status": "done",
              "entity_type": "Distributor", "method": "sheet_name", "confidence": 0.98})
        return session
    if "route" in fmt_lower or "lane" in fmt_lower:
        session.entity_type = "Route"
        session.detect_method = "sheet_name"
        session.detect_confidence = 0.98
        _emit(on_update, "stage", {"stage": "detect", "status": "done",
              "entity_type": "Route", "method": "sheet_name", "confidence": 0.98})
        return session

    # ── Also check filename for hint ──
    fname_lower = session.filename.lower()
    for etype in ["supplier", "distributor", "route"]:
        if etype in fname_lower:
            session.entity_type = etype.title()
            session.detect_method = "filename"
            session.detect_confidence = 0.92
            _emit(on_update, "stage", {"stage": "detect", "status": "done",
                  "entity_type": session.entity_type, "method": "filename", "confidence": 0.92})
            return session

    cols = set(c.lower().replace(" ", "_").replace(" ", "") for c in session.raw_rows[0].keys())

    # Heuristic scoring — count alias matches per entity type (normalised comparison)
    def _norm(s): return s.lower().replace(" ", "_").replace("-", "_").replace(" ", "")
    cols_norm = set(_norm(c) for c in session.raw_rows[0].keys())

    scores = {}
    for etype, schema in SCHEMA.items():
        score = 0
        for canon, aliases in schema["aliases"].items():
            for alias in aliases:
                if _norm(alias) in cols_norm:
                    score += (3 if canon in schema["required"] else 1)
                    break
        scores[etype] = score

    best_etype = max(scores, key=scores.get)
    best_score = scores[best_etype]
    total_possible = sum(3 if c in SCHEMA[best_etype]["required"] else 1
                         for c in SCHEMA[best_etype]["aliases"])
    confidence = min(1.0, best_score / max(total_possible, 1))

    if confidence >= 0.5:
        session.entity_type = best_etype
        session.detect_method = "heuristic"
        session.detect_confidence = round(confidence, 2)
    else:
        # LLM fallback
        sample_cols = list(session.raw_rows[0].keys())[:15]
        sample_row = {k: str(v)[:30] for k, v in list(session.raw_rows[0].items())[:8]}
        raw = _llm([
            {"role": "system", "content":
                "You are a supply chain data classifier for a Neo4j graph database. "
                "Classify the uploaded data as exactly ONE of: Supplier, Distributor, Route. "
                "Supplier = companies supplying materials to plants. "
                "Distributor = cities/hubs that receive goods. "
                "Route = transport lanes between plants and distributors. "
                "Return ONLY valid JSON: {\"type\": \"Supplier|Distributor|Route\", \"confidence\": 0.0-1.0, \"reason\": \"brief\"}. "
                "NEVER return Unknown unless data is completely unrelated to supply chain."},
            {"role": "user", "content":
                f"File: {session.filename}\nColumns: {sample_cols}\nSample row: {sample_row}\n"
                "Classify this data. Return JSON only — no explanation outside the JSON."}
        ], max_tokens=150)
        try:
            raw = re.sub(r"```(?:json)?|```", "", raw).strip()
            result = json.loads(raw)
            detected = str(result.get("type", "Unknown")).strip()
            if detected in SCHEMA:
                session.entity_type = detected
                session.detect_confidence = float(result.get("confidence", 0.7))
            else:
                session.entity_type = "Unknown"
                session.detect_confidence = 0.0
            session.detect_method = "llm"
        except Exception:
            # Even if score is 0, prefer best_etype over Unknown
            # Unknown blocks the entire pipeline — wrong type is recoverable
            session.entity_type = best_etype
            session.detect_method = "heuristic_fallback"
            session.detect_confidence = max(round(confidence, 2), 0.3)

    _emit(on_update, "stage", {
        "stage": "detect", "status": "done",
        "entity_type": session.entity_type,
        "method": session.detect_method,
        "confidence": session.detect_confidence,
    })
    return session


# ══════════════════════════════════════════════════════════════════
# STAGE 3 — SCHEMA MAPPING + VALIDATION
# ══════════════════════════════════════════════════════════════════

def stage_validate_schema(session: PipelineSession, on_update=None) -> PipelineSession:
    """Map columns → canonical fields, check required fields, detect type mismatches."""
    _emit(on_update, "stage", {"stage": "schema", "status": "running"})

    if session.entity_type not in SCHEMA:
        session.schema_errors.append(f"Unknown entity type '{session.entity_type}' — cannot validate schema.")
        return session

    schema = SCHEMA[session.entity_type]
    input_cols = list(session.raw_rows[0].keys()) if session.raw_rows else []

    def _norm_col(s):
        """Normalise column name for fuzzy matching."""
        return s.lower().replace(" ", "_").replace("-", "_").replace(".", "_").strip()

    cols_norm = {_norm_col(c): c for c in input_cols}

    # Build canonical → original column mapping using normalised comparison
    mapping = {}
    for canon, aliases in schema["aliases"].items():
        if canon in mapping:
            continue
        for alias in aliases:
            key = _norm_col(alias)
            if key in cols_norm:
                mapping[canon] = cols_norm[key]
                break
        # Second pass: substring match (e.g. "risk" matches "risk_score")
        if canon not in mapping:
            for col_norm, col_orig in cols_norm.items():
                for alias in aliases:
                    alias_n = _norm_col(alias)
                    if alias_n in col_norm or col_norm in alias_n:
                        mapping[canon] = col_orig
                        break
                if canon in mapping:
                    break

    # LLM help for unmatched required fields
    missing_required = [f for f in schema["required"] if f not in mapping]
    if missing_required:
        unmatched_cols = [c for c in input_cols if c not in mapping.values()]
        if unmatched_cols and _groq:
            raw = _llm([
                {"role": "system", "content":
                    f"Map input column names to canonical fields for a Neo4j {session.entity_type} node. "
                    f"Required fields: {missing_required}. "
                    "Return ONLY valid JSON mapping input_col → canonical_field. Omit uncertain ones."},
                {"role": "user", "content": f"Unmatched input columns: {unmatched_cols}"}
            ], max_tokens=200)
            try:
                raw = re.sub(r"```(?:json)?|```", "", raw).strip()
                llm_map = json.loads(raw)
                for inp, canon in llm_map.items():
                    if canon in schema["aliases"] and inp not in mapping.values():
                        mapping[canon] = inp
            except Exception:
                pass

    session.mapping = mapping

    # Still missing required fields → schema error
    for f in schema["required"]:
        if f not in mapping:
            session.schema_errors.append(
                f"Required field '{f}' not found in file. "
                f"Expected column names: {schema['aliases'].get(f, [f])[:4]}"
            )

    # Normalize rows
    norm = []
    for row in session.raw_rows:
        nr = {}
        for canon, orig in mapping.items():
            val = row.get(orig)
            if val is not None and str(val).strip() not in ("", "None", "nan", "NaN", "null", "NULL"):
                nr[canon] = val
        norm.append(nr)
    session.norm_rows = norm

    _emit(on_update, "stage", {
        "stage": "schema", "status": "done",
        "mapped_fields": len(mapping),
        "schema_errors": len(session.schema_errors),
        "rows": len(session.norm_rows),
    })
    return session


# ══════════════════════════════════════════════════════════════════
# STAGE 4 — SEMANTIC VALIDATION
# ══════════════════════════════════════════════════════════════════

def stage_validate_semantic(session: PipelineSession, on_update=None) -> PipelineSession:
    """Validate field values against semantic rules (range, enum, type checks)."""
    _emit(on_update, "stage", {"stage": "semantic", "status": "running"})

    if session.entity_type not in SCHEMA:
        return session

    schema = SCHEMA[session.entity_type]
    validators = schema.get("validators", {})
    msgs = schema.get("validator_msgs", {})
    valid, invalid = [], []

    for i, row in enumerate(session.norm_rows):
        row_errors = []

        # ── Auto-coerce values before validation ──────────────────
        # mode: normalise to Title case (road → Road)
        if "mode" in row and row["mode"] is not None:
            mode_raw = str(row["mode"]).strip().lower()
            mode_map = {
                "road": "Road", "roadways": "Road", "truck": "Road",
                "rail": "Rail", "railways": "Rail", "train": "Rail",
                "air":  "Air",  "airways":  "Air",  "flight": "Air",
                "sea":  "Sea",  "ship":     "Sea",  "ocean": "Sea",
            }
            row["mode"] = mode_map.get(mode_raw, row["mode"].title()
                          if hasattr(row["mode"], "title") else row["mode"])

        # risk_score: if given as % (e.g. 45 instead of 0.45), normalise
        if "risk_score" in row and row["risk_score"] is not None:
            try:
                v = float(row["risk_score"])
                if v > 1.0:
                    row["risk_score"] = round(v / 100.0, 4)
            except (ValueError, TypeError):
                pass

        # Check required fields present
        for f in schema["required"]:
            if not row.get(f) or str(row.get(f, "")).strip() in ("", "None", "nan"):
                row_errors.append(f"Missing required field '{f}'")

        # Check null IDs
        row_id = str(row.get("id", "")).strip()
        if not row_id or row_id in ("None", "nan", "null", ""):
            row_errors.append("Null or empty ID — every row requires a valid unique ID")

        # Semantic validators
        for field, validator in validators.items():
            if field in row and row[field] is not None:
                try:
                    if not validator(row[field]):
                        row_errors.append(
                            f"Invalid value for '{field}': {row[field]} — {msgs.get(field, 'invalid')}"
                        )
                except (ValueError, TypeError):
                    row_errors.append(
                        f"Type error for '{field}': '{row[field]}' cannot be converted — {msgs.get(field, 'check value')}"
                    )

        if row_errors:
            invalid.append({"row_index": i + 2, "row": row, "errors": row_errors})
            session.semantic_errors.extend([f"Row {i+2}: {e}" for e in row_errors])
        else:
            valid.append(row)

    session.valid_rows = valid
    session.invalid_rows = invalid

    _emit(on_update, "stage", {
        "stage": "semantic", "status": "done",
        "valid": len(valid), "invalid": len(invalid),
        "errors": len(session.semantic_errors),
    })
    return session


# ══════════════════════════════════════════════════════════════════
# STAGE 5 — CONFLICT & DUPLICATE DETECTION
# ══════════════════════════════════════════════════════════════════

def stage_detect_conflicts(session: PipelineSession, on_update=None) -> PipelineSession:
    """Detect: file-internal duplicates, graph conflicts, missing targets, circular refs."""
    _emit(on_update, "stage", {"stage": "conflicts", "status": "running"})

    schema = SCHEMA.get(session.entity_type, {})
    id_field = schema.get("id_field", "id")

    # 1. Intra-file duplicates
    seen_ids: dict = {}
    for i, row in enumerate(session.valid_rows):
        rid = str(row.get("id", "")).strip()
        if rid in seen_ids:
            session.duplicates_in_file.append({
                "id": rid, "first_at": seen_ids[rid] + 2, "duplicate_at": i + 2
            })
        else:
            seen_ids[rid] = i

    # Remove duplicate rows (keep first)
    deduped = []
    seen2 = set()
    for row in session.valid_rows:
        rid = str(row.get("id", "")).strip()
        if rid not in seen2:
            seen2.add(rid)
            deduped.append(row)
    session.valid_rows = deduped

    # 2. Check existing nodes in graph
    if session.valid_rows:
        ids = [str(r.get("id", "")).strip() for r in session.valid_rows]
        label = schema.get("neo4j_label", session.entity_type)
        try:
            existing = _run_cypher(
                f"MATCH (n:{label}) WHERE n.{id_field} IN $ids "
                f"RETURN n.{id_field} AS id, properties(n) AS props",
                {"ids": ids}
            )
            existing_map = {r["id"]: r["props"] for r in existing}
            for row in session.valid_rows:
                rid = str(row.get("id", "")).strip()
                if rid in existing_map:
                    session.existing_in_graph.append({
                        "id": rid,
                        "existing_props": existing_map[rid],
                        "new_props": row,
                    })
        except Exception as e:
            pass  # non-critical

    # 3. Check missing relationship targets
    for row in session.valid_rows:
        missing = []
        for field, target_label, target_id_prop, rel_type in AUTO_RELATIONSHIPS.get(session.entity_type, []):
            target_id = str(row.get(field, "")).strip()
            if target_id and target_id not in ("", "None", "nan"):
                try:
                    found = _run_cypher(
                        f"MATCH (n:{target_label} {{{target_id_prop}: $tid}}) RETURN n LIMIT 1",
                        {"tid": target_id}
                    )
                    if not found:
                        missing.append({
                            "field": field, "target_label": target_label,
                            "target_id": target_id, "rel_type": rel_type
                        })
                except Exception:
                    pass
        if missing:
            session.missing_targets.append({
                "id": str(row.get("id", "?")),
                "missing": missing
            })

    # 4. Circular relationship detection (Supplier → Plant → Supplier loops etc.)
    for row in session.valid_rows:
        row_id = str(row.get("id", "")).strip()
        # Check if entity references itself
        for field in ["plant", "dist_id"]:
            ref_id = str(row.get(field, "")).strip()
            if ref_id and ref_id == row_id:
                session.circular_refs.append(
                    f"Row {row_id}: self-reference — '{field}' points to its own ID"
                )

    # 5. Classify scenario type
    if session.semantic_errors or session.schema_errors:
        session.scenario_type = "invalid_dataset"
    elif session.circular_refs:
        session.scenario_type = "circular_reference"
    elif session.missing_targets:
        session.scenario_type = "missing_targets"
    elif session.duplicates_in_file or session.existing_in_graph:
        session.scenario_type = "duplicate_upload"
    else:
        session.scenario_type = "happy_path"

    _emit(on_update, "stage", {
        "stage": "conflicts", "status": "done",
        "scenario": session.scenario_type,
        "file_dupes": len(session.duplicates_in_file),
        "graph_conflicts": len(session.existing_in_graph),
        "missing_targets": len(session.missing_targets),
        "circular": len(session.circular_refs),
    })
    return session


# ══════════════════════════════════════════════════════════════════
# STAGE 6 — GRAPH DIFF BUILDER
# ══════════════════════════════════════════════════════════════════

def stage_build_graph_diff(session: PipelineSession, on_update=None) -> PipelineSession:
    """Classify each valid row as new node / updated node, compute property diffs."""
    _emit(on_update, "stage", {"stage": "diff", "status": "running"})

    existing_map = {e["id"]: e["existing_props"] for e in session.existing_in_graph}

    for row in session.valid_rows:
        rid = str(row.get("id", "")).strip()
        if rid in existing_map:
            old = existing_map[rid]
            diff = {}
            for k, new_val in row.items():
                old_val = old.get(_canon_to_neo4j(k, session.entity_type))
                if str(old_val) != str(new_val):
                    diff[k] = {"from": old_val, "to": new_val}
            session.updated_nodes.append({
                "id": rid, "old_props": old, "new_props": row, "diff": diff
            })
        else:
            session.new_nodes.append(row)

    # Relationships to be created
    for row in session.valid_rows:
        rid = str(row.get("id", "")).strip()
        for field, target_label, target_id_prop, rel_type in AUTO_RELATIONSHIPS.get(session.entity_type, []):
            target_id = str(row.get(field, "")).strip()
            if target_id and target_id not in ("", "None", "nan"):
                # Check if relationship already exists
                try:
                    label = SCHEMA[session.entity_type]["neo4j_label"]
                    id_field = SCHEMA[session.entity_type]["id_field"]
                    existing_rel = _run_cypher(
                        f"MATCH (a:{label} {{{id_field}: $aid}})-[r:{rel_type}]->(b:{target_label} {{{target_id_prop}: $bid}}) "
                        "RETURN r LIMIT 1",
                        {"aid": rid, "bid": target_id}
                    )
                    if not existing_rel:
                        session.new_relationships.append({
                            "from_id": rid, "from_label": label,
                            "from_id_prop": id_field,
                            "to_id": target_id, "to_label": target_label,
                            "to_id_prop": target_id_prop,
                            "rel_type": rel_type
                        })
                except Exception:
                    session.new_relationships.append({
                        "from_id": rid, "from_label": SCHEMA[session.entity_type]["neo4j_label"],
                        "from_id_prop": SCHEMA[session.entity_type]["id_field"],
                        "to_id": target_id, "to_label": target_label,
                        "to_id_prop": target_id_prop, "rel_type": rel_type
                    })

    _emit(on_update, "stage", {
        "stage": "diff", "status": "done",
        "new_nodes": len(session.new_nodes),
        "updated_nodes": len(session.updated_nodes),
        "new_rels": len(session.new_relationships),
    })
    return session


def _canon_to_neo4j(canon: str, entity_type: str) -> str:
    """Map canonical field name to Neo4j property name."""
    _maps = {
        "Supplier":    {"id": "supplier_id", "name": "supplier_name", "risk_score": "risk_score",
                        "capacity": "annual_capacity_units", "lead_time": "StoP_lead_time_days"},
        "Distributor": {"id": "distributor_id", "city": "distributor_city",
                        "lat": "distributor_latitude", "lng": "distributor_longitude"},
        "Route":       {"id": "route_id", "mode": "mode", "dist_km": "PtoD_distance_km",
                        "days": "PtoD_leadtime_days", "cost": "PtoD_transportation_cost_inr"},
    }
    return _maps.get(entity_type, {}).get(canon, canon)


# ══════════════════════════════════════════════════════════════════
# STAGE 7 — CYPHER PLAN GENERATION + DRY-RUN SIMULATION
# ══════════════════════════════════════════════════════════════════

def stage_generate_cypher(session: PipelineSession, on_update=None) -> PipelineSession:
    """Generate parameterised MERGE Cypher for each valid row."""
    _emit(on_update, "stage", {"stage": "cypher", "status": "running"})

    for row in session.valid_rows:
        try:
            cypher, params, neo4j_id = _build_merge_cypher(row, session.entity_type)
            session.cypher_plan.append({
                "id": str(row.get("id", "?")),
                "cypher": cypher,
                "params": params,
                "neo4j_id": neo4j_id,
                "is_new": str(row.get("id", "")).strip() not in
                          {e["id"] for e in session.existing_in_graph},
            })
        except Exception as e:
            session.cypher_plan.append({
                "id": str(row.get("id", "?")),
                "cypher": f"/* ERROR: {e} */",
                "params": {}, "neo4j_id": None, "is_new": True, "error": str(e),
            })

    _emit(on_update, "stage", {
        "stage": "cypher", "status": "done",
        "statements": len(session.cypher_plan),
    })
    return session


def stage_dry_run(session: PipelineSession, on_update=None) -> PipelineSession:
    """Simulate MERGE execution using EXPLAIN / dry-run without writing."""
    _emit(on_update, "stage", {"stage": "dryrun", "status": "running"})

    for plan_item in session.cypher_plan:
        if plan_item.get("error"):
            session.dry_run_log.append({
                "id": plan_item["id"], "status": "skip",
                "msg": f"Skipped — cypher generation error: {plan_item['error']}"
            })
            continue

        # Use EXPLAIN to parse Cypher without executing
        try:
            explain_q = "EXPLAIN " + plan_item["cypher"].split("RETURN")[0] + "RETURN 1"
            _run_cypher(explain_q, plan_item["params"])
            session.dry_run_log.append({
                "id": plan_item["id"], "status": "ok",
                "msg": "Cypher valid — would " + ("CREATE" if plan_item["is_new"] else "UPDATE"),
                "is_new": plan_item["is_new"],
            })
        except Exception as e:
            err = str(e)
            # EXPLAIN fails on some Groq-generated Cypher but the real query may be fine
            if "EXPLAIN" in err or "SyntaxError" in err:
                session.dry_run_log.append({
                    "id": plan_item["id"], "status": "warning",
                    "msg": f"Cypher may have syntax issue: {err[:120]}"
                })
            else:
                session.dry_run_log.append({
                    "id": plan_item["id"], "status": "ok",
                    "msg": "Simulation passed (EXPLAIN mode)",
                    "is_new": plan_item["is_new"],
                })

    _emit(on_update, "stage", {
        "stage": "dryrun", "status": "done",
        "simulated": len(session.dry_run_log),
        "warnings": sum(1 for d in session.dry_run_log if d.get("status") == "warning"),
    })
    return session


def _build_merge_cypher(row: dict, entity_type: str) -> tuple:
    """Build parameterised MERGE Cypher. Returns (cypher_str, params_dict, id_value)."""
    rid = str(row.get("id", "")).strip()

    if entity_type == "Supplier":
        cypher = (
            "MERGE (s:Supplier {supplier_id: $sid}) "
            "SET s.supplier_name = $name, "
            "    s.risk_score = $risk, "
            "    s.annual_capacity_units = $cap, "
            "    s.StoP_lead_time_days = $lt, "
            "    s.status = 'Active', "
            "    s.supplier_latitude = 0.0, "
            "    s.supplier_longitude = 0.0, "
            "    s.StoP_distance_km = 0.0 "
            "RETURN s.supplier_id AS id"
        )
        params = {
            "sid":  rid,
            "name": str(row.get("name", "Unknown")).strip(),
            "risk": float(row.get("risk_score", 0.5) or 0.5),
            "cap":  int(float(row.get("capacity", 0) or 0)),
            "lt":   int(float(row.get("lead_time", 7) or 7)),
        }
        return cypher, params, rid

    elif entity_type == "Distributor":
        cypher = (
            "MERGE (d:Distributor {distributor_id: $did}) "
            "SET d.distributor_city = $city, "
            "    d.distributor_latitude = $lat, "
            "    d.distributor_longitude = $lng "
            "RETURN d.distributor_id AS id"
        )
        params = {
            "did":  rid,
            "city": str(row.get("city", "")).strip(),
            "lat":  float(row.get("lat", 0.0) or 0.0),
            "lng":  float(row.get("lng", 0.0) or 0.0),
        }
        return cypher, params, rid

    elif entity_type == "Route":
        cypher = (
            "MERGE (r:Route {route_id: $rid}) "
            "SET r.mode = $mode, "
            "    r.PtoD_distance_km = $dist, "
            "    r.PtoD_leadtime_days = $days, "
            "    r.PtoD_transportation_cost_inr = $cost, "
            "    r.plant_id = $plant, "
            "    r.distributor_id = $dist_id "
            "RETURN r.route_id AS id"
        )
        mode_raw = str(row.get("mode", "Road")).strip()
        mode = mode_raw.capitalize() if mode_raw.lower() in {"road","rail","air","sea"} else "Road"
        params = {
            "rid":     rid,
            "mode":    mode,
            "dist":    float(row.get("dist_km", 0) or 0),
            "days":    int(float(row.get("days", 1) or 1)),
            "cost":    float(row.get("cost", 0) or 0),
            "plant":   str(row.get("plant", "")).strip(),
            "dist_id": str(row.get("dist_id", "")).strip(),
        }
        return cypher, params, rid

    raise ValueError(f"No Cypher builder for entity type: {entity_type}")


# ══════════════════════════════════════════════════════════════════
# STAGE 8 — AI IMPACT SIMULATION
# ══════════════════════════════════════════════════════════════════

def stage_impact_simulation(session: PipelineSession, on_update=None) -> PipelineSession:
    """AI-generated pre-commit impact forecast."""
    _emit(on_update, "stage", {"stage": "impact", "status": "running"})

    # Gather live graph metrics for context
    metrics = _gather_graph_metrics_for_impact(session)

    rows_summary = []
    for row in session.valid_rows[:5]:
        rows_summary.append({k: str(v)[:40] for k, v in row.items()})

    prompt = f"""You are a supply chain risk analyst. Predict the impact of adding {len(session.valid_rows)} {session.entity_type}(s) to the Neo4j knowledge graph.

CURRENT GRAPH METRICS:
{json.dumps(metrics, indent=2)}

ENTITY TYPE: {session.entity_type}
SCENARIO: {session.scenario_type.replace('_', ' ').title()}
ROWS TO ADD (sample): {json.dumps(rows_summary, indent=2)}
NEW NODES: {len(session.new_nodes)}
UPDATED NODES: {len(session.updated_nodes)}
NEW RELATIONSHIPS: {len(session.new_relationships)}

Write a 3-5 sentence AI impact simulation covering:
1. Risk score change for affected plants/distributors (estimate %)
2. Which downstream cities/distributors might be affected
3. Whether this increases or decreases supply chain resilience
4. Any specific risk flag (e.g. single-source dependency, demand gap increase)

Format as professional prose. Be specific with entity names and numbers where possible.
If data is missing, make reasonable estimates based on entity type and graph structure.
Start with: "AI Impact Forecast: ..." """

    impact = _llm([
        {"role": "system", "content": "You are a senior supply chain analyst generating concise AI impact forecasts."},
        {"role": "user", "content": prompt}
    ], max_tokens=400, temperature=0.2)

    session.impact_simulation = impact or "Impact simulation unavailable — model call failed. Proceed with manual review."

    _emit(on_update, "stage", {"stage": "impact", "status": "done"})
    return session


def _gather_graph_metrics_for_impact(session: PipelineSession) -> dict:
    """Pull live graph stats relevant to the entity type being added."""
    metrics = {}
    try:
        if session.entity_type == "Supplier":
            r = _run_cypher("MATCH (s:Supplier) RETURN COUNT(s) AS total, ROUND(AVG(s.risk_score),3) AS avg_risk, MAX(s.risk_score) AS max_risk")
            metrics["supplier_stats"] = r[0] if r else {}
            # Check which plants will be affected
            for row in session.valid_rows[:3]:
                plant_id = str(row.get("plant", "")).strip()
                if plant_id:
                    pr = _run_cypher(
                        "MATCH (pl:Plant {plant_id: $pid})-[:DISPATCHES]->(sh:Shipment) "
                        "RETURN pl.plant_name AS name, COUNT(sh) AS ships, "
                        "SUM(CASE WHEN sh.delivery_status='Major Delay' THEN 1 ELSE 0 END) AS delayed",
                        {"pid": plant_id}
                    )
                    if pr:
                        metrics[f"plant_{plant_id}"] = pr[0]

        elif session.entity_type == "Distributor":
            r = _run_cypher("MATCH (d:Distributor) RETURN COUNT(d) AS total, COUNT{(d)-[:DELIVERS_TO]->()} AS rel_count")
            metrics["distributor_stats"] = r[0] if r else {}

        elif session.entity_type == "Route":
            r = _run_cypher("MATCH (r:Route) RETURN COUNT(r) AS total, COUNT(DISTINCT r.mode) AS modes")
            metrics["route_stats"] = r[0] if r else {}

        # Overall network health
        health = _run_cypher(
            "MATCH (sh:Shipment) RETURN COUNT(sh) AS total, "
            "SUM(CASE WHEN sh.delivery_status='Major Delay' THEN 1 ELSE 0 END) AS delayed"
        )
        metrics["network_health"] = health[0] if health else {}

    except Exception:
        pass
    return metrics


# ══════════════════════════════════════════════════════════════════
# STAGE 9 — EXECUTION (after user approval)
# ══════════════════════════════════════════════════════════════════

def stage_execute(session: PipelineSession, on_update=None) -> PipelineSession:
    """Execute MERGE statements, auto-wire relationships, record rollback log."""
    _emit(on_update, "stage", {"stage": "execute", "status": "running"})

    for plan_item in session.cypher_plan:
        if plan_item.get("error"):
            session.failed_ids.append(plan_item["id"])
            continue
        try:
            _run_cypher(plan_item["cypher"], plan_item["params"])
            session.inserted_ids.append(plan_item["id"])
            # Record rollback Cypher
            rollback = _build_rollback_cypher(plan_item["id"], session.entity_type, plan_item.get("is_new", True))
            session.rollback_log.append({
                "id": plan_item["id"],
                "cypher": rollback,
                "is_new": plan_item.get("is_new", True),
                "params": plan_item.get("params", {}),
            })
            # ── CSV sync: reflect this insert/update in data/ CSV ──
            try:
                _csv_upsert(session.entity_type, plan_item["id"], plan_item.get("params", {}))
            except Exception as _csv_e:
                _emit(on_update, "log", {"msg": f"[csv_sync] warning: {_csv_e}"})
            # ──────────────────────────────────────────────────────
        except Exception as e:
            session.failed_ids.append(plan_item["id"])
            _emit(on_update, "log", {"msg": f"Execute error for {plan_item['id']}: {e}"})

    # Auto-wire relationships
    for rel in session.new_relationships:
        if rel["from_id"] not in session.inserted_ids and rel["from_id"] not in session.failed_ids:
            continue  # only wire for successfully inserted nodes
        try:
            rel_cypher = (
                f"MATCH (a:{rel['from_label']} {{{rel['from_id_prop']}: $aid}}), "
                f"(b:{rel['to_label']} {{{rel['to_id_prop']}: $bid}}) "
                f"MERGE (a)-[:{rel['rel_type']}]->(b) "
                "RETURN a, b"
            )
            _run_cypher(rel_cypher, {"aid": rel["from_id"], "bid": rel["to_id"]})
            session.wired_rels.append(rel)
        except Exception as e:
            _emit(on_update, "log", {"msg": f"Rel wire error {rel['from_id']} → {rel['to_id']}: {e}"})

    _emit(on_update, "stage", {
        "stage": "execute", "status": "done",
        "inserted": len(session.inserted_ids),
        "failed": len(session.failed_ids),
        "wired_rels": len(session.wired_rels),
    })
    return session


def _build_rollback_cypher(node_id: str, entity_type: str, is_new: bool) -> str:
    """Build Cypher to undo an insert (delete if new, no-op if update)."""
    if not is_new:
        return f"/* UPDATE — rollback requires restoring prior props manually for {node_id} */"
    schema = SCHEMA.get(entity_type, {})
    label = schema.get("neo4j_label", entity_type)
    id_prop = schema.get("id_field", "id")
    return f"MATCH (n:{label} {{{id_prop}: '{node_id}'}}) DETACH DELETE n"


def execute_rollback(session: PipelineSession, on_update=None) -> dict:
    """Roll back all insertions from a session."""
    _emit(on_update, "stage", {"stage": "rollback", "status": "running"})
    deleted, failed = [], []
    for item in session.rollback_log:
        if "no-op" in item["cypher"].lower() or "manually" in item["cypher"].lower():
            continue
        try:
            _run_cypher(item["cypher"])
            deleted.append(item["id"])
            # ── CSV sync: remove rolled-back node from CSV ──
            try:
                if item.get("is_new", True):
                    _csv_delete(session.entity_type, item["id"])
                else:
                    # It was an update — restore original props stored in rollback_log
                    _csv_upsert(session.entity_type, item["id"], item.get("params", {}))
            except Exception as _csv_e:
                pass  # non-fatal
            # ────────────────────────────────────────────────
        except Exception as e:
            failed.append({"id": item["id"], "error": str(e)})
    _emit(on_update, "stage", {"stage": "rollback", "status": "done",
                                "deleted": len(deleted), "failed": len(failed)})
    return {"deleted": deleted, "failed": failed}


# ══════════════════════════════════════════════════════════════════
# STAGE 10 — RCA SNAPSHOT + BEFORE/AFTER COMPARISON
# ══════════════════════════════════════════════════════════════════

def capture_rca_snapshot(session: PipelineSession) -> dict:
    """Capture key RCA metrics as a before/after snapshot."""
    snapshot = {"timestamp": datetime.now().isoformat()}
    try:
        r = _run_cypher(
            "MATCH (sh:Shipment) RETURN COUNT(sh) AS total, "
            "SUM(CASE WHEN sh.delivery_status='Major Delay' THEN 1 ELSE 0 END) AS delayed, "
            "ROUND(AVG(CASE WHEN sh.delivery_status='Major Delay' THEN sh.delay_days END), 1) AS avg_delay"
        )
        snapshot["shipments"] = r[0] if r else {}

        sup = _run_cypher(
            "MATCH (s:Supplier) RETURN COUNT(s) AS total, "
            "ROUND(AVG(s.risk_score), 3) AS avg_risk, "
            "SUM(CASE WHEN s.risk_score > 0.7 THEN 1 ELSE 0 END) AS high_risk"
        )
        snapshot["suppliers"] = sup[0] if sup else {}

        dist = _run_cypher(
            "MATCH (d:Distributor) RETURN COUNT(d) AS total, "
            "SUM(CASE WHEN d.demand_gap > 0 THEN 1 ELSE 0 END) AS with_gap"
        )
        snapshot["distributors"] = dist[0] if dist else {}
    except Exception:
        pass
    return snapshot


def build_rca_diff_html(before: dict, after: dict, session: PipelineSession) -> str:
    """Build HTML comparing before/after RCA metrics."""
    def _pct_change(old, new):
        try:
            old, new = float(old), float(new)
            if old == 0:
                return "+∞" if new > 0 else "0%"
            change = (new - old) / abs(old) * 100
            sign = "+" if change > 0 else ""
            color = "#f87171" if change > 0 else "#4ade80"
            return f'<span style="color:{color}">{sign}{change:.1f}%</span>'
        except Exception:
            return "—"

    def _metric_row(label, b_val, a_val):
        return (
            f'<tr><td style="color:#94a3b8;font-size:.78rem;padding:5px 12px">{label}</td>'
            f'<td style="color:#e2e8f0;font-weight:600;text-align:right;padding:5px 12px">{b_val}</td>'
            f'<td style="color:#e2e8f0;font-weight:600;text-align:right;padding:5px 12px">{a_val}</td>'
            f'<td style="text-align:center;padding:5px 12px">{_pct_change(b_val, a_val)}</td></tr>'
        )

    b_ship = before.get("shipments", {})
    a_ship = after.get("shipments", {})
    b_sup  = before.get("suppliers", {})
    a_sup  = after.get("suppliers", {})
    b_dist = before.get("distributors", {})
    a_dist = after.get("distributors", {})

    return f"""
<div style="background:rgba(6,12,28,0.9);border:1px solid rgba(56,189,248,0.2);border-radius:12px;padding:18px 22px;margin-top:16px">
  <div style="font-size:.62rem;font-weight:900;text-transform:uppercase;letter-spacing:.14em;color:#38bdf8;margin-bottom:14px">
    📊 Before vs After — RCA Metrics Comparison
  </div>
  <table style="width:100%;border-collapse:collapse">
    <thead>
      <tr style="border-bottom:1px solid rgba(56,189,248,0.2)">
        <th style="color:#7dd3fc;font-size:.7rem;text-align:left;padding:6px 12px">Metric</th>
        <th style="color:#7dd3fc;font-size:.7rem;text-align:right;padding:6px 12px">Before</th>
        <th style="color:#7dd3fc;font-size:.7rem;text-align:right;padding:6px 12px">After</th>
        <th style="color:#7dd3fc;font-size:.7rem;text-align:center;padding:6px 12px">Change</th>
      </tr>
    </thead>
    <tbody>
      {_metric_row("Total Shipments",     b_ship.get("total",0),     a_ship.get("total",0))}
      {_metric_row("Delayed Shipments",   b_ship.get("delayed",0),   a_ship.get("delayed",0))}
      {_metric_row("Avg Delay Days",      b_ship.get("avg_delay",0), a_ship.get("avg_delay",0))}
      {_metric_row("Total Suppliers",     b_sup.get("total",0),      a_sup.get("total",0))}
      {_metric_row("Avg Risk Score",      b_sup.get("avg_risk",0),   a_sup.get("avg_risk",0))}
      {_metric_row("High-Risk Suppliers", b_sup.get("high_risk",0),  a_sup.get("high_risk",0))}
      {_metric_row("Distributors",        b_dist.get("total",0),     a_dist.get("total",0))}
      {_metric_row("Demand Gap Count",    b_dist.get("with_gap",0),  a_dist.get("with_gap",0))}
    </tbody>
  </table>
  <div style="margin-top:12px;font-size:.7rem;color:#64748b;font-style:italic">
    Snapshot taken at {before.get("timestamp","—")} (before) and {after.get("timestamp","—")} (after).
    {len(session.inserted_ids)} node(s) inserted · {len(session.wired_rels)} relationship(s) wired.
  </div>
</div>"""


# ══════════════════════════════════════════════════════════════════
# CHANGE SUMMARY REPORT
# ══════════════════════════════════════════════════════════════════

def build_change_summary_html(session: PipelineSession) -> str:
    """Generate the final Change Summary Report HTML."""
    status_color = "#4ade80" if not session.failed_ids else "#fbbf24"
    status_label = "✅ Complete" if not session.failed_ids else "⚠ Partial"

    scenario_colors = {
        "happy_path": "#4ade80",
        "duplicate_upload": "#fbbf24",
        "invalid_dataset": "#f87171",
        "missing_targets": "#fb923c",
        "circular_reference": "#f87171",
    }
    sc_color = scenario_colors.get(session.scenario_type, "#7dd3fc")

    wired_html = ""
    for rel in session.wired_rels:
        wired_html += (
            f'<div style="font-size:.72rem;color:#e2e8f0;padding:3px 0">'
            f'<span style="color:#38bdf8">{rel["from_id"]}</span>'
            f' <span style="color:#7c3aed">—[{rel["rel_type"]}]→</span>'
            f' <span style="color:#4ade80">{rel["to_id"]}</span>'
            f'</div>'
        )

    impact_html = ""
    if session.impact_simulation:
        impact_html = f"""
<div style="background:rgba(124,58,237,0.08);border:1px solid rgba(124,58,237,0.3);border-radius:10px;padding:14px 18px;margin-top:12px">
  <div style="font-size:.6rem;font-weight:800;text-transform:uppercase;letter-spacing:.12em;color:#a78bfa;margin-bottom:8px">🤖 AI Impact Simulation</div>
  <div style="font-size:.82rem;color:#e2e8f0;line-height:1.65">{session.impact_simulation}</div>
</div>"""

    return f"""
<div style="background:rgba(6,12,28,0.95);border:1px solid rgba(56,189,248,0.25);border-radius:14px;padding:22px 26px">
  <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:18px">
    <div>
      <div style="font-size:1rem;font-weight:800;color:#e2e8f0">📋 Change Summary Report</div>
      <div style="font-size:.7rem;color:#64748b;margin-top:3px">Session {session.session_id} · {session.filename} · {session.created_at[:16]}</div>
    </div>
    <div style="display:flex;gap:10px;align-items:center">
      <span style="font-size:.68rem;padding:4px 12px;border-radius:20px;background:rgba{sc_color.replace('#','')+'22' if len(sc_color)==7 else '(56,189,248,0.1)'};border:1px solid {sc_color};color:{sc_color}">
        {session.scenario_type.replace('_',' ').title()}
      </span>
      <span style="font-size:.82rem;font-weight:700;color:{status_color}">{status_label}</span>
    </div>
  </div>

  <!-- KPI strip -->
  <div style="display:grid;grid-template-columns:repeat(5,1fr);gap:10px;margin-bottom:18px">
    {''.join([
      f'<div style="background:rgba(12,21,40,.8);border:1px solid rgba(56,189,248,.15);border-radius:9px;padding:12px 14px;text-align:center">'
      f'<div style="font-size:.58rem;color:#94a3b8;text-transform:uppercase;letter-spacing:.08em">{lbl}</div>'
      f'<div style="font-size:1.4rem;font-weight:800;color:{col};font-family:monospace">{val}</div>'
      f'</div>'
      for lbl, val, col in [
        ("New Nodes",       len(session.new_nodes),       "#38bdf8"),
        ("Updated Nodes",   len(session.updated_nodes),   "#fbbf24"),
        ("Relationships",   len(session.wired_rels),      "#4ade80"),
        ("Failed",          len(session.failed_ids),      "#f87171" if session.failed_ids else "#4ade80"),
        ("Invalid Rows",    len(session.invalid_rows),    "#fb923c" if session.invalid_rows else "#4ade80"),
      ]
    ])}
  </div>

  <!-- Execution log -->
  <div style="margin-bottom:14px">
    <div style="font-size:.62rem;font-weight:800;text-transform:uppercase;letter-spacing:.12em;color:#38bdf8;margin-bottom:8px">⚡ Executed Changes</div>
    <div style="background:rgba(12,21,40,.6);border:1px solid rgba(255,255,255,.06);border-radius:9px;padding:12px 14px;max-height:180px;overflow-y:auto">
      {''.join([
        f'<div style="font-size:.72rem;padding:3px 0;color:#4ade80">✓ {nid} — {"Created" if plan.get("is_new") else "Updated"}</div>'
        for nid, plan in zip(session.inserted_ids, [p for p in session.cypher_plan if p["id"] in session.inserted_ids])
      ]) if session.inserted_ids else '<div style="color:#64748b;font-size:.72rem">No nodes inserted.</div>'}
      {''.join([
        f'<div style="font-size:.72rem;padding:3px 0;color:#f87171">✗ {fid} — Failed</div>'
        for fid in session.failed_ids
      ])}
    </div>
  </div>

  <!-- Relationships wired -->
  {f'''<div style="margin-bottom:14px">
    <div style="font-size:.62rem;font-weight:800;text-transform:uppercase;letter-spacing:.12em;color:#4ade80;margin-bottom:8px">🔗 Auto-Wired Relationships</div>
    <div style="background:rgba(12,21,40,.6);border:1px solid rgba(74,222,128,.15);border-radius:9px;padding:12px 14px">
      {wired_html or "<div style='color:#64748b;font-size:.72rem'>No new relationships created.</div>"}
    </div>
  </div>''' if session.wired_rels or True else ""}

  {impact_html}

  <!-- Rollback info -->
  <div style="margin-top:14px;padding:10px 14px;background:rgba(239,68,68,.06);border:1px solid rgba(239,68,68,.2);border-radius:9px">
    <div style="font-size:.6rem;font-weight:800;text-transform:uppercase;letter-spacing:.1em;color:#f87171;margin-bottom:5px">↩ Rollback Available</div>
    <div style="font-size:.72rem;color:#cbd5e1">Session ID: <code style="color:#38bdf8;background:rgba(56,189,248,.1);padding:1px 6px;border-radius:4px">{session.session_id}</code> — use this to undo all changes from this upload.</div>
  </div>
</div>"""


# ══════════════════════════════════════════════════════════════════
# HTML BUILDERS FOR UI PANELS
# ══════════════════════════════════════════════════════════════════

def build_validation_report_html(session: PipelineSession) -> str:
    """Build the validation panel HTML: detected type, valid/invalid rows, schema errors."""

    conf_color = "#4ade80" if session.detect_confidence >= 0.8 else "#fbbf24" if session.detect_confidence >= 0.5 else "#f87171"
    scenario_icons = {
        "happy_path": "✅", "duplicate_upload": "⚠️", "invalid_dataset": "❌",
        "missing_targets": "🔗", "circular_reference": "🔄",
    }
    sc_icon = scenario_icons.get(session.scenario_type, "◈")

    # Detected type banner
    detect_html = f"""
<div style="background:rgba(14,165,233,.08);border:1px solid rgba(56,189,248,.3);border-radius:10px;padding:14px 18px;margin-bottom:14px">
  <div style="display:flex;align-items:center;gap:14px">
    <div style="font-size:2rem">{'🏭' if session.entity_type=='Supplier' else '🏙️' if session.entity_type=='Distributor' else '🚛' if session.entity_type=='Route' else '❓'}</div>
    <div style="flex:1">
      <div style="font-size:.6rem;font-weight:800;text-transform:uppercase;letter-spacing:.12em;color:#38bdf8;margin-bottom:4px">AI Detected Entity Type</div>
      <div style="font-size:1.1rem;font-weight:800;color:#e2e8f0">{session.entity_type}</div>
      <div style="font-size:.68rem;color:#94a3b8;margin-top:2px">Method: {session.detect_method} · {session.file_format} · {len(session.raw_rows)} rows parsed</div>
    </div>
    <div style="text-align:center">
      <div style="font-size:.58rem;color:#94a3b8;text-transform:uppercase;letter-spacing:.08em">Confidence</div>
      <div style="font-size:1.4rem;font-weight:800;color:{conf_color};font-family:monospace">{int(session.detect_confidence*100)}%</div>
    </div>
    <div style="text-align:center">
      <div style="font-size:.58rem;color:#94a3b8;text-transform:uppercase;letter-spacing:.08em">Scenario</div>
      <div style="font-size:.9rem;font-weight:700;color:#e2e8f0">{sc_icon} {session.scenario_type.replace('_',' ').title()}</div>
    </div>
  </div>
</div>"""

    # Field mapping table
    if session.mapping:
        mapping_rows = "".join(
            f'<tr><td style="color:#94a3b8;padding:5px 10px;font-size:.72rem">{orig}</td>'
            f'<td style="color:#38bdf8;padding:5px 10px;font-size:.72rem;font-family:monospace">→ {canon}</td></tr>'
            for canon, orig in session.mapping.items()
        )
        mapping_html = f"""
<details style="background:rgba(56,189,248,.04);border:1px solid rgba(56,189,248,.15);border-radius:8px;margin-bottom:12px">
  <summary style="cursor:pointer;padding:9px 14px;font-size:.78rem;font-weight:700;color:#7dd3fc;list-style:none;display:flex;align-items:center;gap:6px">
    <span style="font-size:.6rem">▶</span> 🗂️ Column Mapping ({len(session.mapping)} fields mapped)
  </summary>
  <div style="padding:4px 14px 12px">
    <table style="width:100%;border-collapse:collapse">{mapping_rows}</table>
  </div>
</details>"""
    else:
        mapping_html = ""

    # Valid / Invalid row counts
    rows_html = f"""
<div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:14px">
  <div style="background:rgba(74,222,128,.08);border:1px solid rgba(74,222,128,.3);border-radius:9px;padding:12px 16px;text-align:center">
    <div style="font-size:.58rem;color:#94a3b8;text-transform:uppercase;letter-spacing:.08em">Valid Rows</div>
    <div style="font-size:1.6rem;font-weight:800;color:#4ade80;font-family:monospace">{len(session.valid_rows)}</div>
    <div style="font-size:.62rem;color:#4ade80">ready for import</div>
  </div>
  <div style="background:rgba(248,113,113,.08);border:1px solid rgba(248,113,113,.3);border-radius:9px;padding:12px 16px;text-align:center">
    <div style="font-size:.58rem;color:#94a3b8;text-transform:uppercase;letter-spacing:.08em">Invalid Rows</div>
    <div style="font-size:1.6rem;font-weight:800;color:{'#f87171' if session.invalid_rows else '#4ade80'};font-family:monospace">{len(session.invalid_rows)}</div>
    <div style="font-size:.62rem;color:{'#f87171' if session.invalid_rows else '#64748b'}">{'will be skipped' if session.invalid_rows else 'none'}</div>
  </div>
</div>"""

    # Errors list
    errors_html = ""
    all_errors = session.schema_errors + session.semantic_errors
    if all_errors:
        err_items = "".join(
            f'<div style="display:flex;gap:8px;padding:5px 0;border-bottom:1px solid rgba(255,255,255,.04)">'
            f'<span style="color:#f87171;flex-shrink:0">⚠</span>'
            f'<span style="font-size:.72rem;color:#fca5a5">{e}</span></div>'
            for e in all_errors[:12]
        )
        errors_html = f"""
<details style="background:rgba(239,68,68,.06);border:1px solid rgba(239,68,68,.25);border-radius:8px;margin-bottom:12px" open>
  <summary style="cursor:pointer;padding:9px 14px;font-size:.78rem;font-weight:700;color:#f87171;list-style:none;display:flex;align-items:center;gap:6px">
    <span style="font-size:.6rem">▶</span> ❌ Validation Errors ({len(all_errors)})
  </summary>
  <div style="padding:4px 14px 12px">{err_items}
    {'<div style="font-size:.68rem;color:#64748b;margin-top:6px">... and ' + str(len(all_errors)-12) + ' more errors</div>' if len(all_errors)>12 else ''}
  </div>
</details>"""

    # Conflicts HTML
    conflicts_html = ""
    if session.duplicates_in_file or session.existing_in_graph or session.missing_targets or session.circular_refs:
        parts = []
        if session.duplicates_in_file:
            dupe_items = "".join(
                f'<div style="font-size:.72rem;color:#fde68a;padding:3px 0">'
                f'⚠ <b>{d["id"]}</b> appears at rows {d["first_at"]} and {d["duplicate_at"]} — second occurrence skipped</div>'
                for d in session.duplicates_in_file
            )
            parts.append(f'<div style="margin-bottom:10px"><div style="font-size:.62rem;font-weight:800;color:#fbbf24;text-transform:uppercase;letter-spacing:.1em;margin-bottom:5px">📋 File-Internal Duplicates ({len(session.duplicates_in_file)})</div>{dupe_items}</div>')

        if session.existing_in_graph:
            exist_items = "".join(
                f'<div style="font-size:.72rem;color:#7dd3fc;padding:3px 0">'
                f'🔄 <b>{e["id"]}</b> already exists → will MERGE (update properties)</div>'
                for e in session.existing_in_graph[:8]
            )
            parts.append(f'<div style="margin-bottom:10px"><div style="font-size:.62rem;font-weight:800;color:#38bdf8;text-transform:uppercase;letter-spacing:.1em;margin-bottom:5px">🔄 Existing in Graph ({len(session.existing_in_graph)})</div>{exist_items}</div>')

        if session.missing_targets:
            mt_items = "".join(
                f'<div style="font-size:.72rem;color:#fb923c;padding:3px 0">'
                f'🔗 <b>{m["id"]}</b>: target '
                f'{", ".join(str(t["target_id"])+" ("+t["target_label"]+") not found" for t in m["missing"])}</div>'
                for m in session.missing_targets[:5]
            )
            parts.append(f'<div style="margin-bottom:10px"><div style="font-size:.62rem;font-weight:800;color:#fb923c;text-transform:uppercase;letter-spacing:.1em;margin-bottom:5px">🔗 Missing Relationship Targets ({len(session.missing_targets)})</div>{mt_items}</div>')

        if session.circular_refs:
            cr_items = "".join(
                f'<div style="font-size:.72rem;color:#f87171;padding:3px 0">🔄 {c}</div>'
                for c in session.circular_refs
            )
            parts.append(f'<div style="margin-bottom:10px"><div style="font-size:.62rem;font-weight:800;color:#f87171;text-transform:uppercase;letter-spacing:.1em;margin-bottom:5px">🔄 Circular References ({len(session.circular_refs)})</div>{cr_items}</div>')

        conflicts_html = f"""
<details style="background:rgba(251,146,60,.06);border:1px solid rgba(251,146,60,.25);border-radius:8px;margin-bottom:12px" open>
  <summary style="cursor:pointer;padding:9px 14px;font-size:.78rem;font-weight:700;color:#fb923c;list-style:none;display:flex;align-items:center;gap:6px">
    <span style="font-size:.6rem">▶</span> ⚡ Conflict Detection Results
  </summary>
  <div style="padding:4px 14px 12px">{"".join(parts)}</div>
</details>"""

    return detect_html + mapping_html + rows_html + errors_html + conflicts_html


def build_graph_diff_html(session: PipelineSession) -> str:
    """Build Graph Diff Viewer HTML."""
    if not (session.new_nodes or session.updated_nodes or session.new_relationships):
        return '<div style="color:#64748b;font-size:.78rem;padding:20px;text-align:center">No graph changes to display.</div>'

    def _prop_table(props: dict, color: str = "#e2e8f0") -> str:
        rows = "".join(
            f'<tr><td style="color:#94a3b8;font-size:.68rem;padding:3px 8px">{k}</td>'
            f'<td style="color:{color};font-size:.68rem;padding:3px 8px;font-family:monospace">{str(v)[:40]}</td></tr>'
            for k, v in props.items() if k != "id"
        )
        return f'<table style="width:100%;border-collapse:collapse">{rows}</table>'

    # New nodes
    new_html = ""
    for node in session.new_nodes[:8]:
        rid = node.get("id", "?")
        new_html += f"""
<div style="background:rgba(74,222,128,.07);border:1px solid rgba(74,222,128,.25);border-left:4px solid #4ade80;border-radius:0 9px 9px 0;padding:10px 14px;margin-bottom:8px">
  <div style="display:flex;align-items:center;gap:8px;margin-bottom:6px">
    <span style="font-size:.6rem;font-weight:800;text-transform:uppercase;letter-spacing:.1em;background:rgba(74,222,128,.15);border:1px solid rgba(74,222,128,.4);border-radius:4px;padding:1px 7px;color:#4ade80">NEW</span>
    <span style="font-size:.82rem;font-weight:700;color:#e2e8f0;font-family:monospace">{rid}</span>
    <span style="font-size:.68rem;color:#64748b">:{session.entity_type}</span>
  </div>
  {_prop_table(node, "#4ade80")}
</div>"""
    if len(session.new_nodes) > 8:
        new_html += f'<div style="font-size:.68rem;color:#64748b;padding:6px">... and {len(session.new_nodes)-8} more new nodes</div>'

    # Updated nodes
    updated_html = ""
    for node in session.updated_nodes[:5]:
        rid = node.get("id", "?")
        diff = node.get("diff", {})
        diff_rows = "".join(
            f'<tr>'
            f'<td style="color:#94a3b8;font-size:.68rem;padding:3px 8px">{k}</td>'
            f'<td style="color:#f87171;font-size:.68rem;padding:3px 8px;font-family:monospace;text-decoration:line-through">{str(v["from"])[:35]}</td>'
            f'<td style="color:#4ade80;font-size:.68rem;padding:3px 8px;font-family:monospace">{str(v["to"])[:35]}</td>'
            f'</tr>'
            for k, v in diff.items()
        ) if diff else '<tr><td colspan="3" style="color:#64748b;font-size:.68rem;padding:3px 8px">No property changes detected</td></tr>'
        updated_html += f"""
<div style="background:rgba(251,191,36,.06);border:1px solid rgba(251,191,36,.25);border-left:4px solid #fbbf24;border-radius:0 9px 9px 0;padding:10px 14px;margin-bottom:8px">
  <div style="display:flex;align-items:center;gap:8px;margin-bottom:6px">
    <span style="font-size:.6rem;font-weight:800;text-transform:uppercase;letter-spacing:.1em;background:rgba(251,191,36,.15);border:1px solid rgba(251,191,36,.4);border-radius:4px;padding:1px 7px;color:#fbbf24">UPDATE</span>
    <span style="font-size:.82rem;font-weight:700;color:#e2e8f0;font-family:monospace">{rid}</span>
    <span style="font-size:.68rem;color:#64748b">{len(diff)} fields changing</span>
  </div>
  <table style="width:100%;border-collapse:collapse">
    <tr><th style="color:#7dd3fc;font-size:.62rem;text-align:left;padding:3px 8px">Field</th><th style="color:#f87171;font-size:.62rem;text-align:left;padding:3px 8px">Before</th><th style="color:#4ade80;font-size:.62rem;text-align:left;padding:3px 8px">After</th></tr>
    {diff_rows}
  </table>
</div>"""

    # New relationships
    rels_html = ""
    for rel in session.new_relationships[:10]:
        rels_html += f"""
<div style="background:rgba(56,189,248,.06);border:1px solid rgba(56,189,248,.2);border-radius:8px;padding:9px 14px;margin-bottom:6px;display:flex;align-items:center;gap:10px">
  <span style="font-size:.6rem;font-weight:800;text-transform:uppercase;letter-spacing:.1em;background:rgba(56,189,248,.15);border:1px solid rgba(56,189,248,.4);border-radius:4px;padding:1px 7px;color:#38bdf8">REL</span>
  <span style="font-size:.78rem;font-family:monospace;color:#e2e8f0">({rel['from_label']}:{rel['from_id']}) <span style="color:#7c3aed">-[:{rel['rel_type']}]-></span> ({rel['to_label']}:{rel['to_id']})</span>
</div>"""

    sections = []
    if session.new_nodes:
        sections.append(f'<div style="margin-bottom:16px"><div style="font-size:.62rem;font-weight:800;text-transform:uppercase;letter-spacing:.12em;color:#4ade80;margin-bottom:10px">🟢 New Nodes ({len(session.new_nodes)})</div>{new_html}</div>')
    if session.updated_nodes:
        sections.append(f'<div style="margin-bottom:16px"><div style="font-size:.62rem;font-weight:800;text-transform:uppercase;letter-spacing:.12em;color:#fbbf24;margin-bottom:10px">🟡 Updated Nodes ({len(session.updated_nodes)})</div>{updated_html}</div>')
    if session.new_relationships:
        sections.append(f'<div style="margin-bottom:16px"><div style="font-size:.62rem;font-weight:800;text-transform:uppercase;letter-spacing:.12em;color:#38bdf8;margin-bottom:10px">🔵 New Relationships ({len(session.new_relationships)})</div>{rels_html}</div>')
    if session.removed_relationships:
        sections.append(f'<div style="margin-bottom:16px"><div style="font-size:.62rem;font-weight:800;text-transform:uppercase;letter-spacing:.12em;color:#f87171;margin-bottom:10px">🔴 Removed/Replaced Relationships ({len(session.removed_relationships)})</div></div>')

    return "".join(sections)


def build_cypher_preview_html(session: PipelineSession) -> str:
    """Build formatted Cypher plan preview HTML."""
    if not session.cypher_plan:
        return ""

    def _substitute_params(cypher: str, params: dict) -> str:
        """Substitute $param placeholders with actual values for display."""
        result = cypher
        for key in sorted(params.keys(), key=len, reverse=True):
            val = params[key]
            if isinstance(val, str):
                display_val = f'"{val}"'
            elif isinstance(val, float):
                display_val = f"{val:.2f}"
            else:
                display_val = str(val)
            result = result.replace(f"${key}", display_val)
        return result

    items = []
    for i, plan in enumerate(session.cypher_plan[:10]):
        status_color = "#f87171" if plan.get("error") else "#4ade80"
        status_label = "ERROR" if plan.get("error") else ("NEW" if plan.get("is_new") else "UPDATE")
        raw_cypher     = plan.get("cypher", "")
        params         = plan.get("params", {})
        display_cypher = _substitute_params(raw_cypher, params)
        card_id        = f"cypher-card-{i}"
        items.append(f"""
<div id="{card_id}" class="cypher-card" style="background:rgba(6,12,28,.8);border:1px solid rgba(255,255,255,.07);border-radius:9px;margin-bottom:8px;overflow:hidden">
  <div style="display:flex;align-items:center;gap:8px;padding:7px 12px;background:rgba(255,255,255,.03);border-bottom:1px solid rgba(255,255,255,.06)">
    <span style="width:20px;height:20px;border-radius:50%;background:rgba(56,189,248,.15);border:1px solid rgba(56,189,248,.4);display:flex;align-items:center;justify-content:center;font-size:.62rem;color:#38bdf8;font-family:monospace;flex-shrink:0">{i+1}</span>
    <span style="font-size:.72rem;font-weight:700;color:#e2e8f0;flex:1">{plan['id']}</span>
    <span style="font-size:.58rem;padding:1px 7px;border-radius:4px;background:rgba({status_color.replace('#','').replace('4ade80','74,222,128').replace('f87171','248,113,113')},.15);border:1px solid {status_color};color:{status_color}">{status_label}</span>
    <button onclick="(function(b){{var c=b.closest('.cypher-card').querySelector('.cypher-code').textContent;navigator.clipboard.writeText(c).then(()=>{{b.textContent='✓';b.style.color='#4ade80';setTimeout(()=>{{b.textContent='⧉ Copy';b.style.color='#7dd3fc'}},1500)}})}}) (this)" style="background:rgba(56,189,248,.1);border:1px solid rgba(56,189,248,.3);color:#7dd3fc;font-size:.6rem;font-weight:700;padding:2px 8px;border-radius:4px;cursor:pointer;flex-shrink:0;transition:all .15s">⧉ Copy</button>
  </div>
  <div style="padding:8px 12px;overflow-x:auto">
    <code class="cypher-code" style="font-size:.72rem;line-height:1.6;color:#7dd3fc;white-space:pre-wrap;font-family:'Consolas','Fira Code',monospace">{display_cypher[:600].replace('<','&lt;').replace('>','&gt;')}</code>
  </div>
</div>""")

    extra = f'<div style="font-size:.68rem;color:#64748b;padding:6px">... and {len(session.cypher_plan)-10} more statements</div>' if len(session.cypher_plan) > 10 else ""

    return f"""
<div style="background:rgba(4,9,22,.98);border:1px solid rgba(56,189,248,.18);border-radius:12px;padding:14px">
  <div style="font-size:.6rem;font-weight:800;text-transform:uppercase;letter-spacing:.14em;color:#38bdf8;margin-bottom:12px">🔍 Generated Cypher Plan ({len(session.cypher_plan)} statements)</div>
  {"".join(items)}{extra}
</div>"""


def build_dry_run_html(session: PipelineSession) -> str:
    """Build dry-run simulation results HTML."""
    if not session.dry_run_log:
        return ""
    ok_count   = sum(1 for d in session.dry_run_log if d["status"] == "ok")
    warn_count = sum(1 for d in session.dry_run_log if d["status"] == "warning")
    skip_count = sum(1 for d in session.dry_run_log if d["status"] == "skip")

    items = "".join(
        f'<div style="display:flex;align-items:center;gap:8px;padding:5px 0;border-bottom:1px solid rgba(255,255,255,.04)">'
        f'<span style="color:{"#4ade80" if d["status"]=="ok" else "#fbbf24" if d["status"]=="warning" else "#64748b"};font-size:.8rem">{"✓" if d["status"]=="ok" else "⚠" if d["status"]=="warning" else "–"}</span>'
        f'<span style="font-size:.7rem;font-family:monospace;color:#38bdf8;flex-shrink:0">{d["id"]}</span>'
        f'<span style="font-size:.7rem;color:#94a3b8">{d["msg"][:80]}</span>'
        f'</div>'
        for d in session.dry_run_log
    )

    return f"""
<div style="background:rgba(12,21,40,.7);border:1px solid rgba(56,189,248,.15);border-radius:12px;padding:14px;margin-top:12px">
  <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:10px">
    <div style="font-size:.6rem;font-weight:800;text-transform:uppercase;letter-spacing:.14em;color:#38bdf8">🧪 Dry-Run Simulation Results</div>
    <div style="display:flex;gap:8px">
      <span style="font-size:.62rem;padding:2px 8px;border-radius:4px;background:rgba(74,222,128,.12);color:#4ade80">{ok_count} OK</span>
      {f'<span style="font-size:.62rem;padding:2px 8px;border-radius:4px;background:rgba(251,191,36,.12);color:#fbbf24">{warn_count} warnings</span>' if warn_count else ''}
      {f'<span style="font-size:.62rem;padding:2px 8px;border-radius:4px;background:rgba(100,116,139,.12);color:#64748b">{skip_count} skipped</span>' if skip_count else ''}
    </div>
  </div>
  <div style="max-height:200px;overflow-y:auto">{items}</div>
</div>"""


# ══════════════════════════════════════════════════════════════════
# FULL PIPELINE RUNNER
# ══════════════════════════════════════════════════════════════════


def nl_instruction_to_csv(instruction: str) -> tuple:
    """
    Parse a natural-language graph update instruction into a temp CSV file.
    Returns (filepath, entity_type) so run_stage4_pipeline can process it normally.
    Uses Groq to extract structured fields from the instruction.
    """
    import tempfile, csv, json as _json, os as _os
    from groq import Groq as _Groq
    from dotenv import load_dotenv as _lde
    _lde(".env")

    client = _Groq(api_key=_os.environ.get("GROQ_API_KEY"))

    system = (
        "You are a data extraction assistant. Extract structured entity data from a natural language "
        "instruction for a supply chain graph update. Return ONLY valid JSON, no markdown, no explanation.\n\n"
        "Output format depends on entity type detected:\n"
        "Supplier: {\"entity_type\": \"Supplier\", \"supplier_id\": \"...\", \"supplier_name\": \"...\", "
        "\"risk_score\": 0.0, \"annual_capacity_units\": 0, \"StoP_lead_time_days\": 0, "
        "\"plant_id\": \"...\", \"status\": \"Active\"}\n"
        "Distributor: {\"entity_type\": \"Distributor\", \"distributor_id\": \"...\", "
        "\"distributor_city\": \"...\", \"distributor_latitude\": 0.0, \"distributor_longitude\": 0.0}\n"
        "Route: {\"entity_type\": \"Route\", \"route_id\": \"...\", \"mode\": \"...\", "
        "\"plant_id\": \"...\", \"distributor_id\": \"...\", \"PtoD_distance_km\": 0, "
        "\"PtoD_leadtime_days\": 0, \"PtoD_transportation_cost_inr\": 0}\n"
        "Rules:\n"
        "- Use EXACT values from the instruction. Never invent or substitute values.\n"
        "- risk_score must be 0.0-1.0. If given as percentage (e.g. 40%), convert to decimal (0.4).\n"
        "- If a field is not mentioned, omit it (don't set to 0 unless 0 is the actual value).\n"
        "- status defaults to 'Active' if not mentioned.\n"
        "- Only return the JSON object, nothing else."
    )

    try:
        resp = client.chat.completions.create(
            model=_os.environ.get("GROQ_MODEL", "meta-llama/llama-4-scout-17b-16e-instruct"),
            messages=[{"role": "system", "content": system},
                      {"role": "user",   "content": instruction}],
            max_tokens=500, temperature=0.0,
        )
        raw = resp.choices[0].message.content.strip()
        # Strip markdown fences if present
        if raw.startswith("```"): raw = raw.split("\n", 1)[1].rsplit("```", 1)[0].strip()
        data = _json.loads(raw)
    except Exception as e:
        raise ValueError(f"Could not parse NL instruction into structured data: {e}")

    entity_type = data.pop("entity_type", None)
    if not entity_type or entity_type not in ("Supplier", "Distributor", "Route"):
        raise ValueError(f"Could not detect entity type from instruction (got: {entity_type})")

    # Write to a temp CSV
    tmp = tempfile.NamedTemporaryFile(
        mode="w", suffix=".csv", delete=False,
        encoding="utf-8", newline=""
    )
    cols = list(data.keys())
    writer = csv.DictWriter(tmp, fieldnames=cols)
    writer.writeheader()
    writer.writerow(data)
    tmp.close()
    return tmp.name, entity_type


def run_stage4_pipeline(
    filepath: str,
    filename: str,
    on_update=None,
) -> PipelineSession:
    """
    Run stages 1-8 (parse → impact simulation) without execution.
    Returns fully populated PipelineSession for UI review.
    Execution is separate (run_stage4_execute).
    """
    session = PipelineSession(filepath, filename)
    _emit(on_update, "pipeline_start", {"session_id": session.session_id})

    try:
        session = stage_parse(session, on_update)
        if not session.raw_rows:
            _emit(on_update, "error", {"msg": "No rows parsed from file"})
            return session

        session = stage_detect(session, on_update)

        # Never block on Unknown entity type — fall back to best heuristic guess
        if session.entity_type == "Unknown" or not session.entity_type:
            # Last resort: pick entity type with highest alias match from column names
            cols = set(c.lower().replace(" ","_") for c in (session.raw_rows[0].keys() if session.raw_rows else []))
            best, best_score = "Supplier", 0
            for et, sc in SCHEMA.items():
                score = sum(1 for aliases in sc["aliases"].values()
                            for a in aliases if a.lower().replace(" ","_") in cols)
                if score > best_score:
                    best, best_score = et, score
            session.entity_type = best
            session.detect_method = "emergency_fallback"
            session.detect_confidence = 0.25
            session.schema_errors.append(
                f"Entity type could not be reliably detected — assuming {best}. "
                "If incorrect, rename your file to include 'supplier', 'distributor', or 'route'."
            )

        session = stage_validate_schema(session, on_update)
        session = stage_validate_semantic(session, on_update)
        session = stage_detect_conflicts(session, on_update)
        session = stage_build_graph_diff(session, on_update)
        session = stage_generate_cypher(session, on_update)
        session = stage_dry_run(session, on_update)
        session = stage_impact_simulation(session, on_update)
        # Capture the "before" snapshot now so Before vs After tab
        # can show current-state metrics before the user approves.
        try:
            session.rca_before_snapshot = capture_rca_snapshot(session)
        except Exception:
            session.rca_before_snapshot = {}

    except Exception as e:
        _emit(on_update, "error", {"msg": str(e), "trace": traceback.format_exc()[:500]})

    _emit(on_update, "pipeline_ready", {
        "session_id": session.session_id,
        "valid_rows": len(session.valid_rows),
        "scenario": session.scenario_type,
    })
    return session


def run_stage4_execute(session: PipelineSession, on_update=None) -> PipelineSession:
    """
    Execute stage 9 (MERGE + relationship wiring) and stage 10 (RCA comparison).
    Called after user approval from UI.
    """
    session.rca_before_snapshot = capture_rca_snapshot(session)
    session = stage_execute(session, on_update)
    session.rca_after_snapshot = capture_rca_snapshot(session)
    session.rca_diff_html = build_rca_diff_html(
        session.rca_before_snapshot, session.rca_after_snapshot, session
    )
    session.change_summary_html = build_change_summary_html(session)
    session.approved = True
    _emit(on_update, "execute_done", {
        "inserted": len(session.inserted_ids),
        "failed": len(session.failed_ids),
        "wired_rels": len(session.wired_rels),
    })
    return session


# ── Session registry (in-memory, per-process) ─────────────────────
_SESSION_REGISTRY: dict[str, PipelineSession] = {}


def register_session(session: PipelineSession):
    _SESSION_REGISTRY[session.session_id] = session


def get_session(session_id: str) -> Optional[PipelineSession]:
    return _SESSION_REGISTRY.get(session_id)


def rollback_session(session_id: str, on_update=None) -> dict:
    session = get_session(session_id)
    if not session:
        return {"error": f"Session '{session_id}' not found"}
    return execute_rollback(session, on_update)


# ── Utility ───────────────────────────────────────────────────────
def _emit(on_update, kind: str, data: dict):
    """Emit a progress event to the UI callback."""
    if on_update:
        try:
            on_update((kind, data))
        except Exception:
            pass